from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "263A5B"
BLUE = "2457D6"
LIGHT_BLUE = "EDF3FF"
AMBER = "FFF4DF"
RED = "FDECEC"
LINE = "D9DEE7"
TEXT = "202939"
MUTED = "667085"
WHITE = "FFFFFF"
THIN = Side(style="thin", color=LINE)


def _safe_sheet_name(value: Any, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", str(value or "明细"))[:31] or "明细"
    name, index = base, 2
    while name in used:
        suffix = f"-{index}"
        name = f"{base[:31-len(suffix)]}{suffix}"
        index += 1
    used.add(name)
    return name


def _title(sheet, title: str, subtitle: str, width: int) -> None:
    end = get_column_letter(max(1, width))
    sheet.merge_cells(f"A1:{end}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=16, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32
    sheet.merge_cells(f"A2:{end}2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(size=10, color=BLUE)
    sheet["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 30
    sheet.sheet_view.showGridLines = False


def _header(sheet, row: int, labels: list[str]) -> None:
    for column, label in enumerate(labels, 1):
        cell = sheet.cell(row, column, label)
        cell.font = Font(size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    sheet.row_dimensions[row].height = 27


def _body(sheet, start: int, end: int, width: int) -> None:
    for row in sheet.iter_rows(min_row=start, max_row=max(start, end), min_col=1, max_col=width):
        for cell in row:
            cell.font = Font(size=10, color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)


def _set_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _schedule_groups(item: dict) -> list[dict]:
    schedules = item.get("schedules") or []
    if schedules and all(isinstance(value, dict) and isinstance(value.get("rows"), list) for value in schedules):
        return [{"name": value.get("name") or f"明细{index}", "rows": value["rows"]} for index, value in enumerate(schedules, 1)]
    return [{"name": "明细", "rows": schedules}] if schedules else []


def _add_cover(book: Workbook, workspace: dict, used: set[str]) -> None:
    sheet = book.active
    sheet.title = _safe_sheet_name("使用说明", used)
    _title(
        sheet, "税务申报工作底稿",
        f"由智能财务工作台根据现有业务、账务和税务档案生成。不是已提交申报表；候选值仍需有权人员复核。账面口径：{workspace.get('accounting_basis') or '未注明'}",
        8,
    )
    labels = ["表单", "表号/版本", "期间", "状态", "传输状态", "Agent判断", "复核角色", "官方依据"]
    _header(sheet, 4, labels)
    for row_index, item in enumerate(workspace.get("returns") or [], 5):
        values = [
            item.get("name"), f"{item.get('form_code')} / {item.get('version')}", item.get("period") or workspace.get("period"),
            item.get("status"), item.get("transport"), item.get("agent_position"), item.get("review_role"), item.get("official_source"),
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row_index, column, value)
    _body(sheet, 5, 4 + len(workspace.get("returns") or []), 8)
    workflow_row = max(12, 6 + len(workspace.get("returns") or []))
    sheet.merge_cells(start_row=workflow_row, start_column=1, end_row=workflow_row, end_column=8)
    sheet.cell(workflow_row, 1, "工作流：" + " → ".join(workspace.get("workflow") or []))
    sheet.cell(workflow_row, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet.cell(workflow_row, 1).font = Font(size=10, color=BLUE, bold=True)
    guard_row = workflow_row + 2
    sheet.merge_cells(start_row=guard_row, start_column=1, end_row=guard_row, end_column=8)
    sheet.cell(guard_row, 1, workspace.get("guardrail") or "工作底稿不等于已提交申报表。")
    sheet.cell(guard_row, 1).fill = PatternFill("solid", fgColor=AMBER)
    sheet.cell(guard_row, 1).alignment = Alignment(wrap_text=True, vertical="center")
    _set_widths(sheet, [30, 26, 15, 14, 32, 48, 22, 52])
    sheet.freeze_panes = "A5"


def _add_return(book: Workbook, workspace: dict, item: dict, used: set[str]) -> None:
    sheet = book.create_sheet(_safe_sheet_name(item.get("form_code") or item.get("name"), used))
    _title(
        sheet, item.get("name") or "申报表候选底稿",
        f"{workspace.get('company_name') or '主体待配置'} | {workspace.get('credit_code') or '税号待补'} | 税款所属期 {workspace.get('period_start') or ''} 至 {workspace.get('period_end') or ''} | {item.get('version') or ''}",
        7,
    )
    labels = ["栏次/字段", "项目", "候选填报值", "数据来源", "状态", "Agent说明", "复核结果"]
    _header(sheet, 4, labels)
    fields = item.get("fields") or []
    for row_index, field in enumerate(fields, 5):
        values = [
            field.get("code"), field.get("name"), field.get("value"), field.get("source"), field.get("status"),
            "候选值不等于已申报值；按阻塞项补证据后复核。", "",
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row_index, column, value)
        sheet.cell(row_index, 3).font = Font(size=10, color=BLUE)
        if isinstance(field.get("value"), (int, float)):
            sheet.cell(row_index, 3).number_format = '#,##0.00;[Red](#,##0.00);-'
    _body(sheet, 5, 4 + len(fields), 7)
    blocker_row = max(7, 6 + len(fields))
    sheet.merge_cells(start_row=blocker_row, start_column=1, end_row=blocker_row, end_column=7)
    blockers = item.get("blockers") or []
    sheet.cell(blocker_row, 1, "当前阻塞：" + ("；".join(map(str, blockers)) if blockers else "无硬阻塞，待复核"))
    sheet.cell(blocker_row, 1).fill = PatternFill("solid", fgColor=AMBER if blockers else LIGHT_BLUE)
    sheet.cell(blocker_row, 1).alignment = Alignment(wrap_text=True)
    check_row = blocker_row + 2
    _header(sheet, check_row, ["勾稽检查", "结果", "说明", "修复位置"])
    checks = item.get("checks") or []
    for row_index, check in enumerate(checks, check_row + 1):
        result = "OK" if check.get("passed") is True else "FAIL" if check.get("passed") is False else "WARN"
        values = [check.get("name"), result, check.get("note"), "数据与配置 / 对应业务台账"]
        for column, value in enumerate(values, 1):
            sheet.cell(row_index, column, value)
        sheet.cell(row_index, 2).fill = PatternFill("solid", fgColor=LIGHT_BLUE if result == "OK" else RED if result == "FAIL" else AMBER)
    _body(sheet, check_row + 1, check_row + len(checks), 4)
    _set_widths(sheet, [21, 34, 20, 40, 14, 46, 22])
    sheet.freeze_panes = "A5"

    for group in _schedule_groups(item):
        rows = group["rows"]
        if not rows:
            continue
        detail = book.create_sheet(_safe_sheet_name(f"{item.get('form_code')}-{group['name']}", used))
        keys = list(dict.fromkeys(key for row in rows for key in row))
        _title(detail, f"{item.get('name')} · {group['name']}", "由现有台账生成的候选明细；资料不足时不得直接提交。", len(keys))
        _header(detail, 4, keys)
        for row_index, row in enumerate(rows, 5):
            for column, key in enumerate(keys, 1):
                detail.cell(row_index, column, row.get(key))
        _body(detail, 5, 4 + len(rows), len(keys))
        _set_widths(detail, [28 if "status" in key.lower() else 20 for key in keys])
        detail.freeze_panes = "A5"


def _add_checks(book: Workbook, workspace: dict, used: set[str]) -> None:
    sheet = book.create_sheet(_safe_sheet_name("Checks", used))
    _title(sheet, "Checks", "FAIL/WARN 表示不能把本底稿当作可直接上传的申报文件。", 7)
    _header(sheet, 4, ["检查", "实际", "期望", "差异/问题", "容差", "状态", "修复位置"])
    summary = workspace.get("summary") or {}
    rows = [
        ["主体税号", workspace.get("credit_code") or "", "非空", "" if workspace.get("credit_code") else "税号未配置", 0, "OK" if workspace.get("credit_code") else "FAIL", "数据与配置"],
        ["申报表数量", summary.get("form_count", 0), 5, (summary.get("form_count", 0) or 0) - 5, 0, "OK" if summary.get("form_count") == 5 else "FAIL", "税务工作台"],
        ["可直接上传表", summary.get("direct_upload_ready", 0), 0, 0, 0, "OK", "取得属地模板并完成适配后才改变"],
        ["待补资料表", summary.get("blocked", 0), 0, summary.get("blocked", 0), 0, "OK" if not summary.get("blocked") else "WARN", "各表阻塞事项"],
        ["MODEL STATUS", summary.get("blocked", 0), 0, summary.get("blocked", 0), 0, "PASS" if not summary.get("blocked") else "REVIEW", "先处理FAIL/WARN并由有权人复核"],
    ]
    for row_index, values in enumerate(rows, 5):
        for column, value in enumerate(values, 1):
            sheet.cell(row_index, column, value)
        status = values[5]
        sheet.cell(row_index, 6).fill = PatternFill("solid", fgColor=LIGHT_BLUE if status in {"OK", "PASS"} else RED if status == "FAIL" else AMBER)
    _body(sheet, 5, 9, 7)
    _set_widths(sheet, [27, 21, 17, 28, 12, 14, 42])
    sheet.freeze_panes = "A5"


def build_tax_workbook(workspace: dict, output_path: str | Path, verify_dir: str | Path | None = None) -> Path:
    """Build an auditable tax workpaper using only the public ``openpyxl`` dependency."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    used: set[str] = set()
    _add_cover(book, workspace, used)
    for item in workspace.get("returns") or []:
        _add_return(book, workspace, item, used)
    _add_checks(book, workspace, used)
    book.save(output_path)
    # Re-open once so a truncated/corrupt artifact never reaches the download endpoint.
    check = load_workbook(output_path, read_only=True, data_only=False)
    check.close()
    if verify_dir:
        verify = Path(verify_dir)
        verify.mkdir(parents=True, exist_ok=True)
        (verify / "verification.txt").write_text(
            f"sheets={len(book.sheetnames)}\nforms={len(workspace.get('returns') or [])}\ndirect_upload_ready={(workspace.get('summary') or {}).get('direct_upload_ready', 0)}\n",
            encoding="utf-8",
        )
    return output_path
