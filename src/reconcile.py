from __future__ import annotations

import hashlib
import json
import math
import re
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


SUMMARY_SHEET_HINTS = (
    "对外账单",
    "商店金流账单",
    "三方金流账单",
    "汇总",
    "苹果透视",
)


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    return str(value).replace("\n", "").replace(" ", "").strip()


def _number(value) -> float | None:
    if isinstance(value, bool) or value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("%", "").strip())
    except (TypeError, ValueError):
        return None


def _period(value) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    text = str(value or "").strip()
    match = re.search(r"(20\d{2})[-/.年]?(1[0-2]|0?[1-9])(?!\d)", text)
    return f"{match.group(1)}-{int(match.group(2)):02d}" if match else ""


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", value.lower())


FIELD_ALIASES = {
    "period": ("月份", "账期月份", "結算周期", "结算周期", "年月"),
    "window": ("结算时间", "結算時間"),
    "game": ("游戏名称", "遊戲名稱", "游戏", "项目名称"),
    "platform": ("平台",),
    "channel": ("渠道", "通路别", "通路別"),
    "original_currency": ("原币种", "原幣種", "buyercurrency"),
    "currency": ("结算币种", "結算幣種", "merchantcurrency", "收款币种"),
    "gross_original": ("渠道含税流水（原币种）", "渠道含税流水(原币种)", "amount(buyercurrency)"),
    "gross": ("总流水", "渠道含税流水（结算币种）", "渠道含税流水(结算币种)", "求和项:结算币分成基数流水-税前"),
    "non_share": ("不分成券金额",),
    "refund": ("退款流水", "退款+坏账+测试流水（结算币种）", "退款+坏账+测试流水(结算币种)"),
    "share_base": ("分成基数", "渠道分成基数流水（结算币种）", "渠道分成基数流水(结算币种)", "求和项:结算币分成基数流水-税后"),
    "channel_net": ("渠道后实收（结算币种）", "渠道后实收(结算币种)"),
    "mix": ("travellet渠道占比", "渠道占比"),
    "share_rate": ("分成比例",),
    "fx_rate": ("汇率",),
    "settlement": ("结算金额", "甲方结算金额（结算币种）", "甲方结算金额(结算币种)", "求和项:结算币收入"),
    "withholding_tax": ("预提所得税（结算币种）", "预提所得税(结算币种)", "预扣税"),
    "net": ("甲方实收金额（结算币种）", "甲方实收金额(结算币种)", "我方收入（收款币种）"),
    "transaction_type": ("transactiontype", "salesorreturn"),
    "country": ("buyercountry", "countryofsale", "用户的国家/地区"),
}

FIELD_LABELS = {
    "period": "账期",
    "game": "游戏",
    "platform": "平台",
    "channel": "渠道",
    "original_currency": "原币种",
    "currency": "结算币种",
    "gross_original": "原币流水",
    "gross": "平台/渠道流水",
    "non_share": "不分成金额",
    "refund": "退款/坏账/测试",
    "share_base": "分成基数",
    "channel_net": "渠道后净额",
    "mix": "渠道占比",
    "share_rate": "分成比例",
    "fx_rate": "汇率",
    "settlement": "结算金额",
    "withholding_tax": "代扣税",
    "net": "应收/预计实收",
    "transaction_type": "交易类型",
    "country": "国家/地区",
}

CORE_MAPPING_FIELDS = (
    "period", "game", "platform", "channel", "currency", "gross", "refund",
    "share_base", "channel_net", "share_rate", "settlement", "withholding_tax", "net",
)


def _field_for(header: str) -> str | None:
    clean = _slug(header)
    if not clean:
        return None
    candidates: list[tuple[int, str]] = []
    for field_name, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            alias_clean = _slug(alias)
            generic = alias_clean in {"平台", "渠道", "月份", "游戏", "汇率", "结算币种", "原币种"}
            if (generic and clean == alias_clean) or (not generic and alias_clean in clean):
                candidates.append((len(alias_clean), field_name))
    return max(candidates, default=(0, None))[1]


@dataclass
class SettlementRecord:
    id: str
    source_file: str
    source_sheet: str
    scope: str
    period: str
    game: str
    platform: str
    channel: str
    original_currency: str
    currency: str
    gross_original: float | None
    gross: float | None
    non_share: float
    refunds: float
    taxes: float
    channel_cost: float
    share_base: float | None
    mix: float | None
    share_rate: float | None
    fx_rate: float | None
    settlement_amount: float | None
    withholding_tax: float
    net_receivable: float | None
    country: str = ""
    status: str = "待确认"
    confidence: float = 0.5
    anomalies: list[str] = field(default_factory=list)
    evidence: dict = field(default_factory=dict)


def _infer_context(filename: str, sheet_name: str) -> dict[str, str]:
    combined = f"{filename} {sheet_name}"
    platform = ""
    channel = ""
    if re.search(r"apple|苹果|app store|ios", combined, re.I):
        platform, channel = "iOS", "App Store"
    elif re.search(r"google|谷歌", combined, re.I):
        platform, channel = "Android", "Google Play"
    scope = "国内" if re.search(r"国服|国内|中国区|人民币", combined) else "海外"
    game = ""
    currency = "CNY" if re.search(r"结算币种[:：]?人民币|国服", combined) else ""
    return {"platform": platform, "channel": channel, "scope": scope, "game": game, "currency": currency}


def _header_candidate(row: tuple) -> tuple[int, dict[int, str]]:
    mapping: dict[int, str] = {}
    for index, value in enumerate(row):
        field_name = _field_for(_text(value))
        if field_name and field_name not in mapping.values():
            mapping[index] = field_name
    return len(mapping), mapping


def _header_fingerprint(headers: Iterable[Any]) -> str:
    normalized = "|".join(_slug(_text(value)) for value in headers)
    return hashlib.sha1(normalized.encode("utf-8")).hexdigest()[:16]


def discover_workbook(path: str | Path) -> dict:
    """Return only workbook structure and small previews; never returns full business data."""
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    best: tuple[int, int, int] | None = None
    for sheet_index, sheet in enumerate(workbook.worksheets):
        candidates = []
        preview_rows = []
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 80, 80), values_only=True), 1
        ):
            values = [_text(value)[:80] for value in row[:40]]
            while values and not values[-1]:
                values.pop()
            if row_number <= 12:
                preview_rows.append({"row": row_number, "values": values})
            score, index_mapping = _header_candidate(row)
            non_empty = sum(bool(_text(value)) for value in row)
            if score >= 2 and non_empty >= 3:
                headers = [_text(value)[:120] for value in row[:60]]
                while headers and not headers[-1]:
                    headers.pop()
                field_mapping = {field_name: index for index, field_name in index_mapping.items()}
                candidate = {
                    "row": row_number,
                    "score": score,
                    "headers": headers,
                    "mapping": field_mapping,
                    "fingerprint": _header_fingerprint(headers),
                }
                candidates.append(candidate)
                mapped_fields = set(field_mapping)
                ranking = score
                ranking += 3 if any(hint.lower() in sheet.title.lower() for hint in SUMMARY_SHEET_HINTS) else 0
                ranking += 4 if "settlement" in mapped_fields else 0
                ranking += 2 if "gross" in mapped_fields else 0
                ranking += 1 if "share_base" in mapped_fields else 0
                ranking += 1 if row_number <= 10 else 0
                ranking -= 4 if "country" in mapped_fields else 0
                ranking -= 2 if "transaction_type" in mapped_fields else 0
                if best is None or ranking > best[0]:
                    best = (ranking, sheet_index, len(candidates) - 1)
        sheets.append({
            "name": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "candidates": candidates[:8],
            "preview": preview_rows,
        })
    workbook.close()
    selected = None
    if best:
        _, sheet_index, candidate_index = best
        candidate = sheets[sheet_index]["candidates"][candidate_index]
        selected = {
            "sheet": sheets[sheet_index]["name"],
            "header_row": candidate["row"],
            "mapping": candidate["mapping"],
            "fingerprint": candidate["fingerprint"],
        }
    return {
        "file": path.name,
        "sheets": sheets,
        "selected": selected,
        "fields": [{"key": key, "label": FIELD_LABELS[key]} for key in CORE_MAPPING_FIELDS],
    }


def _record_from_row(
    path: Path,
    sheet_name: str,
    row_number: int,
    row: tuple,
    mapping: dict[int, str],
    context: dict[str, str],
    previous_period: str,
    defaults: dict[str, Any] | None = None,
    formula_mode: str = "declared",
) -> SettlementRecord | None:
    defaults = defaults or {}
    raw = {field_name: row[index] if index < len(row) else None for index, field_name in mapping.items()}
    period = _period(raw.get("period")) or previous_period or _period(defaults.get("period"))
    label = _text(raw.get("period"))
    if label and not _period(raw.get("period")):
        return None
    if not period or label in {"总计", "美元结算合计", "预付款"}:
        return None

    gross = _number(raw.get("gross"))
    settlement = _number(raw.get("settlement"))
    share_base = _number(raw.get("share_base"))
    net = _number(raw.get("net"))
    if gross is None and settlement is None and share_base is None and net is None:
        return None

    transaction_type = _text(raw.get("transaction_type")).lower()
    refunds = _number(raw.get("refund")) or 0.0
    if transaction_type in {"r", "charge refund"}:
        refunds = abs(gross or settlement or 0.0)

    non_share = abs(_number(raw.get("non_share")) or 0.0)
    gross = abs(gross) if gross is not None and transaction_type in {"r", "charge refund"} else gross
    channel_net = _number(raw.get("channel_net"))
    channel_cost = 0.0
    if share_base is not None and channel_net is not None:
        channel_cost = max(0.0, abs(share_base) - abs(channel_net))

    share_rate = _number(raw.get("share_rate"))
    if share_rate is None:
        share_rate = _number(defaults.get("share_rate"))
    if share_rate is not None and 1 < share_rate <= 100:
        share_rate /= 100
    if settlement is None and share_rate is not None:
        if formula_mode == "share_base_x_rate" and share_base is not None:
            settlement = share_base * share_rate
        elif formula_mode == "channel_net_x_rate" and channel_net is not None:
            settlement = channel_net * share_rate

    withholding = _number(raw.get("withholding_tax"))
    if withholding is None:
        withholding = _number(defaults.get("withholding_tax")) or 0.0
    if withholding < 0:
        withholding = abs(withholding)
    if net is None and settlement is not None:
        net = settlement - withholding

    game = _text(raw.get("game")) or _text(defaults.get("game")) or context["game"]
    currency = _text(raw.get("currency")) or _text(defaults.get("currency")) or context.get("currency", "")
    original_currency = _text(raw.get("original_currency")) or currency
    channel = _text(raw.get("channel")) or _text(defaults.get("channel")) or context["channel"]
    platform = _text(raw.get("platform")) or _text(defaults.get("platform")) or context["platform"]
    row_key = f"{path.name}|{sheet_name}|{row_number}|{period}|{game}|{channel}"
    record = SettlementRecord(
        id=hashlib.sha1(row_key.encode("utf-8")).hexdigest()[:12],
        source_file=path.name,
        source_sheet=sheet_name,
        scope=_text(defaults.get("scope")) or context["scope"],
        period=period,
        game=game or "待识别游戏",
        platform=platform or "待识别平台",
        channel=channel or "待识别渠道",
        original_currency=original_currency,
        currency=currency,
        gross_original=_number(raw.get("gross_original")),
        gross=gross,
        non_share=non_share,
        refunds=abs(refunds),
        taxes=0.0,
        channel_cost=channel_cost,
        share_base=share_base,
        mix=_number(raw.get("mix")),
        share_rate=share_rate,
        fx_rate=_number(raw.get("fx_rate")),
        settlement_amount=settlement,
        withholding_tax=withholding,
        net_receivable=net,
        country=_text(raw.get("country")),
        evidence={
            "row": row_number,
            "mapped_fields": sorted(raw),
            "channel_net": channel_net,
            "formula_mode": formula_mode,
        },
    )
    _evaluate(record)
    return record


def parse_workbook_configured(path: str | Path, config: dict) -> list[SettlementRecord]:
    path = Path(path)
    sheet_name = _text(config.get("sheet"))
    header_row = int(config.get("header_row") or 1)
    configured_mapping = config.get("mapping") or {}
    mapping = {
        int(column_index): field_name
        for field_name, column_index in configured_mapping.items()
        if field_name in FIELD_ALIASES and str(column_index).strip() not in {"", "None", "null"}
    }
    if not sheet_name or not mapping:
        raise ValueError("导入配置缺少工作表或字段映射")
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"找不到工作表：{sheet_name}")
    sheet = workbook[sheet_name]
    context = _infer_context(path.name, sheet_name)
    defaults = config.get("defaults") or {}
    formula_mode = config.get("formula_mode") or "declared"
    records = []
    previous_period = _period(defaults.get("period"))
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, max_row=min(sheet.max_row or 5000, 5000), values_only=True),
        header_row + 1,
    ):
        next_header_score, _ = _header_candidate(row)
        if next_header_score >= 4:
            break
        period_column = next((index for index, field_name in mapping.items() if field_name == "period"), None)
        if period_column is not None and period_column < len(row):
            row_period = _period(row[period_column])
            if row_period:
                previous_period = row_period
        record = _record_from_row(
            path, sheet_name, row_number, row, mapping, context, previous_period, defaults, formula_mode
        )
        if record:
            records.append(record)
    workbook.close()
    summary_rows = [record for record in records if not record.country]
    detailed_rows = [record for record in records if record.country]
    return summary_rows if summary_rows and detailed_rows else records


def _evaluate(record: SettlementRecord) -> None:
    anomalies = []
    if not record.currency:
        anomalies.append("缺少结算币种")
    if record.share_rate is not None and not 0 <= record.share_rate <= 1:
        anomalies.append("分成比例超出0–100%")
    if record.gross is not None and record.refunds > abs(record.gross):
        anomalies.append("退款/坏账超过流水")
    if record.share_base is not None and record.share_rate is not None and record.settlement_amount is not None:
        base = record.evidence.get("channel_net")
        if base is None:
            base = record.share_base
        expected = base * record.share_rate
        delta = record.settlement_amount - expected
        tolerance = max(1.0, abs(record.settlement_amount) * 0.001)
        if abs(delta) > tolerance:
            anomalies.append(f"结算金额与分成公式差异 {delta:,.2f}")
        record.evidence["formula_expected"] = round(expected, 4)
        record.evidence["formula_delta"] = round(delta, 4)
    if record.net_receivable is not None and record.settlement_amount is not None:
        expected_net = record.settlement_amount - record.withholding_tax
        if abs(record.net_receivable - expected_net) > max(1.0, abs(expected_net) * 0.001):
            anomalies.append("实收金额与代扣税不勾稽")
    record.anomalies = anomalies
    completeness = sum(
        bool(value)
        for value in (record.period, record.game, record.channel, record.currency)
    ) / 4
    record.confidence = round(max(0.35, min(0.99, completeness - len(anomalies) * 0.08 + 0.15)), 2)
    record.status = "异常" if anomalies else ("已核对" if record.confidence >= 0.85 else "待确认")


def parse_workbook(path: str | Path) -> list[SettlementRecord]:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[SettlementRecord] = []
    candidate_sheets = [
        sheet for sheet in workbook.worksheets
        if any(hint.lower() in sheet.title.lower() for hint in SUMMARY_SHEET_HINTS)
    ]
    external_sheets = [
        sheet for sheet in candidate_sheets
        if any(hint in sheet.title for hint in ("对外账单", "商店金流账单", "三方金流账单"))
    ]
    if external_sheets:
        candidate_sheets = external_sheets
    for sheet in candidate_sheets:
        context = _infer_context(path.name, sheet.title)
        for preview_row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 6, 6), values_only=True):
            preview_text = " ".join(_text(item) for item in preview_row if item is not None)
            if "人民币" in preview_text:
                context["currency"] = "CNY"
            elif "美元" in preview_text and not context.get("currency"):
                context["currency"] = "USD"
        rows = sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 500, 500), values_only=True)
        header_mapping: dict[int, str] = {}
        previous_period = ""
        for row_number, row in enumerate(rows, 1):
            score, candidate = _header_candidate(row)
            if score >= 4 and ("period" in candidate.values() or "game" in candidate.values()):
                header_mapping = candidate
                continue
            if not header_mapping:
                continue
            row_period = _period(row[next((i for i, f in header_mapping.items() if f == "period"), 0)])
            if row_period:
                previous_period = row_period
            record = _record_from_row(path, sheet.title, row_number, row, header_mapping, context, previous_period)
            if record:
                records.append(record)
    workbook.close()
    if records:
        summary_rows = [record for record in records if not record.country]
        detailed_rows = [record for record in records if record.country]
        if summary_rows and detailed_rows:
            records = summary_rows
    return records


def parse_files(paths: Iterable[str | Path]) -> list[SettlementRecord]:
    records: list[SettlementRecord] = []
    for path in paths:
        records.extend(parse_workbook(path))
    return records


def dashboard_payload(records: Iterable[SettlementRecord]) -> dict:
    record_list = list(records)
    currencies: dict[str, dict[str, float | int]] = defaultdict(
        lambda: {"gross": 0.0, "settlement": 0.0, "net": 0.0, "refunds": 0.0, "count": 0}
    )
    periods: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: {"settlement": 0.0, "net": 0.0})
    for record in record_list:
        currency = record.currency or "未知"
        bucket = currencies[currency]
        bucket["gross"] += record.gross or 0.0
        bucket["settlement"] += record.settlement_amount or 0.0
        bucket["net"] += record.net_receivable or 0.0
        bucket["refunds"] += record.refunds
        bucket["count"] += 1
        periods[(record.period, currency)]["settlement"] += record.settlement_amount or 0.0
        periods[(record.period, currency)]["net"] += record.net_receivable or 0.0
    return {
        "records": [asdict(record) for record in record_list],
        "summary": {
            "record_count": len(record_list),
            "file_count": len({record.source_file for record in record_list}),
            "exception_count": sum(record.status == "异常" for record in record_list),
            "pending_count": sum(record.status == "待确认" for record in record_list),
            "currencies": dict(sorted(currencies.items())),
            "periods": [
                {"period": period, "currency": currency, **values}
                for (period, currency), values in sorted(periods.items())
            ],
        },
    }


def save_payload(records: Iterable[SettlementRecord], output: str | Path) -> None:
    Path(output).write_text(json.dumps(dashboard_payload(records), ensure_ascii=False, indent=2), encoding="utf-8")
