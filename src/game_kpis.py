from __future__ import annotations

import hashlib
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ALIASES = {
    "period": ("月份", "期间", "账期", "month", "period"),
    "project_code": ("游戏项目编码", "游戏/项目编码", "项目编码", "project code"),
    "channel": ("渠道", "渠道名称", "channel"),
    "region": ("区域", "国家地区", "国家/地区", "region"),
    "dau": ("dau", "日活"),
    "mau": ("mau", "月活"),
    "new_users": ("新增用户", "新增", "new users"),
    "payers": ("付费用户", "付费人数", "payers"),
    "installs": ("安装数", "激活数", "installs"),
    "gross_bookings": ("流水", "游戏流水", "gross bookings", "gross revenue"),
    "marketing_spend": ("投放金额", "买量金额", "营销费用", "marketing spend"),
    "retention_d1": ("次留", "1日留存", "d1留存", "d1 retention"),
    "retention_d7": ("7日留存", "d7留存", "d7 retention"),
    "retention_d30": ("30日留存", "d30留存", "d30 retention"),
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    return str(value).strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", _text(value).lower())


def _number(value: Any) -> float | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return None if math.isnan(value) or math.isinf(value) else float(value)
    match = re.search(r"-?[\d,.]+", _text(value))
    return float(match.group().replace(",", "")) if match else None


def _period(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    match = re.search(r"(20\d{2})\D?(1[0-2]|0?[1-9])", _text(value))
    return f"{match.group(1)}-{int(match.group(2)):02d}" if match else ""


def _field(value: Any) -> str | None:
    clean = _slug(value)
    matches = []
    for key, aliases in ALIASES.items():
        for alias in aliases:
            normalized = _slug(alias)
            if normalized and (clean == normalized or normalized in clean):
                matches.append((len(normalized), key))
    return max(matches, default=(0, None))[1]


def parse_kpi_workbook(path: str | Path) -> list[dict]:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    output = []
    for sheet in workbook.worksheets:
        if not any(token in _slug(sheet.title) for token in ("经营kpi", "游戏kpi", "经营指标")):
            continue
        mapping = {}
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            candidate = {}
            for index, value in enumerate(row):
                key = _field(value)
                if key and key not in candidate.values():
                    candidate[index] = key
            if len(candidate) >= 4 and "period" in candidate.values() and "project_code" in candidate.values():
                mapping = candidate
                continue
            if not mapping:
                continue
            raw = {key: row[index] if index < len(row) else None for index, key in mapping.items()}
            period, project = _period(raw.get("period")), _text(raw.get("project_code"))
            if not period and not project:
                continue
            if _slug(project).startswith("示例"):
                continue
            anomalies = []
            if not period:
                anomalies.append("缺少有效月份")
            if not project:
                anomalies.append("缺少游戏/项目编码")
            values = {key: _number(raw.get(key)) for key in ALIASES if key not in {"period", "project_code", "channel", "region"}}
            for retention in ("retention_d1", "retention_d7", "retention_d30"):
                value = values.get(retention)
                if value is not None and 1 < value <= 100:
                    values[retention] = value / 100
                if values.get(retention) is not None and not 0 <= values[retention] <= 1:
                    anomalies.append(f"{retention} 应在0到100%之间")
            key = f"{period}|{project}|{_text(raw.get('channel'))}|{_text(raw.get('region'))}"
            output.append({
                "id": hashlib.sha1(key.encode("utf-8")).hexdigest()[:12],
                "period": period, "project_code": project,
                "channel": _text(raw.get("channel")) or "全部渠道",
                "region": _text(raw.get("region")) or "全部区域",
                **values,
                "status": "异常" if anomalies else "可用", "anomalies": anomalies,
                "source_file": path.name, "source_sheet": sheet.title, "source_row": row_number,
            })
    workbook.close()
    return output


def enrich_kpis(rows: list[dict]) -> list[dict]:
    output = []
    for source in rows:
        row = dict(source)
        gross = row.get("gross_bookings")
        dau, mau, payers = row.get("dau"), row.get("mau"), row.get("payers")
        spend, installs = row.get("marketing_spend"), row.get("installs")
        row["arpu"] = round(gross / mau, 2) if gross is not None and mau else None
        row["arppu"] = round(gross / payers, 2) if gross is not None and payers else None
        row["payer_rate"] = round(payers / mau, 4) if payers is not None and mau else None
        row["cpi"] = round(spend / installs, 2) if spend is not None and installs else None
        row["gross_roas"] = round(gross / spend, 4) if gross is not None and spend else None
        row["dau_mau"] = round(dau / mau, 4) if dau is not None and mau else None
        output.append(row)
    return output


def kpi_quality(rows: list[dict], game_codes: set[str] | None = None) -> dict:
    game_codes = game_codes or set()
    issues = [f"{row.get('period')} {row.get('project_code')}：{'、'.join(row.get('anomalies') or [])}" for row in rows if row.get("anomalies")]
    orphaned = sorted({row.get("project_code") for row in rows if game_codes and row.get("project_code") not in game_codes})
    if orphaned:
        issues.append(f"经营KPI项目编码未映射：{'、'.join(orphaned[:8])}")
    return {"record_count": len(rows), "usable_count": sum(row.get("status") == "可用" for row in rows), "issue_count": len(issues), "issues": issues}
