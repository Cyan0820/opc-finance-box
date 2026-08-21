from __future__ import annotations

import hashlib
import math
import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).replace("\n", " ").strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", _text(value).lower())


def _number(value: Any) -> float | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return None if math.isnan(value) or math.isinf(value) else float(value)
    match = re.search(r"-?[\d,.]+", _text(value))
    return float(match.group().replace(",", "")) if match else None


def _period(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    text = _text(value)
    patterns = (
        r"(20\d{2})[-/.年](1[0-2]|0?[1-9])(?:\D|$)",
        r"(20\d{2})(0[1-9]|1[0-2])$",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return f"{match.group(1)}-{int(match.group(2)):02d}"
    return ""


ALIASES = {
    "period": ("月份", "年月", "期间", "账期", "预算月", "month", "period"),
    "scenario": ("情景", "版本", "场景", "scenario", "version"),
    "project": ("游戏", "项目", "产品", "业务线", "project", "game"),
    "category": ("类别", "费用类型", "预算科目", "成本类别", "category"),
    "direction": ("收支方向", "流入流出", "收入支出", "direction", "cash flow"),
    "amount": ("金额", "预算金额", "预测金额", "amount", "budget", "forecast"),
    "currency": ("币种", "currency"),
    "probability": ("概率", "实现概率", "probability"),
    "committed": ("已承诺", "是否承诺", "合同已签", "committed"),
    "note": ("备注", "假设", "说明", "note", "assumption"),
}


def _field(value: Any) -> str | None:
    clean = _slug(value)
    candidates = []
    for field_name, aliases in ALIASES.items():
        for alias in aliases:
            alias_clean = _slug(alias)
            if alias_clean and (clean == alias_clean or alias_clean in clean):
                candidates.append((len(alias_clean), field_name))
    return max(candidates, default=(0, None))[1]


def _category(value: Any) -> str:
    clean = _slug(value)
    rules = (
        ("收入", ("收入", "流水", "营收", "回款", "revenue", "income")),
        ("人力", ("工资", "奖金", "社保", "公积金", "人力", "薪酬", "payroll", "salary")),
        ("外包", ("外包", "美术", "音频", "研发服务", "outsourc")),
        ("投放", ("投放", "买量", "广告", "推广", "marketing", "advert")),
        ("云服务", ("云", "服务器", "带宽", "saas", "software", "cloud")),
        ("税费", ("税", "附加", "tax")),
        ("办公及其他", ("办公", "差旅", "房租", "其他", "office", "other")),
    )
    for category, keywords in rules:
        if any(_slug(keyword) in clean for keyword in keywords):
            return category
    return _text(value) or "办公及其他"


def _direction(value: Any, category: str) -> str:
    clean = _slug(value)
    if any(word in clean for word in ("收入", "流入", "回款", "income", "inflow")):
        return "收入"
    if any(word in clean for word in ("支出", "流出", "付款", "expense", "outflow")):
        return "支出"
    return "收入" if category == "收入" else "支出"


def _boolean(value: Any) -> bool:
    return _slug(value) in {"是", "已承诺", "已签", "yes", "true", "1", "y"}


@dataclass
class PlanLine:
    id: str
    source_file: str
    source_sheet: str
    source_row: int
    period: str
    scenario: str
    project: str
    category: str
    direction: str
    amount: float
    currency: str
    probability: float
    committed: bool
    note: str
    status: str
    anomalies: list[str] = field(default_factory=list)


def _make_line(path: Path, sheet: str, row_number: int, raw: dict[str, Any], period_override: str = "") -> PlanLine | None:
    amount = _number(raw.get("amount"))
    if amount in (None, 0):
        return None
    period = period_override or _period(raw.get("period"))
    category = _category(raw.get("category"))
    direction = _direction(raw.get("direction"), category)
    probability = _number(raw.get("probability"))
    if probability is None:
        probability = 1.0
    if probability > 1 and probability <= 100:
        probability /= 100
    probability = max(0.0, min(1.0, probability))
    anomalies = []
    currency = _text(raw.get("currency")).upper() or "CNY"
    if not period:
        anomalies.append("缺少或无法识别预算月份")
    if currency != "CNY":
        anomalies.append("外币计划需配置预测汇率后才能进入人民币现金预测")
    scenario = _text(raw.get("scenario")) or "基准"
    row_key = f"{path.name}|{sheet}|{row_number}|{period}|{scenario}|{raw.get('project')}|{category}|{amount}"
    return PlanLine(
        id=hashlib.sha1(row_key.encode("utf-8")).hexdigest()[:12],
        source_file=path.name, source_sheet=sheet, source_row=row_number,
        period=period, scenario=scenario, project=_text(raw.get("project")) or "公司公共",
        category=category, direction=direction, amount=round(abs(amount), 2), currency=currency,
        probability=round(probability, 4), committed=_boolean(raw.get("committed")),
        note=_text(raw.get("note")), status="异常" if anomalies else "可用", anomalies=anomalies,
    )


def parse_plan_workbook(path: str | Path) -> list[PlanLine]:
    """同时支持逐行格式和月份横向展开格式的预算/预测表。"""
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    output = []
    planning_sheets = [
        sheet for sheet in workbook.worksheets
        if any(token in _slug(sheet.title) for token in ("预算", "预测", "budget", "forecast", "plan"))
    ]
    for sheet in planning_sheets or workbook.worksheets:
        mapping: dict[int, str] = {}
        month_columns: dict[int, str] = {}
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            candidate = {index: _field(value) for index, value in enumerate(row) if _field(value)}
            candidate = {index: name for index, name in candidate.items() if name}
            months = {index: _period(value) for index, value in enumerate(row) if _period(value)}
            if (len(candidate) >= 3 and ("amount" in candidate.values() or months)) or (
                len(months) >= 2 and any(name in candidate.values() for name in ("category", "project"))
            ):
                mapping, month_columns = candidate, months
                continue
            if not mapping:
                continue
            base = {name: row[index] if index < len(row) else None for index, name in mapping.items()}
            if month_columns and "amount" not in mapping.values():
                for index, period in month_columns.items():
                    raw = dict(base)
                    raw["amount"] = row[index] if index < len(row) else None
                    line = _make_line(path, sheet.title, row_number, raw, period)
                    if line:
                        output.append(line)
            else:
                line = _make_line(path, sheet.title, row_number, base)
                if line:
                    output.append(line)
    workbook.close()
    return output


def planning_payload(lines: Iterable[PlanLine | dict]) -> dict:
    rows = [asdict(line) if isinstance(line, PlanLine) else dict(line) for line in lines]
    return {
        "records": rows,
        "summary": {
            "line_count": len(rows),
            "period_count": len({row.get("period") for row in rows if row.get("period")}),
            "projects": sorted({row.get("project") for row in rows if row.get("project")}),
            "scenarios": sorted({row.get("scenario") or "基准" for row in rows}),
            "income": round(sum(float(row.get("amount") or 0) for row in rows if row.get("direction") == "收入"), 2),
            "expense": round(sum(float(row.get("amount") or 0) for row in rows if row.get("direction") == "支出"), 2),
            "committed_expense": round(sum(
                float(row.get("amount") or 0) for row in rows
                if row.get("direction") == "支出" and row.get("committed")
            ), 2),
            "exception_count": sum(bool(row.get("anomalies")) for row in rows),
        },
    }


def _add_month(period: str, offset: int) -> str:
    year, month = map(int, period.split("-"))
    zero_based = year * 12 + month - 1 + offset
    return f"{zero_based // 12:04d}-{zero_based % 12 + 1:02d}"


def build_planning_analysis(
    lines: Iterable[dict], settlements: Iterable[dict] = (), purchases: Iterable[dict] = (),
    bank_transactions: Iterable[dict] = (), payroll_rows: Iterable[dict] = (),
    company_profile: dict | None = None, as_of_period: str | None = None, scenario: str = "基准",
    collection_actions: Iterable[dict] = (), cash_allocations: Iterable[dict] = (),
) -> dict:
    lines, settlements, purchases = list(lines), list(settlements), list(purchases)
    bank_transactions, payroll_rows = list(bank_transactions), list(payroll_rows)
    collection_actions, cash_allocations = list(collection_actions), list(cash_allocations)
    company_profile = company_profile or {}
    functional_currency = str(company_profile.get("base_currency") or "CNY").upper()
    all_periods = [
        str(row.get("period")) for row in settlements + payroll_rows + lines
        if re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", str(row.get("period") or ""))
    ]
    as_of = (
        as_of_period if re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", str(as_of_period or ""))
        else max(all_periods) if all_periods else datetime.now().strftime("%Y-%m")
    )
    horizon = int((company_profile.get("cash_planning") or {}).get("forecast_months") or 12)
    horizon = max(3, min(24, horizon))
    cash_policy = company_profile.get("cash_planning") or {}
    currency_suffix = functional_currency.lower()
    minimum_buffer = cash_policy.get(f"minimum_buffer_{currency_suffix}")
    if minimum_buffer in (None, ""):
        minimum_buffer = cash_policy.get("minimum_buffer")
    if minimum_buffer in (None, "") and functional_currency == "CNY":
        minimum_buffer = cash_policy.get("minimum_buffer_cny")
    minimum_buffer = float(minimum_buffer or 0)

    functional_bank = [
        row for row in bank_transactions
        if str(row.get("currency") or functional_currency).upper() == functional_currency
        and row.get("balance") is not None
    ]
    functional_bank.sort(key=lambda row: str(row.get("transaction_date") or ""))
    opening_cash = float(functional_bank[-1]["balance"]) if functional_bank else cash_policy.get(
        f"opening_cash_{currency_suffix}"
    )
    if opening_cash in (None, ""):
        opening_cash = cash_policy.get("opening_cash")
    if opening_cash in (None, "") and functional_currency == "CNY":
        opening_cash = cash_policy.get("opening_cash_cny")
    opening_cash = float(opening_cash) if opening_cash not in (None, "") else None

    plan_rows = [
        row for row in lines if (row.get("scenario") or "基准") == scenario
        and str(row.get("currency") or functional_currency).upper() == functional_currency
        and not row.get("anomalies")
    ]
    plan_by_period: dict[str, dict[str, float]] = {}
    for row in plan_rows:
        bucket = plan_by_period.setdefault(row.get("period"), {"inflows": 0.0, "outflows": 0.0, "committed": 0.0})
        amount = float(row.get("amount") or 0) * float(row.get("probability") if row.get("probability") is not None else 1)
        if row.get("direction") == "收入":
            bucket["inflows"] += amount
        else:
            bucket["outflows"] += amount
            if row.get("committed"):
                bucket["committed"] += amount

    first_forecast = _add_month(as_of, 1)
    settlement_by_key = {
        (str(row.get("entity_id") or ""), str(row.get("id") or "")): row
        for row in settlements if row.get("id")
    }
    latest_promises: dict[tuple[str, str], dict] = {}
    latest_dispute_events: dict[tuple[str, str], dict] = {}
    for action in sorted(collection_actions, key=lambda row: str(row.get("recorded_at") or "")):
        action_key = (str(action.get("entity_id") or ""), str(action.get("settlement_id") or ""))
        if action.get("action_type") == "回款承诺":
            latest_promises[action_key] = action
        elif action.get("action_type") in {"争议登记", "争议解除"}:
            latest_dispute_events[action_key] = action
    collection_by_period: dict[str, float] = {}
    missed_collection_commitments = []
    for key, action in latest_promises.items():
        dispute_event = latest_dispute_events.get(key)
        if dispute_event and str(dispute_event.get("recorded_at") or "") > str(action.get("recorded_at") or ""):
            continue
        settlement = settlement_by_key.get(key)
        if not settlement or str(action.get("currency") or functional_currency).upper() != functional_currency:
            continue
        expected = float(settlement.get("net_receivable") or 0)
        allocated = sum(
            float(item.get("amount") or 0) for item in cash_allocations
            if item.get("status") not in {"已撤销", "已退回"}
            and item.get("target_type") == "receivable" and item.get("target_id") == key[1]
            and str(item.get("entity_id") or "") == key[0]
        )
        outstanding = max(0.0, expected - allocated)
        promise_period = str(action.get("promised_date") or "")[:7]
        amount = min(outstanding, float(action.get("promised_amount") or 0))
        if not re.fullmatch(r"\d{4}-\d{2}", promise_period) or amount <= 0:
            continue
        if promise_period <= as_of:
            missed_collection_commitments.append({
                "settlement_id": key[1], "promised_date": action.get("promised_date"),
                "amount": round(amount, 2), "currency": functional_currency,
            })
        else:
            collection_by_period[promise_period] = collection_by_period.get(promise_period, 0) + amount
    unpaid_commitments = 0.0
    for row in purchases:
        if str(row.get("currency") or functional_currency).upper() != functional_currency:
            continue
        expected = float(row.get("invoice_amount") or row.get("accepted_amount") or row.get("ordered_amount") or 0)
        unpaid_commitments += max(0.0, expected - float(row.get("paid_amount") or 0))
    forecast = []
    cash = opening_cash
    cash_with_commitments = opening_cash
    breach_period = None
    for offset in range(horizon):
        period = _add_month(first_forecast, offset)
        bucket = plan_by_period.get(period, {"inflows": 0.0, "outflows": 0.0, "committed": 0.0})
        inflows, outflows = round(bucket["inflows"], 2), round(bucket["outflows"], 2)
        starting = cash
        ending = round(cash + inflows - outflows, 2) if cash is not None else None
        committed_collections = round(collection_by_period.get(period, 0), 2)
        ending_with_commitments = (
            round(cash_with_commitments + inflows + committed_collections - outflows, 2)
            if cash_with_commitments is not None else None
        )
        if breach_period is None and ending is not None and ending < minimum_buffer:
            breach_period = period
        forecast.append({
            "period": period, "starting_cash": starting, "inflows": inflows, "outflows": outflows,
            "net_cash_flow": round(inflows - outflows, 2), "ending_cash": ending,
            "committed_outflows": round(bucket["committed"], 2),
            "collection_commitments": committed_collections,
            "ending_cash_with_commitments": ending_with_commitments,
        })
        cash = ending
        cash_with_commitments = ending_with_commitments

    actuals: dict[tuple[str, str], float] = {}
    for row in settlements:
        if str(row.get("currency") or functional_currency).upper() == functional_currency and row.get("period"):
            actuals[(row["period"], "收入")] = actuals.get((row["period"], "收入"), 0) + float(row.get("settlement_amount") or 0)
    for row in purchases:
        period = str(row.get("order_date") or "")[:7]
        if re.fullmatch(r"\d{4}-\d{2}", period) and str(row.get("currency") or functional_currency).upper() == functional_currency:
            category = _category(row.get("category"))
            actuals[(period, category)] = actuals.get((period, category), 0) + float(
                row.get("accepted_amount") or row.get("ordered_amount") or 0
            )
    for row in payroll_rows:
        if row.get("period") and str(row.get("currency") or functional_currency).upper() == functional_currency:
            actuals[(row["period"], "人力")] = actuals.get((row["period"], "人力"), 0) + float(
                row.get("total_employer_cost") or row.get("gross_salary") or 0
            )
    budget_by_key: dict[tuple[str, str], float] = {}
    for row in plan_rows:
        key = (row.get("period"), row.get("category") or "办公及其他")
        budget_by_key[key] = budget_by_key.get(key, 0) + float(row.get("amount") or 0)
    variance = []
    for key in sorted(set(actuals) | set(budget_by_key)):
        budget, actual = budget_by_key.get(key, 0.0), actuals.get(key, 0.0)
        direction = "收入" if key[1] == "收入" else "支出"
        variance.append({
            "period": key[0], "category": key[1], "direction": direction,
            "budget": round(budget, 2), "actual": round(actual, 2),
            "variance": round(actual - budget, 2),
            "favorable": actual >= budget if direction == "收入" else actual <= budget,
        })

    recommendations = []
    if opening_cash is None:
        recommendations.append({
            "severity": "阻塞", "title": f"补一笔真实的 {functional_currency} 期初现金",
            "recommendation": "优先导入带余额的银行流水；没有可导出余额时，在公司财务档案填写预测起点现金。",
            "tradeoff": "不补余额仍可比较预算与实际，但不能可靠计算资金跑道。",
        })
    if unpaid_commitments and forecast and unpaid_commitments > sum(row["outflows"] for row in forecast[:3]):
        recommendations.append({
            "severity": "高", "title": "现有未付款采购高于未来三个月计划支出",
            "recommendation": "逐笔补付款月份，把已签约采购设为已承诺；先保障工资、税款和核心服务器。",
            "tradeoff": "延后非核心采购可拉长跑道，但可能影响投放节奏或外包交付。",
        })
    if breach_period:
        recommendations.append({
            "severity": "高", "title": f"预计 {breach_period} 低于最低现金缓冲",
            "recommendation": "先做收入回款提前、可变投放后移和外包里程碑拆分三种动作的情景对比。",
            "tradeoff": "压投放保现金会牺牲增长；维持投放则需要更早锁定回款或融资备用方案。",
        })
    if missed_collection_commitments:
        recommendations.append({
            "severity": "高", "title": f"{len(missed_collection_commitments)} 笔承诺回款已过期仍未结清",
            "recommendation": "立即更新对方反馈、争议原因和下一承诺日；基准现金预测不要继续依赖已失效承诺。",
            "tradeoff": "保留原承诺便于追责与复盘；新承诺作为追加记录进入情景层，不覆盖历史事实。",
        })
    if not recommendations:
        recommendations.append({
            "severity": "正常", "title": "当前基准情景未触发资金硬预警",
            "recommendation": "每月滚动更新未来12个月，并单独维护乐观与保守情景。",
            "tradeoff": "滚动预测需要每月更新关键假设，但能显著减少临时资金决策。",
        })
    return {
        "as_of_period": as_of, "scenario": scenario, "functional_currency": functional_currency,
        "opening_cash": opening_cash, "minimum_buffer": minimum_buffer,
        "opening_cash_cny": opening_cash if functional_currency == "CNY" else None,
        "minimum_buffer_cny": minimum_buffer if functional_currency == "CNY" else None,
        "forecast": forecast, "buffer_breach_period": breach_period,
        "runway_months": next((index for index, row in enumerate(forecast) if row["ending_cash"] is not None and row["ending_cash"] < 0), None),
        "unpaid_purchase_commitments": round(unpaid_commitments, 2),
        "collection_commitment_total": round(sum(collection_by_period.values()), 2),
        "missed_collection_commitments": missed_collection_commitments,
        "variance": variance, "recommendations": recommendations,
        "guardrail": f"主体预测只使用功能本位币 {functional_currency}；催收承诺只进入单独情景层，不改法定应收、不静默叠加到基准现金，其他外币在配置预测汇率前不混入现金跑道。",
    }
