from __future__ import annotations

import hashlib
import math
import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value).replace("\n", " ").strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", _text(value).lower())


def _number(value: Any) -> float:
    if value is None or value == "" or isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return 0.0 if math.isnan(value) or math.isinf(value) else float(value)
    match = re.search(r"-?[\d,.]+", _text(value))
    return float(match.group().replace(",", "")) if match else 0.0


ALIASES = {
    "invoice_number": ("发票号码", "发票号", "数电票号码"),
    "invoice_code": ("发票代码",),
    "invoice_date": ("开票日期", "发票日期"),
    "invoice_type": ("发票类型", "票种"),
    "seller_name": ("销售方名称", "销方名称", "供应商名称"),
    "seller_tax_id": ("销售方纳税人识别号", "销方税号", "供应商税号"),
    "buyer_name": ("购买方名称", "购方名称"),
    "buyer_tax_id": ("购买方纳税人识别号", "购方税号"),
    "item": ("项目名称", "货物或应税劳务服务名称", "商品名称"),
    "amount_ex_tax": ("金额", "不含税金额", "合计金额"),
    "tax_rate": ("税率", "征收率"),
    "tax_amount": ("税额", "合计税额"),
    "total_amount": ("价税合计", "含税金额", "价税合计小写"),
    "status": ("发票状态", "状态"),
    "verification_status": ("查验状态", "验真状态"),
    "deduction_status": ("用途确认", "抵扣状态", "勾选状态"),
    "booking_status": ("入账状态", "入账标识"),
    "po_number": ("po编号", "采购订单", "订单号"),
    "project": ("项目", "游戏", "费用归属"),
}


def _field(value: Any) -> str | None:
    clean = _slug(value)
    candidates = []
    for field_name, aliases in ALIASES.items():
        for alias in aliases:
            alias_clean = _slug(alias)
            if alias_clean and (clean == alias_clean or alias_clean in clean):
                candidates.append((len(alias_clean), field_name))
    return max(candidates, default=(0, None))[1]


@dataclass
class InvoiceRecord:
    id: str
    source_file: str
    source_sheet: str
    source_row: int
    invoice_number: str
    invoice_code: str
    invoice_date: str
    invoice_type: str
    seller_name: str
    seller_tax_id_masked: str
    buyer_name: str
    item: str
    amount_ex_tax: float
    tax_rate: float | None
    tax_amount: float
    total_amount: float
    po_number: str
    project: str
    verification_status: str
    deduction_status: str
    booking_status: str
    duplicate_key: str
    status: str
    anomalies: list[str] = field(default_factory=list)


def _mask_tax_id(value: Any) -> str:
    text = _text(value)
    if len(text) <= 8:
        return text
    return f"{text[:4]}****{text[-4:]}"


def parse_invoice_workbook(path: str | Path) -> list[InvoiceRecord]:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = []
    seen: dict[str, str] = {}
    for sheet in workbook.worksheets:
        mapping = {}
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            candidate = {}
            for index, value in enumerate(row):
                name = _field(value)
                if name and name not in candidate.values():
                    candidate[index] = name
            if len(candidate) >= 5 and "invoice_number" in candidate.values():
                mapping = candidate
                continue
            if not mapping:
                continue
            raw = {name: row[index] if index < len(row) else None for index, name in mapping.items()}
            invoice_number = _text(raw.get("invoice_number"))
            if not invoice_number:
                continue
            amount_ex_tax = _number(raw.get("amount_ex_tax"))
            tax_amount = _number(raw.get("tax_amount"))
            total_amount = _number(raw.get("total_amount")) or amount_ex_tax + tax_amount
            rate_raw = _number(raw.get("tax_rate"))
            tax_rate = rate_raw / 100 if 1 < rate_raw <= 100 else (rate_raw if rate_raw else None)
            invoice_code = _text(raw.get("invoice_code"))
            duplicate_key = f"{invoice_code}|{invoice_number}" if invoice_code else invoice_number
            anomalies = []
            if duplicate_key in seen:
                anomalies.append(f"疑似重复发票，已在 {seen[duplicate_key]} 出现")
            else:
                seen[duplicate_key] = f"{sheet.title} 第{row_number}行"
            if abs(total_amount - amount_ex_tax - tax_amount) > max(0.02, abs(total_amount) * 0.001):
                anomalies.append("价税合计与金额+税额不勾稽")
            if tax_rate is not None and amount_ex_tax and abs(tax_amount - amount_ex_tax * tax_rate) > max(0.05, abs(tax_amount) * 0.01):
                anomalies.append("税额与金额×税率不勾稽")
            verification = _text(raw.get("verification_status")) or "待查验"
            booking = _text(raw.get("booking_status")) or "未入账"
            status_text = _text(raw.get("status"))
            if re.search(r"红冲|作废", status_text):
                anomalies.append(f"发票状态为{status_text}，不得按正常蓝票入账")
            if verification not in {"已查验", "查验通过", "有效"}:
                anomalies.append("尚未完成发票查验")
            row_key = f"{path.name}|{sheet.title}|{row_number}|{duplicate_key}"
            rows.append(InvoiceRecord(
                id=hashlib.sha1(row_key.encode("utf-8")).hexdigest()[:12],
                source_file=path.name, source_sheet=sheet.title, source_row=row_number,
                invoice_number=invoice_number, invoice_code=invoice_code,
                invoice_date=_text(raw.get("invoice_date")), invoice_type=_text(raw.get("invoice_type")) or "待识别票种",
                seller_name=_text(raw.get("seller_name")) or "待识别销售方",
                seller_tax_id_masked=_mask_tax_id(raw.get("seller_tax_id")),
                buyer_name=_text(raw.get("buyer_name")), item=_text(raw.get("item")) or "待识别项目",
                amount_ex_tax=round(amount_ex_tax, 2), tax_rate=tax_rate,
                tax_amount=round(tax_amount, 2), total_amount=round(total_amount, 2),
                po_number=_text(raw.get("po_number")), project=_text(raw.get("project")) or "待分配项目",
                verification_status=verification,
                deduction_status=_text(raw.get("deduction_status")) or "待确认用途",
                booking_status=booking, duplicate_key=duplicate_key,
                status="异常" if anomalies else "待匹配采购",
                anomalies=anomalies,
            ))
    workbook.close()
    return rows


VERIFIED_STATUSES = {"已查验", "查验通过", "有效"}
INVOICE_CONTROL_ANOMALIES = (
    "采购订单尚无已验收交付",
    "发票金额超过剩余已验收额度",
)


def _active_matched_invoice(invoice: dict) -> bool:
    if not invoice.get("purchase_match"):
        return False
    status = str(invoice.get("status") or "")
    anomalies = [str(item) for item in invoice.get("anomalies") or []]
    if re.search(r"红冲|作废", status) or any(re.search(r"红冲|作废", item) for item in anomalies):
        return False
    return True


def _invoice_is_payment_eligible(invoice: dict) -> bool:
    match = invoice.get("purchase_match") or {}
    return bool(
        _active_matched_invoice(invoice)
        and match.get("eligible_for_payment", True)
        and invoice.get("verification_status") in VERIFIED_STATUSES
        and not invoice.get("anomalies")
    )


def match_invoices_to_purchases(
    invoices: Iterable[InvoiceRecord], purchases: Iterable[dict], existing_invoices: Iterable[dict] = (),
) -> list[dict]:
    purchase_rows = list(purchases)
    consumed: dict[str, float] = {}
    invoices = list(invoices)
    incoming_ids = {str(getattr(item, "id", "") or (item.get("id") if isinstance(item, dict) else "")) for item in invoices}
    for existing in existing_invoices:
        if str(existing.get("id") or "") in incoming_ids or not _active_matched_invoice(existing):
            continue
        match = existing.get("purchase_match") or {}
        if match.get("eligible_for_payment", True) is False:
            continue
        purchase_id = str(match.get("purchase_id") or "")
        if purchase_id:
            consumed[purchase_id] = consumed.get(purchase_id, 0.0) + _number(existing.get("total_amount"))
    output = []
    for invoice in invoices:
        row = dict(invoice) if isinstance(invoice, dict) else asdict(invoice)
        invoice_total = _number(row.get("total_amount"))
        candidates = []
        for purchase in purchase_rows:
            purchase_id = str(purchase.get("id") or "")
            accepted = _number(purchase.get("accepted_amount"))
            workflow_order = bool(purchase.get("procurement_request_id") or purchase.get("milestones"))
            basis_amount = accepted if (accepted > 0 or workflow_order) else _number(purchase.get("ordered_amount"))
            previously_invoiced = round(consumed.get(purchase_id, 0.0), 2)
            remaining = round(max(0.0, basis_amount - previously_invoiced), 2)
            po_score = 1 if row.get("po_number") and row.get("po_number") == purchase.get("po_number") else 0
            vendor_score = 1 if row.get("seller_name") != "待识别销售方" and row.get("seller_name") == purchase.get("vendor") else 0
            if not basis_amount and not (po_score and vendor_score):
                continue
            amount_score = max(0, 1 - abs(remaining - invoice_total) / max(abs(remaining), invoice_total, 1))
            score = amount_score * 0.6 + po_score * 0.25 + vendor_score * 0.15
            candidates.append((score, po_score, vendor_score, purchase, accepted, basis_amount, previously_invoiced, remaining, workflow_order))
        best = max(candidates, key=lambda item: item[0], default=None)
        if best and (best[0] >= 0.75 or best[1]):
            score, _, _, purchase, accepted, basis_amount, previously_invoiced, remaining, workflow_order = best
            purchase_id = str(purchase.get("id") or "")
            anomalies = list(row.get("anomalies") or [])
            eligible = True
            control_status = "已匹配"
            if workflow_order and accepted <= 0:
                anomalies.append("采购订单尚无已验收交付，发票暂不能进入应付与付款链路")
                eligible = False
                control_status = "待验收"
            elif invoice_total > remaining + 0.01:
                anomalies.append(
                    f"发票金额超过剩余已验收额度 {remaining:,.2f}，超出 {invoice_total - remaining:,.2f}"
                )
                eligible = False
                control_status = "超验收额度"
            row["purchase_match"] = {
                "purchase_id": purchase.get("id"), "po_number": purchase.get("po_number"),
                "procurement_request_id": purchase.get("procurement_request_id"),
                "accepted_delivery_ids": [
                    event.get("delivery_id") for event in purchase.get("acceptance_history") or []
                    if event.get("delivery_id")
                ],
                "target": f"{purchase.get('vendor')} / {purchase.get('item')}",
                "score": round(score, 4),
                "difference": round(invoice_total - remaining, 2),
                "accepted_amount": round(accepted, 2),
                "previously_invoiced_amount": previously_invoiced,
                "remaining_accepted_capacity_before": remaining,
                "amount_basis": "accepted_amount" if accepted > 0 or workflow_order else "legacy_order_amount",
                "eligible_for_payment": eligible,
                "control_status": control_status,
            }
            row["anomalies"] = list(dict.fromkeys(anomalies))
            if eligible:
                consumed[purchase_id] = round(previously_invoiced + invoice_total, 2)
            if not row["anomalies"]:
                row["status"] = "已匹配待入账"
            else:
                row["status"] = "异常"
        else:
            row["purchase_match"] = None
        output.append(row)
    return output


def roll_invoice_totals_to_purchases(purchases: Iterable[dict], invoices: Iterable[dict]) -> list[dict]:
    """回写订单累计开票和可付额度，但不覆盖验收事实。"""
    invoices_by_purchase: dict[str, list[dict]] = {}
    for invoice in invoices:
        match = invoice.get("purchase_match") or {}
        purchase_id = str(match.get("purchase_id") or "")
        if purchase_id and _active_matched_invoice(invoice):
            invoices_by_purchase.setdefault(purchase_id, []).append(invoice)
    output = []
    for source in purchases:
        purchase = dict(source)
        purchase_id = str(purchase.get("id") or "")
        linked = invoices_by_purchase.get(purchase_id, [])
        invoiced = round(sum(_number(item.get("total_amount")) for item in linked), 2)
        verified_clean = round(sum(
            _number(item.get("total_amount")) for item in linked if _invoice_is_payment_eligible(item)
        ), 2)
        accepted = _number(purchase.get("accepted_amount"))
        ordered = _number(purchase.get("ordered_amount"))
        workflow_order = bool(purchase.get("procurement_request_id") or purchase.get("milestones"))
        comparison_amount = accepted if (accepted > 0 or workflow_order) else ordered
        anomalies = [
            str(item) for item in purchase.get("anomalies") or []
            if not any(str(item).startswith(prefix) for prefix in INVOICE_CONTROL_ANOMALIES)
        ]
        if workflow_order and invoiced > 0 and accepted <= 0:
            anomalies.append("采购订单尚无已验收交付，但已关联发票")
        elif comparison_amount and invoiced > comparison_amount + 0.01:
            anomalies.append(f"发票金额超过剩余已验收额度：累计开票 {invoiced:,.2f}，已验收 {comparison_amount:,.2f}")
        invoice_status = (
            "未开票" if not invoiced else "超额开票" if comparison_amount and invoiced > comparison_amount + 0.01
            else "已开票" if comparison_amount and abs(invoiced - comparison_amount) <= 0.01
            else "部分开票"
        )
        purchase.update({
            "invoice_amount": invoiced,
            "invoice_status": invoice_status,
            "payment_eligible_amount": round(min(accepted, verified_clean), 2),
            "invoice_match_summary": {
                "matched_invoice_count": len(linked),
                "matched_invoice_ids": [item.get("id") for item in linked],
                "matched_invoice_amount": invoiced,
                "verified_clean_amount": verified_clean,
                "remaining_accepted_to_invoice": round(max(0.0, accepted - invoiced), 2),
                "procurement_request_id": purchase.get("procurement_request_id"),
                "accepted_delivery_ids": [
                    event.get("delivery_id") for event in purchase.get("acceptance_history") or []
                    if event.get("delivery_id")
                ],
            },
            "anomalies": list(dict.fromkeys(anomalies)),
        })
        output.append(purchase)
    return output


def invoice_payload(records: Iterable[dict]) -> dict:
    rows = list(records)
    return {
        "records": rows,
        "summary": {
            "count": len(rows),
            "amount_ex_tax": round(sum(row.get("amount_ex_tax") or 0 for row in rows), 2),
            "tax_amount": round(sum(row.get("tax_amount") or 0 for row in rows), 2),
            "total_amount": round(sum(row.get("total_amount") or 0 for row in rows), 2),
            "matched_count": sum(bool(row.get("purchase_match")) for row in rows),
            "payment_eligible_count": sum(_invoice_is_payment_eligible(row) for row in rows),
            "exception_count": sum(bool(row.get("anomalies")) for row in rows),
            "unverified_count": sum(row.get("verification_status") not in {"已查验", "查验通过", "有效"} for row in rows),
        },
        "guardrail": "只有查验合法真实、防重复入账且完成必要审签的电子凭证，才能进入会计入账和归档流程。",
    }
