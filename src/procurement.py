from __future__ import annotations

import hashlib
import math
import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


SOURCING_METHODS = {"竞争比价", "框架协议", "单一来源", "平台直采"}
PROCUREMENT_DECISIONS = {"批准", "退回", "取消"}
DELIVERY_ACCEPTANCE_DECISIONS = {"全部验收", "部分验收", "退回整改"}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).replace("\n", " ").strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", _text(value).lower())


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    match = re.search(r"-?[\d,.]+", _text(value))
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


def _date_text(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    text = re.sub(r"\D", "", _text(value))
    if len(text) == 8:
        return f"{text[:4]}-{text[4:6]}-{text[6:]}"
    return _text(value)


def _request_id(entity_id: str, project: str, description: str, actor: str) -> str:
    seed = f"{entity_id}|{project}|{description}|{actor}|{datetime.now(timezone.utc).isoformat()}"
    return f"PR-{hashlib.sha1(seed.encode('utf-8')).hexdigest()[:14].upper()}"


def _same_budget_key(row: dict, *, entity_id: str, project: str, category: str, period: str, currency: str) -> bool:
    return (
        str(row.get("entity_id") or "") == entity_id
        and str(row.get("project") or "").strip().casefold() == project.strip().casefold()
        and str(row.get("category") or "").strip().casefold() == category.strip().casefold()
        and str(row.get("period") or "") == period
        and str(row.get("currency") or "CNY").upper() == currency
    )


def procurement_budget_snapshot(
    plan_lines: Iterable[dict], existing_requests: Iterable[dict], *, entity_id: str,
    project: str, category: str, period: str, currency: str,
) -> dict:
    currency = str(currency or "CNY").upper()
    budget_lines = [
        row for row in plan_lines
        if _same_budget_key(row, entity_id=entity_id, project=project, category=category, period=period, currency=currency)
        and (row.get("scenario") or "基准") == "基准" and row.get("direction") != "收入"
        and not row.get("anomalies")
    ]
    budget = round(sum(float(row.get("amount") or 0) for row in budget_lines), 2)
    reserved = round(sum(
        float(row.get("amount") or 0) for row in existing_requests
        if _same_budget_key(row, entity_id=entity_id, project=project, category=category, period=period, currency=currency)
        and row.get("status") in {"待批准", "已批准", "已下单"}
    ), 2)
    return {
        "entity_id": entity_id, "project": project, "category": category, "period": period,
        "currency": currency, "budget_found": bool(budget_lines), "budget_amount": budget,
        "reserved_amount": reserved, "available_amount": round(budget - reserved, 2),
        "source_line_ids": [row.get("id") for row in budget_lines if row.get("id")],
        "basis": "基准情景同主体、项目、类目、期间及币种预算；待批准、已批准和已下单采购申请持续占用预算。",
    }


def create_procurement_request(
    *, entity_id: str, project: str, category: str, description: str, amount: float,
    currency: str, period: str, needed_by: str, requester: str, sourcing_method: str,
    selected_vendor: str, quotes: Iterable[dict], evidence: Iterable[str], budget_snapshot: dict,
    selection_rationale: str = "", sourcing_exception_reason: str = "",
    budget_exception_reason: str = "", framework_reference: str = "",
) -> dict:
    entity_id = str(entity_id or "").strip()
    project, category, description = map(lambda value: str(value or "").strip(), (project, category, description))
    requester = str(requester or "").strip()
    if not all((entity_id, project, category, description, requester)):
        raise ValueError("采购申请必须填写主体、项目、类目、需求说明和申请人")
    try:
        amount = round(float(amount), 2)
    except (TypeError, ValueError) as exc:
        raise ValueError("采购申请金额无效") from exc
    if amount <= 0:
        raise ValueError("采购申请金额必须大于0")
    currency = str(currency or "CNY").upper()
    if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", str(period or "")):
        raise ValueError("预算期间必须为 YYYY-MM")
    try:
        date.fromisoformat(str(needed_by or ""))
    except ValueError as exc:
        raise ValueError("期望交付日必须为 YYYY-MM-DD") from exc
    sourcing_method = str(sourcing_method or "").strip()
    if sourcing_method not in SOURCING_METHODS:
        raise ValueError("采购方式无效")
    selected_vendor = str(selected_vendor or "").strip()
    if not selected_vendor:
        raise ValueError("请填写拟选供应商")
    evidence_items = list(dict.fromkeys(str(item).strip()[:500] for item in evidence if str(item).strip()))
    quote_rows = []
    for quote in quotes:
        vendor = str(quote.get("vendor") or "").strip()
        quote_currency = str(quote.get("currency") or currency).upper()
        try:
            quote_amount = round(float(quote.get("amount") or 0), 2)
        except (TypeError, ValueError):
            quote_amount = 0.0
        if vendor and quote_amount > 0:
            quote_rows.append({
                "vendor": vendor[:160], "amount": quote_amount, "currency": quote_currency,
                "evidence": str(quote.get("evidence") or "").strip()[:500],
            })
    blockers, warnings = [], []
    budget_exception_reason = str(budget_exception_reason or "").strip()
    available = float(budget_snapshot.get("available_amount") or 0)
    if not budget_snapshot.get("budget_found") or amount > available + 0.01:
        if len(budget_exception_reason) < 8:
            blockers.append("未找到足额同口径预算；请补预算或填写至少8个字的预算例外原因")
        else:
            warnings.append("预算例外待独立审批")
    sourcing_exception_reason = str(sourcing_exception_reason or "").strip()
    selection_rationale = str(selection_rationale or "").strip()
    if sourcing_method == "竞争比价":
        same_currency = [row for row in quote_rows if row["currency"] == currency]
        if len({row["vendor"] for row in same_currency}) < 3:
            if len(sourcing_exception_reason) < 8:
                blockers.append("竞争比价少于3家同币种有效报价；请补报价或说明例外原因")
            else:
                warnings.append("少于三方比价，例外原因待审批")
        selected_quote = next((row for row in same_currency if row["vendor"] == selected_vendor), None)
        if not selected_quote:
            blockers.append("拟选供应商没有同币种有效报价")
        elif selected_quote["amount"] > amount + 0.01:
            blockers.append("拟选供应商报价高于采购申请金额")
        elif same_currency and selected_quote["amount"] > min(row["amount"] for row in same_currency) + 0.01:
            if len(selection_rationale) < 8:
                blockers.append("非最低价中选必须说明交付、质量或综合评分依据")
            else:
                warnings.append("非最低价中选，理由待独立审批")
    elif sourcing_method == "框架协议":
        if len(str(framework_reference or "").strip()) < 4:
            blockers.append("框架采购必须关联有效框架编号或价格表")
    elif sourcing_method == "单一来源":
        if len(sourcing_exception_reason) < 8:
            blockers.append("单一来源必须说明不可替代性、紧急性或连续性交付依据")
        else:
            warnings.append("单一来源理由待独立审批")
    elif sourcing_method == "平台直采" and not evidence_items:
        blockers.append("平台直采至少需要商品页、订单页或价格截图证据")
    status = "阻塞" if blockers else "待批准"
    return {
        "id": _request_id(entity_id, project, description, requester), "entity_id": entity_id,
        "project": project[:160], "category": category[:120], "description": description[:1000],
        "amount": amount, "currency": currency, "period": period, "needed_by": needed_by,
        "requester": requester[:80], "sourcing_method": sourcing_method,
        "selected_vendor": selected_vendor[:160], "quotes": quote_rows,
        "selection_rationale": selection_rationale[:1000],
        "sourcing_exception_reason": sourcing_exception_reason[:1000],
        "budget_exception_reason": budget_exception_reason[:1000],
        "framework_reference": str(framework_reference or "").strip()[:200],
        "budget_snapshot": dict(budget_snapshot), "evidence": evidence_items,
        "status": status, "blockers": blockers, "warnings": warnings,
        "agent_recommendation": (
            "先补齐预算或寻源依据，再进入批准。" if blockers
            else "预算和寻源资料已形成摘要；由独立审批人判断业务必要性、价格与例外是否可接受。"
        ),
        "created_at": datetime.now(timezone.utc).isoformat(), "approval_history": [],
    }


def decide_procurement_request(request: dict, decision: str, actor: str, rationale: str) -> dict:
    decision, actor, rationale = str(decision or "").strip(), str(actor or "").strip(), str(rationale or "").strip()
    if decision not in PROCUREMENT_DECISIONS:
        raise ValueError("采购申请决定无效")
    if not actor:
        raise ValueError("请填写审批人")
    if actor == str(request.get("requester") or ""):
        raise ValueError("申请人不能审批自己的采购申请")
    if len(rationale) < 8:
        raise ValueError("请填写至少8个字的审批依据")
    if decision == "批准" and request.get("blockers"):
        raise ValueError("采购申请仍有阻塞项，不能批准")
    updated = dict(request)
    updated["status"] = {"批准": "已批准", "退回": "已退回", "取消": "已取消"}[decision]
    updated["approval_history"] = [*(request.get("approval_history") or []), {
        "decision": decision, "actor": actor[:80], "rationale": rationale[:1000],
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }]
    return updated


def create_purchase_order_from_request(
    request: dict, *, po_number: str, order_date: str, actor: str,
    milestones: Iterable[dict], evidence: Iterable[str] = (), item: str = "",
) -> tuple[dict, dict]:
    """Turn one approved request into an auditable order with delivery milestones."""
    if request.get("status") != "已批准":
        raise ValueError("只有已批准的采购申请才能生成订单")
    po_number, actor = str(po_number or "").strip(), str(actor or "").strip()
    if not po_number or not actor:
        raise ValueError("生成订单必须填写 PO 编号和经办人")
    try:
        date.fromisoformat(str(order_date or ""))
    except ValueError as exc:
        raise ValueError("订单日期必须为 YYYY-MM-DD") from exc
    order_amount = round(float(request.get("amount") or 0), 2)
    if order_amount <= 0:
        raise ValueError("采购申请金额无效，不能生成订单")
    rows = []
    for index, raw in enumerate(milestones, 1):
        title = str(raw.get("title") or "").strip()
        criteria = str(raw.get("acceptance_criteria") or "").strip()
        owner = str(raw.get("owner") or "").strip()
        due_date = str(raw.get("due_date") or "").strip()
        try:
            milestone_amount = round(float(raw.get("amount") or 0), 2)
            date.fromisoformat(due_date)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"第{index}个里程碑的金额或日期无效") from exc
        if not title or milestone_amount <= 0 or len(criteria) < 4 or not owner:
            raise ValueError(f"第{index}个里程碑必须填写名称、正金额、验收标准和负责人")
        milestone_seed = f"{request.get('id')}|{po_number}|{index}|{title}"
        milestone_id = f"MS-{hashlib.sha1(milestone_seed.encode()).hexdigest()[:12].upper()}"
        rows.append({
            "id": milestone_id, "sequence": index, "title": title[:200],
            "amount": milestone_amount, "currency": str(request.get("currency") or "CNY").upper(),
            "due_date": due_date, "acceptance_criteria": criteria[:1000], "owner": owner[:100],
        })
    if not rows:
        raise ValueError("采购订单至少需要一个交付里程碑")
    milestone_total = round(sum(row["amount"] for row in rows), 2)
    if abs(milestone_total - order_amount) > 0.01:
        raise ValueError("交付里程碑金额合计必须等于已批准采购金额")
    evidence_items = list(dict.fromkeys(str(value).strip()[:500] for value in evidence if str(value).strip()))
    if not evidence_items:
        raise ValueError("生成订单至少需要一项订单或商业确认依据")
    now = datetime.now(timezone.utc).isoformat()
    entity_id = str(request.get("entity_id") or "")
    order_seed = f"{entity_id}|{request.get('id')}|{po_number}"
    order_id = f"PO-{hashlib.sha1(order_seed.encode()).hexdigest()[:14].upper()}"
    order = {
        "id": order_id, "entity_id": entity_id, "procurement_request_id": request.get("id"),
        "po_number": po_number[:160], "order_date": str(order_date),
        "project": request.get("project"), "category": request.get("category"),
        "vendor": request.get("selected_vendor"), "item": (str(item or "").strip() or request.get("description"))[:1000],
        "quantity": None, "unit_price": None, "ordered_amount": order_amount,
        "accepted_amount": 0.0, "invoice_amount": 0.0, "paid_amount": 0.0,
        "currency": str(request.get("currency") or "CNY").upper(), "tax_rate": None,
        "invoice_status": "未开票", "payment_status": "未付款", "delivery_status": "待交付",
        "accounting_status": "缺少交付与验收证据", "acceptance_status": "待交付", "status": "已下单",
        "milestones": rows, "anomalies": [], "evidence": {
            "source": "approved_procurement_request", "procurement_request_id": request.get("id"),
            "budget_snapshot": request.get("budget_snapshot") or {}, "order_evidence": evidence_items,
        },
        "source_file": "智能财务工作台", "source_sheet": "采购订单", "source_row": 0,
        "created_by": actor[:100], "created_at": now, "acceptance_history": [],
    }
    updated_request = dict(request)
    updated_request["status"] = "已下单"
    updated_request["purchase_order_ids"] = [*(request.get("purchase_order_ids") or []), order_id]
    updated_request["fulfillment_history"] = [*(request.get("fulfillment_history") or []), {
        "event": "生成采购订单", "purchase_order_id": order_id, "po_number": po_number,
        "actor": actor[:100], "timestamp": now,
    }]
    return updated_request, order


def record_purchase_delivery(
    purchase: dict, *, milestone_id: str, delivered_amount: float, delivery_date: str,
    delivered_by: str, evidence: Iterable[str], note: str = "", existing_deliveries: Iterable[dict] = (),
) -> dict:
    """Record a delivery event; this does not itself accept cost or allow payment."""
    milestone = next((row for row in purchase.get("milestones") or [] if row.get("id") == milestone_id), None)
    if not milestone:
        raise ValueError("交付必须关联订单中的有效里程碑")
    try:
        delivered_amount = round(float(delivered_amount), 2)
        date.fromisoformat(str(delivery_date or ""))
    except (TypeError, ValueError) as exc:
        raise ValueError("交付金额或日期无效") from exc
    delivered_by = str(delivered_by or "").strip()
    evidence_items = list(dict.fromkeys(str(value).strip()[:500] for value in evidence if str(value).strip()))
    if delivered_amount <= 0 or not delivered_by or not evidence_items:
        raise ValueError("记录交付必须填写正金额、交付人和至少一项交付证据")
    previous = sum(
        float(row.get("delivered_amount") or 0) for row in existing_deliveries
        if row.get("purchase_id") == purchase.get("id") and row.get("milestone_id") == milestone_id
        and row.get("status") not in {"已取消"}
    )
    if previous + delivered_amount > float(milestone.get("amount") or 0) + 0.01:
        raise ValueError("累计交付金额不能超过里程碑金额")
    now = datetime.now(timezone.utc).isoformat()
    seed = f"{purchase.get('id')}|{milestone_id}|{delivery_date}|{previous}|{delivered_amount}|{now}"
    return {
        "id": f"DEL-{hashlib.sha1(seed.encode()).hexdigest()[:14].upper()}",
        "entity_id": purchase.get("entity_id"), "purchase_id": purchase.get("id"),
        "procurement_request_id": purchase.get("procurement_request_id"),
        "po_number": purchase.get("po_number"), "vendor": purchase.get("vendor"),
        "milestone_id": milestone_id, "milestone_title": milestone.get("title"),
        "delivered_amount": delivered_amount, "accepted_amount": 0.0,
        "currency": purchase.get("currency"), "delivery_date": str(delivery_date),
        "delivered_by": delivered_by[:100], "evidence": evidence_items,
        "note": str(note or "").strip()[:1000], "status": "已交付待验收",
        "acceptance_criteria": milestone.get("acceptance_criteria"),
        "acceptance_owner": milestone.get("owner"), "created_at": now,
    }


def apply_delivery_acceptance_decision(
    purchase: dict, delivery: dict, decision: str, actor: str, *,
    accepted_amount: float | None = None, evidence: Iterable[str] = (), note: str = "",
    period: str = "", all_deliveries: Iterable[dict] = (),
) -> tuple[dict, dict]:
    """Accept only a recorded delivery and roll its result up to the purchase order."""
    if decision not in DELIVERY_ACCEPTANCE_DECISIONS:
        raise ValueError("验收决定必须是全部验收、部分验收或退回整改")
    if delivery.get("purchase_id") != purchase.get("id"):
        raise ValueError("交付记录与采购订单不一致")
    if delivery.get("status") != "已交付待验收":
        raise ValueError("该交付记录已处理，不能重复验收")
    actor = str(actor or "").strip()
    if not actor:
        raise ValueError("请填写业务验收人")
    evidence_items = list(dict.fromkeys(str(value).strip()[:500] for value in evidence if str(value).strip()))
    delivered = round(float(delivery.get("delivered_amount") or 0), 2)
    if decision in {"全部验收", "部分验收"} and not evidence_items:
        raise ValueError("确认验收前至少需要一项验收证据")
    if decision == "全部验收":
        target, delivery_status = delivered, "已验收"
    elif decision == "部分验收":
        target = round(float(accepted_amount or 0), 2)
        if target <= 0 or target >= delivered:
            raise ValueError("部分验收金额必须大于0且小于本次交付金额")
        delivery_status = "部分验收"
    else:
        if len(str(note or "").strip()) < 4:
            raise ValueError("退回整改时请填写具体原因")
        target, delivery_status = 0.0, "已退回整改"
    now = datetime.now(timezone.utc).isoformat()
    updated_delivery = dict(delivery)
    updated_delivery.update({
        "accepted_amount": target, "status": delivery_status, "accepted_by": actor[:100],
        "accepted_at": now, "acceptance_evidence": evidence_items,
        "acceptance_note": str(note or "").strip()[:1000], "period": period,
    })
    delivery_rows = [updated_delivery if row.get("id") == delivery.get("id") else row for row in all_deliveries]
    if not any(row.get("id") == delivery.get("id") for row in delivery_rows):
        delivery_rows.append(updated_delivery)
    accepted_total = round(sum(
        float(row.get("accepted_amount") or 0) for row in delivery_rows
        if row.get("purchase_id") == purchase.get("id")
    ), 2)
    ordered = round(float(purchase.get("ordered_amount") or 0), 2)
    if accepted_total > ordered + 0.01:
        raise ValueError("累计验收金额不能超过采购订单金额")
    updated_purchase = dict(purchase)
    order_status = "已验收" if abs(accepted_total - ordered) <= 0.01 else "部分验收" if accepted_total else "待交付"
    updated_purchase.update({
        "accepted_amount": accepted_total, "acceptance_status": order_status,
        "delivery_status": order_status, "accounting_status": (
            "可生成应付凭证" if accepted_total and float(purchase.get("invoice_amount") or 0)
            else "可生成采购暂估凭证" if accepted_total else "验收未通过，阻止付款与入账"
        ),
        "acceptance_history": [*(purchase.get("acceptance_history") or []), {
            "delivery_id": delivery.get("id"), "milestone_id": delivery.get("milestone_id"),
            "decision": decision, "actor": actor[:100], "accepted_amount_after": accepted_total,
            "delivery_accepted_amount": target, "evidence": evidence_items,
            "note": str(note or "").strip()[:1000], "period": period, "timestamp": now,
        }],
    })
    updated_purchase["status"] = "验收完成" if order_status == "已验收" else "履约中"
    updated_purchase["workflow"] = {
        "can_request_invoice": accepted_total > float(purchase.get("invoice_amount") or 0),
        "can_pay": float(purchase.get("invoice_amount") or 0) > float(purchase.get("paid_amount") or 0)
                   and accepted_total + 0.01 >= float(purchase.get("invoice_amount") or 0),
        "accrual_candidate": accepted_total > float(purchase.get("invoice_amount") or 0),
        "remaining_to_accept": round(max(0, ordered - accepted_total), 2),
    }
    return updated_purchase, updated_delivery


def procurement_workflow_payload(
    requests: Iterable[dict], purchases: Iterable[dict], deliveries: Iterable[dict], *, entity_id: str = "",
) -> dict:
    def scoped(rows: Iterable[dict]) -> list[dict]:
        return [dict(row) for row in rows if not entity_id or str(row.get("entity_id") or "") == entity_id]
    request_rows, purchase_rows, delivery_rows = scoped(requests), scoped(purchases), scoped(deliveries)
    return {
        "entity_id": entity_id, "requests": request_rows, "orders": purchase_rows, "deliveries": delivery_rows,
        "summary": {
            "approved_to_order": sum(row.get("status") == "已批准" for row in request_rows),
            "open_orders": sum(row.get("acceptance_status") != "已验收" for row in purchase_rows),
            "delivered_to_accept": sum(row.get("status") == "已交付待验收" for row in delivery_rows),
            "linked_orders": sum(bool(row.get("procurement_request_id")) for row in purchase_rows),
        },
    }


FIELD_ALIASES = {
    "po_number": ("po单", "po编号", "采购订单", "订单编号", "po number"),
    "order_date": ("下单时间", "下单日期", "订单日期", "采购日期"),
    "project": ("项目", "游戏", "费用归属", "项目名称"),
    "vendor": ("供应商", "供应商名称", "收款方"),
    "item": ("名称", "品名", "采购内容", "服务内容", "物料", "条目"),
    "quantity": ("数量", "采购数量"),
    "unit_price": ("单价", "含税单价"),
    "amount": ("金额小计", "总价", "订单金额", "结算金额", "含税金额", "小计"),
    "currency": ("币种", "结算币种"),
    "tax_rate": ("税率", "征收率"),
    "invoice_number": ("发票号码", "发票号"),
    "invoice_amount": ("开票金额", "发票金额"),
    "accepted_amount": ("验收金额", "确认金额"),
    "paid_amount": ("付款金额", "已付金额"),
    "status_note": ("备注", "状态", "订单状态"),
}


def _field(header: Any) -> str | None:
    clean = _slug(header)
    candidates = []
    for field_name, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            alias_clean = _slug(alias)
            if alias_clean and (clean == alias_clean or alias_clean in clean):
                candidates.append((len(alias_clean), field_name))
    return max(candidates, default=(0, None))[1]


@dataclass
class PurchaseRecord:
    id: str
    source_file: str
    source_sheet: str
    source_row: int
    po_number: str
    order_date: str
    project: str
    vendor: str
    category: str
    item: str
    quantity: float | None
    unit_price: float | None
    ordered_amount: float | None
    accepted_amount: float | None
    invoice_amount: float | None
    paid_amount: float | None
    currency: str
    tax_rate: float | None
    invoice_status: str
    payment_status: str
    delivery_status: str
    accounting_status: str
    status: str
    anomalies: list[str] = field(default_factory=list)
    evidence: dict[str, Any] = field(default_factory=dict)
    acceptance_status: str = "待验收"
    accepted_by: str = ""
    accepted_at: str = ""
    acceptance_note: str = ""
    acceptance_evidence: list[str] = field(default_factory=list)
    acceptance_history: list[dict[str, Any]] = field(default_factory=list)


def _header_mapping(row: Iterable[Any]) -> dict[int, str]:
    mapping = {}
    for index, value in enumerate(row):
        field_name = _field(value)
        if field_name and field_name not in mapping.values():
            mapping[index] = field_name
    return mapping


def _infer_category(filename: str, sheet_name: str) -> str:
    combined = f"{filename} {sheet_name}"
    for pattern, category in (
        (r"素材|视频|美术|本地化", "素材制作"),
        (r"广告|买量|投放", "广告投放"),
        (r"活动", "线下活动"),
        (r"周边|定制|礼盒|金钞|闪卡|防伪标", "定制周边"),
        (r"软件|云|服务器", "软件与云服务"),
    ):
        if re.search(pattern, combined, re.I):
            return category
    return "待分类采购"


def parse_purchase_workbook(path: str | Path) -> list[PurchaseRecord]:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    records = []
    purchase_sheets = [
        sheet for sheet in workbook.worksheets
        if any(token in _text(sheet.title).lower() for token in ("采购", "订单", "purchase", "po台账"))
    ]
    for sheet in purchase_sheets or workbook.worksheets:
        mapping: dict[int, str] = {}
        extra_quantity_columns: list[tuple[int, str]] = []
        category = _infer_category(path.name, sheet.title)
        previous_po = ""
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 5000, 5000), values_only=True), 1
        ):
            candidate = _header_mapping(row)
            if len(candidate) >= 3 and ("amount" in candidate.values() or "unit_price" in candidate.values()):
                mapping = candidate
                extra_quantity_columns = [
                    (index, _text(value))
                    for index, value in enumerate(row)
                    if index not in mapping and _text(value) and any(word in _text(value) for word in ("数量", "平标", "卷标"))
                ]
                # Some operational registers use named quantity columns rather than a generic 数量 column.
                if "quantity" not in mapping.values():
                    extra_quantity_columns.extend([
                        (index, _text(value)) for index, value in enumerate(row)
                        if _text(value) in {"平标", "卷标"}
                    ])
                continue
            if not mapping:
                continue
            raw = {name: row[index] if index < len(row) else None for index, name in mapping.items()}
            if not any(_text(value) for value in raw.values()):
                continue
            po_number = _text(raw.get("po_number")) or previous_po
            if _text(raw.get("po_number")):
                previous_po = _text(raw.get("po_number"))
            quantity = _number(raw.get("quantity"))
            item = _text(raw.get("item"))
            if quantity is None:
                named_quantities = []
                seen = set()
                for index, label in extra_quantity_columns:
                    if index in seen or index >= len(row):
                        continue
                    seen.add(index)
                    value = _number(row[index])
                    if value is not None:
                        named_quantities.append((label, value))
                if named_quantities:
                    item, quantity = named_quantities[0]
            unit_price = _number(raw.get("unit_price"))
            amount = _number(raw.get("amount"))
            accepted = _number(raw.get("accepted_amount"))
            invoice = _number(raw.get("invoice_amount"))
            paid = _number(raw.get("paid_amount"))
            if amount is None and quantity is not None and unit_price is not None:
                amount = quantity * unit_price
            if amount is None and quantity is None and unit_price is None:
                continue
            note = _text(raw.get("status_note"))
            anomalies = []
            if not po_number:
                anomalies.append("缺少PO编号")
            if quantity is not None and unit_price is not None and amount is not None:
                expected = quantity * unit_price
                tolerance = max(0.02, abs(amount) * 0.001)
                if abs(amount - expected) > tolerance:
                    anomalies.append(f"数量×单价与订单金额差异 {amount - expected:,.2f}")
            if accepted is not None and amount is not None and accepted > amount * 1.001:
                anomalies.append("验收金额超过PO金额")
            if invoice is not None and accepted is not None and invoice > accepted * 1.001:
                anomalies.append("开票金额超过验收金额")
            if paid is not None and invoice is not None and paid > invoice * 1.001:
                anomalies.append("付款金额超过发票金额")
            invoice_status = "已开票" if invoice else ("未开票" if "未开票" in note else "待确认")
            payment_status = "已付款" if paid else ("未付款" if "未付款" in note else "待确认")
            delivery_status = "已验收" if accepted else ("已交付待确认" if re.search(r"缔结|上传|发货", note) else "待确认")
            accounting_status = "可生成应付凭证" if accepted else "缺少验收证据"
            if "未开票" in note and delivery_status != "待确认":
                anomalies.append("疑似已发生未开票：月结需判断暂估")
            row_key = f"{path.name}|{sheet.title}|{row_number}|{po_number}|{item}|{amount}"
            records.append(PurchaseRecord(
                id=hashlib.sha1(row_key.encode("utf-8")).hexdigest()[:12],
                source_file=path.name,
                source_sheet=sheet.title,
                source_row=row_number,
                po_number=po_number,
                order_date=_date_text(raw.get("order_date")),
                project=_text(raw.get("project")) or "待分配项目",
                vendor=_text(raw.get("vendor")) or "待识别供应商",
                category=category,
                item=item or "采购条目",
                quantity=quantity,
                unit_price=unit_price,
                ordered_amount=amount,
                accepted_amount=accepted,
                invoice_amount=invoice,
                paid_amount=paid,
                currency=_text(raw.get("currency")) or "CNY",
                tax_rate=_number(raw.get("tax_rate")),
                invoice_status=invoice_status,
                payment_status=payment_status,
                delivery_status=delivery_status,
                accounting_status=accounting_status,
                status="异常" if anomalies else "待补证据",
                anomalies=anomalies,
                evidence={"note": note, "mapped_fields": sorted(raw)},
                acceptance_status="已验收" if accepted else ("已交付待验收" if delivery_status == "已交付待确认" else "待验收"),
            ))
    workbook.close()
    return records


def procurement_payload(records: Iterable[PurchaseRecord | dict]) -> dict:
    rows = [asdict(record) if isinstance(record, PurchaseRecord) else dict(record) for record in records]
    ordered = sum(row.get("ordered_amount") or 0 for row in rows)
    accepted = sum(row.get("accepted_amount") or 0 for row in rows)
    invoiced = sum(row.get("invoice_amount") or 0 for row in rows)
    paid = sum(row.get("paid_amount") or 0 for row in rows)
    payment_eligible = sum(row.get("payment_eligible_amount") or 0 for row in rows)
    return {
        "records": rows,
        "summary": {
            "record_count": len(rows),
            "file_count": len({row.get("source_file") for row in rows if row.get("source_file")}),
            "po_count": len({row.get("po_number") for row in rows if row.get("po_number")}),
            "ordered_amount": round(ordered, 2),
            "accepted_amount": round(accepted, 2),
            "invoice_amount": round(invoiced, 2),
            "paid_amount": round(paid, 2),
            "payment_eligible_amount": round(payment_eligible, 2),
            "uninvoiced_exposure": round(max(0, accepted - invoiced), 2),
            "unpaid_exposure": round(max(0, invoiced - paid), 2),
            "pending_acceptance_count": sum(
                (row.get("acceptance_status") or ("已验收" if row.get("accepted_amount") else "待验收"))
                not in {"已验收", "部分验收"} for row in rows
            ),
            "exception_count": sum(bool(row.get("anomalies")) for row in rows),
        },
    }


def apply_acceptance_decision(
    purchase: dict, decision: str, actor: str, *, accepted_amount: float | None = None,
    evidence: Iterable[str] = (), note: str = "", period: str = "",
) -> dict:
    """Apply an auditable business acceptance decision to one purchase record.

    ``accepted_amount`` is the cumulative accepted amount after this decision, not
    the incremental amount for the current event.
    """
    if decision not in {"全部验收", "部分验收", "退回整改"}:
        raise ValueError("验收决定必须是全部验收、部分验收或退回整改")
    actor = str(actor or "").strip()[:100]
    if not actor:
        raise ValueError("请填写业务验收人")
    evidence_items = [str(item).strip()[:500] for item in evidence if str(item).strip()]
    ordered = _number(purchase.get("ordered_amount"))
    invoiced = _number(purchase.get("invoice_amount")) or 0.0
    paid = _number(purchase.get("paid_amount")) or 0.0
    current = _number(purchase.get("accepted_amount")) or 0.0
    if ordered is None or ordered <= 0:
        raise ValueError("采购订单金额缺失，不能完成验收")
    if decision in {"全部验收", "部分验收"} and not evidence_items:
        raise ValueError("确认验收前至少需要一项交付或验收证据")
    if decision == "全部验收":
        target = ordered
        status, delivery = "已验收", "已验收"
    elif decision == "部分验收":
        target = _number(accepted_amount)
        if target is None or target <= 0 or target >= ordered:
            raise ValueError("部分验收累计金额必须大于0且小于订单金额")
        status, delivery = "部分验收", "部分验收"
    else:
        if len(str(note or "").strip()) < 4:
            raise ValueError("退回整改时请填写具体原因")
        target = current
        status = "部分验收，剩余退回整改" if current > 0 else "已退回整改"
        delivery = status
    if target > ordered + max(0.02, ordered * 0.001):
        raise ValueError("验收金额不能超过订单金额")
    if target + 0.01 < invoiced:
        raise ValueError("验收金额不能低于已开票金额，请先处理发票差异")
    if target + 0.01 < paid:
        raise ValueError("验收金额不能低于已付款金额，请先处理付款差异")

    now = datetime.now(timezone.utc).isoformat()
    event = {
        "decision": decision, "actor": actor, "period": period,
        "accepted_amount_before": round(current, 2), "accepted_amount_after": round(target, 2),
        "evidence": evidence_items, "note": str(note or "").strip()[:1000], "timestamp": now,
    }
    updated = dict(purchase)
    updated.update({
        "accepted_amount": round(target, 2), "acceptance_status": status,
        "delivery_status": delivery, "accepted_by": actor, "accepted_at": now,
        "acceptance_note": event["note"], "acceptance_evidence": evidence_items,
        "acceptance_history": [*(purchase.get("acceptance_history") or []), event],
    })
    if target > 0:
        updated["accounting_status"] = "可生成应付凭证" if invoiced else "可生成采购暂估凭证"
    else:
        updated["accounting_status"] = "验收未通过，阻止付款与入账"

    anomalies = [
        item for item in (purchase.get("anomalies") or [])
        if not any(token in item for token in ("缺少验收", "暂估", "开票金额超过验收", "付款金额超过"))
    ]
    if status in {"已退回整改", "部分验收，剩余退回整改"}:
        anomalies.append("交付已退回整改，剩余部分不得付款或入账")
    elif target < ordered:
        anomalies.append(f"部分验收：尚有 {ordered - target:,.2f} 未验收")
    if target > invoiced:
        anomalies.append("疑似已发生未开票：月结需判断暂估")
    if invoiced > target:
        anomalies.append("开票金额超过验收金额")
    if paid > min(invoiced, target):
        anomalies.append("付款金额超过当前可付款金额")
    updated["anomalies"] = anomalies
    updated["status"] = "异常" if anomalies else "验收完成"
    updated["workflow"] = {
        "can_request_invoice": target > invoiced and status not in {"已退回整改"},
        "can_pay": invoiced > 0 and target + 0.01 >= invoiced and paid + 0.01 < invoiced
                   and "退回整改" not in status,
        "accrual_candidate": target > invoiced,
        "remaining_to_accept": round(max(0, ordered - target), 2),
    }
    return updated
