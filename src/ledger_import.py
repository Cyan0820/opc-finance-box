from __future__ import annotations

import csv
import hashlib
import math
import re
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .accounting_import import validate_trial_balance_lines


class LedgerImportError(ValueError):
    """Raised when an external general-ledger export violates the read-only contract."""


ALIASES = {
    "journal_id": ("凭证号", "记账凭证号", "journal id", "journal number", "voucher number"),
    "line_number": ("分录行号", "行号", "line number", "journal line number"),
    "posting_date": ("记账日期", "过账日期", "posting date", "journal date"),
    "period": ("期间", "会计期间", "账期", "period"),
    "currency": ("币种", "本位币", "currency", "functional currency"),
    "account_code": ("科目编码", "科目代码", "account code", "account number"),
    "account_name": ("科目名称", "account name", "account description"),
    "debit": ("借方金额", "借方发生额", "debit", "debit amount"),
    "credit": ("贷方金额", "贷方发生额", "credit", "credit amount"),
    "description": ("摘要", "分录摘要", "description", "memo"),
    "source_document_id": ("原始单据号", "单据号", "source document id", "document number"),
    "project_code": ("项目编码", "项目", "project code", "project"),
    "department": ("部门", "department", "cost center"),
}


STATEMENT_GROUPS = frozenset({"assets", "liabilities", "equity", "revenue", "expenses"})


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).replace("\n", " ").strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", _text(value).casefold())


def _field(value: Any) -> str | None:
    clean = _slug(value)
    matches: list[tuple[int, str]] = []
    for field, aliases in ALIASES.items():
        for alias in aliases:
            token = _slug(alias)
            if token and (clean == token or token in clean):
                matches.append((len(token), field))
    return max(matches, default=(0, None))[1]


def _number(value: Any, field: str) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        raise LedgerImportError(f"{field} must be numeric")
    try:
        number = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError) as exc:
        raise LedgerImportError(f"{field} must be numeric") from exc
    if not math.isfinite(number) or number < 0:
        raise LedgerImportError(f"{field} must be a finite non-negative number")
    return round(number, 2)


def _posting_date(value: Any) -> str:
    text = _text(value)[:10]
    if not re.fullmatch(r"20\d{2}-(0[1-9]|1[0-2])-(0[1-9]|[12]\d|3[01])", text):
        raise LedgerImportError("posting_date must use YYYY-MM-DD")
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise LedgerImportError("posting_date must be a valid calendar date") from exc


def _period(value: Any) -> str:
    text = _text(value)[:7].replace("/", "-")
    if not re.fullmatch(r"20\d{2}-(0[1-9]|1[0-2])", text):
        raise LedgerImportError("period must use YYYY-MM")
    return text


def _currency(value: Any) -> str:
    currency = _text(value).upper()
    if not re.fullmatch(r"[A-Z]{3}", currency):
        raise LedgerImportError("currency must be a three-letter code")
    return currency


@dataclass(frozen=True)
class GeneralLedgerLine:
    journal_line_id: str
    entity_id: str
    journal_id: str
    line_number: str
    posting_date: str
    period: str
    currency: str
    account_code: str
    account_name: str
    debit: float
    credit: float
    description: str
    source_document_id: str
    project_code: str
    department: str
    evidence: dict[str, Any]


def _record(
    raw: dict[str, Any], *, entity_id: str, default_period: str,
    default_currency: str, source_file: str, source_sheet: str,
    source_row: int, batch_id: str,
) -> GeneralLedgerLine:
    journal_id = _text(raw.get("journal_id"))
    line_number = _text(raw.get("line_number"))
    account_code = _text(raw.get("account_code"))
    account_name = _text(raw.get("account_name"))
    if not journal_id or not line_number:
        raise LedgerImportError("journal_id and line_number are required")
    if not account_code or not account_name:
        raise LedgerImportError("account_code and account_name are required")
    posting_date = _posting_date(raw.get("posting_date"))
    period = _period(raw.get("period") or default_period or posting_date[:7])
    if period != posting_date[:7]:
        raise LedgerImportError("period must match posting_date month")
    currency = _currency(raw.get("currency") or default_currency)
    debit = _number(raw.get("debit"), "debit")
    credit = _number(raw.get("credit"), "credit")
    if (debit > 0) == (credit > 0):
        raise LedgerImportError("exactly one of debit or credit must be positive")
    journal_line_id = hashlib.sha256(
        f"{entity_id}|{period}|{currency}|{journal_id}|{line_number}".encode()
    ).hexdigest()[:16]
    return GeneralLedgerLine(
        journal_line_id=journal_line_id,
        entity_id=entity_id,
        journal_id=journal_id,
        line_number=line_number,
        posting_date=posting_date,
        period=period,
        currency=currency,
        account_code=account_code,
        account_name=account_name,
        debit=debit,
        credit=credit,
        description=_text(raw.get("description")),
        source_document_id=_text(raw.get("source_document_id")),
        project_code=_text(raw.get("project_code")),
        department=_text(raw.get("department")),
        evidence={
            "source_file": source_file,
            "source_sheet": source_sheet,
            "source_row": source_row,
            "batch_id": batch_id,
        },
    )


def _header_mapping(row: Iterable[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, value in enumerate(row):
        field = _field(value)
        if field and field not in mapping.values():
            mapping[index] = field
    return mapping


def _consume_rows(
    rows: Iterable[tuple[int, tuple[Any, ...]]], *, entity_id: str,
    default_period: str, default_currency: str, source_file: str,
    source_sheet: str, batch_id: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], bool]:
    mapping: dict[int, str] = {}
    records: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    required = {
        "journal_id", "line_number", "posting_date", "account_code",
        "account_name", "debit", "credit",
    }
    for row_number, values in rows:
        candidate = _header_mapping(values)
        if required <= set(candidate.values()):
            mapping = candidate
            continue
        if not mapping or not any(_text(value) for value in values):
            continue
        raw = {
            field: values[index] if index < len(values) else None
            for index, field in mapping.items()
        }
        try:
            records.append(asdict(_record(
                raw, entity_id=entity_id, default_period=default_period,
                default_currency=default_currency, source_file=source_file,
                source_sheet=source_sheet, source_row=row_number, batch_id=batch_id,
            )))
        except LedgerImportError as exc:
            rejected.append({
                "dataset_type": "finance.general_ledger_lines",
                "row": row_number,
                "source_sheet": source_sheet,
                "reason": str(exc),
            })
    return records, rejected, bool(mapping)


def parse_general_ledger_file(
    path: str | Path, *, entity_id: str, default_period: str = "",
    default_currency: str = "",
) -> dict[str, Any]:
    path = Path(path)
    if path.suffix.casefold() not in {".csv", ".xlsx"}:
        raise LedgerImportError("general ledger must be a .csv or .xlsx file")
    if not entity_id:
        raise LedgerImportError("entity_id is required")
    batch_id = hashlib.sha256(path.read_bytes()).hexdigest()[:16]
    records: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    any_header = False
    if path.suffix.casefold() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [(index, tuple(row)) for index, row in enumerate(csv.reader(handle), 1)]
        accepted, errors, found = _consume_rows(
            rows, entity_id=entity_id, default_period=default_period,
            default_currency=default_currency, source_file=path.name,
            source_sheet="CSV", batch_id=batch_id,
        )
        records.extend(accepted)
        rejected.extend(errors)
        any_header = found
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = (
                    (index, tuple(row))
                    for index, row in enumerate(sheet.iter_rows(values_only=True), 1)
                )
                accepted, errors, found = _consume_rows(
                    rows, entity_id=entity_id, default_period=default_period,
                    default_currency=default_currency, source_file=path.name,
                    source_sheet=sheet.title, batch_id=batch_id,
                )
                records.extend(accepted)
                rejected.extend(errors)
                any_header |= found
        finally:
            workbook.close()
    if not any_header:
        rejected.append({
            "dataset_type": "finance.general_ledger_lines",
            "row": 0,
            "source_sheet": "workbook",
            "reason": (
                "general ledger header requires journal id, line number, posting date, "
                "account code/name and debit/credit"
            ),
        })
    return {
        "batch_id": batch_id,
        "source": {
            "kind": "accounting_general_ledger_export_file",
            "name": path.name,
            "network_access_performed": False,
            "raw_source_rows_retained": False,
        },
        "records": records,
        "rejected_rows": rejected,
    }


def validate_general_ledger_lines(
    rows: Iterable[dict[str, Any]], *, entity_id: str,
) -> dict[str, Any]:
    lines = list(rows)
    issues: list[dict[str, Any]] = []
    if any(str(row.get("entity_id") or "") != entity_id for row in lines):
        issues.append({"severity": "blocking", "type": "cross_entity_general_ledger"})
    line_ids = [str(row.get("journal_line_id") or "") for row in lines]
    duplicate_line_ids = sorted(
        line_id for line_id, count in Counter(line_ids).items() if line_id and count > 1
    )
    if duplicate_line_ids:
        issues.append({
            "severity": "blocking", "type": "duplicate_general_ledger_line",
            "journal_line_ids": duplicate_line_ids,
        })
    journals: list[dict[str, Any]] = []
    journal_keys = sorted({
        (
            str(row.get("entity_id") or ""), str(row.get("period") or ""),
            str(row.get("currency") or ""), str(row.get("journal_id") or ""),
        )
        for row in lines
    })
    for key in journal_keys:
        scoped = [
            row for row in lines
            if (
                str(row.get("entity_id") or ""), str(row.get("period") or ""),
                str(row.get("currency") or ""), str(row.get("journal_id") or ""),
            ) == key
        ]
        debit = round(sum(float(row.get("debit") or 0) for row in scoped), 2)
        credit = round(sum(float(row.get("credit") or 0) for row in scoped), 2)
        difference = round(debit - credit, 2)
        if abs(difference) >= 0.01:
            issues.append({
                "severity": "blocking",
                "type": "unbalanced_journal",
                "entity_id": key[0],
                "period": key[1],
                "currency": key[2],
                "journal_id": key[3],
                "difference": difference,
            })
        journals.append({
            "entity_id": key[0], "period": key[1], "currency": key[2],
            "journal_id": key[3], "line_count": len(scoped),
            "debit": debit, "credit": credit, "difference": difference,
            "balanced": abs(difference) < 0.01,
        })
    scope_summaries = []
    scope_keys = sorted({key[:3] for key in journal_keys})
    for scope in scope_keys:
        scoped = [
            row for row in lines
            if (
                str(row.get("entity_id") or ""), str(row.get("period") or ""),
                str(row.get("currency") or ""),
            ) == scope
        ]
        debit = round(sum(float(row.get("debit") or 0) for row in scoped), 2)
        credit = round(sum(float(row.get("credit") or 0) for row in scoped), 2)
        scope_summaries.append({
            "entity_id": scope[0], "period": scope[1], "currency": scope[2],
            "line_count": len(scoped),
            "journal_count": sum(1 for journal in journals if (
                journal["entity_id"], journal["period"], journal["currency"]
            ) == scope),
            "debit": debit, "credit": credit,
            "difference": round(debit - credit, 2),
            "balanced": abs(debit - credit) < 0.01,
        })
    return {
        "ready": bool(lines) and not issues,
        "entity_id": entity_id,
        "journals": journals,
        "scope_summaries": scope_summaries,
        "issues": issues,
        "candidate_only": True,
        "ledger_modified": False,
        "posting_performed": False,
        "guardrail": (
            "Balanced exported journals are source evidence only; they do not prove account "
            "mapping, completeness, authorization, posting in this Box or period close."
        ),
    }


def reconcile_ledger_to_trial_balance(
    ledger_rows: Iterable[dict[str, Any]],
    trial_balance_rows: Iterable[dict[str, Any]],
    account_mappings: Iterable[dict[str, Any]],
    *,
    entity_id: str,
    period: str,
) -> dict[str, Any]:
    ledger = list(ledger_rows)
    trial = list(trial_balance_rows)
    mappings = list(account_mappings)
    ledger_validation = validate_general_ledger_lines(ledger, entity_id=entity_id)
    trial_validation = validate_trial_balance_lines(trial, entity_id=entity_id)
    issues: list[dict[str, Any]] = [
        *ledger_validation["issues"], *trial_validation["issues"],
    ]
    if any(str(row.get("period") or "") != period for row in ledger + trial):
        issues.append({"severity": "blocking", "type": "cross_period_accounting_export"})

    mapping_by_code: dict[str, dict[str, str]] = {}
    for index, raw in enumerate(mappings, 1):
        if not isinstance(raw, dict):
            issues.append({
                "severity": "blocking", "type": "invalid_account_mapping",
                "mapping_index": index, "reason": "mapping must be an object",
            })
            continue
        mapping = {
            "account_code": _text(raw.get("account_code")),
            "source_account_name": _text(raw.get("source_account_name")),
            "statement_group": _text(raw.get("statement_group")),
            "statement_line_id": _text(raw.get("statement_line_id")),
            "statement_line_name": _text(raw.get("statement_line_name")),
        }
        if (
            not mapping["account_code"]
            or mapping["statement_group"] not in STATEMENT_GROUPS
            or not mapping["statement_line_id"]
            or not mapping["statement_line_name"]
        ):
            issues.append({
                "severity": "blocking", "type": "invalid_account_mapping",
                "mapping_index": index,
                "reason": "account code, allowed statement group and statement line id/name are required",
            })
            continue
        if mapping["account_code"] in mapping_by_code:
            issues.append({
                "severity": "blocking", "type": "duplicate_account_mapping",
                "account_code": mapping["account_code"],
            })
            continue
        mapping_by_code[mapping["account_code"]] = mapping

    ledger_accounts: dict[tuple[str, str], dict[str, Any]] = {}
    for row in ledger:
        key = (str(row.get("currency") or ""), str(row.get("account_code") or ""))
        bucket = ledger_accounts.setdefault(key, {
            "account_name": str(row.get("account_name") or ""),
            "debit": 0.0, "credit": 0.0, "line_count": 0,
        })
        bucket["debit"] += float(row.get("debit") or 0)
        bucket["credit"] += float(row.get("credit") or 0)
        bucket["line_count"] += 1

    trial_accounts = {
        (str(row.get("currency") or ""), str(row.get("account_code") or "")): row
        for row in trial
    }
    reconciliation = []
    all_account_keys = sorted(set(ledger_accounts) | set(trial_accounts))
    for currency, account_code in all_account_keys:
        ledger_account = ledger_accounts.get((currency, account_code), {})
        trial_account = trial_accounts.get((currency, account_code), {})
        ledger_debit = round(float(ledger_account.get("debit") or 0), 2)
        ledger_credit = round(float(ledger_account.get("credit") or 0), 2)
        trial_debit = round(float(trial_account.get("period_debit") or 0), 2)
        trial_credit = round(float(trial_account.get("period_credit") or 0), 2)
        debit_difference = round(ledger_debit - trial_debit, 2)
        credit_difference = round(ledger_credit - trial_credit, 2)
        matched = abs(debit_difference) < 0.01 and abs(credit_difference) < 0.01
        if not matched:
            issues.append({
                "severity": "blocking", "type": "ledger_trial_balance_mismatch",
                "currency": currency, "account_code": account_code,
                "debit_difference": debit_difference,
                "credit_difference": credit_difference,
            })
        reconciliation.append({
            "entity_id": entity_id, "period": period, "currency": currency,
            "account_code": account_code,
            "account_name": str(
                trial_account.get("account_name") or ledger_account.get("account_name") or ""
            ),
            "ledger_debit": ledger_debit, "trial_balance_period_debit": trial_debit,
            "debit_difference": debit_difference,
            "ledger_credit": ledger_credit, "trial_balance_period_credit": trial_credit,
            "credit_difference": credit_difference,
            "matched": matched,
        })

    required_trial_rows = [
        row for row in trial
        if any(abs(float(row.get(field) or 0)) >= 0.01 for field in (
            "opening_debit", "opening_credit", "period_debit", "period_credit",
            "closing_debit", "closing_credit",
        ))
    ]
    required_trial_codes = {str(row.get("account_code") or "") for row in trial}
    for account_code in sorted(set(mapping_by_code) - required_trial_codes):
        issues.append({
            "severity": "blocking", "type": "mapping_account_not_in_trial_balance",
            "account_code": account_code,
        })
    mapped_count = 0
    for row in required_trial_rows:
        code = str(row.get("account_code") or "")
        mapping = mapping_by_code.get(code)
        if mapping is None:
            issues.append({
                "severity": "blocking", "type": "unmapped_trial_balance_account",
                "currency": row.get("currency"), "account_code": code,
                "account_name": row.get("account_name"),
            })
            continue
        if (
            mapping["source_account_name"]
            and mapping["source_account_name"] != str(row.get("account_name") or "")
        ):
            issues.append({
                "severity": "blocking", "type": "account_mapping_name_mismatch",
                "account_code": code,
                "expected_name": mapping["source_account_name"],
                "actual_name": row.get("account_name"),
            })
            continue
        mapped_count += 1

    candidates = []
    currencies = sorted({str(row.get("currency") or "") for row in trial})
    for currency in currencies:
        buckets: dict[tuple[str, str, str], float] = {}
        for row in trial:
            if str(row.get("currency") or "") != currency:
                continue
            mapping = mapping_by_code.get(str(row.get("account_code") or ""))
            if mapping is None:
                continue
            group = mapping["statement_group"]
            if group in {"assets", "expenses"}:
                amount = (
                    float(row.get("closing_debit") or 0) - float(row.get("closing_credit") or 0)
                    if group == "assets"
                    else float(row.get("period_debit") or 0) - float(row.get("period_credit") or 0)
                )
            else:
                amount = (
                    float(row.get("closing_credit") or 0) - float(row.get("closing_debit") or 0)
                    if group in {"liabilities", "equity"}
                    else float(row.get("period_credit") or 0) - float(row.get("period_debit") or 0)
                )
            key = (group, mapping["statement_line_id"], mapping["statement_line_name"])
            buckets[key] = buckets.get(key, 0.0) + amount
        lines = [{
            "statement_group": key[0], "statement_line_id": key[1],
            "statement_line_name": key[2], "amount": round(amount, 2),
        } for key, amount in sorted(buckets.items())]
        totals = {
            group: round(sum(line["amount"] for line in lines if line["statement_group"] == group), 2)
            for group in sorted(STATEMENT_GROUPS)
        }
        current_profit = round(totals["revenue"] - totals["expenses"], 2)
        balance_difference = round(
            totals["assets"] - totals["liabilities"] - totals["equity"] - current_profit,
            2,
        )
        candidates.append({
            "entity_id": entity_id, "period": period, "currency": currency,
            "lines": lines,
            "balance_sheet": {
                "assets": totals["assets"], "liabilities": totals["liabilities"],
                "equity_before_current_profit": totals["equity"],
                "current_profit": current_profit,
                "liabilities_equity_and_current_profit": round(
                    totals["liabilities"] + totals["equity"] + current_profit, 2
                ),
                "difference": balance_difference,
                "balanced": abs(balance_difference) < 0.01,
            },
            "income_statement": {
                "revenue": totals["revenue"], "expenses": totals["expenses"],
                "profit_before_tax_candidate": current_profit,
            },
            "candidate_only": True,
        })
        if abs(balance_difference) >= 0.01:
            issues.append({
                "severity": "blocking", "type": "statement_candidate_unbalanced",
                "currency": currency, "difference": balance_difference,
            })

    unique_issues = []
    seen_issues: set[str] = set()
    for issue in issues:
        fingerprint = repr(sorted(issue.items()))
        if fingerprint not in seen_issues:
            seen_issues.add(fingerprint)
            unique_issues.append(issue)
    denominator = len(required_trial_rows)
    return {
        "ready": bool(ledger) and bool(trial) and not unique_issues,
        "entity_id": entity_id,
        "period": period,
        "ledger_validation": ledger_validation,
        "trial_balance_validation": trial_validation,
        "account_reconciliation": reconciliation,
        "mapping_coverage": {
            "required_account_count": denominator,
            "mapped_account_count": mapped_count,
            "coverage_percent": round((mapped_count / denominator * 100), 2) if denominator else 0.0,
            "mapping_is_explicit": True,
            "mapping_inferred_from_account_name": False,
        },
        "financial_statement_candidates": candidates,
        "issues": unique_issues,
        "candidate_only": True,
        "ledger_modified": False,
        "opening_balances_modified": False,
        "posting_performed": False,
        "period_close_performed": False,
        "external_filing_performed": False,
        "guardrail": (
            "A fully reconciled candidate still requires accounting-policy, mapping, completeness "
            "and close review. It is not a posted ledger, approved financial statement or filing."
        ),
    }
