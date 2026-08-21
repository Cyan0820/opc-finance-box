from __future__ import annotations

import hashlib
import math
import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).replace("\n", " ").strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", _text(value).lower())


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    value = _text(value).replace(",", "").replace("¥", "").replace("$", "")
    match = re.search(r"-?\d+(?:\.\d+)?", value)
    return float(match.group()) if match else None


ALIASES = {
    "transaction_date": ("交易日期", "记账日期", "入账日期", "交易时间", "value date", "transaction date"),
    "transaction_id": ("交易流水号", "银行流水号", "交易编号", "reference", "transaction id"),
    "counterparty": ("对方户名", "对方名称", "交易对手", "收款人", "付款人", "counterparty", "beneficiary"),
    "counterparty_account": ("对方账号", "收款账号", "付款账号", "counterparty account"),
    "summary": ("摘要", "用途", "附言", "交易备注", "remark", "description", "narrative"),
    "debit": ("支出金额", "借方发生额", "付款金额", "debit"),
    "credit": ("收入金额", "贷方发生额", "收款金额", "credit"),
    "amount": ("交易金额", "发生额", "金额", "amount"),
    "direction": ("收支方向", "借贷标志", "交易方向", "direction"),
    "currency": ("币种", "交易币种", "currency"),
    "balance": ("余额", "账户余额", "balance"),
    "account": ("本方账号", "账户", "银行账号", "account number"),
}


def _field(value: Any) -> str | None:
    clean = _slug(value)
    candidates = []
    for field_name, aliases in ALIASES.items():
        for alias in aliases:
            alias_clean = _slug(alias)
            if alias_clean and (clean == alias_clean or alias_clean in clean):
                candidates.append((len(alias_clean), field_name))
    return max(candidates, default=(0, None))[1]


def _mask_account(value: Any) -> str:
    text = re.sub(r"\s+", "", _text(value))
    if len(text) <= 8:
        return text
    return f"{text[:4]}****{text[-4:]}"


@dataclass
class BankTransaction:
    id: str
    source_file: str
    source_sheet: str
    source_row: int
    transaction_date: str
    transaction_id: str
    account_masked: str
    counterparty: str
    counterparty_account_masked: str
    summary: str
    direction: str
    currency: str
    amount: float
    balance: float | None
    status: str = "待认领"
    suggested_match: dict[str, Any] | None = None
    anomalies: list[str] = field(default_factory=list)


def parse_bank_workbook(path: str | Path) -> list[BankTransaction]:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    transactions = []
    for sheet in workbook.worksheets:
        mapping = {}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 100000, 100000), values_only=True), 1
        ):
            candidate = {}
            for index, value in enumerate(row):
                field_name = _field(value)
                if field_name and field_name not in candidate.values():
                    candidate[index] = field_name
            if len(candidate) >= 4 and (
                "amount" in candidate.values()
                or "debit" in candidate.values()
                or "credit" in candidate.values()
            ):
                mapping = candidate
                continue
            if not mapping:
                continue
            raw = {name: row[index] if index < len(row) else None for index, name in mapping.items()}
            debit = _number(raw.get("debit")) or 0.0
            credit = _number(raw.get("credit")) or 0.0
            amount = _number(raw.get("amount"))
            direction_text = _text(raw.get("direction")).lower()
            if credit:
                direction, amount = "收入", abs(credit)
            elif debit:
                direction, amount = "支出", abs(debit)
            elif amount is not None:
                if any(word in direction_text for word in ("支", "借", "debit", "out")) or amount < 0:
                    direction = "支出"
                elif any(word in direction_text for word in ("收", "贷", "credit", "in")):
                    direction = "收入"
                else:
                    direction = "待确认"
                amount = abs(amount)
            else:
                continue
            tx_date = _text(raw.get("transaction_date"))
            transaction_id = _text(raw.get("transaction_id"))
            counterparty = _text(raw.get("counterparty"))
            summary = _text(raw.get("summary"))
            if not any((tx_date, transaction_id, counterparty, summary)):
                continue
            row_key = f"{path.name}|{sheet.title}|{row_number}|{transaction_id}|{amount}"
            anomalies = []
            if direction == "待确认":
                anomalies.append("无法识别收支方向")
            transactions.append(BankTransaction(
                id=hashlib.sha1(row_key.encode("utf-8")).hexdigest()[:12],
                source_file=path.name,
                source_sheet=sheet.title,
                source_row=row_number,
                transaction_date=tx_date,
                transaction_id=transaction_id,
                account_masked=_mask_account(raw.get("account")),
                counterparty=counterparty,
                counterparty_account_masked=_mask_account(raw.get("counterparty_account")),
                summary=summary,
                direction=direction,
                currency=_text(raw.get("currency")).upper() or "CNY",
                amount=round(amount, 2),
                balance=_number(raw.get("balance")),
                anomalies=anomalies,
            ))
    workbook.close()
    return transactions


def _similarity(left: str, right: str) -> float:
    left_slug, right_slug = _slug(left), _slug(right)
    if not left_slug or not right_slug:
        return 0.0
    if left_slug in right_slug or right_slug in left_slug:
        return 1.0
    return SequenceMatcher(None, left_slug, right_slug).ratio()


def suggest_matches(
    transactions: Iterable[dict | BankTransaction],
    settlements: Iterable[dict],
    purchases: Iterable[dict],
    allocations: Iterable[dict] = (),
) -> list[dict]:
    settlement_rows = list(settlements)
    purchase_rows = list(purchases)
    active_allocations = [row for row in allocations if row.get("status") not in {"已撤销", "已退回"}]

    def same_entity(left: dict, right: dict) -> bool:
        return _text(left.get("entity_id")) == _text(right.get("entity_id"))

    def allocated_on_target(target: dict, target_type: str) -> float:
        return sum(float(item.get("amount") or 0) for item in active_allocations if (
            item.get("target_type") == target_type
            and item.get("target_id") == target.get("id")
            and same_entity(item, target)
        ))

    def allocated_on_transaction(transaction: dict) -> float:
        return sum(float(item.get("amount") or 0) for item in active_allocations if (
            item.get("transaction_id") == transaction.get("id") and same_entity(item, transaction)
        ))

    output = []
    for transaction in transactions:
        row = asdict(transaction) if isinstance(transaction, BankTransaction) else dict(transaction)
        transaction_remaining = round(max(0.0, float(row.get("amount") or 0) - allocated_on_transaction(row)), 2)
        row["allocated_amount"] = round(float(row.get("amount") or 0) - transaction_remaining, 2)
        row["remaining_amount"] = transaction_remaining
        candidates = []
        if transaction_remaining <= 0:
            row["suggested_match"] = None
            row["status"] = "已核销"
            output.append(row)
            continue
        if row.get("direction") == "收入":
            for settlement in settlement_rows:
                if settlement.get("release_status") not in {None, "", "released"} or not same_entity(row, settlement):
                    continue
                original_expected = float(settlement.get("net_receivable") or 0)
                allocated_before = allocated_on_target(settlement, "receivable")
                expected = round(max(0.0, original_expected - allocated_before), 2)
                if expected <= 0 or (settlement.get("currency") or "CNY").upper() != _text(row.get("currency")).upper():
                    continue
                amount_score = min(transaction_remaining, expected) / max(transaction_remaining, expected)
                name_score = max(
                    _similarity(row.get("counterparty") or "", settlement.get("channel") or ""),
                    _similarity(row.get("summary") or "", settlement.get("game") or ""),
                )
                score = amount_score * 0.65 + name_score * 0.35
                candidates.append({
                    "type": "应收到账", "target_id": settlement.get("id"),
                    "entity_id": settlement.get("entity_id") or "",
                    "target": f"{settlement.get('game')} / {settlement.get('channel')} / {settlement.get('period')}",
                    "expected_amount": expected, "original_expected_amount": original_expected,
                    "allocated_before": round(allocated_before, 2),
                    "suggested_allocation_amount": round(min(transaction_remaining, expected), 2),
                    "difference": round(transaction_remaining - expected, 2),
                    "score": round(score, 4),
                })
        elif row.get("direction") == "支出":
            for purchase in purchase_rows:
                if not same_entity(row, purchase):
                    continue
                original_expected = float(purchase.get("invoice_amount") or purchase.get("accepted_amount") or purchase.get("ordered_amount") or 0)
                allocated_before = allocated_on_target(purchase, "payable")
                expected = round(max(0.0, original_expected - allocated_before), 2)
                if expected <= 0 or (purchase.get("currency") or "CNY").upper() != _text(row.get("currency")).upper():
                    continue
                amount_score = min(transaction_remaining, expected) / max(transaction_remaining, expected)
                name_score = _similarity(row.get("counterparty") or "", purchase.get("vendor") or "")
                score = amount_score * 0.65 + name_score * 0.35
                candidates.append({
                    "type": "应付付款", "target_id": purchase.get("id"),
                    "entity_id": purchase.get("entity_id") or "",
                    "target": f"{purchase.get('vendor')} / {purchase.get('item')} / {purchase.get('po_number')}",
                    "expected_amount": expected, "original_expected_amount": original_expected,
                    "allocated_before": round(allocated_before, 2),
                    "suggested_allocation_amount": round(min(transaction_remaining, expected), 2),
                    "difference": round(transaction_remaining - expected, 2),
                    "score": round(score, 4),
                })
        best = max(candidates, key=lambda candidate: candidate["score"], default=None)
        if best and best["score"] >= 0.72:
            confidence = best["score"]
            best["confidence"] = confidence
            best["recommendation"] = (
                "金额和交易对手高度吻合；请确认用途后核销，系统不会自动记账。"
                if confidence >= 0.92 and abs(best["difference"]) < 0.01
                else "存在较强匹配，建议按当前剩余应收/应付确认本次核销金额。"
            )
            row["suggested_match"] = best
            row["status"] = "高置信匹配" if confidence >= 0.92 and abs(best["difference"]) < 0.01 else "待确认匹配"
        else:
            row["suggested_match"] = None
            row["status"] = "待认领"
        output.append(row)
    return output


def banking_payload(transactions: Iterable[dict]) -> dict:
    rows = list(transactions)
    receipts = sum(float(row.get("amount") or 0) for row in rows if row.get("direction") == "收入")
    payments = sum(float(row.get("amount") or 0) for row in rows if row.get("direction") == "支出")
    return {
        "transactions": rows,
        "summary": {
            "count": len(rows),
            "receipt_amount": round(receipts, 2),
            "payment_amount": round(payments, 2),
            "high_confidence_count": sum(row.get("status") == "高置信匹配" for row in rows),
            "pending_count": sum(row.get("status") != "高置信匹配" for row in rows),
            "currencies": sorted({row.get("currency") or "未知" for row in rows}),
        },
    }
