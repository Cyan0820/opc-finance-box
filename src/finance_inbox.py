from __future__ import annotations

import hashlib
import csv
import io
import json
import mimetypes
import os
import re
import tempfile
import threading
from dataclasses import asdict, is_dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .banking import banking_payload, parse_bank_workbook, suggest_matches
from .general_ledger import opening_balance_payload, parse_opening_balance_workbook
from .invoices import invoice_payload, match_invoices_to_purchases, parse_invoice_workbook
from .payroll import parse_payroll_workbook, payroll_payload
from .planning import parse_plan_workbook, planning_payload
from .procurement import parse_purchase_workbook, procurement_payload
from .reconcile import dashboard_payload, parse_files
from .document_extraction import (
    bank_records_from_extraction, extract_bank_statement_rows,
    extract_document_text, extract_invoice_fields, invoice_record_from_extraction,
)


SUPPORTED_EXTENSIONS = {".xlsx", ".csv", ".pdf", ".png", ".jpg", ".jpeg", ".webp"}
PERIOD_PATTERN = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")

DOCUMENT_TYPES = {
    "finance_package": {"label": "综合财务资料包", "dataset": None, "capability": "多台账结构化识别"},
    "settlement": {"label": "平台/渠道结算单", "dataset": "settlements", "capability": "结构化识别"},
    "settlement_reconciliation_evidence": {"label": "收入结算核对底稿", "dataset": None, "capability": "证据归档"},
    "bank_statement": {"label": "银行流水/对账单", "dataset": "bank_transactions", "capability": "结构化识别"},
    "bank_statement_document": {"label": "银行 PDF/图片对账单", "dataset": "bank_transactions", "capability": "等待OCR"},
    "invoice_register": {"label": "发票清单", "dataset": "invoices", "capability": "结构化识别"},
    "invoice_document": {"label": "单张发票", "dataset": "invoices", "capability": "等待OCR"},
    "purchase": {"label": "采购/费用/验收资料", "dataset": "purchases", "capability": "结构化识别"},
    "payroll": {"label": "工资薪酬资料", "dataset": "payroll_rows", "capability": "结构化识别"},
    "planning": {"label": "预算与预测", "dataset": "plan_lines", "capability": "结构化识别"},
    "opening_balance": {"label": "期初余额/科目余额表", "dataset": "opening_balances", "capability": "结构化识别"},
    "contract_commercial": {"label": "合同商业条款", "dataset": None, "capability": "等待文字提取"},
    "acceptance_evidence": {"label": "交付/验收证据", "dataset": None, "capability": "证据归档"},
    "unknown": {"label": "待识别资料", "dataset": None, "capability": "人工确认"},
}

HEADER_RULES = {
    "settlement_reconciliation_evidence": {
        "匹配-苹果账单": 5, "剔除退款后差异": 5, "经分数据": 4,
        "付费金额": 3, "核对": 2, "差异": 2,
    },
    "settlement": {
        "结算周期": 3, "游戏名称": 3, "分成比例": 3, "结算金额": 3, "甲方结算金额": 4,
        "buyercurrency": 3, "merchantcurrency": 3, "渠道含税流水": 3, "平台": 1, "渠道": 1,
    },
    "bank_statement": {
        "交易日期": 2, "交易流水号": 3, "银行流水号": 3, "对方户名": 3, "借方发生额": 3,
        "贷方发生额": 3, "账户余额": 2, "本方账号": 2, "counterparty": 3, "value date": 2,
    },
    "invoice_register": {
        "发票号码": 4, "发票代码": 3, "开票日期": 2, "销售方名称": 3, "购买方名称": 2,
        "纳税人识别号": 2, "税额": 2, "价税合计": 3, "查验状态": 3, "数电票号码": 4,
    },
    "purchase": {
        "po编号": 4, "采购订单": 3, "供应商": 3, "采购内容": 2, "订单金额": 2,
        "验收金额": 4, "付款金额": 2, "发票金额": 2, "采购日期": 2,
    },
    "payroll": {
        "工号": 3, "员工编号": 3, "应发工资": 4, "实发工资": 4, "个人所得税": 3,
        "个人社保": 3, "公积金": 2, "专项附加扣除": 3,
    },
    "planning": {
        "预算月": 3, "预算金额": 4, "预测金额": 4, "情景": 3, "scenario": 3,
        "实现概率": 3, "已承诺": 2, "收支方向": 2,
    },
    "opening_balance": {
        "科目编码": 4, "科目名称": 3, "期初借方": 4, "期初贷方": 4,
        "期末借方": 3, "期末贷方": 3, "account code": 3,
    },
}

FILENAME_RULES = {
    "settlement_reconciliation_evidence": ("经分数据核对", "账单数据核对", "收入核对底稿"),
    "settlement": ("结算", "对账单", "statement", "apple", "google", "appstore", "play"),
    "bank_statement": ("银行", "流水", "回单", "bank"),
    "bank_statement_document": ("银行", "对账单", "流水", "回单", "bank", "statement"),
    "invoice_register": ("发票清单", "发票台账", "数电发票"),
    "invoice_document": ("发票", "invoice"),
    "purchase": ("采购", "订单", "验收", "报销", "费用", "purchase", "po"),
    "payroll": ("工资", "薪酬", "个税", "社保", "payroll", "salary"),
    "planning": ("预算", "预测", "forecast", "budget"),
    "opening_balance": ("科目余额", "期初余额", "trialbalance", "openingbalance"),
    "contract_commercial": ("合同", "协议", "contract", "agreement"),
    "acceptance_evidence": ("交付", "验收", "成果", "acceptance", "delivery"),
}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _atomic_write(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
        os.replace(temp_name, path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)


def _slug(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").lower())


def suggest_entity_scope(filename: str, entities: list[dict] | None = None) -> dict:
    """Return a conservative document-level legal-entity routing suggestion.

    A currency, country word or overseas channel is not enough evidence: a Chinese
    company can sell overseas and an SG publisher can hold CNY documents.  We only
    auto-confirm a single-entity Box, or suggest an entity when the configured ID or
    full legal name is present in the filename.  Suggestions still require a person
    to confirm before ledger import.
    """
    entities = [item for item in (entities or []) if item.get("id") and item.get("name")]
    if not entities:
        return {
            "entity_id": "", "entity_name": "", "status": "unassigned",
            "source": "not_configured", "confidence": 0.0,
            "reason": "尚未提供可用的法律主体配置。",
        }
    if len(entities) == 1:
        item = entities[0]
        return {
            "entity_id": str(item["id"]), "entity_name": str(item["name"]),
            "status": "confirmed", "source": "single_entity_box", "confidence": 1.0,
            "reason": "当前 Box 只配置一个法律主体。",
        }
    filename_slug = _slug(Path(str(filename or "")).stem)
    matches = []
    for item in entities:
        entity_id = _slug(item["id"])
        legal_name = _slug(item["name"])
        if (entity_id and entity_id in filename_slug) or (legal_name and legal_name in filename_slug):
            matches.append(item)
    if len(matches) == 1:
        item = matches[0]
        return {
            "entity_id": str(item["id"]), "entity_name": str(item["name"]),
            "status": "suggested", "source": "filename_exact_match", "confidence": 0.9,
            "reason": "文件名包含已配置的主体 ID 或完整名称，入账前仍需确认。",
        }
    return {
        "entity_id": "", "entity_name": "", "status": "unassigned",
        "source": "insufficient_evidence", "confidence": 0.0,
        "reason": "不能仅凭币种、国家或渠道推断法律主体，请人工选择。",
    }


def _period(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    match = re.search(r"(20\d{2})[-/.年]?(1[0-2]|0?[1-9])(?!\d)", str(value or ""))
    return f"{match.group(1)}-{int(match.group(2)):02d}" if match else ""


def _xlsx_evidence(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    texts, sheet_names, periods = [], [], set()
    row_count = 0
    try:
        for sheet in workbook.worksheets[:30]:
            sheet_names.append(sheet.title)
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 40, 40), max_col=60, values_only=True):
                row_count += 1
                for value in row:
                    if value in (None, ""):
                        continue
                    text = str(value)[:200]
                    texts.append(text)
                    if isinstance(value, (date, datetime)) or (
                        isinstance(value, str) and not value.lstrip().startswith("=")
                    ):
                        found = _period(value)
                    else:
                        found = ""
                    if found:
                        periods.add(found)
    finally:
        workbook.close()
    return {
        "text": "\n".join(texts),
        "sheets": sheet_names,
        "periods": sorted(periods),
        "sampled_rows": row_count,
    }


def _read_csv(path: Path) -> tuple[list[list[str]], str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "gb18030", "big5"):
        try:
            text = raw.decode(encoding)
            return list(csv.reader(io.StringIO(text))), encoding
        except UnicodeDecodeError:
            continue
    raise ValueError("CSV 编码无法识别，请另存为 UTF-8、GB18030 或 Big5")


def _csv_evidence(path: Path) -> dict:
    rows, encoding = _read_csv(path)
    sample = rows[:40]
    texts = [str(value)[:200] for row in sample for value in row if str(value).strip()]
    periods = sorted({found for value in texts if (found := _period(value))})
    return {
        "text": "\n".join(texts), "sheets": [f"CSV ({encoding})"],
        "periods": periods, "sampled_rows": len(sample),
    }


def _csv_to_xlsx(path: Path, target: Path) -> None:
    from openpyxl import Workbook
    rows, _ = _read_csv(path)
    book = Workbook(write_only=True)
    sheet = book.create_sheet("CSV明细")
    for row in rows:
        sheet.append(row)
    book.save(target)


def classify_document(path: str | Path, original_filename: str | None = None) -> dict:
    path = Path(path)
    filename = original_filename or path.name
    extension = path.suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        return {
            "document_type": "unknown", "label": DOCUMENT_TYPES["unknown"]["label"],
            "confidence": 0, "alternatives": [], "periods": [], "signals": ["文件格式尚不支持"],
            "capability": "不支持", "destination": None,
        }

    evidence = {"text": "", "sheets": [], "periods": [], "sampled_rows": 0}
    if extension == ".xlsx":
        try:
            evidence = _xlsx_evidence(path)
        except Exception as error:
            return {
                "document_type": "unknown", "label": DOCUMENT_TYPES["unknown"]["label"],
                "confidence": 0.05, "alternatives": [], "periods": [],
                "signals": [f"工作簿无法读取：{error}"], "capability": "人工确认", "destination": None,
            }
    elif extension == ".csv":
        try:
            evidence = _csv_evidence(path)
        except Exception as error:
            return {
                "document_type": "unknown", "label": DOCUMENT_TYPES["unknown"]["label"],
                "confidence": 0.05, "alternatives": [], "periods": [],
                "signals": [f"CSV 无法读取：{error}"], "capability": "人工确认", "destination": None,
            }

    binary_document = extension in {".pdf", ".png", ".jpg", ".jpeg", ".webp"}
    filename_slug = _slug(filename)
    if binary_document:
        direct_type = None
        if any(token in filename_slug for token in ("发票", "invoice")):
            direct_type = "invoice_document"
        elif any(token in filename_slug for token in ("银行", "对账单", "流水", "回单", "bankstatement")):
            direct_type = "bank_statement_document"
        elif any(token in filename_slug for token in ("合同", "协议", "contract", "agreement")):
            direct_type = "contract_commercial"
        elif any(token in filename_slug for token in ("验收", "交付", "成果", "acceptance", "delivery")):
            direct_type = "acceptance_evidence"
        if direct_type:
            definition = DOCUMENT_TYPES[direct_type]
            return {
                "document_type": direct_type, "label": definition["label"], "confidence": 0.7,
                "alternatives": [], "periods": [],
                "signals": ["依据文件名进入文字识别/证据队列，尚未读取正文"],
                "capability": definition["capability"], "destination": definition["dataset"], "sheets": [],
            }

    haystack = _slug(f"{filename}\n{' '.join(evidence['sheets'])}\n{evidence['text']}")
    scores: dict[str, float] = {name: 0 for name in DOCUMENT_TYPES if name != "unknown"}
    signals: dict[str, list[str]] = {name: [] for name in scores}
    for document_type, keywords in FILENAME_RULES.items():
        for keyword in keywords:
            if _slug(keyword) in _slug(filename):
                scores[document_type] += 1.5
                signals[document_type].append(f"文件名包含“{keyword}”")
    for document_type, rules in HEADER_RULES.items():
        for keyword, weight in rules.items():
            if _slug(keyword) in haystack:
                scores[document_type] += weight
                signals[document_type].append(f"识别到“{keyword}”")

    ranked = sorted(scores.items(), key=lambda item: item[1], reverse=True)
    top_type, top_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else 0
    strong_types = [(name, score) for name, score in ranked if score >= 8]
    if extension == ".xlsx" and len(evidence["sheets"]) >= 3 and len(strong_types) >= 3:
        top_type = "finance_package"
        confidence = min(0.99, 0.8 + len(strong_types) * 0.025)
        signals["finance_package"] = [
            f"同一工作簿识别到 {DOCUMENT_TYPES[name]['label']}（得分 {round(score, 1)}）"
            for name, score in strong_types[:8]
        ]
    elif top_score < 3:
        top_type = "unknown"
        confidence = 0.2 if top_score else 0.05
    else:
        confidence = min(0.99, 0.48 + top_score * 0.025 + max(0, top_score - second_score) * 0.02)

    # 单张 PDF/图片发票通常没有可扫描表头，文件名线索足以进入 OCR 队列但不足以自动入账。
    if binary_document and top_type == "invoice_register":
        top_type = "invoice_document"
        confidence = min(confidence, 0.75)
    definition = DOCUMENT_TYPES[top_type]
    alternatives = [
        {"document_type": name, "label": DOCUMENT_TYPES[name]["label"], "score": round(score, 2)}
        for name, score in ranked if name != top_type and score > 0
    ][:3]
    return {
        "document_type": top_type,
        "label": definition["label"],
        "confidence": round(confidence, 4),
        "alternatives": alternatives,
        "periods": evidence["periods"],
        "signals": signals.get(top_type, [])[:12] or ["缺少足够的结构或文件名线索"],
        "capability": definition["capability"],
        "destination": definition["dataset"],
        "sheets": evidence["sheets"],
    }


def _plain_records(records: list[Any]) -> list[dict]:
    return [asdict(item) if is_dataclass(item) else dict(item) for item in records]


def recognize_structured_document(
    path: str | Path,
    document_type: str,
    period: str = "",
    datasets: dict[str, list[dict]] | None = None,
    company_profile: dict | None = None,
) -> dict:
    path = Path(path)
    datasets = datasets or {}
    if path.suffix.lower() == ".csv":
        with tempfile.TemporaryDirectory(prefix="finance-inbox-csv-") as temp_dir:
            converted = Path(temp_dir) / f"{path.stem}.xlsx"
            _csv_to_xlsx(path, converted)
            return recognize_structured_document(converted, document_type, period, datasets, company_profile)
    if document_type == "finance_package":
        batches = []
        errors = []
        for nested_type in (
            "settlement", "purchase", "bank_statement", "invoice_register",
            "payroll", "opening_balance", "planning",
        ):
            try:
                result = recognize_structured_document(path, nested_type, period, datasets, company_profile)
                if result.get("records"):
                    batches.append({"document_type": nested_type, **result})
            except Exception as error:
                errors.append({"document_type": nested_type, "error": str(error)})
        if not batches:
            raise ValueError("综合资料包未解析出可用台账")
        return {
            "dataset": None, "records": [], "batches": batches, "errors": errors,
            "preview": {
                "batch_count": len(batches),
                "record_count": sum(len(batch["records"]) for batch in batches),
                "datasets": [batch["dataset"] for batch in batches],
            },
        }
    if document_type == "invoice_document":
        extraction = extract_document_text(path)
        fields = extract_invoice_fields(extraction, path.name)
        record = invoice_record_from_extraction(fields, path.name)
        return {
            "dataset": "invoices", "records": [record],
            "extraction": extraction, "field_extraction": fields,
            "preview": {"record_count": 1, "confidence": fields["confidence"], "missing_fields": fields["missing_fields"]},
        }
    if document_type == "bank_statement_document":
        extraction = extract_document_text(path)
        fields = extract_bank_statement_rows(extraction, path.name)
        records = suggest_matches(
            bank_records_from_extraction(fields, path.name),
            datasets.get("settlements") or [], datasets.get("purchases") or [],
        )
        for record in records:
            record["match_suggestion_status"] = record.get("status") or "待认领"
            record["status"] = "待人工确认"
        return {
            "dataset": "bank_transactions", "records": records,
            "extraction": extraction, "field_extraction": fields,
            "preview": {
                "record_count": len(records), "confidence": fields["confidence"],
                "missing_field_count": sum(len(item.get("missing_fields") or []) for item in fields.get("rows") or []),
            },
        }
    if document_type == "settlement_reconciliation_evidence":
        rows = extract_settlement_reconciliation_evidence(path)
        return {
            "dataset": None, "records": [], "evidence_only": True,
            "reconciliation_evidence": rows,
            "preview": {
                "record_count": len(rows),
                "exception_count": sum(bool(item.get("is_exception")) for item in rows),
                "periods": sorted({item["period"] for item in rows if item.get("period")}),
            },
        }
    if document_type in {"contract_commercial", "acceptance_evidence"}:
        extraction = extract_document_text(path)
        return {
            "dataset": None, "records": [], "extraction": extraction,
            "preview": {"page_count": extraction["page_count"], "confidence": extraction["confidence"]},
            "evidence_only": True,
        }
    if document_type == "settlement":
        payload = dashboard_payload(parse_files([path]))
        return {"dataset": "settlements", "records": payload.get("records") or [], "preview": payload}
    if document_type == "bank_statement":
        records = suggest_matches(
            parse_bank_workbook(path), datasets.get("settlements") or [], datasets.get("purchases") or [],
        )
        payload = banking_payload(records)
        return {"dataset": "bank_transactions", "records": payload["transactions"], "preview": payload}
    if document_type == "invoice_register":
        records = match_invoices_to_purchases(parse_invoice_workbook(path), datasets.get("purchases") or [])
        payload = invoice_payload(records)
        return {"dataset": "invoices", "records": payload["records"], "preview": payload}
    if document_type == "purchase":
        payload = procurement_payload(parse_purchase_workbook(path))
        return {"dataset": "purchases", "records": payload["records"], "preview": payload}
    if document_type == "payroll":
        if not PERIOD_PATTERN.fullmatch(period):
            raise ValueError("工资资料需要指定 YYYY-MM 账期")
        profile = company_profile or {}
        parsed = parse_payroll_workbook(
            path, period, str(profile.get("jurisdiction") or "CN"),
            str(profile.get("functional_currency") or profile.get("base_currency") or "CNY"),
        )
        payload = payroll_payload(parsed, period)
        return {"dataset": "payroll_rows", "records": payload["records"], "preview": payload}
    if document_type == "planning":
        payload = planning_payload(parse_plan_workbook(path))
        return {"dataset": "plan_lines", "records": payload["records"], "preview": payload}
    if document_type == "opening_balance":
        payload = opening_balance_payload(parse_opening_balance_workbook(path, period))
        return {"dataset": "opening_balances", "records": payload["records"], "preview": payload}
    raise ValueError("该资料目前只能归档或等待文字识别，不能自动写入台账")


def extract_settlement_reconciliation_evidence(path: str | Path) -> list[dict]:
    """Extract compact iOS/platform-vs-operational reconciliation evidence.

    The result is evidence only: it never creates settlement revenue records.
    """
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    aliases = {
        "period": ("年月", "月份", "账期"), "game": ("游戏", "游戏名称"),
        "date": ("日期", "交易日期"), "platform": ("平台",), "channel": ("渠道",),
        "operational_amount": ("付费金额", "经分金额", "业务流水"),
        "platform_amount": ("匹配-苹果账单", "平台账单金额", "苹果账单金额"),
        "difference": ("差异",), "refund": ("退款",),
        "net_difference": ("剔除退款后差异", "退款后差异"),
    }

    def field(value: Any) -> str | None:
        clean = _slug(value)
        candidates = []
        for name, names in aliases.items():
            for alias in names:
                alias_clean = _slug(alias)
                if alias_clean and (clean == alias_clean or alias_clean in clean):
                    candidates.append((len(alias_clean), name))
        return max(candidates, default=(0, None))[1]

    def number(value: Any) -> float | None:
        if value in {None, ""} or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return round(float(value), 2)
        match = re.search(r"-?[\d,]+(?:\.\d+)?", str(value))
        return round(float(match.group(0).replace(",", "")), 2) if match else None

    result = []
    try:
        for sheet in workbook.worksheets:
            mapping: dict[int, str] = {}
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 10000, 10000), values_only=True), 1,
            ):
                candidate = {}
                for index, value in enumerate(row):
                    name = field(value)
                    if name and name not in candidate.values():
                        candidate[index] = name
                if len(candidate) >= 6 and {"period", "operational_amount", "difference"}.issubset(candidate.values()):
                    mapping = candidate
                    continue
                if not mapping:
                    continue
                raw = {name: row[index] if index < len(row) else None for index, name in mapping.items()}
                period = _period(raw.get("period") or raw.get("date"))
                operational = number(raw.get("operational_amount"))
                platform = number(raw.get("platform_amount"))
                difference = number(raw.get("difference"))
                refund = number(raw.get("refund"))
                net_difference = number(raw.get("net_difference"))
                if not period or operational is None:
                    continue
                materiality = max(1.0, abs(operational) * 0.001)
                check_value = net_difference if net_difference is not None else difference
                result.append({
                    "source_sheet": sheet.title, "source_row": row_number,
                    "period": period, "game": str(raw.get("game") or "").strip(),
                    "date": str(raw.get("date") or "").strip(),
                    "platform": str(raw.get("platform") or "").strip(),
                    "channel": str(raw.get("channel") or "").strip(),
                    "operational_amount": operational, "platform_amount": platform,
                    "difference": difference, "refund": refund, "net_difference": net_difference,
                    "is_exception": check_value is None or abs(check_value) > materiality,
                    "materiality": round(materiality, 2),
                })
    finally:
        workbook.close()
    return result[:5000]


class FinanceInboxStore:
    def __init__(self, root: str | Path):
        self.root = Path(root)
        self.documents = self.root / "documents"
        self.blobs = self.root / "blobs"
        self.events_file = self.root / "events.jsonl"
        self._lock = threading.RLock()

    def ingest(
        self, filename: str, body: bytes, actor: str = "财务工作台用户",
        entity_id: str = "", entities: list[dict] | None = None,
    ) -> dict:
        safe_name = Path(str(filename or "未命名文件")).name[:240]
        extension = Path(safe_name).suffix.lower()
        if extension not in SUPPORTED_EXTENSIONS:
            raise ValueError("目前支持 xlsx、csv、pdf、png、jpg、jpeg、webp")
        if not body:
            raise ValueError("文件内容为空")
        digest = hashlib.sha256(body).hexdigest()
        document_id = f"DOC-{digest[:16].upper()}"
        existing_path = self.documents / f"{document_id}.json"
        if existing_path.exists():
            existing = self.load(document_id)
            if entity_id:
                existing = self.assign_entity_scope(document_id, entity_id, actor, entities or [])
            self.append_event(document_id, "DUPLICATE_RECEIVED", actor, {"filename": safe_name})
            existing["duplicate"] = True
            return existing

        blob = self.blobs / f"{digest}{extension}"
        blob.parent.mkdir(parents=True, exist_ok=True)
        if not blob.exists():
            fd, temp_name = tempfile.mkstemp(prefix=".upload.", dir=blob.parent)
            try:
                with os.fdopen(fd, "wb") as handle:
                    handle.write(body)
                os.replace(temp_name, blob)
            finally:
                if os.path.exists(temp_name):
                    os.unlink(temp_name)
        classification = classify_document(blob, safe_name)
        status = "已识别待确认" if classification["confidence"] >= 0.75 else "待确认类型"
        if classification["capability"] in {"等待OCR", "等待文字提取"}:
            status = "等待文字识别"
        entity_scope = suggest_entity_scope(safe_name, entities)
        if entity_id:
            selected = next((item for item in (entities or []) if item.get("id") == entity_id), None)
            if not selected:
                raise ValueError("资料所属法律主体不在当前 Box 配置中")
            entity_scope = {
                "entity_id": entity_id, "entity_name": str(selected.get("name") or entity_id),
                "status": "confirmed", "source": "uploader_selected", "confidence": 1.0,
                "reason": "上传人已明确选择资料所属法律主体。",
                "confirmed_by": str(actor or "财务工作台用户")[:80], "confirmed_at": _now(),
            }
        document = {
            "id": document_id,
            "original_filename": safe_name,
            "sha256": digest,
            "size": len(body),
            "extension": extension,
            "mime_type": mimetypes.guess_type(safe_name)[0] or "application/octet-stream",
            "blob_path": str(blob),
            "status": status,
            "classification": classification,
            "entity_scope": entity_scope,
            "created_by": str(actor or "财务工作台用户")[:80],
            "created_at": _now(),
            "updated_at": _now(),
            "recognition": None,
            "commit": None,
            "duplicate": False,
        }
        with self._lock:
            _atomic_write(existing_path, document)
            self.append_event(document_id, "DOCUMENT_RECEIVED", actor, {
                "filename": safe_name, "sha256": digest, "size": len(body),
            })
            self.append_event(document_id, "DOCUMENT_CLASSIFIED", "Agent", classification)
            self.append_event(document_id, "ENTITY_SCOPE_ROUTED", actor if entity_id else "Agent", entity_scope)
        return document

    def assign_entity_scope(
        self, document_id: str, entity_id: str, actor: str,
        entities: list[dict], note: str = "",
    ) -> dict:
        actor = str(actor or "").strip()
        if not actor:
            raise ValueError("请填写主体确认人")
        selected = next((item for item in entities if item.get("id") == entity_id), None)
        if not selected:
            raise ValueError("资料所属法律主体不在当前 Box 配置中")
        document = self.load(document_id)
        current = document.get("entity_scope") or {}
        if document.get("status") == "已入台账" and current.get("entity_id") != entity_id:
            raise ValueError("资料已入台账，不能在收件箱中改换法律主体")
        scope = {
            "entity_id": entity_id, "entity_name": str(selected.get("name") or entity_id),
            "status": "confirmed", "source": "human_confirmation", "confidence": 1.0,
            "reason": str(note or "用户在识别或入账前确认资料所属主体。")[:500],
            "confirmed_by": actor[:80], "confirmed_at": _now(),
            "previous": {
                "entity_id": current.get("entity_id") or "", "status": current.get("status") or "unassigned",
                "source": current.get("source") or "",
            },
        }
        document["entity_scope"] = scope
        self.save(document)
        self.append_event(document_id, "ENTITY_SCOPE_CONFIRMED", actor, scope)
        return document

    def load(self, document_id: str) -> dict:
        document_id = str(document_id or "").strip()
        if not re.fullmatch(r"DOC-[A-F0-9]{16}", document_id):
            raise ValueError("资料编号无效")
        path = self.documents / f"{document_id}.json"
        if not path.exists():
            raise ValueError("资料不存在")
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as error:
            raise ValueError("资料记录无法读取") from error

    def save(self, document: dict) -> dict:
        document["updated_at"] = _now()
        self.load(document["id"])
        with self._lock:
            _atomic_write(self.documents / f"{document['id']}.json", document)
        return document

    def list(self, limit: int = 200) -> list[dict]:
        if not self.documents.exists():
            return []
        result = []
        for path in self.documents.glob("DOC-*.json"):
            try:
                result.append(json.loads(path.read_text(encoding="utf-8")))
            except (OSError, json.JSONDecodeError):
                continue
        result.sort(key=lambda item: item.get("updated_at") or "", reverse=True)
        return result[:max(1, min(int(limit), 1000))]

    def recognize(
        self, document_id: str, document_type: str, period: str,
        datasets: dict[str, list[dict]], actor: str = "Agent",
        entity_id: str = "", entity_name: str = "",
        company_profile: dict | None = None,
    ) -> dict:
        document = self.load(document_id)
        if document_type not in DOCUMENT_TYPES:
            raise ValueError("资料类型无效")
        result = recognize_structured_document(
            document["blob_path"], document_type, period, datasets, company_profile,
        )
        records = result.pop("records")
        batches = result.pop("batches", [])
        if entity_id:
            for record in records:
                record["entity_id"] = entity_id
                record["source_document_id"] = document_id
            for batch in batches:
                for record in batch.get("records") or []:
                    record["entity_id"] = entity_id
                    record["source_document_id"] = document_id
        document["classification"]["document_type"] = document_type
        document["classification"]["label"] = DOCUMENT_TYPES[document_type]["label"]
        document["classification"]["destination"] = result["dataset"]
        document["recognition"] = {
            **result,
            "period": period,
            "entity_id": entity_id,
            "entity_name": entity_name,
            "record_count": (
                int((result.get("preview") or {}).get("record_count") or 0)
                if result.get("evidence_only")
                else len(records) + sum(len(batch.get("records") or []) for batch in batches)
            ),
            "records": records,
            "batches": batches,
            "recognized_at": _now(),
        }
        if result.get("evidence_only"):
            document["status"] = "已提取待归档"
        else:
            document["status"] = "已解析待入账" if records or batches else "未识别到可用记录"
        self.save(document)
        self.append_event(document_id, "DOCUMENT_RECOGNIZED", actor, {
            "document_type": document_type, "period": period,
            "entity_id": entity_id,
            "dataset": result["dataset"],
            "datasets": [batch.get("dataset") for batch in batches],
            "record_count": document["recognition"]["record_count"],
        })
        return document

    def mark_committed(
        self, document_id: str, actor: str, dataset: str | None, record_count: int,
        datasets: list[str] | None = None, entity_id: str = "",
    ) -> dict:
        document = self.load(document_id)
        document["status"] = "已入台账"
        document["commit"] = {
            "dataset": dataset, "datasets": datasets or ([dataset] if dataset else []),
            "record_count": record_count,
            "entity_id": entity_id,
            "actor": str(actor or "财务工作台用户")[:80], "timestamp": _now(),
        }
        # 正式台账已经保存标准记录；收件箱不再重复持有完整解析明细。
        if document.get("recognition"):
            document["recognition"].pop("records", None)
            for batch in document["recognition"].get("batches") or []:
                batch.pop("records", None)
        self.save(document)
        self.append_event(document_id, "DOCUMENT_COMMITTED", actor, document["commit"])
        return document

    def commit_blockers(self, document_id: str) -> list[str]:
        """Return fail-closed blockers for writing a recognition preview to ledgers."""
        document = self.load(document_id)
        recognition = document.get("recognition") or {}
        records = recognition.get("records") or []
        document_type = str((document.get("classification") or {}).get("document_type") or "")
        blockers = []
        if document_type in {"invoice_document", "bank_statement_document"}:
            confirmed = any(
                bool(item.get("confirmed_against_original"))
                for item in recognition.get("corrections") or []
            )
            if not confirmed:
                blockers.append("OCR/文字层结果必须由人工对照原件确认后才能入台账")
        if document_type == "invoice_document":
            required = ("invoice_number", "invoice_date", "seller_name", "total_amount")
            for index, record in enumerate(records, 1):
                missing = [name for name in required if record.get(name) in {None, "", 0}]
                if missing:
                    blockers.append(f"第{index}张发票缺少关键字段：{'、'.join(missing)}")
        if document_type == "bank_statement_document":
            required = ("transaction_date", "transaction_id", "direction", "currency", "amount")
            for index, record in enumerate(records, 1):
                missing = [name for name in required if record.get(name) in {None, "", 0, "待确认"}]
                if missing:
                    blockers.append(f"第{index}条银行流水缺少关键字段：{'、'.join(missing)}")
        return blockers

    def link_to_business_record(
        self, document_id: str, *, target_type: str, target_id: str,
        entity_id: str, actor: str, note: str = "",
    ) -> dict:
        """Archive evidence against a business record without creating accounting facts."""
        if target_type not in {"purchase", "settlement"}:
            raise ValueError("当前证据只能关联采购/验收或收入结算记录")
        document = self.load(document_id)
        scope = document.get("entity_scope") or {}
        if scope.get("status") != "confirmed" or scope.get("entity_id") != entity_id:
            raise ValueError("证据资料与采购记录的法律主体不一致")
        document_type = str((document.get("classification") or {}).get("document_type") or "")
        allowed_links = {
            "purchase": {"acceptance_evidence", "contract_commercial"},
            "settlement": {"settlement_reconciliation_evidence", "contract_commercial"},
        }
        if document_type not in allowed_links.get(target_type, set()):
            raise ValueError("资料类型不能关联到所选业务记录")
        actor = str(actor or "").strip()
        if not actor:
            raise ValueError("请填写证据关联人")
        link = {
            "target_type": target_type, "target_id": str(target_id or ""),
            "entity_id": entity_id, "document_type": document_type,
            "actor": actor[:80], "note": str(note or "")[:1000], "linked_at": _now(),
        }
        existing = [
            item for item in document.get("business_links") or []
            if not (item.get("target_type") == target_type and item.get("target_id") == target_id)
        ]
        document["business_links"] = [*existing, link]
        document["status"] = "已归档并关联"
        self.save(document)
        self.append_event(document_id, "DOCUMENT_LINKED", actor, link)
        return document

    def correct(
        self, document_id: str, patches: list[dict], actor: str,
        confirmed_against_original: bool = False, note: str = "",
    ) -> dict:
        actor = str(actor or "").strip()
        if not actor:
            raise ValueError("请填写校正人")
        if not isinstance(patches, list) or not patches:
            raise ValueError("请提供需要校正的字段")
        document = self.load(document_id)
        recognition = document.get("recognition") or {}
        records = recognition.get("records") or []
        if not records:
            raise ValueError("该资料没有可校正的单条识别记录")
        protected = {
            "id", "source_file", "source_sheet", "source_row", "sha256",
            "booking_status", "verification_status", "duplicate_key", "entity_id",
            "source_document_id",
        }
        allowed_by_type = {
            "invoice_document": {
                "invoice_number", "invoice_code", "invoice_date", "invoice_type",
                "seller_name", "buyer_name", "item", "amount_ex_tax", "tax_rate",
                "tax_amount", "total_amount", "po_number", "project",
                "deduction_status", "seller_tax_id_masked",
            },
            "bank_statement_document": {
                "transaction_date", "transaction_id", "counterparty", "summary",
                "direction", "currency", "amount", "balance", "account_masked",
                "counterparty_account_masked",
            },
        }
        allowed = allowed_by_type.get(str((document.get("classification") or {}).get("document_type") or ""))
        changes = []
        for patch in patches:
            index = int(patch.get("index", 0))
            fields = patch.get("fields") or {}
            if index < 0 or index >= len(records) or not isinstance(fields, dict):
                raise ValueError("校正记录位置或字段无效")
            forbidden = sorted(protected.intersection(fields))
            if forbidden:
                raise ValueError(f"不能通过字段校正修改：{'、'.join(forbidden)}")
            if allowed is not None:
                unexpected = sorted(set(fields) - allowed)
                if unexpected:
                    raise ValueError(f"该识别类型不允许校正：{'、'.join(unexpected)}")
            before = {key: records[index].get(key) for key in fields}
            records[index].update(fields)
            changes.append({"index": index, "before": before, "after": fields})
        if confirmed_against_original:
            recognized_type = str((document.get("classification") or {}).get("document_type") or "")
            for record in records:
                record["anomalies"] = [
                    item for item in (record.get("anomalies") or [])
                    if "尚未由人工对照原票确认" not in item
                    and "尚未由人工对照原件确认" not in item
                ]
                if record.get("status") == "待人工确认":
                    record["status"] = (
                        record.get("match_suggestion_status") or "已核原件待认领"
                        if recognized_type == "bank_statement_document"
                        else "已核原件待查验"
                    )
        correction = {
            "actor": actor[:80], "timestamp": _now(), "changes": changes,
            "confirmed_against_original": bool(confirmed_against_original), "note": str(note or "")[:1000],
        }
        recognition.setdefault("corrections", []).append(correction)
        document["status"] = "已解析待入账"
        self.save(document)
        self.append_event(document_id, "DOCUMENT_CORRECTED", actor, correction)
        return document

    def append_event(self, document_id: str, event_type: str, actor: str, detail: dict | None = None) -> dict:
        event = {
            "id": hashlib.sha1(f"{document_id}|{event_type}|{_now()}".encode()).hexdigest()[:16],
            "document_id": document_id, "type": event_type, "actor": str(actor or "Agent")[:80],
            "timestamp": _now(), "detail": detail or {},
        }
        self.root.mkdir(parents=True, exist_ok=True)
        with self._lock:
            with self.events_file.open("a", encoding="utf-8") as handle:
                handle.write(json.dumps(event, ensure_ascii=False) + "\n")
        return event

    def events(self, document_id: str, limit: int = 200) -> list[dict]:
        self.load(document_id)
        if not self.events_file.exists():
            return []
        result = []
        for line in self.events_file.read_text(encoding="utf-8").splitlines():
            try:
                event = json.loads(line)
            except json.JSONDecodeError:
                continue
            if event.get("document_id") == document_id:
                result.append(event)
        return list(reversed(result[-max(1, min(int(limit), 1000)):]))
