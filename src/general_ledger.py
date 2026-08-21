from __future__ import annotations

import hashlib
import math
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", _text(value).lower())


def _number(value: Any) -> float:
    if value is None or value == "" or isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return 0.0 if math.isnan(value) or math.isinf(value) else float(value)
    match = re.search(r"-?[\d,.]+", _text(value))
    return float(match.group().replace(",", "")) if match else 0.0


ALIASES = {
    "period": ("期间", "账期", "年月", "period"),
    "account_code": ("科目编码", "科目代码", "account code", "code"),
    "account_name": ("科目名称", "会计科目", "account name", "account"),
    "opening_debit": ("期初借方", "期初借方余额", "opening debit"),
    "opening_credit": ("期初贷方", "期初贷方余额", "opening credit"),
    "closing_debit": ("期末借方", "期末借方余额", "closing debit"),
    "closing_credit": ("期末贷方", "期末贷方余额", "closing credit"),
}


def _field(value: Any) -> str | None:
    clean = _slug(value)
    candidates = []
    for field_name, aliases in ALIASES.items():
        for alias in aliases:
            alias_clean = _slug(alias)
            if alias_clean and (clean == alias_clean or alias_clean in clean):
                candidates.append((len(alias_clean), field_name))
    return max(candidates, default=(0, None))[1]


def _account_category(code: str, name: str) -> str:
    if code.startswith("1"):
        return "资产"
    if code.startswith("2"):
        return "负债"
    if code.startswith("3"):
        return "权益"
    if code.startswith("5"):
        if code.startswith(("5001", "5051")):
            return "收入"
        return "成本费用"
    if any(word in name for word in ("银行", "现金", "应收", "资产", "存货", "待摊")):
        return "资产"
    if any(word in name for word in ("应付", "应交", "借款", "负债")):
        return "负债"
    return "待映射"


@dataclass
class OpeningBalance:
    id: str
    source_file: str
    source_sheet: str
    source_row: int
    period: str
    account_code: str
    account_name: str
    account: str
    category: str
    opening_debit: float
    opening_credit: float
    status: str
    anomalies: list[str] = field(default_factory=list)


def parse_opening_balance_workbook(path: str | Path, period: str = "") -> list[OpeningBalance]:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = []
    seen = set()
    for sheet in workbook.worksheets:
        mapping = {}
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            candidate = {index: _field(value) for index, value in enumerate(row) if _field(value)}
            if len(candidate) >= 3 and "account_name" in candidate.values():
                mapping = candidate
                continue
            if not mapping:
                continue
            raw = {name: row[index] if index < len(row) else None for index, name in mapping.items()}
            name, code = _text(raw.get("account_name")), _text(raw.get("account_code"))
            if not name and not code:
                continue
            row_period = _text(raw.get("period"))[:7] or period
            debit = _number(raw.get("opening_debit")) or _number(raw.get("closing_debit"))
            credit = _number(raw.get("opening_credit")) or _number(raw.get("closing_credit"))
            anomalies = []
            key = (row_period, code or name)
            if key in seen:
                anomalies.append("同一期间科目重复")
            seen.add(key)
            if debit and credit:
                anomalies.append("同一科目同时存在借方和贷方期初余额")
            if not row_period:
                anomalies.append("缺少期初余额所属期间")
            account = f"{code} {name}".strip()
            identity = f"{path.name}|{sheet.title}|{row_number}|{row_period}|{account}"
            rows.append(OpeningBalance(
                id=hashlib.sha1(identity.encode("utf-8")).hexdigest()[:12],
                source_file=path.name, source_sheet=sheet.title, source_row=row_number,
                period=row_period, account_code=code, account_name=name, account=account,
                category=_account_category(code, name), opening_debit=round(abs(debit), 2),
                opening_credit=round(abs(credit), 2), status="异常" if anomalies else "可用", anomalies=anomalies,
            ))
    workbook.close()
    return rows


def opening_balance_payload(rows: Iterable[OpeningBalance]) -> dict:
    items = list(rows)
    return {"records": [asdict(item) for item in items], "summary": {
        "count": len(items), "periods": sorted({item.period for item in items if item.period}),
        "total_debit": round(sum(item.opening_debit for item in items), 2),
        "total_credit": round(sum(item.opening_credit for item in items), 2),
        "difference": round(sum(item.opening_debit - item.opening_credit for item in items), 2),
        "exception_count": sum(bool(item.anomalies) for item in items),
    }}


def build_financial_statements(opening_balances: Iterable[dict], trial_balance: dict, period: str) -> dict:
    opening = [row for row in opening_balances if row.get("period") == period]
    accounts: dict[str, dict] = {}
    for row in opening:
        account = row.get("account") or f"{row.get('account_code', '')} {row.get('account_name', '')}".strip()
        accounts[account] = {
            "account": account, "category": row.get("category") or _account_category(
                str(row.get("account_code") or ""), str(row.get("account_name") or "")
            ),
            "opening_debit": float(row.get("opening_debit") or 0),
            "opening_credit": float(row.get("opening_credit") or 0), "period_debit": 0.0, "period_credit": 0.0,
        }
    for row in trial_balance.get("rows") or []:
        account = row["account"]
        code, _, name = account.partition(" ")
        bucket = accounts.setdefault(account, {
            "account": account, "category": row.get("category") or _account_category(code, name),
            "opening_debit": 0.0, "opening_credit": 0.0, "period_debit": 0.0, "period_credit": 0.0,
        })
        if row.get("category"):
            bucket["category"] = row["category"]
        bucket["period_debit"] += float(row.get("debit") or 0)
        bucket["period_credit"] += float(row.get("credit") or 0)
    detail = []
    for bucket in accounts.values():
        net = bucket["opening_debit"] - bucket["opening_credit"] + bucket["period_debit"] - bucket["period_credit"]
        bucket["closing_debit"] = round(max(0, net), 2)
        bucket["closing_credit"] = round(max(0, -net), 2)
        detail.append(bucket)
    assets = sum(row["closing_debit"] - row["closing_credit"] for row in detail if row["category"] == "资产")
    liabilities = sum(row["closing_credit"] - row["closing_debit"] for row in detail if row["category"] == "负债")
    equity = sum(row["closing_credit"] - row["closing_debit"] for row in detail if row["category"] == "权益")
    revenue = sum(row["period_credit"] - row["period_debit"] for row in detail if row["category"] == "收入")
    expenses = sum(row["period_debit"] - row["period_credit"] for row in detail if row["category"] == "成本费用")
    profit = revenue - expenses
    balance_difference = round(assets - liabilities - equity - profit, 2)
    return {
        "period": period, "opening_available": bool(opening), "detail": sorted(detail, key=lambda row: row["account"]),
        "balance_sheet": {"assets": round(assets, 2), "liabilities": round(liabilities, 2),
                          "equity_before_current_profit": round(equity, 2), "current_profit": round(profit, 2),
                          "liabilities_and_equity": round(liabilities + equity + profit, 2),
                          "difference": balance_difference, "balanced": abs(balance_difference) < 0.01},
        "income_statement": {"revenue": round(revenue, 2), "expenses": round(expenses, 2),
                             "profit_before_tax": round(profit, 2)},
        "guardrail": "无期初余额时仅能生成本期发生额利润草稿；资产负债表必须接入上期经确认的期末余额。",
    }
