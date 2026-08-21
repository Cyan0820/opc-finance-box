from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "263A5B"
BLUE = "2457D6"
LIGHT = "EDF3FF"
AMBER = "FFF4DF"
LINE = "D9DEE7"
TEXT = "202939"
WHITE = "FFFFFF"
THIN = Side(style="thin", color=LINE)


def _table(sheet, headers: list[str], rows: Iterable[Iterable], widths: dict[int, int] | None = None) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.append(headers)
    for row in rows:
        sheet.append(list(row))
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True, size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    sheet.row_dimensions[1].height = 27
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color=TEXT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
    for index, header in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(index)].width = (widths or {}).get(index, min(28, max(12, len(header) + 4)))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _validate(path: Path) -> Path:
    check = load_workbook(path, read_only=True, data_only=False)
    check.close()
    return path


def build_demo_workbook(config: dict, output_path: str | Path) -> Path:
    """Generate one complete fictional scenario workbook with public dependencies only."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    intro = book.active
    intro.title = "使用说明"
    intro.sheet_view.showGridLines = False
    intro.merge_cells("A1:H1")
    intro["A1"] = f"智能财务工作台｜{config['label']}"
    intro["A1"].fill = PatternFill("solid", fgColor=NAVY)
    intro["A1"].font = Font(size=16, bold=True, color=WHITE)
    intro["A1"].alignment = Alignment(vertical="center")
    intro.row_dimensions[1].height = 34
    info = [
        ("公司", config["company"]), ("场景", config["label"]), ("数据期间", "2026-01 至 2026-02"),
        ("用途", "产品演示、导入测试和财务流程体验"), ("真实性", "全部为虚构示例，不对应真实公司或个人"),
        ("导入顺序", "结算对账 → 采购 → 银行 → 发票 → 工资 → 期初 → 预算/KPI"),
        ("币种提示", "原币独立核对；只有配置对应期间汇率后才折算人民币" if config["label"] == "海外示例" else "CNY独立核对"),
        ("安全边界", "不会自动付款，也不会未经确认直接申报"),
    ]
    for row_index, (label, value) in enumerate(info, 3):
        intro.cell(row_index, 1, label).fill = PatternFill("solid", fgColor=LIGHT)
        intro.cell(row_index, 1).font = Font(color=BLUE, bold=True)
        intro.cell(row_index, 2, value)
        for column in (1, 2):
            intro.cell(row_index, column).border = Border(bottom=THIN)
            intro.cell(row_index, column).alignment = Alignment(wrap_text=True, vertical="center")
        intro.row_dimensions[row_index].height = 27
    intro.column_dimensions["A"].width = 18
    intro.column_dimensions["B"].width = 78
    intro.merge_cells("A13:H13")
    intro["A13"] = "提示：示例用于展示系统能力；凭证、税务候选值和申报资料仍应由有权限人员按真实资料复核。"
    intro["A13"].fill = PatternFill("solid", fgColor=AMBER)
    intro["A13"].alignment = Alignment(wrap_text=True, vertical="center")

    profile = config["profile"]
    _table(book.create_sheet("公司档案"), ["配置项", "示例值"], [
        ["公司名称", profile.get("company_name")], ["统一社会信用代码", profile.get("credit_code")],
        ["注册地", profile.get("registered_city")], ["会计准则", profile.get("accounting_standard")],
        ["增值税纳税人类型", profile.get("vat_taxpayer_type")], ["期初现金", (profile.get("cash_planning") or {}).get("opening_cash_cny")],
        ["最低现金缓冲", (profile.get("cash_planning") or {}).get("minimum_buffer_cny")],
        ["月末USD/CNY", (((profile.get("fx_policy") or {}).get("month_end_rates") or {}).get("2026-02") or {}).get("USD") or "不适用"],
    ], {1: 28, 2: 56})
    settlement_headers = ["结算周期", "游戏名称", "平台", "渠道", "结算币种", "渠道含税流水（结算币种）", "退款流水", "分成基数", "分成比例", "结算金额", "预提所得税（结算币种）", "甲方实收金额（结算币种）", "国家/地区"]
    purchase_headers = ["PO编号", "下单日期", "项目", "供应商", "采购内容", "数量", "含税单价", "订单金额", "验收金额", "开票金额", "付款金额", "币种", "税率", "状态/备注"]
    bank_headers = ["交易日期", "交易流水号", "本方账号", "对方户名", "对方账号", "摘要", "收支方向", "币种", "金额", "账户余额"]
    invoice_headers = ["发票号码", "开票日期", "发票类型", "销售方名称", "销售方税号", "购买方名称", "项目名称", "不含税金额", "税率", "税额", "价税合计", "PO编号", "项目", "查验状态", "抵扣状态", "入账状态", "状态"]
    payroll_headers = ["人员编号", "姓名", "部门", "项目", "应发工资", "个人社保", "个人公积金", "专项附加扣除", "其他扣除", "累计收入", "累计扣除", "累计已预扣税额", "实发工资", "研发工时占比"]
    opening_headers = ["期间", "科目编码", "科目名称", "期初借方", "期初贷方"]
    plan_headers = ["月份", "情景", "项目", "类别", "收支方向", "金额", "币种", "概率", "已承诺", "备注"]
    kpi_headers = ["月份", "游戏项目编码", "渠道", "区域", "DAU", "MAU", "新增用户", "付费用户", "安装数", "游戏流水", "投放金额", "D1留存", "D7留存", "D30留存"]
    settlement_name = "对外账单-国服" if config["label"] == "国服示例" else "商店金流账单-海外"
    for name, headers, rows in (
        (settlement_name, settlement_headers, config.get("settlements") or []),
        ("采购台账", purchase_headers, config.get("purchases") or []),
        ("银行流水", bank_headers, config.get("bank") or []),
        ("发票台账", invoice_headers, config.get("invoices") or []),
        ("工资表", payroll_headers, config.get("payroll") or []),
        ("期初余额", opening_headers, config.get("opening") or []),
        ("预算预测", plan_headers, config.get("plans") or []),
        ("经营KPI", kpi_headers, config.get("kpis") or []),
    ):
        _table(book.create_sheet(name), headers, rows)
    book.save(output_path)
    return _validate(output_path)


def build_onboarding_template(output_path: str | Path) -> Path:
    """Create the first-use master-data and opening template without private examples."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    intro = book.active
    intro.title = "初始化清单"
    _table(intro, ["顺序", "工作表", "谁来填", "填写内容", "是否必须"], [
        [1, "主体配置", "负责人/会计服务机构", "主体、准则、税务类型、申报周期、现金假设", "是"],
        [2, "游戏项目", "业务负责人", "稳定项目编码、阶段、负责人", "是"],
        [3, "渠道规则", "发行/运营", "平台、区域、币种、分成、结算公式、证据引用和回款天数", "是"],
        [4, "组织映射", "负责人", "匿名人员或岗位到项目的归属比例", "是"],
        [5, "供应商", "采购/业务", "供应商编码、类别、默认项目和账期", "建议"],
        [6, "期初科目余额", "会计服务机构", "上一期已确认期末余额", "关账前必须"],
        [7, "经营KPI", "运营", "月度活跃、付费、流水、投放和留存", "分析建议"],
    ])
    _table(book.create_sheet("主体配置"), ["配置项", "填写值", "说明"], [
        ["公司名称", "", "营业执照主体"], ["统一社会信用代码", "", "18位代码"],
        ["注册及主管税务地区", "", "省/市"], ["执行会计准则", "小企业会计准则", "由会计复核"],
        ["增值税纳税人类型", "", "一般纳税人/小规模纳税人"], ["增值税申报周期", "", "月度/季度"],
        ["外部会计/代账机构", "", "可留空"], ["复核联系人", "", "可留空"],
        ["预测起点现金", "", "人民币金额"], ["最低现金缓冲", "", "人民币金额"],
    ], {1: 30, 2: 36, 3: 46})
    _table(book.create_sheet("游戏项目"), ["项目编码", "游戏项目名称", "项目阶段", "上线日期", "负责人", "部门", "预算单元", "成本中心", "是否启用"], [])
    _table(book.create_sheet("渠道规则"), ["渠道编码", "渠道名称", "游戏项目编码", "平台", "区域", "结算币种", "收入模式", "分成比例", "结算公式", "合同证据引用", "生效月份", "结算周期", "回款天数", "是否启用"], [])
    _table(book.create_sheet("组织映射"), ["人员/岗位编码", "部门", "预算单元", "成本中心", "游戏项目编码", "分摊比例", "生效月份", "是否启用"], [])
    _table(book.create_sheet("供应商"), ["供应商编码", "供应商名称", "供应商类别", "默认币种", "归属项目编码", "付款天数", "是否启用"], [])
    _table(book.create_sheet("期初科目余额"), ["期间", "科目编码", "科目名称", "期初借方", "期初贷方"], [])
    _table(book.create_sheet("经营KPI"), ["月份", "游戏项目编码", "渠道", "区域", "DAU", "MAU", "新增用户", "付费用户", "安装数", "游戏流水", "投放金额", "D1留存", "D7留存", "D30留存"], [])
    _table(book.create_sheet("Checks"), ["检查", "要求", "当前状态"], [
        ["隐私", "组织映射使用匿名人员或岗位编码", "待填写"],
        ["币种", "不同币种不直接相加，汇率另在公司档案配置", "待填写"],
        ["期初", "借方合计应等于贷方合计", "待填写"],
        ["结论", "业务同学填事实，会计税务结论由Agent建议并交有权人复核", "说明"],
    ])
    book.save(output_path)
    return _validate(output_path)


def build_shadow_close_template(
    output_path: str | Path,
    *,
    entity_ids: Iterable[str] = (),
) -> Path:
    """Create a one-entity, one-period baseline workbook for read-only close comparison."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    intro = book.active
    intro.title = "使用说明"
    _table(intro, ["项目", "说明"], [
        ["用途", "导入人工已复核的关账基准，与 Agent 候选结果做只读对比"],
        ["范围", "每份工作簿只能包含一个法律主体和一个期间"],
        ["必填", "至少填写基准总账、基准报表或基准税务中的一张表"],
        ["容差", "绝对容差默认 1.00；百分比容差默认 0.1%；显式填 0 表示不使用该容差"],
        ["安全边界", "导入、差异解释和签认都不会覆盖台账、凭证、税表、银行或期间状态"],
    ], {1: 18, 2: 86})
    _table(book.create_sheet("基准总账"), ["主体ID", "期间", "科目编码", "科目名称", "期末借方", "期末贷方", "来源", "证据说明", "绝对容差", "百分比容差"], [])
    _table(book.create_sheet("基准报表"), ["主体ID", "期间", "指标编码", "指标名称", "金额", "来源", "证据说明", "绝对容差", "百分比容差"], [])
    _table(book.create_sheet("基准税务"), ["主体ID", "期间", "表单编码", "字段编码", "字段名称", "金额", "来源", "证据说明", "绝对容差", "百分比容差"], [])
    configured_entities = [str(item).strip() for item in entity_ids if str(item).strip()]
    entity_requirement = (
        "主体ID必须是当前 Box 已配置的法律主体：" + "、".join(configured_entities)
        if configured_entities
        else "主体ID必须是已配置的法律主体，如 cn_studio 或 sg_publisher"
    )
    _table(book.create_sheet("Checks"), ["检查", "要求"], [
        ["主体", entity_requirement],
        ["期间", "格式为 YYYY-MM，所有数据行必须一致"],
        ["唯一性", "同一域内科目或指标键不能重复"],
        ["复核", "基准数应来自人工已复核工作底稿，并保留来源和证据说明"],
    ], {1: 18, 2: 86})
    book.save(output_path)
    return _validate(output_path)
