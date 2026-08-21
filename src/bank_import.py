from __future__ import annotations

import csv
import hashlib
import math
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


class BankImportError(ValueError):
    """Raised when a bank export cannot satisfy the strict import contract."""


ALIASES = {
    "transaction_date": ("交易日期", "记账日期", "入账日期", "交易时间", "value date", "transaction date", "date"),
    "bank_transaction_id": ("交易流水号", "银行流水号", "交易编号", "reference", "transaction id", "流水号"),
    "account": ("本方账号", "银行账号", "账户号码", "account number", "account"),
    "counterparty": ("对方户名", "对方名称", "交易对手", "收款人", "付款人", "counterparty", "beneficiary"),
    "counterparty_account": ("对方账号", "收款账号", "付款账号", "counterparty account"),
    "summary": ("摘要", "用途", "附言", "交易备注", "remark", "description", "narrative"),
    "debit": ("支出金额", "借方发生额", "付款金额", "debit"),
    "credit": ("收入金额", "贷方发生额", "收款金额", "credit"),
    "amount": ("交易金额", "发生额", "金额", "amount"),
    "direction": ("收支方向", "借贷标志", "交易方向", "direction"),
    "currency": ("币种", "交易币种", "currency"),
    "balance": ("余额", "账户余额", "balance"),
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).replace("\n", " ").strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", _text(value).casefold())


def _field(value: Any) -> str | None:
    clean = _slug(value)
    candidates: list[tuple[int, str]] = []
    for field, aliases in ALIASES.items():
        for alias in aliases:
            alias_clean = _slug(alias)
            if alias_clean and (clean == alias_clean or alias_clean in clean):
                candidates.append((len(alias_clean), field))
    return max(candidates, default=(0, None))[1]


def _number(value: Any, field: str, *, required: bool = False) -> float | None:
    if value in (None, ""):
        if required:
            raise BankImportError(f"{field} is required")
        return None
    if isinstance(value, bool):
        raise BankImportError(f"{field} must be numeric")
    try:
        result = float(str(value).replace(",", "").replace("¥", "").replace("$", "").strip())
    except (TypeError, ValueError) as exc:
        raise BankImportError(f"{field} must be numeric") from exc
    if not math.isfinite(result):
        raise BankImportError(f"{field} must be finite")
    return result


def _iso_date(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    text = _text(value)
    match = re.fullmatch(r"(20\d{2})[-/.年](0?[1-9]|1[0-2])[-/.月](0?[1-9]|[12]\d|3[01])日?", text)
    if not match:
        raise BankImportError("transaction_date must use YYYY-MM-DD")
    return f"{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"


def mask_account(value: Any) -> str:
    text = re.sub(r"\s+", "", _text(value))
    if not text:
        return ""
    if "*" in text:
        return text[:40]
    if len(text) <= 8:
        return text
    return f"{text[:4]}****{text[-4:]}"


def mask_embedded_account_numbers(value: Any) -> str:
    text = _text(value)
    return re.sub(
        r"(?<!\d)\d{9,}(?!\d)",
        lambda match: f"{match.group(0)[:4]}****{match.group(0)[-4:]}",
        text,
    )[:160]


def _currency(value: Any) -> str:
    currency = _text(value).upper()
    if not re.fullmatch(r"[A-Z]{3}", currency):
        raise BankImportError("currency must be a three-letter code")
    return currency


def _amount_and_direction(raw: dict[str, Any]) -> tuple[float, str, str]:
    debit = _number(raw.get("debit"), "debit") or 0.0
    credit = _number(raw.get("credit"), "credit") or 0.0
    amount = _number(raw.get("amount"), "amount")
    direction_text = _text(raw.get("direction")).casefold()
    if debit and credit:
        raise BankImportError("debit and credit cannot both be non-zero")
    if credit:
        value, direction, code = abs(credit), "收入", "inflow"
    elif debit:
        value, direction, code = abs(debit), "支出", "outflow"
    elif amount is not None:
        if amount < 0 or any(word in direction_text for word in ("支", "借", "debit", "out")):
            value, direction, code = abs(amount), "支出", "outflow"
        elif any(word in direction_text for word in ("收", "贷", "credit", "in")):
            value, direction, code = abs(amount), "收入", "inflow"
        else:
            raise BankImportError("signed amount or explicit direction is required")
    else:
        raise BankImportError("debit, credit or amount is required")
    if value <= 0:
        raise BankImportError("transaction amount must be positive")
    return round(value, 2), direction, code


@dataclass(frozen=True)
class StandardBankTransaction:
    id: str
    bank_transaction_id: str
    transaction_id: str
    entity_id: str
    transaction_date: str
    account_masked: str
    counterparty: str
    counterparty_account_masked: str
    summary: str
    direction: str
    direction_code: str
    currency: str
    amount: float
    balance: float | None
    source_file: str
    source_sheet: str
    source_row: int
    status: str
    evidence: dict[str, Any]


def _record(
    raw: dict[str, Any], *, entity_id: str, default_currency: str,
    default_account: str, source_file: str, source_sheet: str,
    source_row: int, batch_id: str,
) -> StandardBankTransaction:
    transaction_id = _text(raw.get("bank_transaction_id"))
    if not transaction_id:
        raise BankImportError("bank_transaction_id is required for idempotent import")
    account_masked = mask_account(raw.get("account") or default_account)
    if not account_masked:
        raise BankImportError("account reference is required")
    amount, direction, direction_code = _amount_and_direction(raw)
    balance = _number(raw.get("balance"), "balance")
    currency = _currency(raw.get("currency") or default_currency)
    record_id = hashlib.sha256(f"{entity_id}|{transaction_id}".encode()).hexdigest()[:16]
    evidence = {
        "source_file": source_file, "source_sheet": source_sheet,
        "source_row": source_row, "batch_id": batch_id,
    }
    return StandardBankTransaction(
        id=record_id, bank_transaction_id=transaction_id, transaction_id=transaction_id,
        entity_id=entity_id, transaction_date=_iso_date(raw.get("transaction_date")),
        account_masked=account_masked, counterparty=_text(raw.get("counterparty")),
        counterparty_account_masked=mask_account(raw.get("counterparty_account")),
        summary=_text(raw.get("summary")), direction=direction, direction_code=direction_code,
        currency=currency, amount=amount, balance=round(balance, 2) if balance is not None else None,
        source_file=source_file, source_sheet=source_sheet, source_row=source_row,
        status="待人工确认", evidence=evidence,
    )


def _header_mapping(values: Iterable[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, value in enumerate(values):
        field = _field(value)
        if field and field not in mapping.values():
            mapping[index] = field
    return mapping


def _consume_rows(
    rows: Iterable[tuple[int, tuple[Any, ...]]], *, entity_id: str,
    default_currency: str, default_account: str, source_file: str,
    source_sheet: str, batch_id: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], bool]:
    mapping: dict[int, str] = {}
    records: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    for row_number, values in rows:
        candidate = _header_mapping(values)
        if (
            "transaction_date" in candidate.values()
            and "bank_transaction_id" in candidate.values()
            and bool({"amount", "debit", "credit"} & set(candidate.values()))
        ):
            mapping = candidate
            continue
        if not mapping or not any(_text(value) for value in values):
            continue
        raw = {field: values[index] if index < len(values) else None for index, field in mapping.items()}
        try:
            records.append(asdict(_record(
                raw, entity_id=entity_id, default_currency=default_currency,
                default_account=default_account, source_file=source_file,
                source_sheet=source_sheet, source_row=row_number, batch_id=batch_id,
            )))
        except BankImportError as exc:
            rejected.append({
                "dataset_type": "finance.bank_transactions", "row": row_number,
                "source_sheet": source_sheet, "reason": str(exc),
            })
    return records, rejected, bool(mapping)


def parse_bank_statement_file(
    path: str | Path, *, entity_id: str, default_currency: str = "",
    account_reference: str = "",
) -> dict[str, Any]:
    path = Path(path)
    if path.suffix.casefold() not in {".csv", ".xlsx"}:
        raise BankImportError("bank statement must be a .csv or .xlsx file")
    if not entity_id:
        raise BankImportError("entity_id is required")
    batch_id = hashlib.sha256(path.read_bytes()).hexdigest()[:16]
    safe_file_name = mask_embedded_account_numbers(path.name)
    records: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    if path.suffix.casefold() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [(index, tuple(row)) for index, row in enumerate(csv.reader(handle), 1)]
        accepted, errors, header_found = _consume_rows(
            rows, entity_id=entity_id, default_currency=default_currency,
            default_account=account_reference, source_file=safe_file_name,
            source_sheet="CSV", batch_id=batch_id,
        )
        records.extend(accepted)
        rejected.extend(errors)
        if not header_found:
            rejected.append({
                "dataset_type": "finance.bank_transactions", "row": 0,
                "source_sheet": "CSV",
                "reason": "bank statement header requires date, transaction id and amount/debit/credit",
            })
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        any_header = False
        try:
            for sheet in workbook.worksheets:
                rows = (
                    (index, tuple(row)) for index, row in enumerate(
                        sheet.iter_rows(values_only=True), 1
                    )
                )
                accepted, errors, header_found = _consume_rows(
                    rows, entity_id=entity_id, default_currency=default_currency,
                    default_account=account_reference, source_file=safe_file_name,
                    source_sheet=mask_embedded_account_numbers(sheet.title), batch_id=batch_id,
                )
                records.extend(accepted)
                rejected.extend(errors)
                any_header = any_header or header_found
        finally:
            workbook.close()
        if not any_header:
            rejected.append({
                "dataset_type": "finance.bank_transactions", "row": 0,
                "source_sheet": "workbook",
                "reason": "bank statement header requires date, transaction id and amount/debit/credit",
            })
    return {
        "batch_id": batch_id,
        "source": {
            "kind": "bank_statement_file", "name": safe_file_name,
            "network_access_performed": False, "raw_account_numbers_retained": False,
        },
        "records": records,
        "rejected_rows": rejected,
    }
