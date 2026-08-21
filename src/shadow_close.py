from __future__ import annotations

import hashlib
import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PERIOD_PATTERN = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")
ACCOUNT_PATTERN = re.compile(r"^([A-Za-z0-9._-]+)\s*(.*)$")
DECISIONS = {"验证通过", "接受差异", "退回补数"}
EXCEPTION_CLASSIFICATIONS = {
    "mapping", "cutoff", "accounting_policy", "source_evidence",
    "timing", "foreign_exchange", "accepted_scope", "system_defect",
}
COMPARISON_STATUSES = {"一致", "需解释", "Agent 缺项", "人工基准缺项"}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _number(value: Any, *, decimals: int = 2) -> float | None:
    if value in {None, ""}:
        return None
    if isinstance(value, bool):
        raise ValueError(f"金额必须是数字：{value}")
    try:
        parsed = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"金额必须是数字：{value}") from exc
    if not math.isfinite(parsed):
        raise ValueError(f"金额必须是有限数字：{value}")
    return round(parsed, decimals)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _headers(sheet, required: Iterable[str]) -> tuple[int, dict[str, int]]:
    """Find the table header so templates may include a short title/instruction block."""
    expected = set(required)
    max_row = min(sheet.max_row, 10) if sheet.max_row else 10
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row), 1):
        values = [_text(cell.value) for cell in row]
        headers = {value: index for index, value in enumerate(values) if value}
        if expected.issubset(headers):
            return row_number, headers
    return 1, {}


def _required(headers: dict[str, int], sheet_name: str, names: Iterable[str]) -> None:
    missing = [name for name in names if name not in headers]
    if missing:
        raise ValueError(f"{sheet_name} 缺少列：{'、'.join(missing)}")


def _cell(row: tuple, headers: dict[str, int], name: str) -> Any:
    index = headers.get(name)
    return row[index].value if index is not None and index < len(row) else None


def parse_shadow_close_workbook(path: str | Path) -> dict:
    """Read a human-produced close baseline without changing the finance ledger."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows: list[dict] = []
        specs = {
            "基准总账": ("trial_balance", ["主体ID", "期间", "科目编码", "科目名称", "期末借方", "期末贷方"]),
            "基准报表": ("statement", ["主体ID", "期间", "指标编码", "指标名称", "金额"]),
            "基准税务": ("tax", ["主体ID", "期间", "表单编码", "字段编码", "字段名称", "金额"]),
        }
        for sheet_name, (domain, required) in specs.items():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            header_row, headers = _headers(sheet, required)
            _required(headers, sheet_name, required)
            for source_row, row in enumerate(sheet.iter_rows(min_row=header_row + 1), header_row + 1):
                entity_id = _text(_cell(row, headers, "主体ID"))
                period = _text(_cell(row, headers, "期间"))
                if not entity_id and not period:
                    continue
                if not entity_id or not PERIOD_PATTERN.fullmatch(period):
                    raise ValueError(f"{sheet_name} 第{source_row}行主体ID或期间无效")
                absolute_tolerance = _number(
                    _cell(row, headers, "绝对容差"), decimals=6,
                )
                percent_tolerance = _number(
                    _cell(row, headers, "百分比容差"), decimals=6,
                )
                if (
                    absolute_tolerance is not None and absolute_tolerance < 0
                ) or (
                    percent_tolerance is not None and percent_tolerance < 0
                ):
                    raise ValueError(f"{sheet_name} 第{source_row}行容差不能为负数")
                base = {
                    "domain": domain, "entity_id": entity_id, "period": period,
                    "source": _text(_cell(row, headers, "来源")),
                    "evidence": _text(_cell(row, headers, "证据说明")),
                    "absolute_tolerance": 1.0 if absolute_tolerance is None else absolute_tolerance,
                    "percent_tolerance": 0.001 if percent_tolerance is None else percent_tolerance,
                    "source_sheet": sheet_name, "source_row": source_row,
                }
                if domain == "trial_balance":
                    code = _text(_cell(row, headers, "科目编码"))
                    if not code:
                        raise ValueError(f"{sheet_name} 第{source_row}行缺少科目编码")
                    debit = _number(_cell(row, headers, "期末借方")) or 0.0
                    credit = _number(_cell(row, headers, "期末贷方")) or 0.0
                    if debit < 0 or credit < 0 or (debit and credit):
                        raise ValueError(f"{sheet_name} 第{source_row}行借贷余额必须非负且不能同时有值")
                    base.update({"key": code, "name": _text(_cell(row, headers, "科目名称")), "value": round(debit - credit, 2)})
                elif domain == "statement":
                    code = _text(_cell(row, headers, "指标编码"))
                    if not code:
                        raise ValueError(f"{sheet_name} 第{source_row}行缺少指标编码")
                    base.update({"key": code, "name": _text(_cell(row, headers, "指标名称")), "value": _number(_cell(row, headers, "金额"))})
                else:
                    form = _text(_cell(row, headers, "表单编码"))
                    field = _text(_cell(row, headers, "字段编码"))
                    if not form or not field:
                        raise ValueError(f"{sheet_name} 第{source_row}行缺少表单或字段编码")
                    base.update({"key": f"{form}:{field}", "form_code": form, "field_code": field, "name": _text(_cell(row, headers, "字段名称")), "value": _number(_cell(row, headers, "金额"))})
                rows.append(base)
    finally:
        workbook.close()
    if not rows:
        raise ValueError("未找到基准总账、基准报表或基准税务数据")
    scopes = {(row["entity_id"], row["period"]) for row in rows}
    if len(scopes) != 1:
        raise ValueError("一份 shadow close 基准只能包含一个法律主体和一个期间")
    entity_id, period = next(iter(scopes))
    seen: set[tuple[str, str]] = set()
    for row in rows:
        identity = (row["domain"], row["key"])
        if identity in seen:
            raise ValueError(f"同一域存在重复基准键：{row['key']}")
        seen.add(identity)
    canonical = json.dumps(rows, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    fingerprint = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    return {
        "id": f"SHADOW-{entity_id}-{period}", "entity_id": entity_id, "period": period,
        "rows": rows, "row_count": len(rows), "source_fingerprint": fingerprint,
        "imported_at": _now(), "status": "待对比",
        "guardrail": "人工关账基准只用于只读对比，不覆盖台账、凭证、税表或期间状态。",
    }


def _account_key(account: str) -> tuple[str, str]:
    match = ACCOUNT_PATTERN.match(_text(account))
    return (match.group(1), match.group(2).strip()) if match else (_text(account), _text(account))


def _agent_rows(finance: dict) -> dict[str, dict[str, dict]]:
    result: dict[str, dict[str, dict]] = {"trial_balance": {}, "statement": {}, "tax": {}}
    for row in (finance.get("financial_statements") or {}).get("detail") or []:
        code, name = _account_key(row.get("account"))
        closing = (_number(row.get("closing_debit")) or 0) - (
            _number(row.get("closing_credit")) or 0
        )
        result["trial_balance"][code] = {"key": code, "name": name or row.get("account"), "value": round(closing, 2)}
    statements = finance.get("financial_statements") or {}
    balance = statements.get("balance_sheet") or {}
    income = statements.get("income_statement") or {}
    statement_values = {
        "BS_ASSETS": ("资产总额", balance.get("assets")),
        "BS_LIABILITIES": ("负债总额", balance.get("liabilities")),
        "BS_EQUITY": ("权益（含本期利润）", None if balance.get("liabilities_and_equity") is None or balance.get("liabilities") is None else float(balance.get("liabilities_and_equity")) - float(balance.get("liabilities"))),
        "IS_REVENUE": ("营业收入", income.get("revenue")),
        "IS_EXPENSES": ("成本费用", income.get("expenses")),
        "IS_PROFIT": ("利润总额", income.get("profit_before_tax")),
    }
    for code, (name, value) in statement_values.items():
        result["statement"][code] = {"key": code, "name": name, "value": _number(value)}
    workspace = (finance.get("tax_pack") or {}).get("returns_workspace") or {}
    for form in workspace.get("returns") or []:
        form_code = _text(form.get("form_code"))
        for field in form.get("fields") or []:
            key = f"{form_code}:{_text(field.get('code'))}"
            result["tax"][key] = {"key": key, "name": _text(field.get("name")), "value": _number(field.get("value")), "status": field.get("status")}
    return result


def shadow_close_report_fingerprint(report: dict) -> str:
    """Recompute the digest that binds one report to its exact scope and values."""
    fingerprint_payload = {
        "baseline_id": report.get("baseline_id"),
        "entity_id": report.get("entity_id"),
        "period": report.get("period"),
        "baseline": report.get("baseline_source_fingerprint"),
        "rows": [
            {
                key: row.get(key)
                for key in (
                    "domain", "key", "manual_value", "agent_value", "difference",
                    "allowed_tolerance", "status",
                )
            }
            for row in report.get("comparisons") or []
        ],
    }
    runtime_fingerprint = _text(report.get("runtime_fingerprint"))
    if runtime_fingerprint:
        fingerprint_payload["runtime_fingerprint"] = runtime_fingerprint
    canonical = json.dumps(
        fingerprint_payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def validate_shadow_close_report(report: dict) -> dict:
    """Fail closed before a human signs a serialized Shadow Close report."""
    if not isinstance(report, dict):
        raise ValueError("Shadow close 报告必须是 JSON 对象")
    for field in (
        "baseline_id", "entity_id", "period", "baseline_source_fingerprint",
        "report_fingerprint",
    ):
        if not _text(report.get(field)):
            raise ValueError(f"Shadow close 报告缺少字段：{field}")
    if not PERIOD_PATTERN.fullmatch(_text(report.get("period"))):
        raise ValueError("Shadow close 报告期间必须是 YYYY-MM")
    runtime_fingerprint = report.get("runtime_fingerprint")
    if runtime_fingerprint is not None and not re.fullmatch(
        r"[0-9a-f]{64}", _text(runtime_fingerprint),
    ):
        raise ValueError("Shadow close 报告 runtime_fingerprint 无效")
    comparisons = report.get("comparisons")
    if not isinstance(comparisons, list) or not comparisons:
        raise ValueError("Shadow close 报告必须包含至少一项比较")
    seen: set[tuple[str, str]] = set()
    for row in comparisons:
        if not isinstance(row, dict):
            raise ValueError("Shadow close 比较项必须是 JSON 对象")
        identity = (_text(row.get("domain")), _text(row.get("key")))
        if not all(identity) or identity in seen:
            raise ValueError("Shadow close 比较项的域和编码必须存在且唯一")
        seen.add(identity)
        if identity[0] not in {"trial_balance", "statement", "tax"}:
            raise ValueError("Shadow close 比较项的域无效")
        if row.get("status") not in COMPARISON_STATUSES:
            raise ValueError("Shadow close 比较项状态无效")
        tolerance = _number(row.get("allowed_tolerance"), decimals=6)
        if tolerance is None or tolerance < 0:
            raise ValueError("Shadow close 比较项容差必须是有限非负数")
        manual_value = _number(row.get("manual_value"))
        agent_value = _number(row.get("agent_value"))
        difference = _number(row.get("difference"))
        if manual_value is not None and agent_value is not None:
            expected_difference = round(agent_value - manual_value, 2)
            if difference != expected_difference:
                raise ValueError("Shadow close 比较项差异与人工/Agent 金额不一致")
            expected_status = "一致" if abs(expected_difference) <= tolerance else "需解释"
            if row.get("status") != expected_status:
                raise ValueError("Shadow close 比较项状态与明确容差不一致")
        elif difference is not None or row.get("status") == "一致":
            raise ValueError("Shadow close 缺项比较的差异或状态无效")
    matched = sum(row.get("status") == "一致" for row in comparisons)
    exceptions = len(comparisons) - matched
    if report.get("comparison_count") != len(comparisons):
        raise ValueError("Shadow close 报告比较总数与明细不一致")
    if report.get("matched_count") != matched or report.get("exception_count") != exceptions:
        raise ValueError("Shadow close 报告匹配/差异计数与明细不一致")
    expected_status = "一致待签认" if not exceptions else "存在差异"
    if report.get("status") != expected_status:
        raise ValueError("Shadow close 报告状态与比较明细不一致")
    supplied_domain_summary = report.get("domain_summary")
    if not isinstance(supplied_domain_summary, list):
        raise ValueError("Shadow close 报告缺少分域汇总")
    labels = {"trial_balance": "总账", "statement": "报表", "tax": "税务"}
    expected_domain_summary = []
    for domain in ("trial_balance", "statement", "tax"):
        scoped = [row for row in comparisons if row.get("domain") == domain]
        if scoped:
            expected_domain_summary.append({
                "domain": domain,
                "label": labels[domain],
                "count": len(scoped),
                "matched": sum(row.get("status") == "一致" for row in scoped),
                "exceptions": sum(row.get("status") != "一致" for row in scoped),
            })
    if supplied_domain_summary != expected_domain_summary:
        raise ValueError("Shadow close 报告分域汇总与比较明细不一致")
    expected_fingerprint = shadow_close_report_fingerprint(report)
    if report.get("report_fingerprint") != expected_fingerprint:
        raise ValueError("Shadow close 报告指纹与当前范围或比较值不一致")
    review, review_current = report.get("review"), report.get("review_current")
    if review_current:
        if not isinstance(review, dict):
            raise ValueError("Shadow close 当前签认缺少复核记录")
        for field, expected in (
            ("baseline_id", report.get("baseline_id")),
            ("entity_id", report.get("entity_id")),
            ("period", report.get("period")),
            ("report_fingerprint", expected_fingerprint),
        ):
            if review.get(field) != expected:
                raise ValueError(f"Shadow close 当前签认的 {field} 与报告不一致")
        if review.get("decision") not in DECISIONS:
            raise ValueError("Shadow close 当前签认决定无效")
        if not _text(review.get("actor")) or len(_text(review.get("rationale"))) < 6:
            raise ValueError("Shadow close 当前签认缺少有效复核人或依据")
    elif review is not None:
        raise ValueError("Shadow close 非当前签认不能附在当前报告上")
    return {
        "valid": True,
        "report_fingerprint": expected_fingerprint,
        "comparison_count": len(comparisons),
        "matched_count": matched,
        "exception_count": exceptions,
    }


def compare_shadow_close(
    baseline: dict,
    finance: dict,
    reviews: Iterable[dict] = (),
    *,
    runtime_fingerprint: str | None = None,
) -> dict:
    agent = _agent_rows(finance)
    comparisons = []
    baseline_by_domain: dict[str, dict[str, dict]] = {"trial_balance": {}, "statement": {}, "tax": {}}
    for row in baseline.get("rows") or []:
        baseline_by_domain[row["domain"]][row["key"]] = row
    included_domains = [domain for domain in ("trial_balance", "statement", "tax") if baseline_by_domain[domain]]
    for domain in included_domains:
        # Tax workbooks intentionally allow a verified subset of fields; absent tax fields are
        # out of the validation scope, not missing manual evidence. Ledger/report baselines remain
        # completeness checks and therefore compare the union.
        keys = sorted(
            set(baseline_by_domain[domain])
            if domain == "tax"
            else set(baseline_by_domain[domain]) | set(agent[domain])
        )
        for key in keys:
            manual = baseline_by_domain[domain].get(key)
            candidate = agent[domain].get(key)
            manual_value = manual.get("value") if manual else None
            agent_value = candidate.get("value") if candidate else None
            tolerance_value = (manual or {}).get("absolute_tolerance")
            percent_value = (manual or {}).get("percent_tolerance")
            try:
                tolerance = 1.0 if tolerance_value is None else float(tolerance_value)
                percent_tolerance = 0.001 if percent_value is None else float(percent_value)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"Shadow close 容差必须是数字：{domain}/{key}") from exc
            if (
                not math.isfinite(tolerance) or tolerance < 0
                or not math.isfinite(percent_tolerance) or percent_tolerance < 0
            ):
                raise ValueError(f"Shadow close 容差必须是有限非负数：{domain}/{key}")
            allowed = round(max(
                tolerance,
                abs(float(manual_value)) * percent_tolerance
                if manual_value is not None else tolerance,
            ), 6)
            if manual is None:
                status, reason, difference = "人工基准缺项", "Agent 有结果，但人工基准未提供对应行", None
            elif candidate is None or agent_value is None:
                status, reason, difference = "Agent 缺项", "人工基准有金额，但 Agent 当前无法形成候选值", None
            elif manual_value is None:
                status, reason, difference = "人工基准缺项", "人工基准未填写金额", None
            else:
                difference = round(float(agent_value) - float(manual_value), 2)
                if abs(difference) <= allowed:
                    status, reason = "一致", f"差异在容差 {allowed:,.2f} 内"
                else:
                    status, reason = "需解释", f"差异超过容差 {allowed:,.2f}"
            comparisons.append({
                "domain": domain, "key": key, "name": (manual or candidate or {}).get("name") or key,
                "manual_value": manual_value, "agent_value": agent_value, "difference": difference,
                "allowed_tolerance": allowed,
                "status": status, "reason": reason, "source": (manual or {}).get("source"),
                "evidence": (manual or {}).get("evidence"),
            })
    material = [row for row in comparisons if row["status"] != "一致"]
    domain_summary = []
    labels = {"trial_balance": "总账", "statement": "报表", "tax": "税务"}
    for domain in included_domains:
        scoped = [row for row in comparisons if row["domain"] == domain]
        domain_summary.append({
            "domain": domain, "label": labels[domain], "count": len(scoped),
            "matched": sum(row["status"] == "一致" for row in scoped),
            "exceptions": sum(row["status"] != "一致" for row in scoped),
        })
    fingerprint_input = {
        "baseline_id": baseline.get("id"),
        "entity_id": baseline.get("entity_id"),
        "period": baseline.get("period"),
        "baseline_source_fingerprint": baseline.get("source_fingerprint"),
        "comparisons": comparisons,
    }
    if runtime_fingerprint is not None:
        fingerprint_input["runtime_fingerprint"] = runtime_fingerprint
    report_fingerprint = shadow_close_report_fingerprint(fingerprint_input)
    current_review = next((row for row in reversed(list(reviews)) if row.get("baseline_id") == baseline.get("id") and row.get("report_fingerprint") == report_fingerprint), None)
    report = {
        "baseline_id": baseline.get("id"), "entity_id": baseline.get("entity_id"), "period": baseline.get("period"),
        "baseline_source_fingerprint": baseline.get("source_fingerprint"),
        "status": "一致待签认" if not material else "存在差异", "comparison_count": len(comparisons),
        "matched_count": len(comparisons) - len(material), "exception_count": len(material),
        "domain_summary": domain_summary, "comparisons": comparisons,
        "report_fingerprint": report_fingerprint, "review": current_review,
        "review_current": bool(current_review),
        "recommendation": "差异为零或在容差内，可由独立复核人签认验证通过。" if not material else "先逐项解释口径、截止、映射或缺失证据；不自动把人工数覆盖到 Agent。",
        "guardrail": "Shadow close 是只读验证：差异和签认不改变法定账、税表、银行流水或审批记录。",
    }
    if runtime_fingerprint is not None:
        report["runtime_fingerprint"] = runtime_fingerprint
    return report


def review_shadow_close(
    report: dict,
    decision: str,
    actor: str,
    rationale: str,
    evidence: Iterable[str],
    exception_resolutions: Iterable[dict] = (),
) -> dict:
    validate_shadow_close_report(report)
    if decision not in DECISIONS:
        raise ValueError("签认决定只能是验证通过、接受差异或退回补数")
    actor, rationale = _text(actor), _text(rationale)
    evidence = [_text(item)[:500] for item in evidence if _text(item)]
    if not actor or len(rationale) < 6:
        raise ValueError("请填写复核人和至少6个字的签认依据")
    if decision == "验证通过" and report.get("exception_count"):
        raise ValueError("仍有差异或缺项，不能签认验证通过")
    if decision == "接受差异" and not evidence:
        raise ValueError("接受差异必须附至少一项差异解释证据")
    material = {
        (str(row.get("domain") or ""), str(row.get("key") or ""))
        for row in report.get("comparisons") or [] if row.get("status") != "一致"
    }
    resolutions = []
    for item in exception_resolutions:
        if not isinstance(item, dict) or set(item) != {
            "domain", "key", "classification", "rationale", "evidence_references",
        }:
            raise ValueError("差异处置必须包含域、编码、分类、说明和证据引用")
        identity = (str(item.get("domain") or ""), str(item.get("key") or ""))
        classification = str(item.get("classification") or "")
        resolution_rationale = _text(item.get("rationale"))
        resolution_evidence = [
            _text(reference)[:500]
            for reference in item.get("evidence_references") or [] if _text(reference)
        ]
        if identity not in material:
            raise ValueError("差异处置指向了当前报告中不存在的差异")
        if classification not in EXCEPTION_CLASSIFICATIONS:
            raise ValueError("差异分类无效")
        if len(resolution_rationale) < 12 or not resolution_evidence:
            raise ValueError("每项差异处置必须有至少12个字说明和证据引用")
        resolutions.append({
            "domain": identity[0], "key": identity[1],
            "classification": classification,
            "rationale": resolution_rationale[:1000],
            "evidence_references": resolution_evidence,
        })
    resolution_identities = {(item["domain"], item["key"]) for item in resolutions}
    if len(resolution_identities) != len(resolutions):
        raise ValueError("同一差异不能重复处置")
    if decision == "接受差异" and resolution_identities != material:
        raise ValueError("接受差异必须逐项提供完整处置和证据")
    if decision != "接受差异" and resolutions:
        raise ValueError("只有“接受差异”可以附加差异处置")
    review_seed = f"{report.get('report_fingerprint')}|{actor}|{_now()}"
    return {
        "id": f"SHADOW-REVIEW-{hashlib.sha1(review_seed.encode()).hexdigest()[:16]}",
        "baseline_id": report.get("baseline_id"), "entity_id": report.get("entity_id"), "period": report.get("period"),
        "report_fingerprint": report.get("report_fingerprint"), "decision": decision,
        "actor": actor[:80], "rationale": rationale[:1000], "evidence": evidence,
        "exception_resolutions": resolutions, "reviewed_at": _now(),
        "scope_note": "该签认只对当前基准和当前 Agent 结果指纹有效；任一数据变化后自动失效。",
    }
