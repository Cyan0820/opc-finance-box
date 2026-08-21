from __future__ import annotations

import hashlib
import math
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


IIT_BRACKETS = (
    (36_000, 0.03, 0),
    (144_000, 0.10, 2_520),
    (300_000, 0.20, 16_920),
    (420_000, 0.25, 31_920),
    (660_000, 0.30, 52_920),
    (960_000, 0.35, 85_920),
    (float("inf"), 0.45, 181_920),
)


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", _text(value).lower())


def _number(value: Any) -> float:
    if value is None or value == "" or isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return 0.0 if math.isnan(value) or math.isinf(value) else float(value)
    match = re.search(r"-?[\d,.]+", _text(value))
    return float(match.group().replace(",", "")) if match else 0.0


ALIASES = {
    "employee_id": ("工号", "员工编号", "人员编号", "employee id", "staff id"),
    "name": ("姓名", "员工姓名", "employee name", "staff name"),
    "department": ("部门", "成本中心", "department", "cost centre", "cost center"),
    "project": ("项目", "游戏", "研发项目", "project", "game"),
    "gross_salary": ("应发工资", "应发合计", "税前工资", "工资薪金", "gross salary", "gross pay"),
    "social_security": ("个人社保", "社保个人", "社保扣款"),
    "housing_fund": ("个人公积金", "公积金个人", "公积金扣款"),
    "special_deduction": ("专项附加扣除", "专项附加"),
    "other_deduction": ("其他扣除", "依法确定其他扣除"),
    "tax_exempt": ("免税收入",),
    "cumulative_income": ("累计收入", "本年累计收入"),
    "cumulative_deductions": ("累计扣除", "本年累计扣除"),
    "tax_paid_ytd": ("累计已预扣税额", "已预缴个税", "累计个税"),
    "iit": ("本月个税", "个人所得税", "个税"),
    "withholding_tax": ("withholding tax", "employee tax", "tax withheld", "代扣税"),
    "employee_deductions": ("employee deductions", "employee contribution", "员工扣款", "雇员扣款"),
    "employer_contributions": ("employer cpf", "employer contribution", "employer contributions", "雇主供款"),
    "employer_levies": ("sdl", "employer levy", "employer levies", "雇主征费"),
    "other_employer_cost": ("employer benefits", "other employer cost", "其他雇主成本"),
    "currency": ("币种", "currency"),
    "net_salary": ("实发工资", "实发合计", "net salary", "net pay"),
    "rd_ratio": ("研发工时占比", "研发比例", "研发占比"),
    "allocation_ratio": ("项目分摊比例", "项目投入比例", "project allocation ratio"),
    "timesheet_hours": ("本项目工时", "项目工时", "project hours"),
    "total_hours": ("本期总工时", "总工时", "total hours"),
    "allocation_evidence": ("工时证据", "分摊证据", "timesheet evidence", "allocation evidence"),
    "allocation_evidence_type": ("证据类型", "工时证据类型", "evidence type"),
    "activity_type": ("活动性质", "研发活动类型", "activity type"),
}


def _field(value: Any) -> str | None:
    clean = _slug(value)
    candidates = []
    for field_name, aliases in ALIASES.items():
        for alias in aliases:
            alias_clean = _slug(alias)
            if alias_clean and (clean == alias_clean or alias_clean in clean):
                candidates.append((len(alias_clean), field_name))
    return max(candidates, default=(0, None))[1]


def cumulative_withholding_tax(
    cumulative_income: float,
    cumulative_tax_exempt: float,
    cumulative_basic_deduction: float,
    cumulative_special: float,
    cumulative_special_additional: float,
    cumulative_other: float,
    cumulative_relief: float,
    tax_paid_ytd: float,
) -> dict[str, float]:
    taxable = max(0.0, cumulative_income - cumulative_tax_exempt - cumulative_basic_deduction
                  - cumulative_special - cumulative_special_additional - cumulative_other)
    upper, rate, quick = next(bracket for bracket in IIT_BRACKETS if taxable <= bracket[0])
    cumulative_tax = max(0.0, taxable * rate - quick - cumulative_relief)
    current_tax = max(0.0, cumulative_tax - tax_paid_ytd)
    return {
        "cumulative_taxable_income": round(taxable, 2),
        "rate": rate,
        "quick_deduction": quick,
        "cumulative_tax": round(cumulative_tax, 2),
        "current_tax": round(current_tax, 2),
    }


@dataclass
class PayrollRecord:
    id: str
    source_file: str
    source_sheet: str
    source_row: int
    period: str
    employee_masked: str
    department: str
    project: str
    gross_salary: float
    social_security: float
    housing_fund: float
    special_deduction: float
    other_deduction: float
    calculated_iit: float
    declared_iit: float | None
    net_salary: float
    rd_ratio: float
    rd_salary_candidate: float
    status: str
    anomalies: list[str] = field(default_factory=list)
    currency: str = "CNY"
    jurisdiction: str = "CN"
    employee_deductions: float = 0.0
    withholding_tax: float = 0.0
    employer_contributions: float = 0.0
    employer_levies: float = 0.0
    other_employer_cost: float = 0.0
    total_employer_cost: float = 0.0
    payroll_basis: str = "CN_CUMULATIVE_WITHHOLDING_CANDIDATE"
    statutory_calculation_status: str = "system_candidate"
    allocation_ratio: float = 0.0
    timesheet_hours: float = 0.0
    total_hours: float = 0.0
    allocation_evidence: list[str] = field(default_factory=list)
    allocation_evidence_type: str = ""
    activity_type: str = ""


def parse_payroll_workbook(
    path: str | Path, period: str, jurisdiction: str = "CN", functional_currency: str = "CNY",
) -> list[PayrollRecord]:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    records = []
    month = int(period.split("-")[1])
    jurisdiction = _text(jurisdiction).upper() or "CN"
    functional_currency = _text(functional_currency).upper() or ("CNY" if jurisdiction == "CN" else "USD")
    for sheet in workbook.worksheets:
        mapping = {}
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            candidate = {}
            for index, value in enumerate(row):
                name = _field(value)
                if name and name not in candidate.values():
                    candidate[index] = name
            if len(candidate) >= 4 and "gross_salary" in candidate.values():
                mapping = candidate
                continue
            if not mapping:
                continue
            raw = {name: row[index] if index < len(row) else None for index, name in mapping.items()}
            gross = _number(raw.get("gross_salary"))
            if gross == 0 and not _text(raw.get("name")):
                continue
            social = abs(_number(raw.get("social_security")))
            fund = abs(_number(raw.get("housing_fund")))
            special_additional = abs(_number(raw.get("special_deduction")))
            other = abs(_number(raw.get("other_deduction")))
            currency = _text(raw.get("currency")).upper() or functional_currency
            declared_source = raw.get("iit") if raw.get("iit") not in (None, "") else raw.get("withholding_tax")
            declared = abs(_number(declared_source)) if declared_source not in (None, "") else None
            net = _number(raw.get("net_salary"))
            ratio = _number(raw.get("rd_ratio"))
            if ratio > 1 and ratio <= 100:
                ratio /= 100
            ratio = max(0, min(1, ratio))
            allocation_ratio = _number(raw.get("allocation_ratio"))
            if allocation_ratio > 1 and allocation_ratio <= 100:
                allocation_ratio /= 100
            allocation_ratio = max(0, min(1, allocation_ratio))
            timesheet_hours = max(0, _number(raw.get("timesheet_hours")))
            total_hours = max(0, _number(raw.get("total_hours")))
            allocation_evidence = list(dict.fromkeys(
                item.strip()[:200] for item in re.split(r"[；;\n]+", _text(raw.get("allocation_evidence")))
                if item.strip()
            ))
            anomalies = []
            if jurisdiction == "CN":
                tax_exempt = abs(_number(raw.get("tax_exempt")))
                cumulative_income = _number(raw.get("cumulative_income")) or gross * month
                cumulative_deductions = abs(_number(raw.get("cumulative_deductions")))
                if cumulative_deductions:
                    cumulative_special = cumulative_deductions
                    basic_deduction = 5000 * month
                else:
                    cumulative_special = (social + fund) * month
                    basic_deduction = 5000 * month
                tax_paid_ytd = abs(_number(raw.get("tax_paid_ytd")))
                tax = cumulative_withholding_tax(
                    cumulative_income, tax_exempt, basic_deduction, cumulative_special,
                    special_additional * month, other * month, 0, tax_paid_ytd,
                )
                withholding = tax["current_tax"]
                employee_deductions = social + fund
                employer_contributions = employer_levies = other_employer_cost = 0.0
                payroll_basis = "CN_CUMULATIVE_WITHHOLDING_CANDIDATE"
                calculation_status = "system_candidate"
                if declared is not None and abs(declared - withholding) > 1:
                    anomalies.append(f"表内个税与累计预扣试算差异 {declared - withholding:,.2f}")
            else:
                # Overseas payroll is an imported local-payroll result. The workbench deliberately
                # does not infer statutory rates because eligibility and rates depend on employee facts.
                withholding = declared or 0.0
                explicit_deductions = abs(_number(raw.get("employee_deductions")))
                employee_deductions = explicit_deductions or (social + fund + other)
                employer_contributions = abs(_number(raw.get("employer_contributions")))
                employer_levies = abs(_number(raw.get("employer_levies")))
                other_employer_cost = abs(_number(raw.get("other_employer_cost")))
                payroll_basis = "IMPORTED_LOCAL_PAYROLL"
                calculation_status = "imported_pending_local_review"
            if not net:
                net = gross - employee_deductions - withholding
                if jurisdiction != "CN":
                    anomalies.append("未提供实发工资，当前由应发减员工扣款及代扣税倒推，需当地 payroll 报告确认")
            payroll_gap = round(gross - employee_deductions - withholding - net, 2)
            if abs(payroll_gap) > max(1, abs(gross) * 0.001):
                anomalies.append(f"应发、员工扣款、代扣税与实发未勾稽，差异 {payroll_gap:,.2f}")
            if ratio and not _text(raw.get("project")):
                anomalies.append("填报研发占比但缺少研发项目")
            employee_key = _text(raw.get("employee_id")) or _text(raw.get("name"))
            masked = hashlib.sha1(employee_key.encode("utf-8")).hexdigest()[:8] if employee_key else "匿名人员"
            row_key = f"{path.name}|{sheet.title}|{row_number}|{employee_key}|{gross}"
            records.append(PayrollRecord(
                id=hashlib.sha1(row_key.encode("utf-8")).hexdigest()[:12],
                source_file=path.name, source_sheet=sheet.title, source_row=row_number,
                period=period,
                employee_masked=f"员工-{masked}", department=_text(raw.get("department")) or "待分配部门",
                project=_text(raw.get("project")), gross_salary=round(gross, 2),
                social_security=round(social, 2), housing_fund=round(fund, 2),
                special_deduction=round(special_additional, 2), other_deduction=round(other, 2),
                calculated_iit=withholding if jurisdiction == "CN" else 0.0, declared_iit=declared,
                net_salary=round(net, 2), rd_ratio=round(ratio, 4),
                rd_salary_candidate=round(gross * ratio, 2),
                status="异常" if anomalies else "待复核", anomalies=anomalies,
                currency=currency, jurisdiction=jurisdiction,
                employee_deductions=round(employee_deductions, 2), withholding_tax=round(withholding, 2),
                employer_contributions=round(employer_contributions, 2), employer_levies=round(employer_levies, 2),
                other_employer_cost=round(other_employer_cost, 2),
                total_employer_cost=round(gross + employer_contributions + employer_levies + other_employer_cost, 2),
                payroll_basis=payroll_basis, statutory_calculation_status=calculation_status,
                allocation_ratio=round(allocation_ratio, 4), timesheet_hours=round(timesheet_hours, 2),
                total_hours=round(total_hours, 2), allocation_evidence=allocation_evidence,
                allocation_evidence_type=_text(raw.get("allocation_evidence_type"))[:100],
                activity_type=_text(raw.get("activity_type"))[:100],
            ))
    workbook.close()
    return records


def payroll_payload(records: Iterable[PayrollRecord], period: str) -> dict:
    rows = list(records)
    jurisdictions = sorted({row.jurisdiction for row in rows})
    currencies = sorted({row.currency for row in rows})
    is_cn = jurisdictions == ["CN"] or not jurisdictions
    return {
        "period": period,
        "jurisdiction": jurisdictions[0] if len(jurisdictions) == 1 else "MIXED",
        "currency": currencies[0] if len(currencies) == 1 else "MIXED",
        "records": [asdict(row) for row in rows],
        "summary": {
            "employee_count": len(rows),
            "gross_salary": round(sum(row.gross_salary for row in rows), 2),
            "calculated_iit": round(sum(row.calculated_iit for row in rows), 2),
            "withholding_tax": round(sum(row.withholding_tax for row in rows), 2),
            "employee_deductions": round(sum(row.employee_deductions for row in rows), 2),
            "employer_contributions": round(sum(row.employer_contributions for row in rows), 2),
            "employer_levies": round(sum(row.employer_levies for row in rows), 2),
            "total_employer_cost": round(sum(row.total_employer_cost for row in rows), 2),
            "net_salary": round(sum(row.net_salary for row in rows), 2),
            "rd_salary_candidate": round(sum(row.rd_salary_candidate for row in rows), 2),
            "exception_count": sum(bool(row.anomalies) for row in rows),
        },
        "decision_support": {
            "recommendation": (
                "中国主体使用累计预扣法生成候选税额；研发工资候选额仍需项目、人员角色和工时记录佐证。"
                if is_cn else
                "海外主体只导入并校验当地已核准 payroll 结果；系统不猜法定供款、征费或代扣税比例，提交前由当地服务机构复核。"
            ),
            "business_questions": ["这个人本月是否在职并应取得这笔工资？", "研发占比是否有工时或项目记录支持？"],
        },
    }
