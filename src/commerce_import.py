from __future__ import annotations

import csv
import hashlib
import io
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .commerce import (
    CommerceDataError,
    CommerceOrder,
    CommerceReturn,
    CommerceReturnReceipt,
    CommerceSettlement,
)
from .commerce_import_costs import CommerceImportCost, ImportCostDataError


ORDER_FIELDS = (
    "order_id", "entity_id", "period", "channel", "destination_country", "currency",
    "merchandise_gross_ex_tax", "discounts_ex_tax", "shipping_income_ex_tax",
    "tax_collected", "refunds_ex_tax", "refunded_tax", "cogs", "fulfillment_cost",
    "shipping_cost",
)
SETTLEMENT_FIELDS = (
    "settlement_id", "entity_id", "period", "channel", "currency",
    "reported_order_inflow", "channel_and_payment_fees", "tax_withheld_or_remitted",
    "other_adjustments", "payout",
)
RETURN_FIELDS = (
    "return_id", "order_id", "entity_id", "period", "channel", "sku", "currency",
    "authorized_quantity", "refunded_quantity", "refund_amount_ex_tax", "refunded_tax",
)
RETURN_RECEIPT_FIELDS = (
    "receipt_id", "return_id", "entity_id", "period", "sku", "warehouse",
    "received_quantity", "disposition",
)
IMPORT_COST_FIELDS = (
    "entry_line_id", "import_entry_id", "entity_id", "period", "sku", "warehouse",
    "origin_country", "destination_country", "currency", "quantity", "declared_value",
    "inbound_freight", "insurance", "customs_duty", "import_tax", "brokerage",
)
ORDER_REQUIRED = {
    "order_id", "entity_id", "period", "channel", "destination_country", "currency",
    "merchandise_gross_ex_tax",
}
SETTLEMENT_REQUIRED = {
    "settlement_id", "entity_id", "period", "channel", "currency",
    "reported_order_inflow", "payout",
}
RETURN_REQUIRED = {
    "return_id", "order_id", "entity_id", "period", "channel", "sku", "currency",
    "authorized_quantity",
}
RETURN_RECEIPT_REQUIRED = set(RETURN_RECEIPT_FIELDS)
IMPORT_COST_REQUIRED = set(IMPORT_COST_FIELDS[:9]) | {"quantity"}
NUMERIC_FIELDS = (
    set(ORDER_FIELDS[6:]) | set(SETTLEMENT_FIELDS[5:])
    | set(RETURN_FIELDS[7:]) | {"received_quantity"} | set(IMPORT_COST_FIELDS[9:])
)
TABLE_FIELDS = {
    "orders": ORDER_FIELDS,
    "settlements": SETTLEMENT_FIELDS,
    "returns": RETURN_FIELDS,
    "return_receipts": RETURN_RECEIPT_FIELDS,
    "import_costs": IMPORT_COST_FIELDS,
}
TABLE_REQUIRED = {
    "orders": ORDER_REQUIRED,
    "settlements": SETTLEMENT_REQUIRED,
    "returns": RETURN_REQUIRED,
    "return_receipts": RETURN_RECEIPT_REQUIRED,
    "import_costs": IMPORT_COST_REQUIRED,
}
TABLE_PARSERS = {
    "orders": CommerceOrder.from_dict,
    "settlements": CommerceSettlement.from_dict,
    "returns": CommerceReturn.from_dict,
    "return_receipts": CommerceReturnReceipt.from_dict,
    "import_costs": CommerceImportCost.from_dict,
}
TABLE_MINIMUM_SCORES = {
    "orders": 5, "settlements": 4, "returns": 6, "return_receipts": 6, "import_costs": 8,
}

ALIASES = {
    "order_id": ("order_id", "订单id", "订单号", "订单编号", "order number", "name"),
    "settlement_id": ("settlement_id", "结算id", "结算编号", "打款批次", "payout id"),
    "return_id": ("return_id", "退货id", "退货单号", "退货授权号", "return number", "rma"),
    "receipt_id": ("receipt_id", "收货事件id", "退货入库单号", "收货单号", "receipt id"),
    "entity_id": ("entity_id", "法律主体id", "主体id", "公司主体", "legal entity"),
    "period": ("period", "期间", "月份", "账期", "结算期间"),
    "channel": ("channel", "渠道", "销售渠道", "店铺", "平台"),
    "destination_country": ("destination_country", "目的地国家", "收货国家", "国家地区", "country"),
    "currency": ("currency", "币种", "交易币种", "结算币种"),
    "merchandise_gross_ex_tax": (
        "merchandise_gross_ex_tax", "商品原价不含税", "商品销售额不含税", "gross sales ex tax"
    ),
    "discounts_ex_tax": ("discounts_ex_tax", "折扣不含税", "discount ex tax"),
    "shipping_income_ex_tax": ("shipping_income_ex_tax", "运费收入不含税", "shipping income ex tax"),
    "tax_collected": ("tax_collected", "已收税额", "销售税", "tax collected"),
    "refunds_ex_tax": ("refunds_ex_tax", "退款不含税", "refund ex tax"),
    "refunded_tax": ("refunded_tax", "退回税额", "refunded tax"),
    "cogs": ("cogs", "商品成本", "销售成本", "product cost"),
    "fulfillment_cost": ("fulfillment_cost", "履约成本", "仓储履约成本"),
    "shipping_cost": ("shipping_cost", "物流成本", "配送成本", "shipping cost"),
    "reported_order_inflow": (
        "reported_order_inflow", "渠道报告订单净流入", "订单资金净流入", "reported order inflow"
    ),
    "channel_and_payment_fees": (
        "channel_and_payment_fees", "渠道及支付费用", "支付手续费", "payment fees"
    ),
    "tax_withheld_or_remitted": (
        "tax_withheld_or_remitted", "渠道代扣代缴税额", "代扣税额", "tax withheld"
    ),
    "other_adjustments": ("other_adjustments", "其他调整", "adjustments"),
    "payout": ("payout", "实际打款", "结算到账", "实付金额"),
    "sku": ("sku", "商品sku", "库存编码", "商品编码"),
    "authorized_quantity": ("authorized_quantity", "授权退货数量", "退货授权数量"),
    "refunded_quantity": ("refunded_quantity", "已退款数量", "退款数量"),
    "refund_amount_ex_tax": ("refund_amount_ex_tax", "退款金额不含税", "退货退款不含税"),
    "warehouse": ("warehouse", "仓库", "收货仓", "仓库编码"),
    "received_quantity": ("received_quantity", "实收数量", "退货入库数量", "收货数量"),
    "disposition": ("disposition", "处置状态", "质检处置", "库存处置"),
    "entry_line_id": ("entry_line_id", "进口明细行id", "报关明细行号"),
    "import_entry_id": ("import_entry_id", "进口批次号", "报关单号", "进口申报号"),
    "origin_country": ("origin_country", "原产国", "起运国家"),
    "quantity": ("quantity", "进口数量", "申报数量"),
    "declared_value": ("declared_value", "申报货值", "报关货值"),
    "inbound_freight": ("inbound_freight", "进口运费", "入境运费"),
    "insurance": ("insurance", "保险费", "运输保险"),
    "customs_duty": ("customs_duty", "关税", "关税金额"),
    "import_tax": ("import_tax", "进口税", "进口增值税"),
    "brokerage": ("brokerage", "报关服务费", "清关服务费"),
}


def _slug(value: Any) -> str:
    return re.sub(r"[\s_\-/（）()]+", "", str(value or "").strip().lower())


ALIAS_LOOKUP = {
    _slug(alias): field
    for field, aliases in ALIASES.items()
    for alias in aliases
}


def _period(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    text = str(value or "").strip()
    match = re.search(r"(20\d{2})[-/.年]?(1[0-2]|0?[1-9])", text)
    return f"{match.group(1)}-{int(match.group(2)):02d}" if match else text


def _decode_csv(path: Path) -> tuple[list[list[Any]], str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "gb18030", "big5"):
        try:
            text = raw.decode(encoding)
            return [list(row) for row in csv.reader(io.StringIO(text))], encoding
        except UnicodeDecodeError:
            continue
    raise CommerceDataError("CSV encoding must be UTF-8, GB18030 or Big5")


def _header_mapping(row: Iterable[Any]) -> dict[int, str]:
    mapping = {}
    used = set()
    for index, value in enumerate(row):
        field = ALIAS_LOOKUP.get(_slug(value))
        if field and field not in used:
            mapping[index] = field
            used.add(field)
    return mapping


def _detect_table(rows: list[list[Any]], record_type: str | None = None) -> tuple[int, str, dict[int, str]]:
    candidates = []
    for row_index, row in enumerate(rows[:30]):
        mapping = _header_mapping(row)
        for table_type, fields in TABLE_FIELDS.items():
            score = len(set(mapping.values()) & set(fields))
            candidates.append((score, row_index, table_type, mapping))
    if record_type:
        if record_type not in TABLE_FIELDS:
            raise CommerceDataError(
                "record_type must be orders, settlements, returns, return_receipts or import_costs"
            )
        candidates = [item for item in candidates if item[2] == record_type]
    score, row_index, detected_type, mapping = max(candidates, default=(0, 0, "", {}))
    minimum = TABLE_MINIMUM_SCORES.get(detected_type, 99)
    if score < minimum:
        raise CommerceDataError("Cannot locate a supported commerce header row")
    return row_index, detected_type, mapping


def _normalize_rows(
    rows: list[list[Any]],
    *,
    source_file: str,
    source_sheet: str,
    batch_id: str,
    record_type: str | None = None,
    default_entity_id: str | None = None,
    default_channel: str | None = None,
) -> dict[str, Any]:
    header_index, detected_type, mapping = _detect_table(rows, record_type)
    required = TABLE_REQUIRED[detected_type]
    fields = TABLE_FIELDS[detected_type]
    records = []
    rejected = []
    for row_index, values in enumerate(rows[header_index + 1:], header_index + 2):
        if not any(value not in (None, "") for value in values):
            continue
        record = {
            field: values[column] if column < len(values) else None
            for column, field in mapping.items()
        }
        if default_entity_id and not record.get("entity_id"):
            record["entity_id"] = default_entity_id
        if default_channel and not record.get("channel"):
            record["channel"] = default_channel
        if "period" in record:
            record["period"] = _period(record["period"])
        for field in NUMERIC_FIELDS:
            if field in fields:
                if record.get(field) in (None, ""):
                    record[field] = 0
        evidence = {
            "source_file": source_file,
            "source_sheet": source_sheet,
            "source_row": row_index,
            "batch_id": batch_id,
        }
        record["evidence"] = evidence
        missing = sorted(field for field in required if record.get(field) in (None, ""))
        if missing:
            rejected.append({"row": row_index, "reason": f"missing required fields: {', '.join(missing)}"})
            continue
        try:
            TABLE_PARSERS[detected_type](record)
        except (CommerceDataError, ImportCostDataError) as exc:
            rejected.append({"row": row_index, "reason": str(exc)})
            continue
        records.append(record)
    return {
        "record_type": detected_type,
        "header_row": header_index + 1,
        "records": records,
        "rejected_rows": rejected,
    }


def _batch_id(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()[:16]


def parse_commerce_csv(
    path: str | Path,
    *,
    record_type: str | None = None,
    default_entity_id: str | None = None,
    default_channel: str | None = None,
) -> dict[str, Any]:
    path = Path(path)
    rows, encoding = _decode_csv(path)
    batch_id = _batch_id(path)
    result = _normalize_rows(
        rows,
        source_file=path.name,
        source_sheet=f"CSV ({encoding})",
        batch_id=batch_id,
        record_type=record_type,
        default_entity_id=default_entity_id,
        default_channel=default_channel,
    )
    return _batch_payload(path, batch_id, [result])


def parse_commerce_workbook(
    path: str | Path,
    *,
    default_entity_id: str | None = None,
    default_channel: str | None = None,
) -> dict[str, Any]:
    path = Path(path)
    batch_id = _batch_id(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    results = []
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if not rows:
                continue
            preferred_type = (
                "import_costs" if any(token in sheet.title for token in ("进口成本", "关税", "报关"))
                else "return_receipts" if any(
                    token in sheet.title for token in ("退货入库", "退回入库", "退货收货")
                )
                else "returns" if any(token in sheet.title for token in ("退货", "退款"))
                else "orders" if "订单" in sheet.title
                else "settlements" if any(token in sheet.title for token in ("结算", "打款"))
                else None
            )
            try:
                results.append(_normalize_rows(
                    rows,
                    source_file=path.name,
                    source_sheet=sheet.title,
                    batch_id=batch_id,
                    record_type=preferred_type,
                    default_entity_id=default_entity_id,
                    default_channel=default_channel,
                ))
            except CommerceDataError:
                continue
    finally:
        workbook.close()
    if not results:
        raise CommerceDataError(
            "Workbook contains no supported commerce table"
        )
    return _batch_payload(path, batch_id, results)


def _batch_payload(path: Path, batch_id: str, results: list[dict[str, Any]]) -> dict[str, Any]:
    orders = [record for result in results if result["record_type"] == "orders" for record in result["records"]]
    settlements = [
        record for result in results if result["record_type"] == "settlements" for record in result["records"]
    ]
    returns = [
        record for result in results if result["record_type"] == "returns" for record in result["records"]
    ]
    return_receipts = [
        record for result in results if result["record_type"] == "return_receipts"
        for record in result["records"]
    ]
    import_costs = [
        record for result in results if result["record_type"] == "import_costs"
        for record in result["records"]
    ]
    rejected = [
        {**row, "record_type": result["record_type"]}
        for result in results for row in result["rejected_rows"]
    ]
    order_keys = [f"order|{row['entity_id']}|{row['order_id']}" for row in orders]
    settlement_keys = [f"settlement|{row['entity_id']}|{row['settlement_id']}" for row in settlements]
    return_keys = [
        f"return|{row['entity_id']}|{row['return_id']}|{row['sku']}" for row in returns
    ]
    receipt_keys = [
        f"return_receipt|{row['entity_id']}|{row['receipt_id']}" for row in return_receipts
    ]
    import_cost_keys = [
        f"import_cost|{row['entity_id']}|{row['entry_line_id']}" for row in import_costs
    ]
    all_keys = order_keys + settlement_keys + return_keys + receipt_keys + import_cost_keys
    duplicates = sorted({key for key in all_keys if all_keys.count(key) > 1})
    return {
        "batch_id": batch_id,
        "source_file": path.name,
        "orders": orders,
        "settlements": settlements,
        "returns": returns,
        "return_receipts": return_receipts,
        "import_costs": import_costs,
        "quality": {
            "ready": not rejected and not duplicates and bool(
                orders or settlements or returns or return_receipts or import_costs
            ),
            "order_count": len(orders),
            "settlement_count": len(settlements),
            "return_count": len(returns),
            "return_receipt_count": len(return_receipts),
            "import_cost_count": len(import_costs),
            "rejected_count": len(rejected),
            "rejected_rows": rejected,
            "duplicate_business_keys": duplicates,
        },
    }


def parse_commerce_file(path: str | Path, **kwargs: Any) -> dict[str, Any]:
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return parse_commerce_csv(path, **kwargs)
    if path.suffix.lower() == ".xlsx":
        kwargs.pop("record_type", None)
        return parse_commerce_workbook(path, **kwargs)
    raise CommerceDataError("Commerce connector supports .csv and .xlsx files")
