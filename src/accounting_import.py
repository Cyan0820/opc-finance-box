from __future__ import annotations

import csv
import hashlib
import math
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


class AccountingImportError(ValueError):
    """Raised when an accounting export violates the trial-balance contract."""


ALIASES = {
    "account_code": ("科目编码", "科目代码", "account code", "account number"),
    "account_name": ("科目名称", "account name", "account description"),
    "period": ("期间", "会计期间", "账期", "period"),
    "currency": ("币种", "currency"),
    "opening_debit": ("期初借方", "opening debit"),
    "opening_credit": ("期初贷方", "opening credit"),
    "period_debit": ("本期借方", "本期借方发生额", "period debit", "debit movement"),
    "period_credit": ("本期贷方", "本期贷方发生额", "period credit", "credit movement"),
    "closing_debit": ("期末借方", "closing debit", "ending debit"),
    "closing_credit": ("期末贷方", "closing credit", "ending credit"),
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).replace("\n", " ").strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", _text(value).casefold())


def _field(value: Any) -> str | None:
    clean = _slug(value)
    matches: list[tuple[int, str]] = []
    for field, aliases in ALIASES.items():
        for alias in aliases:
            token = _slug(alias)
            if token and (clean == token or token in clean):
                matches.append((len(token), field))
    return max(matches, default=(0, None))[1]


def _number(value: Any, field: str) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        raise AccountingImportError(f"{field} must be numeric")
    try:
        result = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError) as exc:
        raise AccountingImportError(f"{field} must be numeric") from exc
    if not math.isfinite(result) or result < 0:
        raise AccountingImportError(f"{field} must be a finite non-negative number")
    return round(result, 2)


def _period(value: Any) -> str:
    text = _text(value)[:7].replace("/", "-")
    if not re.fullmatch(r"20\d{2}-(0[1-9]|1[0-2])", text):
        raise AccountingImportError("period must use YYYY-MM")
    return text


def _currency(value: Any) -> str:
    currency = _text(value).upper()
    if not re.fullmatch(r"[A-Z]{3}", currency):
        raise AccountingImportError("currency must be a three-letter code")
    return currency


@dataclass(frozen=True)
class TrialBalanceLine:
    line_id: str
    entity_id: str
    period: str
    currency: str
    account_code: str
    account_name: str
    opening_debit: float
    opening_credit: float
    period_debit: float
    period_credit: float
    closing_debit: float
    closing_credit: float
    evidence: dict[str, Any]


def _record(
    raw: dict[str, Any], *, entity_id: str, default_period: str,
    default_currency: str, source_file: str, source_sheet: str,
    source_row: int, batch_id: str,
) -> TrialBalanceLine:
    account_code = _text(raw.get("account_code"))
    account_name = _text(raw.get("account_name"))
    if not account_code or not account_name:
        raise AccountingImportError("account_code and account_name are required")
    values = {
        field: _number(raw.get(field), field)
        for field in (
            "opening_debit", "opening_credit", "period_debit", "period_credit",
            "closing_debit", "closing_credit",
        )
    }
    if values["opening_debit"] and values["opening_credit"]:
        raise AccountingImportError("opening debit and credit cannot both be non-zero")
    if values["closing_debit"] and values["closing_credit"]:
        raise AccountingImportError("closing debit and credit cannot both be non-zero")
    period = _period(raw.get("period") or default_period)
    currency = _currency(raw.get("currency") or default_currency)
    line_id = hashlib.sha256(
        f"{entity_id}|{period}|{currency}|{account_code}".encode()
    ).hexdigest()[:16]
    return TrialBalanceLine(
        line_id=line_id, entity_id=entity_id, period=period, currency=currency,
        account_code=account_code, account_name=account_name, **values,
        evidence={
            "source_file": source_file, "source_sheet": source_sheet,
            "source_row": source_row, "batch_id": batch_id,
        },
    )


def _header_mapping(row: Iterable[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, value in enumerate(row):
        field = _field(value)
        if field and field not in mapping.values():
            mapping[index] = field
    return mapping


def _consume_rows(
    rows: Iterable[tuple[int, tuple[Any, ...]]], *, entity_id: str,
    default_period: str, default_currency: str, source_file: str,
    source_sheet: str, batch_id: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], bool]:
    mapping: dict[int, str] = {}
    records: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    for row_number, values in rows:
        candidate = _header_mapping(values)
        if {
            "account_code", "account_name", "closing_debit", "closing_credit",
        }.issubset(candidate.values()):
            mapping = candidate
            continue
        if not mapping or not any(_text(value) for value in values):
            continue
        raw = {field: values[index] if index < len(values) else None for index, field in mapping.items()}
        try:
            records.append(asdict(_record(
                raw, entity_id=entity_id, default_period=default_period,
                default_currency=default_currency, source_file=source_file,
                source_sheet=source_sheet, source_row=row_number, batch_id=batch_id,
            )))
        except AccountingImportError as exc:
            rejected.append({
                "dataset_type": "finance.trial_balance_lines", "row": row_number,
                "source_sheet": source_sheet, "reason": str(exc),
            })
    return records, rejected, bool(mapping)


def parse_trial_balance_file(
    path: str | Path, *, entity_id: str, default_period: str = "",
    default_currency: str = "",
) -> dict[str, Any]:
    path = Path(path)
    if path.suffix.casefold() not in {".csv", ".xlsx"}:
        raise AccountingImportError("trial balance must be a .csv or .xlsx file")
    if not entity_id:
        raise AccountingImportError("entity_id is required")
    batch_id = hashlib.sha256(path.read_bytes()).hexdigest()[:16]
    records: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    any_header = False
    if path.suffix.casefold() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [(index, tuple(row)) for index, row in enumerate(csv.reader(handle), 1)]
        accepted, errors, header_found = _consume_rows(
            rows, entity_id=entity_id, default_period=default_period,
            default_currency=default_currency, source_file=path.name,
            source_sheet="CSV", batch_id=batch_id,
        )
        records.extend(accepted); rejected.extend(errors); any_header = header_found
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = ((index, tuple(row)) for index, row in enumerate(sheet.iter_rows(values_only=True), 1))
                accepted, errors, header_found = _consume_rows(
                    rows, entity_id=entity_id, default_period=default_period,
                    default_currency=default_currency, source_file=path.name,
                    source_sheet=sheet.title, batch_id=batch_id,
                )
                records.extend(accepted); rejected.extend(errors); any_header |= header_found
        finally:
            workbook.close()
    if not any_header:
        rejected.append({
            "dataset_type": "finance.trial_balance_lines", "row": 0,
            "source_sheet": "workbook",
            "reason": "trial balance header requires account code/name and closing debit/credit",
        })
    return {
        "batch_id": batch_id,
        "source": {"kind": "accounting_export_file", "name": path.name, "network_access_performed": False},
        "records": records, "rejected_rows": rejected,
    }


def validate_trial_balance_lines(
    rows: Iterable[dict[str, Any]], *, entity_id: str,
) -> dict[str, Any]:
    lines = list(rows)
    scopes = sorted({
        (str(row.get("entity_id") or ""), str(row.get("period") or ""), str(row.get("currency") or ""))
        for row in lines
    })
    issues: list[dict[str, Any]] = []
    if any(scope[0] != entity_id for scope in scopes):
        issues.append({"severity": "blocking", "type": "cross_entity_trial_balance"})
    summaries = []
    for scope in scopes:
        scoped = [row for row in lines if (
            str(row.get("entity_id")), str(row.get("period")), str(row.get("currency"))
        ) == scope]
        closing_debit = round(sum(float(row.get("closing_debit") or 0) for row in scoped), 2)
        closing_credit = round(sum(float(row.get("closing_credit") or 0) for row in scoped), 2)
        opening_debit = round(sum(float(row.get("opening_debit") or 0) for row in scoped), 2)
        opening_credit = round(sum(float(row.get("opening_credit") or 0) for row in scoped), 2)
        period_debit = round(sum(float(row.get("period_debit") or 0) for row in scoped), 2)
        period_credit = round(sum(float(row.get("period_credit") or 0) for row in scoped), 2)
        difference = round(closing_debit - closing_credit, 2)
        if abs(difference) >= 0.01:
            issues.append({
                "severity": "blocking", "type": "unbalanced_trial_balance",
                "entity_id": scope[0], "period": scope[1], "currency": scope[2],
                "difference": difference,
            })
        roll_forward_checked = any(
            float(row.get(field) or 0) != 0
            for row in scoped
            for field in ("opening_debit", "opening_credit", "period_debit", "period_credit")
        )
        roll_forward_difference = round(
            (opening_debit - opening_credit) + (period_debit - period_credit)
            - (closing_debit - closing_credit),
            2,
        )
        if roll_forward_checked and abs(roll_forward_difference) >= 0.01:
            issues.append({
                "severity": "blocking", "type": "trial_balance_roll_forward_mismatch",
                "entity_id": scope[0], "period": scope[1], "currency": scope[2],
                "difference": roll_forward_difference,
            })
        summaries.append({
            "entity_id": scope[0], "period": scope[1], "currency": scope[2],
            "line_count": len(scoped), "opening_debit": opening_debit,
            "opening_credit": opening_credit, "period_debit": period_debit,
            "period_credit": period_credit, "closing_debit": closing_debit,
            "closing_credit": closing_credit, "difference": difference,
            "balanced": abs(difference) < 0.01,
            "roll_forward_checked": roll_forward_checked,
            "roll_forward_difference": roll_forward_difference if roll_forward_checked else None,
            "roll_forward_consistent": (
                abs(roll_forward_difference) < 0.01 if roll_forward_checked else None
            ),
        })
    return {
        "ready": bool(lines) and not issues,
        "entity_id": entity_id,
        "summaries": summaries,
        "issues": issues,
        "candidate_only": True,
        "ledger_or_opening_balances_modified": False,
        "posting_performed": False,
        "guardrail": "A balanced accounting export is evidence for review; it does not prove account mapping, completeness, posting or period close.",
    }
