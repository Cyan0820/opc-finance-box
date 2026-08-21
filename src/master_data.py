from __future__ import annotations

import hashlib
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SHEET_TYPES = {
    "游戏项目": "game",
    "项目主数据": "game",
    "渠道规则": "channel",
    "渠道主数据": "channel",
    "组织映射": "organization",
    "组织主数据": "organization",
    "供应商": "vendor",
    "供应商主数据": "vendor",
}

ALIASES = {
    "code": ("项目编码", "渠道编码", "人员岗位编码", "人员/岗位编码", "供应商编码", "编码", "code"),
    "name": ("游戏项目名称", "游戏/项目名称", "渠道名称", "供应商名称", "名称", "name"),
    "stage": ("项目阶段", "阶段", "stage"),
    "launch_date": ("上线日期", "预计上线日期", "launch date"),
    "owner": ("负责人", "owner"),
    "department": ("部门", "归属部门", "department"),
    "budget_unit": ("预算单元", "预算部门", "budget unit"),
    "cost_center": ("成本中心", "cost center"),
    "project_code": ("游戏项目编码", "游戏/项目编码", "归属项目编码", "项目编码", "project code"),
    "platform": ("平台", "platform"),
    "region": ("区域", "国家地区", "国家/地区", "region"),
    "currency": ("默认币种", "结算币种", "币种", "currency"),
    "revenue_model": ("收入模式", "商业模式", "revenue model"),
    "share_rate": ("分成比例", "我方分成比例", "share rate"),
    "settlement_formula": ("结算公式", "分成公式", "settlement formula"),
    "contract_reference": ("合同证据引用", "协议证据引用", "平台政策引用", "contract reference"),
    "settlement_cycle": ("结算周期", "账期规则", "settlement cycle"),
    "payment_days": ("回款天数", "付款天数", "账期天数", "payment days"),
    "category": ("供应商类别", "采购类别", "类别", "category"),
    "allocation_rate": ("分摊比例", "项目分摊比例", "allocation rate"),
    "effective_period": ("生效月份", "生效期间", "effective period"),
    "active": ("是否启用", "启用", "active"),
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value).replace("\n", " ").strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", _text(value).lower())


def _number(value: Any) -> float | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return None if math.isnan(value) or math.isinf(value) else float(value)
    match = re.search(r"-?[\d,.]+", _text(value))
    return float(match.group().replace(",", "")) if match else None


def _field(value: Any) -> str | None:
    clean = _slug(value)
    matches = []
    for key, aliases in ALIASES.items():
        for alias in aliases:
            normalized = _slug(alias)
            if normalized and (clean == normalized or normalized in clean):
                matches.append((len(normalized), key))
    return max(matches, default=(0, None))[1]


def _record_type(sheet_name: str) -> str | None:
    clean = _slug(sheet_name)
    for label, record_type in SHEET_TYPES.items():
        if _slug(label) in clean:
            return record_type
    return None


def _active(value: Any) -> bool:
    return _slug(value) not in {"否", "停用", "false", "0", "n", "no"}


def parse_master_workbook(path: str | Path) -> list[dict]:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    for sheet in workbook.worksheets:
        record_type = _record_type(sheet.title)
        if not record_type:
            continue
        mapping: dict[int, str] = {}
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            candidate = {}
            for index, value in enumerate(row):
                key = _field(value)
                if key and key not in candidate.values():
                    candidate[index] = key
            if len(candidate) >= 3 and ("code" in candidate.values() or "project_code" in candidate.values()):
                mapping = candidate
                continue
            if not mapping:
                continue
            raw = {key: row[index] if index < len(row) else None for index, key in mapping.items()}
            if not any(_text(value) for value in raw.values()):
                continue
            code = _text(raw.get("code"))
            if record_type == "game" and not code:
                code = _text(raw.get("project_code"))
            name = _text(raw.get("name"))
            if not code and not name:
                continue
            if _slug(code).startswith("示例") or _slug(name).startswith("示例"):
                continue
            anomalies = []
            if not code:
                anomalies.append("缺少唯一编码")
            if record_type in {"game", "channel", "vendor"} and not name:
                anomalies.append("缺少名称")
            if record_type in {"channel", "organization"} and not _text(raw.get("project_code")):
                anomalies.append("缺少游戏/项目编码映射")
            allocation = _number(raw.get("allocation_rate"))
            share_rate = _number(raw.get("share_rate"))
            for key, value in (("allocation_rate", allocation), ("share_rate", share_rate)):
                if value is not None and 1 < value <= 100:
                    value /= 100
                raw[key] = value
            key = f"{record_type}|{code}|{name}"
            records.append({
                "id": hashlib.sha1(key.encode("utf-8")).hexdigest()[:12],
                "record_type": record_type,
                "code": code,
                "name": name,
                "stage": _text(raw.get("stage")),
                "launch_date": _text(raw.get("launch_date")),
                "owner": _text(raw.get("owner")),
                "department": _text(raw.get("department")),
                "budget_unit": _text(raw.get("budget_unit")),
                "cost_center": _text(raw.get("cost_center")),
                "project_code": "" if record_type == "game" else _text(raw.get("project_code")),
                "platform": _text(raw.get("platform")),
                "region": _text(raw.get("region")),
                "currency": _text(raw.get("currency")).upper() or "CNY",
                "revenue_model": _text(raw.get("revenue_model")),
                "share_rate": raw.get("share_rate"),
                "settlement_formula": _text(raw.get("settlement_formula")) or "declared",
                "contract_reference": _text(raw.get("contract_reference")),
                "settlement_cycle": _text(raw.get("settlement_cycle")),
                "payment_days": _number(raw.get("payment_days")),
                "category": _text(raw.get("category")),
                "allocation_rate": raw.get("allocation_rate"),
                "effective_period": _text(raw.get("effective_period"))[:7],
                "active": _active(raw.get("active")),
                "status": "异常" if anomalies else "可用",
                "anomalies": anomalies,
                "source_file": path.name,
                "source_sheet": sheet.title,
                "source_row": row_number,
            })
    workbook.close()
    return records


PROFILE_FIELDS = {
    "公司名称": ("company_name",),
    "统一社会信用代码": ("credit_code",),
    "注册及主管税务地区": ("registered_city",),
    "执行会计准则": ("accounting_standard",),
    "增值税纳税人类型": ("vat_taxpayer_type",),
    "增值税申报周期": ("vat_filing_frequency",),
    "外部会计/代账机构": ("external_accountant", "provider"),
    "复核联系人": ("external_accountant", "contact"),
    "预测起点现金": ("cash_planning", "opening_cash_cny"),
    "最低现金缓冲": ("cash_planning", "minimum_buffer_cny"),
}


def parse_profile_workbook(path: str | Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    patch: dict = {}
    sheet = next((item for item in workbook.worksheets if "主体配置" in item.title), None)
    if sheet:
        for row in sheet.iter_rows(min_row=1, max_col=4, values_only=True):
            label = _text(row[0] if row else None)
            value = row[1] if len(row) > 1 else None
            path_keys = PROFILE_FIELDS.get(label)
            if not path_keys or value in (None, ""):
                continue
            if label in {"预测起点现金", "最低现金缓冲"}:
                value = _number(value)
            target = patch
            for key in path_keys[:-1]:
                target = target.setdefault(key, {})
            target[path_keys[-1]] = value
    workbook.close()
    return patch


def master_quality(records: list[dict]) -> dict:
    by_type = {key: [] for key in ("game", "channel", "organization", "vendor")}
    for row in records:
        by_type.setdefault(row.get("record_type") or "unknown", []).append(row)
    game_codes = {row.get("code") for row in by_type["game"] if row.get("code")}
    issues = []
    for record_type, rows in by_type.items():
        codes = [row.get("code") for row in rows if row.get("code")]
        duplicates = sorted({code for code in codes if codes.count(code) > 1})
        if duplicates:
            issues.append(f"{record_type} 存在重复编码：{'、'.join(duplicates[:5])}")
    orphaned = sorted({
        row.get("project_code") for kind in ("channel", "organization", "vendor") for row in by_type[kind]
        if row.get("project_code") and row.get("project_code") not in game_codes
    })
    if orphaned:
        issues.append(f"项目映射找不到游戏项目主数据：{'、'.join(orphaned[:8])}")
    issues.extend(
        f"{row.get('record_type')} {row.get('code') or row.get('name')}：{'、'.join(row.get('anomalies') or [])}"
        for row in records if row.get("anomalies")
    )
    return {
        "counts": {key: len(value) for key, value in by_type.items()},
        "usable_count": sum(row.get("status") == "可用" for row in records),
        "issue_count": len(issues),
        "issues": issues,
    }
