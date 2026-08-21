import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from src.finance_inbox import FinanceInboxStore, classify_document, suggest_entity_scope, _period
from src.finance_inbox import extract_settlement_reconciliation_evidence


ENTITIES = [
    {"id": "cn_studio", "name": "星火游戏（上海）有限公司"},
    {"id": "sg_publisher", "name": "Spark Game Pte. Ltd."},
]


def workbook_bytes(headers, row, title="明细"):
    book = Workbook()
    sheet = book.active
    sheet.title = title
    sheet.append(headers)
    sheet.append(row)
    body = BytesIO()
    book.save(body)
    return body.getvalue()


class FinanceInboxTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.store = FinanceInboxStore(Path(self.temp.name) / "inbox")

    def tearDown(self):
        self.temp.cleanup()

    def test_bank_statement_is_classified_from_structure_not_only_filename(self):
        body = workbook_bytes(
            ["交易日期", "交易流水号", "对方户名", "贷方发生额", "账户余额", "币种"],
            ["2026-01-02", "T001", "某渠道", 1000, 5000, "CNY"],
        )
        path = Path(self.temp.name) / "随便一个名字.xlsx"
        path.write_bytes(body)
        result = classify_document(path)
        self.assertEqual(result["document_type"], "bank_statement")
        self.assertGreaterEqual(result["confidence"], 0.75)

    def test_period_parser_keeps_two_digit_months(self):
        self.assertEqual(_period("2025-12"), "2025-12")
        self.assertEqual(_period("2025年10月"), "2025-10")

    def test_invoice_register_can_be_received_recognized_and_deduplicated(self):
        body = workbook_bytes(
            ["发票号码", "发票代码", "开票日期", "销售方名称", "购买方名称", "金额", "税额", "价税合计", "查验状态"],
            ["001", "3100", "2026-01-03", "供应商", "游戏公司", 100, 6, 106, "已查验"],
            "发票台账",
        )
        document = self.store.ingest("1月资料.xlsx", body)
        self.assertEqual(document["classification"]["document_type"], "invoice_register")
        recognized = self.store.recognize(document["id"], "invoice_register", "2026-01", {"purchases": []})
        self.assertEqual(recognized["recognition"]["record_count"], 1)
        duplicate = self.store.ingest("改了文件名.xlsx", body)
        self.assertEqual(duplicate["id"], document["id"])
        self.assertTrue(duplicate["duplicate"])

    def test_pdf_invoice_enters_ocr_queue_instead_of_claiming_it_was_read(self):
        document = self.store.ingest("供应商发票.pdf", b"%PDF-1.4 fake test content")
        self.assertEqual(document["classification"]["document_type"], "invoice_document")
        self.assertEqual(document["status"], "等待文字识别")

    def test_bank_pdf_requires_correction_and_original_confirmation_before_commit(self):
        document = self.store.ingest(
            "sg_publisher 银行对账单.pdf", b"%PDF-1.4 fake test content",
            "上传人", "sg_publisher", ENTITIES,
        )
        extraction = {
            "method": "pdf_text", "page_count": 1, "pages": [], "confidence": 0.95,
            "text": (
                "2026-08-01 Reference: TX001 Counterparty: Apple Currency: USD "
                "Credit: 1250.50 Balance: 5000.50"
            ),
        }
        with patch("src.finance_inbox.extract_document_text", return_value=extraction):
            recognized = self.store.recognize(
                document["id"], "bank_statement_document", "2026-08", {}, "Agent",
                "sg_publisher", "Spark Game Pte. Ltd.",
            )
        record = recognized["recognition"]["records"][0]
        self.assertEqual(record["entity_id"], "sg_publisher")
        self.assertEqual(record["source_document_id"], document["id"])
        self.assertIn("对照原件", self.store.commit_blockers(document["id"])[0])
        corrected = self.store.correct(
            document["id"], [{"index": 0, "fields": {}}], "财务负责人", True,
            "已逐页对照银行原文件",
        )
        self.assertEqual(self.store.commit_blockers(document["id"]), [])
        self.assertFalse(any("对照原件确认" in item for item in corrected["recognition"]["records"][0]["anomalies"]))

    def test_evidence_document_links_only_to_same_entity_purchase(self):
        document = self.store.ingest(
            "美术外包验收证明.pdf", b"%PDF-1.4 fake test content",
            "上传人", "cn_studio", ENTITIES,
        )
        document["classification"]["document_type"] = "acceptance_evidence"
        document["recognition"] = {"evidence_only": True, "record_count": 0}
        document["status"] = "已提取待归档"
        self.store.save(document)
        linked = self.store.link_to_business_record(
            document["id"], target_type="purchase", target_id="PO-1",
            entity_id="cn_studio", actor="项目负责人", note="证明角色立绘已交付",
        )
        self.assertEqual(linked["status"], "已归档并关联")
        self.assertEqual(linked["business_links"][0]["target_id"], "PO-1")
        with self.assertRaisesRegex(ValueError, "法律主体"):
            self.store.link_to_business_record(
                document["id"], target_type="purchase", target_id="PO-2",
                entity_id="sg_publisher", actor="项目负责人",
            )

    def test_ios_operational_reconciliation_is_evidence_not_settlement_revenue(self):
        body = workbook_bytes(
            ["年月", "游戏", "日期", "平台", "渠道", "付费金额", "匹配-苹果账单", "差异", "退款", "剔除退款后差异"],
            ["2026-01", "ROR", "2026-01-01", "ios", "国内", 1000, 998, -2, 1, -1],
            "核对-经分数据&IOS账单数据",
        )
        path = Path(self.temp.name) / "IOS账单&经分数据核对-202601.xlsx"
        path.write_bytes(body)
        classified = classify_document(path)
        self.assertEqual(classified["document_type"], "settlement_reconciliation_evidence")
        rows = extract_settlement_reconciliation_evidence(path)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["operational_amount"], 1000)
        document = self.store.ingest(
            path.name, body, "上传人", "cn_studio", ENTITIES,
        )
        recognized = self.store.recognize(
            document["id"], "settlement_reconciliation_evidence", "2026-01", {},
            "Agent", "cn_studio", ENTITIES[0]["name"],
        )
        self.assertEqual(recognized["status"], "已提取待归档")
        self.assertEqual(recognized["recognition"]["record_count"], 1)
        self.assertEqual(recognized["recognition"]["records"], [])
        linked = self.store.link_to_business_record(
            document["id"], target_type="settlement", target_id="SET-1",
            entity_id="cn_studio", actor="财务负责人", note="核对iOS平台账单与经分数据",
        )
        self.assertEqual(linked["business_links"][0]["target_type"], "settlement")

    def test_unknown_file_cannot_be_recognized_as_a_ledger_without_confirmation(self):
        body = workbook_bytes(["栏目A", "栏目B"], ["文本", 123])
        document = self.store.ingest("未知.xlsx", body)
        self.assertEqual(document["classification"]["document_type"], "unknown")
        self.assertEqual(document["status"], "待确认类型")

    def test_multi_sheet_workbook_is_a_finance_package_not_forced_into_one_type(self):
        book = Workbook()
        settlement = book.active
        settlement.title = "渠道结算"
        settlement.append(["结算周期", "游戏名称", "平台", "渠道", "结算币种", "渠道含税流水（结算币种）", "分成比例", "结算金额"])
        settlement.append(["2026-01", "游戏A", "iOS", "App Store", "CNY", 1000, 0.7, 700])
        bank = book.create_sheet("银行流水")
        bank.append(["交易日期", "交易流水号", "对方户名", "贷方发生额", "账户余额", "币种"])
        bank.append(["2026-01-02", "T001", "渠道", 700, 5000, "CNY"])
        invoice = book.create_sheet("发票台账")
        invoice.append(["发票号码", "发票代码", "开票日期", "销售方名称", "购买方名称", "金额", "税额", "价税合计", "查验状态"])
        invoice.append(["001", "3100", "2026-01-03", "供应商", "游戏公司", 100, 6, 106, "已查验"])
        body = BytesIO()
        book.save(body)
        document = self.store.ingest("一月财务包.xlsx", body.getvalue())
        self.assertEqual(document["classification"]["document_type"], "finance_package")
        recognized = self.store.recognize(document["id"], "finance_package", "2026-01", {})
        self.assertGreaterEqual(len(recognized["recognition"]["batches"]), 3)

    def test_document_id_rejects_path_traversal(self):
        with self.assertRaises(ValueError):
            self.store.load("../../etc/passwd")

    def test_csv_bank_statement_is_classified_and_parsed(self):
        body = (
            "交易日期,交易流水号,对方户名,贷方发生额,账户余额,币种\n"
            "2026-01-02,T001,某渠道,1000,5000,CNY\n"
        ).encode("utf-8-sig")
        document = self.store.ingest("银行导出.csv", body)
        self.assertEqual(document["classification"]["document_type"], "bank_statement")
        recognized = self.store.recognize(document["id"], "bank_statement", "2026-01", {})
        self.assertEqual(recognized["recognition"]["record_count"], 1)

    def test_human_can_correct_ocr_fields_but_not_verification_status(self):
        body = workbook_bytes(
            ["发票号码", "发票代码", "开票日期", "销售方名称", "购买方名称", "金额", "税额", "价税合计", "查验状态"],
            ["001", "3100", "2026-01-03", "供应商", "游戏公司", 100, 6, 106, "已查验"],
        )
        document = self.store.ingest("发票清单.xlsx", body)
        document = self.store.recognize(document["id"], "invoice_register", "2026-01", {})
        corrected = self.store.correct(
            document["id"], [{"index": 0, "fields": {"seller_name": "正确供应商"}}],
            "财务负责人", True, "已对照原件",
        )
        self.assertEqual(corrected["recognition"]["records"][0]["seller_name"], "正确供应商")
        self.assertEqual(corrected["recognition"]["corrections"][0]["actor"], "财务负责人")
        with self.assertRaises(ValueError):
            self.store.correct(
                document["id"], [{"index": 0, "fields": {"verification_status": "已查验"}}], "财务负责人",
            )

    def test_multi_entity_filename_is_not_guessed_from_currency_or_channel(self):
        scope = suggest_entity_scope("Google Play USD 2026-01 对账单.xlsx", ENTITIES)
        self.assertEqual(scope["status"], "unassigned")
        self.assertEqual(scope["entity_id"], "")
        self.assertIn("不能仅凭币种", scope["reason"])

    def test_exact_legal_name_only_creates_suggestion_not_confirmation(self):
        scope = suggest_entity_scope("Spark Game Pte. Ltd. 2026-01 invoice.xlsx", ENTITIES)
        self.assertEqual(scope["entity_id"], "sg_publisher")
        self.assertEqual(scope["status"], "suggested")
        self.assertEqual(scope["source"], "filename_exact_match")

    def test_human_entity_confirmation_is_attached_to_every_recognized_record(self):
        body = workbook_bytes(
            ["交易日期", "交易流水号", "对方户名", "贷方发生额", "账户余额", "币种"],
            ["2026-01-02", "T001", "某渠道", 1000, 5000, "USD"],
        )
        document = self.store.ingest("Google Play USD 对账单.xlsx", body, entities=ENTITIES)
        self.assertEqual(document["entity_scope"]["status"], "unassigned")
        confirmed = self.store.assign_entity_scope(
            document["id"], "sg_publisher", "财务负责人", ENTITIES, "银行账户归属已核对",
        )
        self.assertEqual(confirmed["entity_scope"]["status"], "confirmed")
        recognized = self.store.recognize(
            document["id"], "bank_statement", "2026-01", {}, "Agent",
            "sg_publisher", "Spark Game Pte. Ltd.",
        )
        record = recognized["recognition"]["records"][0]
        self.assertEqual(record["entity_id"], "sg_publisher")
        self.assertEqual(record["source_document_id"], document["id"])
        events = self.store.events(document["id"])
        self.assertTrue(any(event["type"] == "ENTITY_SCOPE_CONFIRMED" for event in events))

    def test_entity_id_cannot_be_changed_through_record_correction(self):
        body = workbook_bytes(
            ["发票号码", "发票代码", "开票日期", "销售方名称", "购买方名称", "金额", "税额", "价税合计", "查验状态"],
            ["001", "3100", "2026-01-03", "供应商", "游戏公司", 100, 6, 106, "已查验"],
        )
        document = self.store.ingest("发票清单.xlsx", body, "上传人", "cn_studio", ENTITIES)
        document = self.store.recognize(
            document["id"], "invoice_register", "2026-01", {}, "Agent",
            "cn_studio", "星火游戏（上海）有限公司",
        )
        with self.assertRaisesRegex(ValueError, "entity_id"):
            self.store.correct(
                document["id"], [{"index": 0, "fields": {"entity_id": "sg_publisher"}}], "财务负责人",
            )


if __name__ == "__main__":
    unittest.main()
