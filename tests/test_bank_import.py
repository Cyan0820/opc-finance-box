import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from src.bank_import import parse_bank_statement_file
from src.box_runtime import BoxRuntime
from src.default_connectors import build_default_connector_registry


ROOT = Path(__file__).resolve().parents[1]


class BankImportTests(unittest.TestCase):
    def setUp(self):
        self.runtime = BoxRuntime(ROOT / "examples" / "boxes" / "cn_dtc_store.json", ROOT / "packs")

    @staticmethod
    def _workbook(path: Path, rows: list[list[object]]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "银行流水"
        for row in rows:
            sheet.append(row)
        workbook.save(path)

    def test_xlsx_import_masks_accounts_and_preserves_batch_evidence(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "6222021234567890-bank.xlsx"
            self._workbook(path, [
                ["交易日期", "交易流水号", "本方账号", "对方户名", "对方账号", "贷方发生额", "账户余额", "币种"],
                ["2026-08-01", "TX-001", "6222021234567890", "渠道A", "4333123456789012", 700, 5700, "CNY"],
            ])
            result = parse_bank_statement_file(path, entity_id="cn_dtc_company")
        self.assertEqual(len(result["records"]), 1)
        row = result["records"][0]
        self.assertEqual(row["account_masked"], "6222****7890")
        self.assertEqual(row["counterparty_account_masked"], "4333****9012")
        self.assertNotIn("6222021234567890", str(result))
        self.assertEqual(result["source"]["name"], "6222****7890-bank.xlsx")
        self.assertEqual(row["direction_code"], "inflow")
        self.assertEqual(row["evidence"]["batch_id"], result["batch_id"])
        self.assertFalse(result["source"]["raw_account_numbers_retained"])

    def test_csv_connector_is_entity_scoped_and_ready(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bank.csv"
            path.write_text(
                "交易日期,交易流水号,摘要,交易金额,收支方向,币种\n"
                "2026/08/02,TX-002,供应商付款,120.50,支出,CNY\n",
                encoding="utf-8",
            )
            result = build_default_connector_registry().dispatch(
                self.runtime,
                "file.bank_statement",
                {
                    "path": str(path), "default_entity_id": "cn_dtc_company",
                    "account_reference": "MAIN-CNY",
                },
            )
        self.assertTrue(result["batch"]["quality"]["ready"], result)
        row = result["batch"]["datasets"]["finance.bank_transactions"][0]
        self.assertEqual(row["entity_id"], "cn_dtc_company")
        self.assertEqual(row["direction"], "支出")
        self.assertEqual(row["amount"], 120.5)

    def test_ambiguous_direction_and_missing_transaction_id_fail_closed(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bad.xlsx"
            self._workbook(path, [
                ["交易日期", "交易流水号", "本方账号", "交易金额", "币种"],
                ["2026-08-03", "", "MAIN", 50, "CNY"],
                ["2026-08-04", "TX-004", "MAIN", 50, "CNY"],
            ])
            result = build_default_connector_registry().dispatch(
                self.runtime,
                "file.bank_statement",
                {"path": str(path), "default_entity_id": "cn_dtc_company"},
            )
        self.assertFalse(result["batch"]["quality"]["ready"])
        self.assertEqual(result["batch"]["quality"]["record_count"], 0)
        reasons = " ".join(item["reason"] for item in result["batch"]["quality"]["rejected_rows"])
        self.assertIn("bank_transaction_id is required", reasons)
        self.assertIn("explicit direction is required", reasons)

    def test_duplicate_business_key_does_not_enter_dataset_twice(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "duplicate.csv"
            path.write_text(
                "交易日期,交易流水号,本方账号,贷方发生额,币种\n"
                "2026-08-01,TX-001,MAIN,10,CNY\n"
                "2026-08-02,TX-001,MAIN,20,CNY\n",
                encoding="utf-8",
            )
            result = build_default_connector_registry().dispatch(
                self.runtime,
                "file.bank_statement",
                {"path": str(path), "default_entity_id": "cn_dtc_company"},
            )
        self.assertFalse(result["batch"]["quality"]["ready"])
        self.assertEqual(result["batch"]["quality"]["record_count"], 1)
        self.assertEqual(len(result["batch"]["quality"]["duplicate_business_keys"]), 1)

    def test_xlsx_ignores_cover_sheet_when_another_sheet_has_bank_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "multi-sheet.xlsx"
            workbook = Workbook()
            workbook.active.title = "说明"
            workbook.active.append(["银行导出文件", "请勿编辑"])
            sheet = workbook.create_sheet("流水")
            sheet.append(["交易日期", "交易流水号", "本方账号", "贷方发生额", "币种"])
            sheet.append(["2026-08-05", "TX-005", "MAIN", 88, "CNY"])
            workbook.save(path)
            result = parse_bank_statement_file(path, entity_id="cn_dtc_company")
        self.assertEqual(len(result["records"]), 1)
        self.assertEqual(result["rejected_rows"], [])


if __name__ == "__main__":
    unittest.main()
