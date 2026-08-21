import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.tax_export import build_tax_workbook


class TaxExportTests(unittest.TestCase):
    def test_public_python_export_contains_guardrails_forms_and_checks(self):
        workspace = {
            "company_name": "虚构游戏公司", "credit_code": "91310000DEMO000001",
            "period": "2026-02", "period_start": "2026-02-01", "period_end": "2026-02-28",
            "accounting_basis": "已过账总账", "workflow": ["生成候选值", "人工复核", "有权人申报"],
            "guardrail": "工作底稿不是已提交申报表。",
            "summary": {"form_count": 1, "direct_upload_ready": 0, "blocked": 1},
            "returns": [{
                "form_code": "VAT-RETURN", "name": "增值税申报表", "version": "2026年第6号",
                "period": "2026-02", "status": "待复核", "transport": "不可直接上传",
                "agent_position": "候选口径", "review_role": "有权申报人", "official_source": "https://example.test",
                "fields": [{"code": "VAT-1", "name": "销售额候选", "value": 100, "source": "已过账总账", "status": "候选"}],
                "blockers": ["待复核"], "checks": [{"name": "收入勾稽", "passed": None, "note": "待核对"}],
                "schedules": [{"name": "销项候选", "rows": [{"channel": "App Store", "amount": 100}]}],
            }],
        }
        with tempfile.TemporaryDirectory() as temp:
            target = build_tax_workbook(workspace, Path(temp) / "tax.xlsx", Path(temp) / "verify")
            book = load_workbook(target, read_only=True, data_only=False)
            self.assertIn("使用说明", book.sheetnames)
            self.assertIn("VAT-RETURN", book.sheetnames)
            self.assertIn("VAT-RETURN-销项候选", book.sheetnames)
            self.assertIn("Checks", book.sheetnames)
            self.assertIn("不是已提交申报表", book["使用说明"]["A2"].value)
            self.assertEqual(book["Checks"]["B7"].value, 0)
            book.close()


if __name__ == "__main__":
    unittest.main()
