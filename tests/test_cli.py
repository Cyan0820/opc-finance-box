import io
import json
import tempfile
import unittest
import os
import hashlib
import hmac
import zipfile
from datetime import datetime, timezone
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from src.cli import main
from src.pipeline_scheduler import pipeline_request_fingerprint, schedule_job_approval_fingerprint
from src.box_compiler import compile_box_file
from src.box_runtime import BoxRuntime
from src.airwallex_webhooks import AirwallexWebhookStore
from src.tax_applicability_artifacts import (
    build_tax_applicability_workpaper,
    import_tax_applicability_review,
    review_tax_applicability_workpaper,
)
from tests.test_distribution_verify import DistributionVerifyTests


ROOT = Path(__file__).resolve().parents[1]


class CliTests(unittest.TestCase):
    def _run(self, args):
        output = io.StringIO()
        error = io.StringIO()
        with redirect_stdout(output), redirect_stderr(error):
            code = main(args)
        return code, output.getvalue(), error.getvalue()

    def test_options_lists_installed_jurisdictions(self):
        code, output, error = self._run(["options"])
        self.assertEqual(code, 0, error)
        payload = json.loads(output)
        codes = {item["country_code"] for item in payload["result"]["jurisdictions"]}
        self.assertEqual(codes, {"AE", "AU", "CA", "CN", "DE", "FR", "GB", "HK", "IE", "JP", "KR", "NL", "NZ", "SG", "US"})
        self.assertIn("shopify_stripe", {
            item["id"] for item in payload["result"]["integration_presets"]
        })
        self.assertIn("paypal", {
            item["id"] for item in payload["result"]["integration_presets"]
        })
        self.assertIn("woocommerce", {
            item["id"] for item in payload["result"]["integration_presets"]
        })
        self.assertIn("amazon_seller", {
            item["id"] for item in payload["result"]["integration_presets"]
        })

    def test_box_starters_lists_every_contract_checked_product_country_pair(self):
        code, output, error = self._run(["box-starters"])
        self.assertEqual(code, 0, error)
        catalog = json.loads(output)["result"]
        self.assertTrue(catalog["complete"])
        self.assertEqual(catalog["ready_combination_count"], 45)
        self.assertEqual(catalog["unavailable_combinations"], [])
        self.assertEqual(
            {item["profile_id"] for item in catalog["entries"]},
            {"game", "dtc", "marketplace"},
        )
        self.assertEqual(
            {item["country_code"] for item in catalog["entries"]},
            {"AE", "AU", "CA", "CN", "DE", "FR", "GB", "HK", "IE", "JP", "KR", "NL", "NZ", "SG", "US"},
        )
        self.assertTrue(all(item["contract_checked"] for item in catalog["entries"]))
        self.assertTrue(all(item["filing_ready"] is False for item in catalog["entries"]))

    def test_cfo_metrics_evaluate_runs_entity_bound_deterministic_contract(self):
        config = ROOT / "examples" / "boxes" / "us_dtc_shopify_stripe_c_corp.json"
        runtime = BoxRuntime(config, ROOT / "packs")
        request = {
            "runtime_fingerprint": runtime.snapshot()["fingerprint"],
            "period": "2026-07",
            "currency": "USD",
            "metric_type_ids": ["dtc_net_sales"],
            "operand_values": {
                "gross_order_sales_ex_tax_including_shipping": "500",
                "discounts_and_refunds_ex_tax": "42.5",
            },
            "confirmed_control_type_ids": [
                "order_and_refund_period_scope_aligned",
                "tax_inclusive_policy_confirmed",
            ],
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            request_path = Path(temp_dir) / "cfo-metric-request.json"
            request_path.write_text(json.dumps(request), encoding="utf-8")
            code, output, error = self._run([
                "cfo-metrics-evaluate",
                str(config),
                str(request_path),
                "--entity", "us_store",
            ])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertEqual(result["service"]["service_id"], "core.evaluate_cfo_metrics")
        self.assertEqual(result["service"]["entity_ids"], ["us_store"])
        self.assertEqual(result["output"]["metric_results"][0]["value"], "457.5")
        self.assertFalse(result["output"]["external_actions_performed"])

    def test_starter_init_creates_verified_dtc_country_workspace(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workspace = Path(temp_dir).resolve() / "nl-dtc"
            code, output, error = self._run([
                "starter-init", str(workspace),
                "--profile", "dtc",
                "--country", "NL",
                "--integration", "shopify_stripe",
                "--name", "NL DTC Starter",
                "--entity-id", "nl_dtc_entity",
                "--entity-name", "NL DTC entity (confirm)",
                "--actor", "cli-starter-recipient",
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertTrue(result["initialized"])
            self.assertEqual(result["starter_id"], "dtc.nl")
            self.assertEqual(result["selected_integrations"], ["shopify_stripe"])
            self.assertEqual(result["compiled_file_count"], 42)
            self.assertEqual(result["workspace_file_count"], 56)
            self.assertTrue(result["workspace_verified"])
            self.assertFalse(result["filing_ready"])
            self.assertFalse(result["destination_path_returned"])
            self.assertTrue((workspace / "box.json").is_file())
            self.assertTrue((workspace / "compiled" / "box.lock.json").is_file())
            code, output, error = self._run([
                "handoff-unpack-verify", str(workspace),
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["valid"])
            code, output, error = self._run([
                "starter-init", str(workspace),
                "--profile", "dtc", "--country", "NL",
                "--actor", "cli-starter-recipient",
            ])
            self.assertEqual(code, 2)
            self.assertEqual(output, "")
            self.assertIn("already exists", error)

    def test_trial_init_and_verify_create_one_runnable_separated_demo_workspace(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workspace = Path(temp_dir).resolve() / "amazon-trial"
            code, output, error = self._run([
                "trial-init", str(workspace),
                "--profile", "marketplace",
                "--country", "US",
                "--integration", "amazon_seller",
                "--name", "Amazon Seller Trial",
                "--actor", "cli-trial-founder",
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertTrue(result["initialized"])
            self.assertTrue(result["workspace_verified"])
            self.assertTrue(result["ready_to_run_locally"])
            self.assertEqual(result["starter_id"], "marketplace.us")
            self.assertEqual(result["selected_integrations"], ["amazon_seller"])
            self.assertTrue(result["box_workspace_immutable"])
            self.assertTrue(result["runtime_data_separate"])
            self.assertFalse(result["credentials_persisted"])
            self.assertFalse(result["destination_path_returned"])
            self.assertTrue((workspace / "box" / "box.json").is_file())
            self.assertTrue((workspace / "runtime-data" / "runtime-data-manifest.json").is_file())
            code, output, error = self._run(["trial-verify", str(workspace)])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertFalse(verified["paths_returned"])
            code, output, error = self._run(["trial-onboarding", str(workspace)])
            self.assertEqual(code, 0, error)
            onboarding = json.loads(output)["result"]
            self.assertEqual(
                onboarding["artifact_type"],
                "opc_finance_box_trial_onboarding_plan",
            )
            self.assertEqual(onboarding["current_stage_id"], "explore_local_demo")
            self.assertEqual(onboarding["summary"]["journey_stage_count"], 5)
            self.assertFalse(onboarding["summary"]["production_ready"])
            self.assertTrue(onboarding["control_boundary"]["commands_are_templates_only"])

    def test_starter_compose_creates_verified_multi_entity_workspace(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workspace = Path(temp_dir).resolve() / "multi-entity-game"
            code, output, error = self._run([
                "starter-compose", str(workspace),
                "--profile", "game",
                "--entity", "CN=cn_studio",
                "--entity", "SG=sg_publisher",
                "--entity-name", "cn_studio=China game studio (confirm)",
                "--entity-integration", "sg_publisher=xero",
                "--reporting-currency", "CNY",
                "--name", "Global game OPC",
                "--actor", "cli-multi-entity-recipient",
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertEqual(result["profile_id"], "game")
            self.assertEqual(result["entity_count"], 2)
            self.assertEqual(result["starter_ids"], ["game.cn", "game.sg"])
            self.assertEqual(result["selected_integrations"], ["xero"])
            self.assertEqual(result["entity_integrations"], [{
                "entity_id": "sg_publisher", "integration": "xero",
            }])
            self.assertTrue(result["cross_currency"])
            self.assertFalse(result["cross_currency_aggregation_authorized"])
            self.assertTrue(result["workspace_verified"])
            config = json.loads((workspace / "box.json").read_text())
            self.assertIn("feature.multi_entity", config["features"])
            self.assertEqual([item["id"] for item in config["entities"]], [
                "cn_studio", "sg_publisher",
            ])
            self.assertEqual(config["connector_bindings"], [
                {
                    "connector_pack": "connector.file_import",
                    "entity_ids": ["cn_studio", "sg_publisher"],
                },
                {"connector_pack": "connector.xero", "entity_ids": ["sg_publisher"]},
            ])
            code, output, error = self._run([
                "handoff-unpack-verify", str(workspace),
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["valid"])

    def test_handoff_bundle_cli_writes_first_customer_activation_guide_exclusively(self):
        spec = ROOT / "examples" / "box_specs" / "dtc_cn.json"
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir).resolve() / "customer-handoff.zip"
            code, output, error = self._run([
                "handoff-bundle", str(spec), "--output", str(output_path),
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertTrue(result["written"])
            self.assertTrue(result["activation_guide_included"])
            self.assertFalse(result["secret_values_included"])
            self.assertEqual(
                result["sha256"], hashlib.sha256(output_path.read_bytes()).hexdigest(),
            )
            with zipfile.ZipFile(output_path) as archive:
                self.assertIn("ACTIVATION.md", archive.namelist())
                activation = archive.read("ACTIVATION.md").decode("utf-8")
                self.assertIn("activation-runbook-verify", activation)
                manifest = json.loads(archive.read("bundle-manifest.json"))
                self.assertTrue(manifest["activation_guide_included"])
            code, output, error = self._run([
                "handoff-verify", str(output_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertTrue(verified["reproducible_with_installed_packs"])
            self.assertEqual(verified["member_count"], 55)
            self.assertEqual(verified["manifest_file_count"], 54)
            self.assertFalse(verified["archive_extracted"])
            self.assertFalse(verified["paths_returned"])
            receipt_path = Path(temp_dir).resolve() / "customer-handoff.browser-receipt.json"
            receipt_path.write_text(json.dumps({
                "schema_version": 1,
                "verification_status": "passed",
                "filename": output_path.name,
                "size_bytes": verified["size_bytes"],
                "sha256": verified["bundle_sha256"],
                "runtime_fingerprint": verified["runtime_fingerprint"],
                "manifest_schema_version": verified["manifest_schema_version"],
                "manifest_file_count": verified["manifest_file_count"],
                "browser_bytes_verified": True,
                "receipt_is_digital_signature": False,
                "archive_members_executed": False,
                "active_runtime_changed": False,
                "external_actions_performed": False,
            }, indent=2) + "\n", encoding="utf-8")
            if os.name != "nt":
                receipt_path.chmod(0o600)
            code, output, error = self._run([
                "handoff-receipt-verify", str(output_path), str(receipt_path),
            ])
            self.assertEqual(code, 0, error)
            receipt_verified = json.loads(output)["result"]
            self.assertTrue(receipt_verified["bundle_receipt_match"])
            self.assertFalse(receipt_verified["browser_execution_attested"])
            self.assertFalse(receipt_verified["receipt_is_digital_signature"])
            fork_root = Path(temp_dir).resolve() / "private-fork"
            code, output, error = self._run([
                "handoff-unpack", str(output_path), str(fork_root),
                "--actor", "cli-handoff-recipient",
            ])
            self.assertEqual(code, 0, error)
            unpacked = json.loads(output)["result"]
            self.assertTrue(unpacked["unpacked"])
            self.assertTrue(unpacked["installed_tree_verified"])
            self.assertEqual(unpacked["installed_file_count"], 56)
            self.assertFalse(unpacked["destination_path_returned"])
            self.assertFalse(unpacked["actor_returned"])
            code, output, error = self._run([
                "handoff-unpack-verify", str(fork_root),
            ])
            self.assertEqual(code, 0, error)
            installed = json.loads(output)["result"]
            self.assertTrue(installed["valid"])
            self.assertTrue(installed["installed_pack_reproducible"])
            self.assertFalse(installed["source_bundle_required"])
            self.assertFalse(installed["receipt_is_digital_signature"])
            code, output, error = self._run([
                "handoff-bundle", str(spec), "--output", str(output_path),
            ])
            self.assertEqual(code, 2)
            self.assertEqual(output, "")
            self.assertIn("refusing to overwrite", error)

    def test_source_kit_cli_writes_and_reproduces_fork_ready_archive(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir).resolve() / "source-kit.zip"
            code, output, error = self._run([
                "source-kit-bundle", "--output", str(output_path),
            ])
            self.assertEqual(code, 0, error)
            written = json.loads(output)["result"]
            self.assertTrue(written["written"])
            self.assertTrue(written["fork_ready_source_tree"])
            self.assertTrue(written["tests_included"])
            self.assertFalse(written["git_history_included"])
            self.assertFalse(written["runtime_data_included"])
            self.assertFalse(written["output_path_returned"])
            self.assertEqual(
                written["sha256"], hashlib.sha256(output_path.read_bytes()).hexdigest(),
            )
            code, output, error = self._run([
                "source-kit-verify", str(output_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertTrue(verified["reproducible_from_installed_source"])
            self.assertTrue(verified["fork_ready_source_tree"])
            self.assertFalse(verified["archive_extracted"])
            self.assertFalse(verified["paths_returned"])
            fork_root = Path(temp_dir).resolve() / "source-fork"
            code, output, error = self._run([
                "source-kit-unpack", str(output_path), str(fork_root),
                "--actor", "cli-source-recipient",
            ])
            self.assertEqual(code, 0, error)
            unpacked = json.loads(output)["result"]
            self.assertTrue(unpacked["unpacked"])
            self.assertTrue(unpacked["installed_tree_verified"])
            self.assertFalse(unpacked["destination_path_returned"])
            self.assertFalse(unpacked["actor_returned"])
            self.assertFalse(unpacked["archive_members_executed"])
            code, output, error = self._run([
                "source-kit-unpack-verify", str(fork_root),
            ])
            self.assertEqual(code, 0, error)
            installed = json.loads(output)["result"]
            self.assertTrue(installed["valid"])
            self.assertTrue(installed["installed_source_reproducible"])
            self.assertFalse(installed["source_archive_required"])
            self.assertFalse(installed["receipt_is_digital_signature"])
            code, output, error = self._run([
                "source-kit-bundle", "--output", str(output_path),
            ])
            self.assertEqual(code, 2)
            self.assertEqual(output, "")
            self.assertIn("refusing to overwrite", error)

    def test_connector_preflight_cli_groups_providers_and_returns_safe_next_action(self):
        config = str(
            ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
        )
        with patch.dict(os.environ, {
            "OPC_SHOPIFY_ADMIN_TOKEN": "cli-private-shopify",
            "OPC_STRIPE_RESTRICTED_KEY": "cli-private-stripe",
        }, clear=False):
            code, output, error = self._run(["connector-preflight", config])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertEqual(result["summary"]["provider_group_count"], 3)
        self.assertEqual(result["summary"]["network_provider_group_count"], 2)
        self.assertEqual(result["summary"]["blocked_provider_group_count"], 0)
        groups = {item["pack_id"]: item for item in result["provider_groups"]}
        self.assertIn(
            "connector-access-request-init",
            groups["connector.stripe"]["next_action"]["command_template"],
        )
        self.assertNotIn("cli-private-shopify", output)
        self.assertNotIn("cli-private-stripe", output)
        self.assertFalse(result["control_boundary"]["network_access_performed"])
        self.assertFalse(result["control_boundary"]["commands_executed"])

    def test_connector_access_cli_initializes_verifies_and_requires_network_opt_in(self):
        config = str(
            ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            request_path = Path(temp_dir).resolve() / "private-shopify-access.json"
            code, output, error = self._run([
                "connector-access-request-init",
                config,
                "--pack", "connector.shopify",
                "--entity", "cn_dtc_company",
                "--output", str(request_path),
            ])
            self.assertEqual(code, 0, error)
            initialized = json.loads(output)["result"]
            self.assertTrue(initialized["template_only"])
            self.assertFalse(initialized["ready_for_network_probe"])
            private = json.loads(request_path.read_text(encoding="utf-8"))
            private["account_binding"]["shop_domain"] = "cli-private.myshopify.com"
            request_path.write_text(json.dumps(private), encoding="utf-8")
            os.chmod(request_path, 0o600)

            code, output, error = self._run([
                "connector-access-request-verify", config, str(request_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertNotIn("cli-private", output)

            with patch.dict(os.environ, {
                "OPC_SHOPIFY_ADMIN_TOKEN": "cli-probe-private-token",
            }, clear=False):
                code, output, error = self._run([
                    "connector-access-probe", config, str(request_path),
                ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertEqual(result["status"], "network_authorization_required")
            self.assertFalse(result["control_boundary"]["network_access_performed"])
            self.assertNotIn("cli-private", output)
            self.assertNotIn("cli-probe-private-token", output)

            code, output, error = self._run([
                "connector-access-probe", config, str(request_path), "--allow-network",
            ])
            self.assertEqual(code, 2)
            self.assertIn("require --output", error)

            receipt_path = Path(temp_dir).resolve() / "shopify-access-receipt.json"
            safe_written = {
                "written": True,
                "status": "passed",
                "ready_for_private_shadow_request": True,
                "receipt_is_digital_signature": False,
            }
            with patch(
                "src.cli.write_connector_access_probe_receipt",
                return_value=safe_written,
            ) as write_receipt:
                code, output, error = self._run([
                    "connector-access-probe", config, str(request_path),
                    "--allow-network", "--output", str(receipt_path),
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], safe_written)
            write_receipt.assert_called_once()

            safe_verified = {
                "valid": True,
                "ready_for_private_shadow_request": True,
                "receipt_is_digital_signature": False,
            }
            with patch(
                "src.cli.verify_private_connector_access_probe_receipt",
                return_value=safe_verified,
            ) as verify_receipt:
                code, output, error = self._run([
                    "connector-access-receipt-verify", config,
                    str(request_path), str(receipt_path),
                    "--as-of", "2026-08-16T12:00:00+00:00",
                    "--maximum-age-days", "30",
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], safe_verified)
            verify_receipt.assert_called_once()

            safe_renewed = {
                "renewed": True,
                "superseded_receipt_retained": True,
                "archive_path_returned": False,
            }
            with patch(
                "src.cli.renew_connector_access_probe_receipt",
                return_value=safe_renewed,
            ) as renew_receipt:
                code, output, error = self._run([
                    "connector-access-receipt-renew", config,
                    str(request_path), str(receipt_path), "--allow-network",
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], safe_renewed)
            renew_receipt.assert_called_once()
            self.assertTrue(renew_receipt.call_args.kwargs["allow_network"])

    def test_connector_access_cli_initializes_environment_bound_wise_request(self):
        config = str(
            ROOT / "examples" / "boxes" / "sg_dtc_wise_store.json"
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            request_path = Path(temp_dir).resolve() / "private-wise-access.json"
            code, output, error = self._run([
                "connector-access-request-init",
                config,
                "--pack", "connector.wise",
                "--entity", "sg_store",
                "--output", str(request_path),
            ])
            self.assertEqual(code, 0, error)
            initialized = json.loads(output)["result"]
            self.assertEqual(initialized["operator_edits_required"], [])
            self.assertFalse(initialized["credentials_included"])
            request = json.loads(request_path.read_text(encoding="utf-8"))
            self.assertEqual(
                request["account_binding"],
                {"mode": "entity_environment_binding"},
            )
            code, output, error = self._run([
                "connector-access-request-verify", config, str(request_path),
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["valid"])
            self.assertNotIn("profile_id", output)

    def test_production_readiness_cli_aggregates_gates_without_claiming_release(self):
        code, output, error = self._run([
            "production-readiness",
            str(ROOT / "examples" / "boxes" / "global_game_studio_xero.json"),
            "--as-of", "2026-08-14",
        ])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertEqual(result["summary"]["stage_count"], 11)
        self.assertTrue(result["summary"]["ready_for_internal_demo"])
        self.assertFalse(result["summary"]["ready_for_bounded_shadow"])
        self.assertFalse(result["summary"]["ready_for_stable_promotion"])
        self.assertFalse(result["summary"]["ready_for_external_filing"])
        self.assertFalse(result["control_boundary"]["external_actions_performed"])

    def test_activation_status_cli_returns_current_operator_wave(self):
        code, output, error = self._run([
            "activation-status",
            str(ROOT / "examples" / "boxes" / "global_game_studio_xero.json"),
            "--as-of", "2026-08-14",
        ])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertEqual(
            result["summary"]["current_wave_stage_ids"],
            ["tax_applicability", "connector_configuration"],
        )
        self.assertTrue(result["control_boundary"]["commands_are_templates_only"])
        self.assertFalse(result["control_boundary"]["commands_executed"])
        self.assertFalse(result["summary"]["ready_for_external_filing"])

    def test_activation_workspace_cli_initializes_verifies_and_evaluates_privately(self):
        config = str(ROOT / "examples" / "boxes" / "global_game_studio.json")
        with tempfile.TemporaryDirectory() as temp_dir:
            workspace = Path(temp_dir).resolve() / "first customer private"
            code, output, error = self._run([
                "activation-init", config, str(workspace),
                "--period", "2026-08",
                "--facts-as-of", "2026-08-14",
                "--prepared-by", "activation-preparer",
            ])
            self.assertEqual(code, 0, error)
            initialized = json.loads(output)["result"]
            self.assertTrue(initialized["initialized"])
            self.assertEqual(initialized["tax_workpaper_count"], 2)
            self.assertEqual(initialized["review_artifact_count"], 0)
            self.assertFalse(initialized["credentials_included"])

            code, output, error = self._run([
                "activation-workspace-verify", config, str(workspace),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertFalse(verified["paths_returned"])

            code, output, error = self._run([
                "activation-workspace-status", config, str(workspace),
                "--as-of", "2026-08-14",
            ])
            self.assertEqual(code, 0, error)
            status = json.loads(output)["result"]
            self.assertEqual(
                status["activation"]["summary"]["current_wave_stage_ids"],
                ["tax_applicability"],
            )
            self.assertFalse(status["control_boundary"]["private_paths_returned"])
            self.assertNotIn(str(workspace), json.dumps(status))

            code, output, error = self._run([
                "connector-access-alerts", config, str(workspace),
                "--as-of", "2026-08-14",
            ])
            self.assertEqual(code, 0, error)
            access_alerts = json.loads(output)["result"]
            self.assertEqual(access_alerts["artifact_type"], "connector_access_alert_candidates")
            self.assertEqual(
                access_alerts["alert_count"],
                status["connector_access_alerts"]["alert_count"],
            )
            self.assertEqual(
                access_alerts["warning_count"],
                status["connector_access_alerts"]["warning_count"],
            )
            self.assertFalse(access_alerts["notifications_sent"])
            self.assertNotIn(str(workspace), json.dumps(access_alerts))

            code, output, error = self._run([
                "activation-runbook-record", config, str(workspace),
                "tax-workpaper-complete:cn_studio",
                "--outcome", "reported-complete",
                "--actor", "activation-operator",
                "--rationale", "底稿已在私有目录完成，但不替代权威税务复核",
            ])
            self.assertEqual(code, 0, error)
            recorded = json.loads(output)["result"]
            self.assertTrue(recorded["recorded"])
            self.assertFalse(recorded["authoritative_completion"])
            self.assertNotIn("activation-operator", output)
            self.assertNotIn(str(workspace), output)

            code, output, error = self._run([
                "activation-runbook-status", config, str(workspace),
            ])
            self.assertEqual(code, 0, error)
            runbook = json.loads(output)["result"]
            self.assertEqual(runbook["event_count"], 1)
            self.assertEqual(runbook["reported_complete_count"], 1)
            self.assertFalse(runbook["authoritative_completion_inferred"])
            self.assertFalse(runbook["evidence_gates_unlocked"])
            self.assertNotIn(str(workspace), output)

            code, output, error = self._run([
                "activation-runbook-verify", config, str(workspace),
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["valid"])

    def test_connector_shadow_status_cli_inspects_private_registry_safely(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            review_dir = Path(temp_dir) / "connector-shadow-reviews"
            review_dir.mkdir(mode=0o700)
            code, output, error = self._run([
                "connector-shadow-status",
                str(ROOT / "examples" / "boxes" / "sg_dtc_shopify_stripe_wise_store.json"),
                "--review-dir", str(review_dir),
                "--as-of", "2026-08-14",
            ])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertEqual(result["artifact_type"], "connector_shadow_registry_workspace")
        self.assertEqual(result["summary"]["activation_status"], "incomplete")
        self.assertEqual(result["summary"]["required_network_pack_count"], 3)
        self.assertFalse(result["summary"]["ready_for_connector_shadow_evidence"])
        self.assertFalse(result["control_boundary"]["paths_returned"])

    def test_pilot_shadow_run_cli_parses_exact_entity_attempt_bindings(self):
        config = ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            expected = {
                "valid": True, "registration_id": "a" * 24,
                "external_actions_performed": False,
            }
            with patch(
                "src.cli.register_pilot_shadow_run", return_value=expected,
            ) as register:
                code, output, error = self._run([
                    "pilot-shadow-run-register", str(config),
                    str(root / "handoff.json"), str(root / "readiness.json"),
                    "--runs-root", str(root / "runs"),
                    "--entity-attempt", "cn_dtc_company=" + "b" * 24,
                    "--actor", "shadow-registrar",
                    "--rationale", "Independent CLI registration rationale.",
                    "--evidence-reference", "evidence://shadow/cli",
                    "--output", str(root / "registration.json"),
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], expected)
            self.assertEqual(
                register.call_args.args[4], {"cn_dtc_company": "b" * 24}
            )

            code, output, error = self._run([
                "pilot-shadow-run-register", str(config),
                str(root / "handoff.json"), str(root / "readiness.json"),
                "--entity-attempt", "not-a-binding",
                "--actor", "shadow-registrar",
                "--rationale", "Invalid binding must fail safely.",
                "--evidence-reference", "evidence://shadow/cli",
                "--output", str(root / "invalid.json"),
            ])
            self.assertEqual(code, 2)
            self.assertEqual(output, "")
            self.assertIn("entity_id=attempt_id", error)

    def test_auth_token_hash_and_policy_validation_never_echo_raw_token(self):
        token = "cli-token-abcdefghijklmnopqrstuvwxyz-12345678"
        with patch.dict(os.environ, {"TEST_OPC_AUTH_TOKEN": token}):
            code, output, error = self._run([
                "auth-token-hash", "--token-env", "TEST_OPC_AUTH_TOKEN",
            ])
        self.assertEqual(code, 0, error)
        hashed = json.loads(output)["result"]
        self.assertRegex(hashed["token_sha256"], r"^[0-9a-f]{64}$")
        self.assertFalse(hashed["raw_token_returned"])
        self.assertNotIn(token, output)
        with tempfile.TemporaryDirectory() as temp_dir:
            policy = Path(temp_dir) / "auth.json"
            policy.write_text(json.dumps({
                "schema_version": 1,
                "principals": [{
                    "principal_id": "cli_operator",
                    "token_sha256": hashed["token_sha256"],
                    "roles": ["operator"],
                }],
            }), encoding="utf-8")
            policy.chmod(0o600)
            code, output, error = self._run(["auth-policy-validate", str(policy)])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertTrue(result["valid"])
        self.assertEqual(result["policy"]["principals"][0]["principal_id"], "cli_operator")
        self.assertNotIn(hashed["token_sha256"], output)

    def test_distribution_verify_cli_rejects_placeholder_build_and_accepts_product_wheel(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "opc_finance_box-0.1.0-py3-none-any.whl"
            DistributionVerifyTests._wheel(path)
            code, output, error = self._run(["distribution-verify", str(path)])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["valid"])
            unknown = Path(temp_dir) / "UNKNOWN-0.0.0-py3-none-any.whl"
            DistributionVerifyTests._wheel(unknown, name="UNKNOWN", version="0.0.0")
            code, output, error = self._run(["distribution-verify", str(unknown)])
            self.assertEqual(code, 2)
            self.assertEqual(output, "")
            self.assertIn("unexpected project name", json.loads(error)["error"])

    def test_deployment_assets_and_isolated_smoke_cli(self):
        code, output, error = self._run([
            "deployment-assets-verify", str(ROOT / "deployment"),
        ])
        self.assertEqual(code, 0, error)
        self.assertTrue(json.loads(output)["result"]["valid"])
        code, output, error = self._run([
            "--packs", str(ROOT / "packs"),
            "deployment-smoke",
            str(ROOT / "examples" / "boxes" / "cn_dtc_store.json"),
            "--timeout-seconds", "15",
        ])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertTrue(result["passed"])
        self.assertTrue(result["server_process_terminated"])
        self.assertFalse(result["external_actions_performed"])

    def test_runtime_data_cli_initializes_backs_up_verifies_and_restores(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            runtime = base / "runtime"
            code, output, error = self._run([
                "runtime-data-init", str(runtime), "--actor", "CLI 部署人",
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["initialized"])
            code, output, error = self._run([
                "runtime-data-upgrade-preflight", str(runtime),
            ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"]["decision"], "no_change")
            backup = base / "backup"
            code, output, error = self._run([
                "runtime-data-backup", str(runtime), str(backup),
                "--actor", "CLI 备份人", "--service-stopped-confirmed",
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["valid"])
            code, output, error = self._run(["runtime-data-backup-verify", str(backup)])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["valid"])
            restored = base / "restored"
            code, output, error = self._run([
                "runtime-data-restore", str(backup), str(restored),
                "--actor", "CLI 恢复人",
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["restored"])

    def test_connector_sync_cli_builds_strict_plan_and_reads_empty_control_state(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            request_base = root / "request-base.json"
            request_base.write_text(
                json.dumps({"shop_domain": "opc-demo.myshopify.com"}), encoding="utf-8",
            )
            config = ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
            sync_root = root / "connector-sync"
            code, output, error = self._run([
                "connector-sync-plan", str(config), "shopify.orders",
                "--entity", "cn_dtc_company", "--stream", "primary-orders",
                "--mode", "incremental", "--window-start", "2026-08-01T00:00:00Z",
                "--window-end", "2026-08-02T00:00:00Z",
                "--request-base", str(request_base), "--sync-root", str(sync_root),
            ])
            self.assertEqual(code, 0, error)
            plan = json.loads(output)["result"]
            self.assertEqual(plan["request"]["mode"], "fetch")
            self.assertEqual(plan["expected_checkpoint_event_hash"], "GENESIS")
            self.assertFalse(plan["secret_values_included"])
            code, output, error = self._run([
                "connector-sync-status", str(config), "--sync-root", str(sync_root),
            ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"]["counts"]["attempts"], 0)
            code, output, error = self._run([
                "connector-sync-verify", "--sync-root", str(sync_root),
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["valid"])

    def test_airwallex_webhook_cli_processes_private_claim_without_exposing_expense_id(self):
        config = ROOT / "examples" / "boxes" / "sg_dtc_shopify_stripe_wise_airwallex_store.json"
        runtime = BoxRuntime(config, ROOT / "packs")
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            webhook_root = root / "webhooks"
            request_base = root / "request-base.json"
            request_base.write_text(json.dumps({
                "currency_minor_units": {"SGD": 2, "USD": 2},
            }), encoding="utf-8")
            body = json.dumps({
                "id": "evt_cli_001", "name": "spend.expense.updated",
                "account_id": "acct_sg_demo",
                "data": {
                    "id": "exp_cli_private_001", "legal_entity_id": "le_sg_demo",
                    "account_id": "acct_sg_demo", "status": "APPROVED",
                },
                "created_at": datetime.now(timezone.utc).isoformat(), "version": "2026-07-17",
            }, separators=(",", ":")).encode()
            secret = "cli-webhook-secret"
            timestamp = str(int(datetime.now(timezone.utc).timestamp() * 1000))
            signature = hmac.new(
                secret.encode(), timestamp.encode() + body, hashlib.sha256,
            ).hexdigest()
            store = AirwallexWebhookStore(webhook_root)
            store.receive(
                body, timestamp=timestamp, signature=signature, secret=secret,
                entity_bindings_json=json.dumps({
                    "sg_store": {
                        "legal_entity_id": "le_sg_demo", "account_id": "acct_sg_demo",
                        "environment": "sandbox",
                    },
                }),
                allowed_entity_ids={"sg_store"},
                runtime_fingerprint=runtime.snapshot()["fingerprint"],
            )
            code, output, error = self._run([
                "airwallex-webhook-status", str(config),
                "--webhook-root", str(webhook_root),
            ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"]["counts"]["pending"], 1)
            self.assertNotIn("exp_cli_private_001", output)

            code, output, error = self._run([
                "airwallex-webhook-process", str(config),
                "--request-base", str(request_base), "--actor", "cli-worker",
                "--shadow-output", str(root / "invalid-multi-shadow.json"),
                "--webhook-root", str(webhook_root),
            ])
            self.assertNotEqual(code, 0)
            self.assertIn("requires --limit 1", error)

            fake_result = {
                "ready": True, "blocked_at": None,
                "founder_briefing": {"record_count": 1, "state_change_count": 0},
                "network_access_performed": True,
                "external_actions_performed": False,
            }
            shadow_output = root / "private-shadow-observation.json"
            shadow_summary = {
                "output": str(shadow_output),
                "observation_fingerprint": "d" * 64,
                "expense_record_count": 1,
                "state_change_candidate_count": 0,
                "raw_source_values_returned": False,
                "financial_amounts_returned": False,
                "external_actions_performed": False,
            }
            with (
                patch("src.cli.dispatch_box_pipeline_request", return_value=fake_result) as dispatch,
                patch(
                    "src.cli.write_airwallex_shadow_observation",
                    return_value=shadow_summary,
                ) as write_observation,
            ):
                code, output, error = self._run([
                    "airwallex-webhook-process", str(config),
                    "--request-base", str(request_base), "--actor", "cli-worker",
                    "--limit", "1", "--shadow-output", str(shadow_output),
                    "--webhook-root", str(webhook_root),
                ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertEqual(result["succeeded_count"], 1)
            self.assertFalse(result["external_actions_performed"])
            self.assertEqual(result["shadow_observation"], shadow_summary)
            write_observation.assert_called_once()
            self.assertNotIn("exp_cli_private_001", output)
            connector_request = dispatch.call_args.args[1]["payload"]["connector_request"]
            self.assertEqual(connector_request["expense_ids"], ["exp_cli_private_001"])
            self.assertEqual(
                connector_request["webhook_contexts"][0]["event_name"],
                "spend.expense.updated",
            )
            self.assertEqual(
                connector_request["webhook_contexts"][0]["runtime_fingerprint"],
                runtime.snapshot()["fingerprint"],
            )
            self.assertNotIn(
                "exp_cli_private_001",
                json.dumps(connector_request["webhook_contexts"]),
            )
            code, output, error = self._run([
                "airwallex-webhook-verify", "--webhook-root", str(webhook_root),
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["valid"])

    def test_airwallex_webhook_cli_resolves_quarantine_with_review_evidence(self):
        config = ROOT / "examples" / "boxes" / "sg_dtc_shopify_stripe_wise_airwallex_store.json"
        runtime = BoxRuntime(config, ROOT / "packs")
        with tempfile.TemporaryDirectory() as temp_dir:
            webhook_root = Path(temp_dir) / "webhooks"
            store = AirwallexWebhookStore(webhook_root)
            body = json.dumps({
                "id": "evt_cli_quarantine", "name": "spend.expense.updated",
                "account_id": "acct_sg_demo",
                "data": {
                    "id": "exp_cli_quarantine", "legal_entity_id": "le_sg_demo",
                    "account_id": "acct_sg_demo", "status": "APPROVED",
                },
                "created_at": datetime.now(timezone.utc).isoformat(), "version": "2026-07-17",
            }, separators=(",", ":")).encode()
            secret = "cli-quarantine-secret"
            timestamp = str(int(datetime.now(timezone.utc).timestamp() * 1000))
            signature = hmac.new(secret.encode(), timestamp.encode() + body, hashlib.sha256).hexdigest()
            receipt = store.receive(
                body, timestamp=timestamp, signature=signature, secret=secret,
                entity_bindings_json=json.dumps({
                    "sg_store": {
                        "legal_entity_id": "le_sg_demo", "account_id": "acct_sg_demo",
                        "environment": "sandbox",
                    },
                }), allowed_entity_ids={"sg_store"},
                runtime_fingerprint=runtime.snapshot()["fingerprint"],
            )
            for _ in range(3):
                claim = store.claim_next(
                    runtime_fingerprint=runtime.snapshot()["fingerprint"], actor="worker",
                )
                store.record_failure(claim, "read-only provider failed", actor="worker")
            code, output, error = self._run([
                "airwallex-webhook-quarantine-resolve", str(config), receipt["receipt_id"],
                "--resolution", "retry", "--actor", "independent-reviewer",
                "--rationale", "Binding and read-only permissions were independently reviewed",
                "--evidence-reference", "review://airwallex/cli-quarantine",
                "--webhook-root", str(webhook_root),
            ])
            self.assertEqual(code, 0, error)
            resolved = json.loads(output)["result"]
            self.assertEqual(resolved["status"], "pending")
            self.assertEqual(resolved["attempt_count"], 0)
            self.assertEqual(resolved["resolution"]["resolution"], "retry")
            self.assertNotIn("exp_cli_quarantine", output)

    def test_connector_shadow_cli_builds_and_seals_real_anonymized_baseline(self):
        config = ROOT / "examples" / "boxes" / "sg_dtc_shopify_stripe_wise_airwallex_store.json"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workpaper_path = root / "real-shadow-workpaper.json"
            baseline_path = root / "real-shadow-baseline.json"
            code, output, error = self._run([
                "connector-shadow-baseline-init", str(config),
                "--pipeline", "finance.expense_evidence_review",
                "--entity", "sg_store", "--period", "2026-08",
                "--prepared-by", "independent-source-preparer",
                "--output", str(workpaper_path),
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["template_only"])
            workpaper = json.loads(workpaper_path.read_text(encoding="utf-8"))
            workpaper["source_independence"] = {
                "prepared_from_independent_source": True,
                "pipeline_output_used_as_baseline": False,
                "source_scope_confirmed": True,
            }
            workpaper["anonymization"]["private_source_evidence_retained"] = True
            workpaper["source_expectations"][0]["expected_record_count"] = 2
            workpaper["source_expectations"][0]["evidence_references"] = [
                "private-export://airwallex/sg-store/2026-08",
            ]
            controls = {
                "pipeline_ready": True, "expense_record_count": 2,
                "receipt_missing_count": 1, "business_purpose_missing_count": 0,
                "uncleared_count": 0, "accounting_mapping_missing_count": 1,
                "state_change_candidate_count": 0,
                "network_refetch_performed": True,
                "webhook_refetch_basis": True,
                "external_actions_disabled": True,
            }
            for item in workpaper["control_expectations"]:
                item["expected_value"] = controls[item["control_id"]]
            workpaper["evidence_references"] = [
                "workpaper://airwallex/sg-store/2026-08/counts",
            ]
            workpaper["finalization_ready"] = True
            workpaper_path.write_text(json.dumps(workpaper), encoding="utf-8")
            code, output, error = self._run([
                "connector-shadow-baseline-finalize", str(config), str(workpaper_path),
                "--output", str(baseline_path),
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertTrue(result["real_sample_evidence"])
            baseline = json.loads(baseline_path.read_text(encoding="utf-8"))
            self.assertEqual(baseline["schema_version"], 2)
            self.assertEqual(baseline["sample_classification"], "real_anonymized")
            self.assertNotIn("runtime_fingerprint", baseline)

    def test_xero_shadow_observe_runs_in_memory_and_returns_only_safe_summary(self):
        config = ROOT / "examples" / "boxes" / "global_game_studio_xero.json"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            request_path = root / "xero-live-request.json"
            output_path = root / "xero-observation.json"
            code, output, error = self._run([
                "xero-shadow-request-init", str(config),
                "--entity", "cn_studio", "--period", "2026-07",
                "--output", str(request_path),
            ])
            self.assertEqual(code, 0, error)
            initialized = json.loads(output)["result"]
            self.assertTrue(initialized["request_contract_complete"])
            self.assertEqual(initialized["operator_edits_required"], [])
            code, output, error = self._run([
                "xero-shadow-request-verify", str(config), str(request_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertFalse(verified["network_access_performed"])
            request = json.loads(request_path.read_text(encoding="utf-8"))
            private_result = {"private_balance": 987654.32}
            safe_summary = {
                "output": str(output_path),
                "observation_fingerprint": "a" * 64,
                "private_pipeline_result_sha256": "b" * 64,
                "trial_balance_line_count": 12,
                "scope_count": 1,
                "pipeline_ready": True,
                "raw_source_values_returned": False,
                "financial_amounts_returned": False,
                "external_actions_performed": False,
            }
            with (
                patch(
                    "src.cli._verified_connector_access_binding",
                    return_value={"mode": "entity_environment_binding"},
                ) as access,
                patch("src.cli.dispatch_box_pipeline_request", return_value=private_result) as dispatch,
                patch(
                    "src.cli.write_xero_shadow_observation", return_value=safe_summary,
                ) as observe,
            ):
                code, output, error = self._run([
                    "xero-shadow-observe", str(config), str(request_path),
                    "--access-request", str(root / "xero-access.json"),
                    "--access-receipt", str(root / "xero-receipt.json"),
                    "--output", str(output_path),
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], safe_summary)
            self.assertNotIn("987654.32", output)
            dispatch.assert_called_once()
            observe.assert_called_once()
            self.assertEqual(access.call_args.kwargs["expected_pack_id"], "connector.xero")
            self.assertEqual(access.call_args.kwargs["expected_entity_id"], "cn_studio")
            self.assertEqual(dispatch.call_args.args[1], request)
            self.assertEqual(observe.call_args.args[2], output_path)

    def test_wise_shadow_observe_runs_in_memory_and_returns_only_safe_summary(self):
        config = ROOT / "examples" / "boxes" / "sg_dtc_wise_store.json"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            request_path = root / "wise-live-request.json"
            output_path = root / "wise-observation.json"
            code, output, error = self._run([
                "wise-shadow-request-init", str(config),
                "--entity", "sg_store", "--period", "2026-07",
                "--output", str(request_path),
            ])
            self.assertEqual(code, 0, error)
            initialized = json.loads(output)["result"]
            self.assertTrue(initialized["request_contract_complete"])
            self.assertEqual(initialized["operator_edits_required"], [])
            code, output, error = self._run([
                "wise-shadow-request-verify", str(config), str(request_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertFalse(verified["network_access_performed"])
            request = json.loads(request_path.read_text(encoding="utf-8"))
            private_result = {"private_closing_balance": 987654.32}
            safe_summary = {
                "output": str(output_path),
                "observation_fingerprint": "a" * 64,
                "private_pipeline_result_sha256": "b" * 64,
                "bank_transaction_count": 12,
                "account_scope_count": 1,
                "pipeline_ready": True,
                "raw_source_values_returned": False,
                "financial_amounts_returned": False,
                "external_actions_performed": False,
            }
            with (
                patch(
                    "src.cli._verified_connector_access_binding",
                    return_value={"mode": "entity_environment_binding"},
                ) as access,
                patch("src.cli.dispatch_box_pipeline_request", return_value=private_result) as dispatch,
                patch("src.cli.write_wise_shadow_observation", return_value=safe_summary) as observe,
            ):
                code, output, error = self._run([
                    "wise-shadow-observe", str(config), str(request_path),
                    "--access-request", str(root / "wise-access.json"),
                    "--access-receipt", str(root / "wise-receipt.json"),
                    "--output", str(output_path),
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], safe_summary)
            self.assertNotIn("987654.32", output)
            dispatch.assert_called_once()
            observe.assert_called_once()
            self.assertEqual(access.call_args.kwargs["expected_pack_id"], "connector.wise")
            self.assertEqual(access.call_args.kwargs["expected_entity_id"], "sg_store")
            self.assertEqual(dispatch.call_args.args[1], request)
            self.assertEqual(observe.call_args.args[2], output_path)

    def test_paypal_shadow_request_and_observe_keep_private_values_out_of_stdout(self):
        config = ROOT / "examples" / "boxes" / "us_dtc_paypal_c_corp.json"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            request_path = root / "paypal-live-request.json"
            output_path = root / "paypal-observation.json"
            code, output, error = self._run([
                "paypal-shadow-request-init", str(config),
                "--entity", "us_dtc_company", "--period", "2026-08",
                "--output", str(request_path),
            ])
            self.assertEqual(code, 0, error)
            initialized = json.loads(output)["result"]
            self.assertTrue(initialized["request_contract_complete"])
            self.assertEqual(initialized["operator_edits_required"], [])
            code, output, error = self._run([
                "paypal-shadow-request-verify", str(config), str(request_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertFalse(verified["network_access_performed"])
            request = json.loads(request_path.read_text(encoding="utf-8"))
            private_result = {
                "private_amount": "987654.32",
                "private_customer": "private@example.invalid",
            }
            safe_summary = {
                "output": str(output_path),
                "observation_fingerprint": "a" * 64,
                "private_pipeline_result_sha256": "b" * 64,
                "transaction_count": 3,
                "refund_candidate_count": 1,
                "reversal_candidate_count": 0,
                "pipeline_ready": True,
                "raw_source_values_returned": False,
                "customer_values_returned": False,
                "financial_amounts_returned": False,
                "external_actions_performed": False,
            }
            with (
                patch(
                    "src.cli._verified_connector_access_binding",
                    return_value={"mode": "entity_environment_binding"},
                ) as access,
                patch("src.cli.dispatch_box_pipeline_request", return_value=private_result) as dispatch,
                patch(
                    "src.cli.write_paypal_shadow_observation", return_value=safe_summary,
                ) as observe,
            ):
                code, output, error = self._run([
                    "paypal-shadow-observe", str(config), str(request_path),
                    "--access-request", str(root / "paypal-access.json"),
                    "--access-receipt", str(root / "paypal-receipt.json"),
                    "--output", str(output_path),
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], safe_summary)
            self.assertNotIn("987654.32", output)
            self.assertNotIn("private@example.invalid", output)
            dispatch.assert_called_once()
            observe.assert_called_once()
            self.assertEqual(access.call_args.kwargs["expected_pack_id"], "connector.paypal")
            self.assertEqual(access.call_args.kwargs["expected_entity_id"], "us_dtc_company")
            self.assertEqual(dispatch.call_args.args[1], request)
            self.assertEqual(observe.call_args.args[2], output_path)

    def test_woocommerce_shadow_request_and_observe_keep_private_values_out_of_stdout(self):
        config = ROOT / "examples" / "boxes" / "us_dtc_woocommerce_c_corp.json"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            request_path = root / "woocommerce-live-request.json"
            output_path = root / "woocommerce-observation.json"
            code, output, error = self._run([
                "woocommerce-shadow-request-init", str(config),
                "--entity", "us_dtc_company", "--period", "2026-08",
                "--output", str(request_path),
            ])
            self.assertEqual(code, 0, error)
            initialized = json.loads(output)["result"]
            self.assertTrue(initialized["request_contract_complete"])
            self.assertEqual(initialized["operator_edits_required"], [])
            code, output, error = self._run([
                "woocommerce-shadow-request-verify", str(config), str(request_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertFalse(verified["network_access_performed"])
            request = json.loads(request_path.read_text(encoding="utf-8"))
            private_result = {
                "private_amount": "987654.32",
                "private_customer": "private@example.invalid",
                "private_site": "https://private.example",
            }
            safe_summary = {
                "output": str(output_path),
                "observation_fingerprint": "a" * 64,
                "private_pipeline_result_sha256": "b" * 64,
                "order_count": 3,
                "refund_event_count": 1,
                "pipeline_ready": True,
                "raw_source_values_returned": False,
                "site_origin_returned": False,
                "customer_or_product_values_returned": False,
                "financial_amounts_returned": False,
                "external_actions_performed": False,
            }
            with (
                patch(
                    "src.cli._verified_connector_access_binding",
                    return_value={"mode": "entity_environment_binding"},
                ) as access,
                patch("src.cli.dispatch_box_pipeline_request", return_value=private_result) as dispatch,
                patch(
                    "src.cli.write_woocommerce_shadow_observation", return_value=safe_summary,
                ) as observe,
            ):
                code, output, error = self._run([
                    "woocommerce-shadow-observe", str(config), str(request_path),
                    "--access-request", str(root / "woocommerce-access.json"),
                    "--access-receipt", str(root / "woocommerce-receipt.json"),
                    "--output", str(output_path),
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], safe_summary)
            self.assertNotIn("987654.32", output)
            self.assertNotIn("private@example.invalid", output)
            self.assertNotIn("private.example", output)
            dispatch.assert_called_once()
            observe.assert_called_once()
            self.assertEqual(
                access.call_args.kwargs["expected_pack_id"], "connector.woocommerce",
            )
            self.assertEqual(
                access.call_args.kwargs["expected_entity_id"], "us_dtc_company",
            )
            self.assertEqual(dispatch.call_args.args[1], request)
            self.assertEqual(observe.call_args.args[2], output_path)

    def test_shipbob_shadow_request_and_observe_keep_private_values_out_of_stdout(self):
        config = ROOT / "examples" / "boxes" / "us_dtc_shopify_stripe_shipbob_c_corp.json"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            request_path = root / "shipbob-live-request.json"
            output_path = root / "shipbob-observation.json"
            code, output, error = self._run([
                "shipbob-shadow-request-init", str(config),
                "--entity", "us_dtc_company", "--period", "2026-08",
                "--output", str(request_path),
            ])
            self.assertEqual(code, 0, error)
            initialized = json.loads(output)["result"]
            self.assertTrue(initialized["request_contract_complete"])
            self.assertEqual(initialized["operator_edits_required"], [])
            code, output, error = self._run([
                "shipbob-shadow-request-verify", str(config), str(request_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertFalse(verified["network_access_performed"])
            request = json.loads(request_path.read_text(encoding="utf-8"))
            private_result = {
                "private_amount": "987654.32",
                "private_sku": "PRIVATE-SKU",
                "private_warehouse": "PRIVATE-WAREHOUSE",
            }
            safe_summary = {
                "output": str(output_path),
                "observation_fingerprint": "a" * 64,
                "private_pipeline_result_sha256": "b" * 64,
                "order_count": 3,
                "shipment_count": 2,
                "return_count": 1,
                "return_item_count": 1,
                "pipeline_ready": True,
                "raw_source_values_returned": False,
                "merchant_account_values_returned": False,
                "customer_or_inventory_values_returned": False,
                "financial_amounts_returned": False,
                "external_actions_performed": False,
            }
            with (
                patch(
                    "src.cli._verified_connector_access_binding",
                    return_value={"mode": "entity_environment_binding"},
                ) as access,
                patch("src.cli.dispatch_box_pipeline_request", return_value=private_result) as dispatch,
                patch(
                    "src.cli.write_shipbob_shadow_observation", return_value=safe_summary,
                ) as observe,
            ):
                code, output, error = self._run([
                    "shipbob-shadow-observe", str(config), str(request_path),
                    "--access-request", str(root / "shipbob-access-request.json"),
                    "--access-receipt", str(root / "shipbob-access-receipt.json"),
                    "--output", str(output_path),
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], safe_summary)
            self.assertNotIn("987654.32", output)
            self.assertNotIn("PRIVATE-SKU", output)
            self.assertNotIn("PRIVATE-WAREHOUSE", output)
            dispatch.assert_called_once()
            observe.assert_called_once()
            self.assertEqual(
                access.call_args.kwargs["expected_pack_id"], "connector.shipbob",
            )
            self.assertEqual(
                access.call_args.kwargs["expected_entity_id"], "us_dtc_company",
            )
            self.assertEqual(dispatch.call_args.args[1], request)
            self.assertEqual(observe.call_args.args[2], output_path)

    def test_amazon_seller_shadow_request_and_observe_keep_private_values_out_of_stdout(self):
        config = ROOT / "examples" / "boxes" / "us_marketplace_amazon_seller_c_corp.json"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            request_path = root / "amazon-live-request.json"
            output_path = root / "amazon-observation.json"
            code, output, error = self._run([
                "amazon-seller-shadow-request-init", str(config),
                "--entity", "us_amazon_marketplace_company", "--period", "2026-07",
                "--marketplace-id", "ATVPDKIKX0DER", "--output", str(request_path),
            ])
            self.assertEqual(code, 0, error)
            initialized = json.loads(output)["result"]
            self.assertTrue(initialized["request_contract_complete"])
            self.assertEqual(initialized["operator_edits_required"], [])
            self.assertNotIn("ATVPDKIKX0DER", output)
            code, output, error = self._run([
                "amazon-seller-shadow-request-verify", str(config), str(request_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertFalse(verified["network_access_performed"])
            self.assertNotIn("ATVPDKIKX0DER", output)
            request = json.loads(request_path.read_text(encoding="utf-8"))
            private_result = {
                "private_amount": "987654.32",
                "private_seller": "PRIVATE-SELLER",
                "private_marketplace": "ATVPDKIKX0DER",
            }
            safe_summary = {
                "output": str(output_path),
                "observation_fingerprint": "a" * 64,
                "private_pipeline_result_sha256": "b" * 64,
                "order_count": 3,
                "inventory_sku_count": 2,
                "transaction_count": 4,
                "pipeline_ready": True,
                "raw_source_values_returned": False,
                "seller_region_or_marketplace_values_returned": False,
                "buyer_product_or_inventory_values_returned": False,
                "financial_amounts_returned": False,
                "external_actions_performed": False,
            }
            with (
                patch(
                    "src.cli._verified_connector_access_binding",
                    return_value={"mode": "entity_environment_binding"},
                ) as access,
                patch("src.cli.dispatch_box_pipeline_request", return_value=private_result) as dispatch,
                patch(
                    "src.cli.write_amazon_seller_shadow_observation", return_value=safe_summary,
                ) as observe,
            ):
                code, output, error = self._run([
                    "amazon-seller-shadow-observe", str(config), str(request_path),
                    "--access-request", str(root / "amazon-access-request.json"),
                    "--access-receipt", str(root / "amazon-access-receipt.json"),
                    "--output", str(output_path),
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], safe_summary)
            self.assertNotIn("987654.32", output)
            self.assertNotIn("PRIVATE-SELLER", output)
            self.assertNotIn("ATVPDKIKX0DER", output)
            dispatch.assert_called_once()
            observe.assert_called_once()
            self.assertEqual(
                access.call_args.kwargs["expected_pack_id"],
                "connector.amazon_seller",
            )
            self.assertEqual(
                access.call_args.kwargs["expected_entity_id"],
                "us_amazon_marketplace_company",
            )
            self.assertEqual(dispatch.call_args.args[1], request)
            self.assertEqual(observe.call_args.args[2], output_path)

    def test_shopify_monthly_shadow_observe_runs_in_memory_and_returns_only_safe_summary(self):
        config = ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            request_path = root / "shopify-monthly-live-request.json"
            output_path = root / "shopify-monthly-observation.json"
            code, output, error = self._run([
                "shopify-monthly-shadow-request-init", str(config),
                "--entity", "cn_dtc_company", "--period", "2026-07",
                "--output", str(request_path),
            ])
            self.assertEqual(code, 0, error)
            initialized = json.loads(output)["result"]
            self.assertTrue(initialized["template_only"])
            request = json.loads(request_path.read_text(encoding="utf-8"))
            request["payload"]["currency_minor_units"] = {"USD": 2}
            request["payload"]["shopify_monthly_request"]["shop_domain"] = (
                "private-store.myshopify.com"
            )
            request["payload"]["processor_links"] = [{
                "entity_id": "cn_dtc_company",
                "shopify_transaction_id": "gid://shopify/OrderTransaction/7201",
                "stripe_source_object_id": "ch_month_7001",
                "evidence": {
                    "source_file": "private-export://processor-links/2026-07",
                    "batch_id": "approved-links-2026-07",
                },
            }]
            request_path.write_text(json.dumps(request), encoding="utf-8")
            if os.name != "nt":
                request_path.chmod(0o600)
            code, output, error = self._run([
                "shopify-monthly-shadow-request-verify", str(config),
                str(request_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertNotIn("private-store.myshopify.com", output)
            self.assertNotIn("gid://", output)
            private_result = {
                "private_order_amount": "987654.32",
                "shop_domain": "private-store.myshopify.com",
                "raw_order_id": "gid://shopify/Order/987654",
            }
            safe_summary = {
                "output": str(output_path),
                "observation_fingerprint": "a" * 64,
                "private_pipeline_result_sha256": "b" * 64,
                "sample_period": "2026-07",
                "shopify_record_count": 12,
                "stripe_record_count": 8,
                "pipeline_ready": True,
                "raw_source_values_returned": False,
                "financial_amounts_returned": False,
                "external_actions_performed": False,
            }
            with (
                patch("src.cli.dispatch_box_pipeline_request", return_value=private_result) as dispatch,
                patch(
                    "src.cli._verified_connector_access_binding",
                    side_effect=[
                        {"mode": "store_domain", "shop_domain": "private-store.myshopify.com"},
                        {"mode": "own_account", "account_id": "acct_123456789ABC"},
                    ],
                ) as access_gate,
                patch(
                    "src.cli.write_shopify_stripe_monthly_shadow_observation",
                    return_value=safe_summary,
                ) as observe,
            ):
                code, output, error = self._run([
                    "shopify-monthly-shadow-observe", str(config), str(request_path),
                    "--output", str(output_path),
                    "--shopify-access-request", str(root / "shopify-access.json"),
                    "--shopify-access-receipt", str(root / "shopify-receipt.json"),
                    "--stripe-access-request", str(root / "stripe-access.json"),
                    "--stripe-access-receipt", str(root / "stripe-receipt.json"),
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], safe_summary)
            self.assertNotIn("987654.32", output)
            self.assertNotIn("private-store.myshopify.com", output)
            self.assertNotIn("gid://", output)
            dispatch.assert_called_once()
            self.assertEqual(access_gate.call_count, 2)
            observe.assert_called_once()
            self.assertEqual(dispatch.call_args.args[1], request)
            self.assertEqual(observe.call_args.args[2], output_path)

    def test_stripe_shadow_request_and_observe_keep_bank_evidence_private(self):
        config = ROOT / "examples" / "boxes" / "cn_dtc_stripe_store.json"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            request_path = root / "stripe-live-request.json"
            output_path = root / "stripe-observation.json"
            code, output, error = self._run([
                "stripe-shadow-request-init", str(config),
                "--entity", "cn_dtc_company", "--period", "2026-08",
                "--output", str(request_path),
            ])
            self.assertEqual(code, 0, error)
            initialized = json.loads(output)["result"]
            self.assertTrue(initialized["template_only"])
            request = json.loads(request_path.read_text(encoding="utf-8"))
            request["payload"]["bank_transactions"] = [{
                "bank_transaction_id": "private_bank_txn_8101",
                "entity_id": "cn_dtc_company",
                "amount_minor": 7180,
                "currency": "USD",
                "direction": "inflow",
                "transaction_date": "2026-08-15",
                "reference": "private Stripe payout evidence",
                "evidence": {
                    "source_file": "private-export://bank/2026-08",
                    "batch_id": "approved-bank-evidence-2026-08",
                },
            }]
            request_path.write_text(json.dumps(request), encoding="utf-8")
            if os.name != "nt":
                request_path.chmod(0o600)
            code, output, error = self._run([
                "stripe-shadow-request-verify", str(config), str(request_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertNotIn("7180", output)
            self.assertNotIn("private_bank_txn_8101", output)
            self.assertNotIn("private Stripe payout evidence", output)

            private_result = {
                "private_amount": "987654.32",
                "raw_payout_id": "po_private_987",
                "bank_reference": "private Stripe payout evidence",
            }
            safe_summary = {
                "output": str(output_path),
                "observation_fingerprint": "a" * 64,
                "private_pipeline_result_sha256": "b" * 64,
                "balance_transaction_count": 12,
                "payout_count": 3,
                "payout_bank_candidate_count": 3,
                "pipeline_ready": True,
                "raw_source_values_returned": False,
                "bank_references_returned": False,
                "financial_amounts_returned": False,
                "external_actions_performed": False,
            }
            with (
                patch(
                    "src.cli.dispatch_box_pipeline_request", return_value=private_result,
                ) as dispatch,
                patch(
                    "src.cli._verified_connector_access_binding",
                    return_value={
                        "mode": "connected_account", "account_id": "acct_123456789ABC",
                    },
                ) as access_gate,
                patch(
                    "src.cli.write_stripe_shadow_observation", return_value=safe_summary,
                ) as observe,
            ):
                code, output, error = self._run([
                    "stripe-shadow-observe", str(config), str(request_path),
                    "--output", str(output_path),
                    "--access-request", str(root / "stripe-access.json"),
                    "--access-receipt", str(root / "stripe-receipt.json"),
                ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"], safe_summary)
            self.assertNotIn("987654.32", output)
            self.assertNotIn("po_private_987", output)
            self.assertNotIn("private Stripe payout evidence", output)
            dispatch.assert_called_once()
            access_gate.assert_called_once()
            observe.assert_called_once()
            dispatched_request = dispatch.call_args.args[1]
            self.assertEqual(
                dispatched_request["payload"]["balance_request"]["stripe_account"],
                "acct_123456789ABC",
            )
            self.assertEqual(
                dispatched_request["payload"]["payout_request"]["stripe_account"],
                "acct_123456789ABC",
            )
            self.assertNotIn("acct_123456789ABC", output)
            self.assertEqual(observe.call_args.args[2], output_path)

    def test_stable_promotion_cli_reads_empty_scoped_control_state(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            promotion_root = Path(temp_dir) / "release-promotion"
            config = ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
            code, output, error = self._run([
                "promotion-status", str(config), "--limit", "25",
                "--promotion-root", str(promotion_root),
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertEqual(result["counts"]["assessments"], 0)
            self.assertEqual(result["list_limit"], 25)
            self.assertFalse(result["raw_financial_values_included"])
            code, output, error = self._run([
                "promotion-verify", "--promotion-root", str(promotion_root),
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["valid"])

    def test_stable_promotion_cli_writes_bound_template_without_overwrite(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "evidence.json"
            config = ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
            code, output, error = self._run([
                "promotion-template", str(config), "core.finance",
                "--output", str(output_path),
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            template = json.loads(output_path.read_text())
            self.assertTrue(result["template_only"])
            self.assertFalse(result["assessment_ready"])
            self.assertEqual(result["runtime_fingerprint"], template["runtime_fingerprint"])
            self.assertEqual(template["pack_id"], "core.finance")
            self.assertFalse(template["sample"]["representative"])
            code, _, error = self._run([
                "promotion-template", str(config), "core.finance",
                "--output", str(output_path),
            ])
            self.assertEqual(code, 2)
            self.assertIn("refusing to overwrite", error)

    def test_pilot_readiness_cli_initializes_reviews_and_verifies_safely(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            config = ROOT / "examples" / "boxes" / "global_game_studio.json"
            workpaper = root / "pilot-workpaper.json"
            reviewed = root / "pilot-reviewed.json"
            code, output, error = self._run([
                "pilot-readiness-init", str(config), "--period", "2026-07",
                "--prepared-by", "pilot-preparer", "--output", str(workpaper),
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["template_only"])
            value = json.loads(workpaper.read_text(encoding="utf-8"))
            value["operator_principal"] = "pilot-operator"
            for entity in value["entities"]:
                for domain in entity["data_domains"]:
                    domain.update({
                        "status": "ready",
                        "acquisition_mode": "file_export",
                        "period_coverage": ["2026-07"],
                        "read_only_confirmed": True,
                        "mapping_approved_by": "mapping-reviewer",
                        "evidence_references": [
                            f"evidence://cli/{entity['entity_id']}/{domain['domain']}"
                        ],
                    })
            entity_ids = [item["entity_id"] for item in value["entities"]]
            for connector in value["network_connectors"]:
                connector.update({
                    "status": "approved_file_fallback", "entity_ids": entity_ids,
                    "credential_reference_configured": False,
                    "provider_contract_passed": False,
                    "bounded_read_window_confirmed": False,
                    "checkpoint_owner": "checkpoint-owner",
                    "mapping_approved_by": "connector-reviewer",
                    "evidence_references": ["evidence://cli/connector/fallback"],
                })
            value["shadow_close_plan"].update({
                "planned": True, "baseline_owner": "baseline-owner",
                "evidence_references": ["workpaper://cli/shadow-plan"],
            })
            workpaper.write_text(json.dumps(value), encoding="utf-8")
            workpaper.chmod(0o600)
            code, output, error = self._run([
                "pilot-readiness-review", str(config), str(workpaper),
                "--actor", "pilot-reviewer",
                "--rationale", "Independent review confirms the bounded pilot controls.",
                "--evidence-reference", "advisor://cli/pilot-review",
                "--output", str(reviewed),
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["ready_for_bounded_shadow"])
            code, output, error = self._run([
                "pilot-readiness-verify", str(config), str(reviewed),
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertTrue(result["ready_for_bounded_shadow"])
            self.assertFalse(result["ready_for_external_filing"])
            self.assertFalse(result["actors_returned"])
            self.assertNotIn("pilot-reviewer", output)
            self.assertNotIn("advisor://", output)
            review_due_at = result["review_due_at"]
            expires_at = result["expires_at"]
            code, output, error = self._run([
                "pilot-readiness-alerts", str(config),
                "--review", str(reviewed), "--as-of", review_due_at,
            ])
            self.assertEqual(code, 0, error)
            alerts = json.loads(output)["result"]
            self.assertEqual(alerts["status"], "review_due")
            self.assertEqual(alerts["warning_count"], 1)
            self.assertFalse(alerts["notifications_sent"])
            self.assertFalse(alerts["schedule_installed"])
            code, output, error = self._run([
                "pilot-readiness-alerts", str(config),
                "--review", str(reviewed), "--as-of", expires_at,
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["ready_for_bounded_shadow"])

    def test_shadow_close_cli_builds_compares_and_signs_private_artifacts(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            config = ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
            workbook_path = root / "shadow-baseline.xlsx"
            code, output, error = self._run([
                "shadow-close-template", str(config), "--output", str(workbook_path),
            ])
            self.assertEqual(code, 0, error)
            template_result = json.loads(output)["result"]
            self.assertEqual(template_result["entity_ids"], ["cn_dtc_company"])
            self.assertFalse(template_result["contains_financial_values"])
            code, _, error = self._run([
                "shadow-close-template", str(config), "--output", str(workbook_path),
            ])
            self.assertEqual(code, 2)
            self.assertIn("refusing to overwrite", error)

            book = load_workbook(workbook_path)
            self.assertIn(
                "cn_dtc_company",
                " ".join(
                    str(cell.value or "")
                    for row in book["Checks"].iter_rows()
                    for cell in row
                ),
            )
            book["基准总账"].append([
                "cn_dtc_company", "2026-07", "1001", "Cash", 100, 0,
                "signed close", "audit://trial-balance", 0, 0,
            ])
            for code_value, name, amount in (
                ("BS_ASSETS", "Assets", 100),
                ("BS_LIABILITIES", "Liabilities", 20),
                ("BS_EQUITY", "Equity", 80),
                ("IS_REVENUE", "Revenue", 50),
                ("IS_EXPENSES", "Expenses", 30),
                ("IS_PROFIT", "Profit", 20),
            ):
                book["基准报表"].append([
                    "cn_dtc_company", "2026-07", code_value, name, amount,
                    "signed close", "audit://statements", 0, 0,
                ])
            book.save(workbook_path)
            book.close()
            finance_path = root / "finance.json"
            finance_path.write_text(json.dumps({
                "entity_id": "cn_dtc_company",
                "period": "2026-07",
                "financial_statements": {
                    "detail": [{
                        "account": "1001 Cash", "closing_debit": 100,
                        "closing_credit": 0,
                    }],
                    "balance_sheet": {
                        "assets": 100, "liabilities": 20,
                        "liabilities_and_equity": 100,
                    },
                    "income_statement": {
                        "revenue": 50, "expenses": 30,
                        "profit_before_tax": 20,
                    },
                },
                "tax_pack": {},
            }), encoding="utf-8")
            report_path = root / "shadow-report.json"
            code, output, error = self._run([
                "shadow-close-compare", str(config), str(workbook_path),
                str(finance_path), "--output", str(report_path),
            ])
            self.assertEqual(code, 0, error)
            comparison = json.loads(output)["result"]
            self.assertEqual(comparison["comparison_count"], 7)
            self.assertEqual(comparison["exception_count"], 0)
            self.assertFalse(comparison["raw_financial_values_returned"])
            self.assertNotIn("manual_value", output)
            self.assertEqual(report_path.stat().st_mode & 0o777, 0o600)
            reviewed_path = root / "shadow-reviewed.json"
            code, output, error = self._run([
                "shadow-close-review", str(config), str(report_path),
                "--decision", "passed", "--actor", "independent-reviewer",
                "--rationale", "Verified signed ledger and statements",
                "--evidence-reference", "audit://review-record",
                "--output", str(reviewed_path),
            ])
            self.assertEqual(code, 0, error)
            signed = json.loads(output)["result"]
            self.assertTrue(signed["review_current"])
            self.assertEqual(signed["decision"], "验证通过")
            self.assertFalse(signed["raw_financial_values_returned"])
            reviewed = json.loads(reviewed_path.read_text(encoding="utf-8"))
            self.assertEqual(
                reviewed["review"]["report_fingerprint"],
                reviewed["report_fingerprint"],
            )
            self.assertEqual(reviewed_path.stat().st_mode & 0o777, 0o600)
            code, output, error = self._run([
                "shadow-close-verify", str(config), str(reviewed_path),
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["valid"])
            self.assertTrue(verified["review_current"])
            self.assertEqual(verified["decision"], "验证通过")
            self.assertFalse(verified["raw_financial_values_returned"])
            self.assertNotIn("manual_value", output)
            code, _, error = self._run([
                "shadow-close-verify",
                str(ROOT / "examples" / "boxes" / "cn_dtc_store.json"),
                str(reviewed_path),
            ])
            self.assertEqual(code, 2)
            self.assertIn("different Box runtime fingerprint", error)

            finance = json.loads(finance_path.read_text(encoding="utf-8"))
            finance["financial_statements"]["income_statement"]["revenue"] = 45
            finance_path.write_text(json.dumps(finance), encoding="utf-8")
            exception_report_path = root / "shadow-report-with-difference.json"
            code, _, error = self._run([
                "shadow-close-compare", str(config), str(workbook_path),
                str(finance_path), "--output", str(exception_report_path),
            ])
            self.assertEqual(code, 0, error)
            exception_report = json.loads(
                exception_report_path.read_text(encoding="utf-8")
            )
            exceptions = [
                {
                    "domain": item["domain"],
                    "key": item["key"],
                    "classification": "cutoff",
                    "rationale": "Historical close used a reviewed cutoff convention",
                    "evidence_references": ["audit://difference-resolution"],
                }
                for item in exception_report["comparisons"]
                if item["status"] != "一致"
            ]
            resolutions_path = root / "resolutions.json"
            resolutions_path.write_text(json.dumps(exceptions), encoding="utf-8")
            accepted_path = root / "shadow-accepted.json"
            code, output, error = self._run([
                "shadow-close-review", str(config), str(exception_report_path),
                "--decision", "accepted-differences",
                "--actor", "independent-reviewer",
                "--rationale", "Reviewed every difference and its evidence",
                "--evidence-reference", "audit://accepted-difference-review",
                "--resolutions", str(resolutions_path),
                "--output", str(accepted_path),
            ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"]["decision"], "接受差异")
            self.assertEqual(
                json.loads(accepted_path.read_text(encoding="utf-8"))["review"]
                ["exception_resolutions"],
                exceptions,
            )

    def test_shadow_close_cli_rejects_scope_mismatch_and_tampered_report(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            config = ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
            workbook_path = root / "baseline.xlsx"
            book = Workbook()
            book.active.title = "基准报表"
            book.active.append([
                "主体ID", "期间", "指标编码", "指标名称", "金额", "来源",
                "证据说明", "绝对容差", "百分比容差",
            ])
            book.active.append([
                "cn_dtc_company", "2026-07", "IS_REVENUE", "Revenue", 50,
                "signed close", "audit://statement", 0, 0,
            ])
            book.save(workbook_path)
            finance_path = root / "finance.json"
            finance_path.write_text(json.dumps({
                "entity_id": "other_entity", "period": "2026-07",
                "financial_statements": {"detail": []},
            }), encoding="utf-8")
            report_path = root / "report.json"
            code, _, error = self._run([
                "shadow-close-compare", str(config), str(workbook_path),
                str(finance_path), "--output", str(report_path),
            ])
            self.assertEqual(code, 2)
            self.assertIn("entity_id does not match", error)
            self.assertFalse(report_path.exists())

            finance_path.write_text(json.dumps({
                "entity_id": "cn_dtc_company", "period": "2026-07",
                "financial_statements": {
                    "detail": [], "balance_sheet": {},
                    "income_statement": {"revenue": 50},
                },
            }), encoding="utf-8")
            code, _, error = self._run([
                "shadow-close-compare", str(config), str(workbook_path),
                str(finance_path), "--output", str(report_path),
            ])
            self.assertEqual(code, 0, error)
            report = json.loads(report_path.read_text(encoding="utf-8"))
            report["comparisons"][0]["manual_value"] = 999
            report_path.write_text(json.dumps(report), encoding="utf-8")
            code, _, error = self._run([
                "shadow-close-verify", str(config), str(report_path),
            ])
            self.assertEqual(code, 2)
            self.assertRegex(error, "差异与人工/Agent|指纹")
            code, _, error = self._run([
                "shadow-close-review", str(config), str(report_path),
                "--decision", "needs-correction", "--actor", "reviewer",
                "--rationale", "Source values need correction",
                "--output", str(root / "reviewed.json"),
            ])
            self.assertEqual(code, 2)
            self.assertIn("指纹", error)

    def test_pack_audit_reports_contract_and_implementation_separately(self):
        code, output, error = self._run(["pack-audit"])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertTrue(result["contract_valid"])
        self.assertTrue(result["complete_implementation"])
        self.assertEqual(result["coverage_counts"]["declared_only"], 0)

    def test_create_validate_compile_flow(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            box_path = Path(temp_dir) / "box.json"
            compiled_path = Path(temp_dir) / "compiled"
            code, _, error = self._run([
                "create",
                str(ROOT / "examples" / "box_specs" / "dtc_cn.json"),
                "--output",
                str(box_path),
            ])
            self.assertEqual(code, 0, error)
            code, output, error = self._run(["validate", str(box_path)])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"]["name"], "我的中国独立站 OPC")
            code, _, error = self._run(["compile", str(box_path), "--output", str(compiled_path)])
            self.assertEqual(code, 0, error)
            self.assertTrue((compiled_path / "box.lock.json").exists())

    def test_create_shopify_stripe_stack_with_selected_tax_country(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            box_path = Path(temp_dir) / "sg-box.json"
            compiled_path = Path(temp_dir) / "compiled"
            code, _, error = self._run([
                "create", str(ROOT / "examples" / "box_specs" / "shopify_stripe_sg.json"),
                "--output", str(box_path),
            ])
            self.assertEqual(code, 0, error)
            box = json.loads(box_path.read_text(encoding="utf-8"))
            self.assertEqual(box["entities"][0]["tax_pack"], "jurisdiction.sg")
            self.assertEqual(box["connectors"], [
                "connector.file_import", "connector.shopify", "connector.stripe",
            ])
            code, _, error = self._run([
                "compile", str(box_path), "--output", str(compiled_path),
            ])
            self.assertEqual(code, 0, error)
            pipelines = json.loads(
                (compiled_path / "pipeline-catalog.json").read_text(encoding="utf-8")
            )
            combined = next(
                item for item in pipelines
                if item["pipeline_id"] == "dtc.shopify_stripe_daily_close"
            )
            self.assertEqual(combined["implementation_status"], "executable")

    def test_installed_cli_can_scaffold_a_new_tax_country_pack(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            code, output, error = self._run([
                "jurisdiction-init",
                "--output-root", temp_dir,
                "--slug", "us_federal",
                "--country-code", "US",
                "--display-name", "US Federal Tax Design Pack",
                "--source-authority", "Internal Revenue Service",
                "--source-title", "Official business tax guide",
                "--source-url", "https://www.irs.gov/businesses",
                "--verified-at", "2026-08-13",
                "--rules-effective-at", "2026-01-01",
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertEqual(result["pack_id"], "jurisdiction.us_federal")
            self.assertEqual(result["tax_readiness"], "design")
            self.assertTrue((Path(temp_dir) / "us_federal" / "rules.json").exists())

    def test_cli_scaffolds_a_secret_safe_network_connector_pack(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            code, output, error = self._run([
                "connector-init", "--output-root", temp_dir,
                "--slug", "sample_store", "--display-name", "Sample Store API",
                "--secret-env", "OPC_SAMPLE_STORE_TOKEN",
                "--base-url", "https://api.example.test/v1/finance",
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            destination = Path(result["destination"])
            self.assertEqual(result["pack_id"], "connector.sample_store")
            self.assertTrue((destination / "provider_contract_test.py").exists())
            contract = json.loads((destination / "provider-contract.json").read_text(encoding="utf-8"))
            self.assertTrue(contract["network_fetch_implemented"])
            self.assertEqual(contract["secret_env"], "OPC_SAMPLE_STORE_TOKEN")

    def test_invalid_service_request_returns_safe_error(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            request = Path(temp_dir) / "request.json"
            request.write_text("{}", encoding="utf-8")
            code, output, error = self._run([
                "dispatch",
                str(ROOT / "examples" / "boxes" / "global_game_studio.json"),
                str(request),
            ])
            self.assertEqual(code, 2)
            self.assertEqual(output, "")
            payload = json.loads(error)
            self.assertFalse(payload["ok"])
            self.assertEqual(payload["error"], "service_id is required")

    def test_doctor_returns_distinct_code_when_runtime_is_unsupported(self):
        code, output, error = self._run([
            "doctor",
            str(ROOT / "examples" / "boxes" / "global_game_studio.json"),
        ])
        payload = json.loads(output)
        self.assertEqual(error, "")
        self.assertIn(code, {0, 3})
        self.assertEqual(code == 0, payload["result"]["ready"])

    def test_tax_rule_status_uses_explicit_clock_and_never_releases_filing(self):
        code, output, error = self._run([
            "tax-rule-status",
            str(ROOT / "examples" / "boxes" / "global_game_studio.json"),
            "--as-of", "2027-02-10",
        ])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertEqual(result["counts"]["expired"], 2)
        self.assertFalse(result["calendar_release_allowed"])
        self.assertFalse(result["external_filing_release_allowed"])
        self.assertFalse(result["tax_calculation_performed"])
        self.assertFalse(result["external_actions_performed"])

        code, output, error = self._run([
            "doctor", str(ROOT / "examples" / "boxes" / "global_game_studio.json"),
            "--as-of", "2027-02-10",
        ])
        self.assertEqual(code, 0, error)
        lifecycle = next(
            item for item in json.loads(output)["result"]["checks"]
            if item["check_id"] == "tax.rule_lifecycle"
        )
        self.assertEqual(lifecycle["status"], "warning")

    def test_tax_applicability_cli_creates_reviews_and_verifies_one_entity(self):
        config = str(ROOT / "examples" / "boxes" / "global_game_studio.json")
        with tempfile.TemporaryDirectory() as temp_dir:
            workpaper = Path(temp_dir) / "cn-workpaper.json"
            review_dir = Path(temp_dir) / "reviews"
            review_dir.mkdir()
            reviewed = Path(temp_dir) / "cn-reviewed-source.json"
            code, output, error = self._run([
                "tax-applicability-init", config,
                "--entity", "cn_studio",
                "--prepared-by", "tax-operator",
                "--facts-as-of", "2026-08-14",
                "--output", str(workpaper),
            ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"]["unanswered_count"], 5)
            value = json.loads(workpaper.read_text(encoding="utf-8"))
            answers = {
                "legal_form_and_pack_scope": "confirmed_in_scope",
                "tax_residency_and_permanent_establishment": "confirmed_in_scope",
                "direct_and_indirect_tax_registrations": "confirmed_complete",
                "fiscal_year_and_return_periods": "confirmed",
                "special_cross_border_and_group_regimes": "reviewed_no_additional_scope",
            }
            for question in value["entity"]["questions"]:
                question["answer"] = answers[question["question_id"]]
                question["evidence_references"] = [
                    f"evidence://cn-review/{question['question_id']}"
                ]
            workpaper.write_text(
                json.dumps(value, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            code, output, error = self._run([
                "tax-applicability-review", config, str(workpaper),
                "--decision", "approved-in-scope",
                "--actor", "cn-local-tax-reviewer",
                "--rationale", "已完成当地适用性复核。",
                "--evidence-reference", "advisor://cn-review/memo",
                "--output", str(reviewed),
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["applicability_gate_passed"])
            self.assertNotIn("confirmed_in_scope", output)
            code, output, error = self._run([
                "tax-applicability-import", config, str(reviewed),
                "--review-dir", str(review_dir), "--as-of", "2026-08-14",
            ])
            self.assertEqual(code, 0, error)
            imported = json.loads(output)["result"]
            self.assertTrue(imported["imported"])
            self.assertEqual(imported["entity_id"], "cn_studio")
            self.assertFalse(imported["overwrite_performed"])
            self.assertFalse(imported["paths_returned"])
            self.assertTrue((review_dir / "cn_studio.json").is_file())
            code, output, error = self._run([
                "tax-applicability-import", config, str(reviewed),
                "--review-dir", str(review_dir), "--as-of", "2026-08-14",
            ])
            self.assertEqual(code, 2)
            self.assertEqual(output, "")
            self.assertIn("refusing to overwrite", error)
            code, output, error = self._run([
                "tax-applicability-verify", config, str(reviewed),
                "--as-of", "2026-08-14",
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertEqual(result["entity_id"], "cn_studio")
            self.assertTrue(result["applicability_gate_passed"])
            self.assertFalse(result["answers_returned"])
            self.assertNotIn("已完成当地适用性复核", output)
            code, output, error = self._run([
                "tax-applicability-portfolio-verify", config, str(reviewed),
                "--as-of", "2026-08-14",
            ])
            self.assertEqual(code, 2)
            self.assertEqual(output, "")
            self.assertIn("missing=['sg_publisher']", json.loads(error)["error"])
            code, output, error = self._run([
                "tax-applicability-status", config,
                "--review-dir", str(review_dir), "--as-of", "2026-08-14",
            ])
            self.assertEqual(code, 0, error)
            registry = json.loads(output)["result"]
            self.assertEqual(registry["counts"]["current"], 1)
            self.assertEqual(registry["counts"]["missing"], 1)
            self.assertFalse(registry["ready_for_calendar_release"])
            self.assertFalse(registry["paths_returned"])
            code, output, error = self._run([
                "doctor", config, "--as-of", "2026-08-14",
                "--tax-applicability-review", str(reviewed),
            ])
            self.assertEqual(code, 0, error)
            applicability = next(
                item for item in json.loads(output)["result"]["checks"]
                if item["check_id"] == "tax.applicability_reviews"
            )
            self.assertEqual(applicability["details"]["approved_entity_ids"], ["cn_studio"])
            self.assertEqual(applicability["details"]["missing_entity_ids"], ["sg_publisher"])
            self.assertFalse(
                json.loads(output)["result"]["ready_for_tax_calendar_release"]
            )

    def test_tax_applicability_registry_receipt_cli_seals_and_verifies(self):
        config_path = ROOT / "examples" / "boxes" / "global_game_studio.json"
        runtime = BoxRuntime(config_path, ROOT / "packs")
        answers = {
            "legal_form_and_pack_scope": "confirmed_in_scope",
            "tax_residency_and_permanent_establishment": "confirmed_in_scope",
            "direct_and_indirect_tax_registrations": "confirmed_complete",
            "fiscal_year_and_return_periods": "confirmed",
            "special_cross_border_and_group_regimes": "reviewed_no_additional_scope",
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            review_dir = root / "reviews"
            review_dir.mkdir()
            for entity_id in ("cn_studio", "sg_publisher"):
                workpaper = build_tax_applicability_workpaper(
                    runtime, entity_id, prepared_by=f"{entity_id}-preparer",
                    facts_as_of="2026-08-14",
                )
                for question in workpaper["entity"]["questions"]:
                    question["answer"] = answers[question["question_id"]]
                    question["evidence_references"] = [
                        f"evidence://cli-receipt/{entity_id}/{question['question_id']}"
                    ]
                workpaper_path = root / f"{entity_id}-workpaper.json"
                source_review = root / f"{entity_id}-review.json"
                workpaper_path.write_text(json.dumps(workpaper), encoding="utf-8")
                workpaper_path.chmod(0o600)
                review_tax_applicability_workpaper(
                    runtime, workpaper_path, source_review,
                    decision="approved-in-scope", actor=f"{entity_id}-reviewer",
                    rationale="Private CLI registry receipt test.",
                    evidence_references=[f"advisor://cli-receipt/{entity_id}"],
                )
                import_tax_applicability_review(
                    runtime, source_review, review_dir, as_of="2026-08-14",
                )
            receipt = root / "registry-receipt.json"
            code, output, error = self._run([
                "tax-applicability-registry-seal", str(config_path),
                "--review-dir", str(review_dir), "--actor", "registry-controller",
                "--as-of", "2026-08-14", "--output", str(receipt),
            ])
            self.assertEqual(code, 0, error)
            created = json.loads(output)["result"]
            self.assertTrue(created["receipt_created"])
            self.assertFalse(created["paths_returned"])
            self.assertTrue(receipt.is_file())
            code, output, error = self._run([
                "tax-applicability-registry-verify", str(config_path), str(receipt),
                "--review-dir", str(review_dir), "--as-of", "2027-07-15",
            ])
            self.assertEqual(code, 0, error)
            verified = json.loads(output)["result"]
            self.assertTrue(verified["registry_unchanged"])
            self.assertEqual(verified["counts"]["review_due"], 2)
            self.assertFalse(verified["digital_signature_verified"])
            code, output, error = self._run([
                "tax-applicability-alerts", str(config_path),
                "--review-dir", str(review_dir), "--receipt", str(receipt),
                "--as-of", "2027-07-15",
            ])
            self.assertEqual(code, 0, error)
            alerts = json.loads(output)["result"]
            self.assertEqual(alerts["warning_count"], 2)
            self.assertEqual(alerts["critical_count"], 0)
            self.assertFalse(alerts["notifications_sent"])
            self.assertFalse(alerts["paths_returned"])

    def test_upgrade_check_returns_distinct_code_for_blocking_contract_change(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            baseline = compile_box_file(
                ROOT / "examples" / "boxes" / "global_game_studio.json", ROOT / "packs"
            )
            baseline["entities"][0]["functional_currency"] = "USD"
            path = Path(temp_dir) / "previous-box.lock.json"
            path.write_text(json.dumps(baseline), encoding="utf-8")
            code, output, error = self._run([
                "upgrade-check",
                str(ROOT / "examples" / "boxes" / "global_game_studio.json"),
                str(path),
            ])
            self.assertEqual(code, 4, error)
            result = json.loads(output)["result"]
            self.assertFalse(result["compatible"])
            self.assertGreater(result["counts"]["blocking"], 0)

    def test_eval_cli_runs_bundled_finance_boundary_suite(self):
        code, output, error = self._run([
            "eval", str(ROOT / "evals" / "core_packs.json"), "--project-root", str(ROOT),
        ])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertTrue(result["passed"])
        self.assertEqual(result["counts"]["total"], 48)

    def test_pipeline_cli_can_record_and_query_secret_free_attempt(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            runs_root = Path(temp_dir) / "runs"
            config = ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
            request = ROOT / "examples" / "pipelines" / "shopify_stripe_daily_close_fixture.json"
            code, output, error = self._run([
                "pipeline", str(config), str(request), "--record",
                "--runs-root", str(runs_root), "--actor", "CLI 测试负责人",
            ])
            self.assertEqual(code, 0, error)
            payload = json.loads(output)["result"]
            self.assertTrue(payload["pipeline_result"]["ready"])
            attempt_id = payload["run_record"]["attempt_id"]
            self.assertFalse(payload["run_record"]["secret_values_persisted"])

            code, output, error = self._run([
                "pipeline-runs", str(config), "--runs-root", str(runs_root),
                "--pipeline-id", "dtc.shopify_stripe_daily_close", "--entity", "cn_dtc_company",
            ])
            self.assertEqual(code, 0, error)
            runs = json.loads(output)["result"]["runs"]
            self.assertEqual([item["attempt_id"] for item in runs], [attempt_id])

            code, output, error = self._run([
                "pipeline-run-show", str(config), attempt_id, "--runs-root", str(runs_root),
            ])
            self.assertEqual(code, 0, error)
            shown = json.loads(output)["result"]["run"]
            self.assertEqual(shown["attempt_id"], attempt_id)
            self.assertEqual(shown["review_status"], "pending_review")

            code, output, error = self._run([
                "pipeline-review-queue", str(config), "--runs-root", str(runs_root),
            ])
            self.assertEqual(code, 0, error)
            self.assertEqual(len(json.loads(output)["result"]["review_tasks"]), 3)

            for gate in shown["required_review_gates"]:
                code, output, error = self._run([
                    "pipeline-run-review", str(config), attempt_id,
                    "--runs-root", str(runs_root), "--gate", gate,
                    "--decision", "approved", "--actor", "CLI 复核人",
                    "--rationale", "证据已完成复核",
                    "--evidence-reference", f"evidence://{gate}",
                ])
                self.assertEqual(code, 0, error)
            reviewed = json.loads(output)["result"]["run"]
            self.assertTrue(reviewed["review_complete"])
            self.assertTrue(reviewed["release_candidate"])
            self.assertFalse(reviewed["release_candidate_is_external_authorization"])
            code, output, error = self._run([
                "pipeline-review-queue", str(config), "--runs-root", str(runs_root),
            ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"]["review_tasks"], [])
            code, output, error = self._run([
                "pipeline-runs-verify", str(config), "--runs-root", str(runs_root),
            ])
            self.assertEqual(code, 0, error)
            integrity = json.loads(output)["result"]["integrity"]
            self.assertTrue(integrity["valid"])
            self.assertEqual(integrity["review_event_count_for_box"], 3)
            backup = Path(temp_dir) / "backup"
            code, output, error = self._run([
                "pipeline-runs-backup", str(backup), "--runs-root", str(runs_root),
                "--actor", "CLI 备份人",
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["backup"]["valid"])
            code, output, error = self._run(["pipeline-backup-verify", str(backup)])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"]["backup"]["event_count"], 4)
            restored_root = Path(temp_dir) / "restored-runs"
            code, output, error = self._run([
                "pipeline-runs-restore", str(backup), "--runs-root", str(restored_root),
                "--actor", "CLI 恢复人",
            ])
            self.assertEqual(code, 0, error)
            self.assertTrue(json.loads(output)["result"]["restore"]["restored"])
            serialized = (runs_root / "pipeline_runs.jsonl").read_text(encoding="utf-8")
            self.assertNotIn("Demo Shopify payment", serialized)
            self.assertNotIn("processor-links-demo.csv", serialized)

    def test_pipeline_schedule_cli_inspects_and_runs_one_approved_due_job(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            now = datetime.now(timezone.utc).replace(second=0, microsecond=0)
            request = root / "request.json"
            request.write_bytes(
                (ROOT / "examples" / "pipelines" / "shopify_stripe_daily_close_fixture.json").read_bytes()
            )
            schedule = root / "schedule.json"
            schedule_job = {
                    "job_id": "cli-dtc-close", "enabled": True,
                    "pipeline_id": "dtc.shopify_stripe_daily_close",
                    "entity_id": "cn_dtc_company", "request_file": "request.json",
                    "request_fingerprint": pipeline_request_fingerprint(
                        json.loads(request.read_text(encoding="utf-8"))
                    ),
                    "cadence": {"kind": "daily", "local_time": now.strftime("%H:%M")},
                    "execution_window_minutes": 60, "max_attempts": 2,
                    "retry_delay_minutes": 15, "lease_seconds": 900,
                    "operator": "cli_scheduler", "alert_owner": "finance_owner",
                    "approved_by": "schedule_reviewer", "approved_at": now.isoformat(),
                    "approval_fingerprint": None,
            }
            schedule_job["approval_fingerprint"] = schedule_job_approval_fingerprint(schedule_job)
            schedule.write_text(json.dumps({
                "schema_version": 2,
                "timezone": "UTC",
                "jobs": [schedule_job],
            }), encoding="utf-8")
            runs_root = root / "runs"
            config = ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
            common = [
                str(config), str(schedule), "--runs-root", str(runs_root),
                "--now", now.isoformat(),
            ]
            code, output, error = self._run(["pipeline-schedule-inspect", *common])
            self.assertEqual(code, 0, error)
            inspected = json.loads(output)["result"]
            self.assertEqual(inspected["jobs"][0]["status"], "due")
            self.assertFalse(inspected["dispatch_performed"])
            code, output, error = self._run([
                "pipeline-schedule-run", *common, "--actor", "cli_scheduler",
            ])
            self.assertEqual(code, 0, error)
            run = json.loads(output)["result"]
            self.assertEqual(run["counts"]["dispatched"], 1)
            self.assertEqual(run["outcomes"][0]["status"], "ready")
            code, output, error = self._run(["pipeline-schedule-inspect", *common])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"]["jobs"][0]["status"], "completed")

    def test_pipeline_request_fingerprint_cli_returns_only_digest(self):
        request = ROOT / "examples" / "pipelines" / "shopify_stripe_daily_close_fixture.json"
        code, output, error = self._run(["pipeline-request-fingerprint", str(request)])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertRegex(result["request_fingerprint"], r"^[0-9a-f]{64}$")
        self.assertFalse(result["raw_request_returned"])
        self.assertNotIn("Demo Shopify payment", output)

    def test_pipeline_observability_cli_exports_safe_json_and_prometheus(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
            common = [
                "pipeline-observability", str(config),
                "--runs-root", str(Path(temp_dir) / "runs"),
                "--now", "2026-08-14T00:00:00+00:00",
            ]
            code, output, error = self._run(common)
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertFalse(result["schedule_configured"])
            self.assertFalse(result["secret_values_included"])
            code, output, error = self._run([*common, "--prometheus"])
            self.assertEqual(code, 0, error)
            metrics = json.loads(output)["result"]["metrics"]
            self.assertIn("opc_finance_pipeline_ledger_integrity 1", metrics)
            self.assertNotIn(str(temp_dir), metrics)

    def test_game_pipeline_cli_records_two_independent_review_tasks(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            statement = root / "App Store结算.xlsx"
            request_path = root / "request.json"
            book = Workbook()
            sheet = book.active
            sheet.title = "商店金流账单-App Store"
            sheet.append([
                "账期月份", "游戏名称", "平台", "渠道", "总流水", "退款流水",
                "结算金额", "预提所得税（结算币种）", "甲方实收金额（结算币种）", "结算币种",
            ])
            sheet.append(["2026-07", "G1", "iOS", "App Store", 1000, 0, 700, 70, 630, "USD"])
            book.save(statement)
            request_path.write_text(json.dumps({
                "pipeline_id": "game.channel_settlement_close",
                "payload": {
                    "entity_id": "sg_publisher",
                    "connector_id": "file.app_store_settlements",
                    "connector_request": {"path": str(statement)},
                    "contract_mappings": [{
                        "entity_id": "sg_publisher", "period": "2026-07", "game": "G1",
                        "channel": "App Store", "currency": "USD",
                        "contract_basis": 1000, "contract_rate": 0.7,
                        "evidence": {
                            "source_reference": "contract://app-store/G1/2026",
                            "captured_at": "2026-08-13",
                        },
                    }],
                },
            }), encoding="utf-8")
            runs_root = root / "runs"
            config = ROOT / "examples" / "boxes" / "global_game_studio.json"
            code, output, error = self._run([
                "pipeline", str(config), str(request_path), "--record",
                "--runs-root", str(runs_root), "--actor", "游戏月结执行人",
            ])
            self.assertEqual(code, 0, error)
            result = json.loads(output)["result"]
            self.assertTrue(result["pipeline_result"]["ready"])
            self.assertEqual(result["run_record"]["required_review_gates"], [
                "channel_contract_mapping", "game_principal_agent_assessment",
            ])
            code, output, error = self._run([
                "pipeline-review-queue", str(config), "--runs-root", str(runs_root),
            ])
            self.assertEqual(code, 0, error)
            tasks = json.loads(output)["result"]["review_tasks"]
            self.assertEqual({task["gate"] for task in tasks}, {
                "channel_contract_mapping", "game_principal_agent_assessment",
            })

    def test_pipeline_run_cli_is_scoped_to_box_runtime_fingerprint(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            runs_root = Path(temp_dir) / "runs"
            shopify_config = ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json"
            request = ROOT / "examples" / "pipelines" / "shopify_stripe_daily_close_fixture.json"
            code, output, error = self._run([
                "pipeline", str(shopify_config), str(request), "--record",
                "--runs-root", str(runs_root),
            ])
            self.assertEqual(code, 0, error)
            attempt_id = json.loads(output)["result"]["run_record"]["attempt_id"]
            code, output, error = self._run([
                "pipeline-runs",
                str(ROOT / "examples" / "boxes" / "cn_dtc_stripe_store.json"),
                "--runs-root", str(runs_root),
            ])
            self.assertEqual(code, 0, error)
            self.assertEqual(json.loads(output)["result"]["runs"], [])
            code, _, error = self._run([
                "pipeline-run-show",
                str(ROOT / "examples" / "boxes" / "cn_dtc_stripe_store.json"),
                attempt_id, "--runs-root", str(runs_root),
            ])
            self.assertEqual(code, 2)
            self.assertIn("not found for this Box", error)

    def test_connector_cli_imports_standard_commerce_batch(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "orders.csv"
            request = Path(temp_dir) / "request.json"
            source.write_text(
                "订单ID,法律主体ID,期间,渠道,目的地国家,币种,商品原价不含税\n"
                "DTC-1,cn_dtc_company,2026-07,DTC,US,USD,100\n",
                encoding="utf-8",
            )
            request.write_text(json.dumps({"path": str(source)}), encoding="utf-8")
            code, output, error = self._run([
                "import",
                str(ROOT / "examples" / "boxes" / "cn_dtc_store.json"),
                "file.commerce",
                str(request),
            ])
        self.assertEqual(code, 0, error)
        self.assertTrue(json.loads(output)["result"]["batch"]["quality"]["ready"])

    def test_commerce_pipeline_cli_reaches_reconciled_analysis(self):
        code, output, error = self._run([
            "commerce-pipeline",
            str(ROOT / "examples" / "boxes" / "cn_dtc_api_store.json"),
            "example.commerce_api_payload",
            str(ROOT / "examples" / "connectors" / "commerce_api_payload.json"),
        ])
        self.assertEqual(code, 0, error)
        result = json.loads(output)["result"]
        self.assertTrue(result["ready"])
        self.assertEqual(result["lineage"]["service_id"], "commerce.analyze")


if __name__ == "__main__":
    unittest.main()
