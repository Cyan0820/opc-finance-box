import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from src.master_data import master_quality, parse_master_workbook, parse_profile_workbook


class MasterDataTests(unittest.TestCase):
    def _workbook(self, path: Path):
        wb = Workbook()
        profile = wb.active
        profile.title = "主体配置"
        profile.append(["字段", "填写值"])
        profile.append(["公司名称", "正现金流游戏公司"])
        game = wb.create_sheet("游戏项目")
        game.append(["项目编码", "游戏项目名称", "项目阶段", "负责人"])
        game.append(["G001", "星海", "运营", "制作人A"])
        channel = wb.create_sheet("渠道规则")
        channel.append(["渠道编码", "渠道名称", "游戏项目编码", "平台", "分成比例"])
        channel.append(["C001", "App Store", "G001", "iOS", 0.7])
        org = wb.create_sheet("组织映射")
        org.append(["人员/岗位编码", "部门", "预算单元", "成本中心", "游戏项目编码", "分摊比例"])
        org.append(["ROLE-1", "研发", "BU-RD", "CC-RD", "G001", 1])
        wb.save(path)

    def test_parse_front_loaded_master_data_and_profile(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "上线包.xlsx"
            self._workbook(path)
            rows = parse_master_workbook(path)
            profile = parse_profile_workbook(path)
        self.assertEqual(len(rows), 3)
        self.assertEqual(profile["company_name"], "正现金流游戏公司")
        self.assertEqual(master_quality(rows)["issue_count"], 0)

    def test_orphan_project_mapping_is_blocked_by_quality_check(self):
        rows = [
            {"record_type": "game", "code": "G001", "name": "星海", "status": "可用", "anomalies": []},
            {"record_type": "channel", "code": "C001", "name": "渠道", "project_code": "BAD", "status": "可用", "anomalies": []},
        ]
        quality = master_quality(rows)
        self.assertTrue(any("找不到游戏项目" in issue for issue in quality["issues"]))


if __name__ == "__main__":
    unittest.main()
