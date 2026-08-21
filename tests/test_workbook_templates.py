import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.demo_scenarios import load_demo_scenarios
from src.server import DEMO_SCENARIOS
from src.shadow_close import parse_shadow_close_workbook
from src.workbook_templates import build_demo_workbook, build_onboarding_template, build_shadow_close_template


class WorkbookTemplateTests(unittest.TestCase):
    def test_two_demo_workbooks_and_onboarding_template_are_generated_publicly(self):
        scenarios = load_demo_scenarios(DEMO_SCENARIOS)
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            for key in ("domestic", "overseas"):
                target = build_demo_workbook(scenarios[key], root / scenarios[key]["filename"])
                book = load_workbook(target, read_only=True, data_only=True)
                self.assertIn("采购台账", book.sheetnames)
                self.assertIn("银行流水", book.sheetnames)
                self.assertIn("经营KPI", book.sheetnames)
                if key == "domestic":
                    self.assertIn("对外账单-国服", book.sheetnames)
                    ios_rows = list(book["对外账单-国服"].iter_rows(min_row=2, values_only=True))
                    self.assertTrue(any(row[2] == "iOS" and row[3] == "App Store 中国区" for row in ios_rows))
                book.close()
            onboarding = build_onboarding_template(root / "onboarding.xlsx")
            book = load_workbook(onboarding, read_only=True, data_only=True)
            self.assertEqual(set(("主体配置", "游戏项目", "渠道规则", "组织映射", "供应商", "期初科目余额", "经营KPI", "Checks")) - set(book.sheetnames), set())
            book.close()

    def test_shadow_close_template_round_trips_with_one_entity_and_period(self):
        with tempfile.TemporaryDirectory() as temp:
            target = build_shadow_close_template(Path(temp) / "shadow-close.xlsx")
            book = load_workbook(target)
            for name in ("基准总账", "基准报表", "基准税务"):
                book[name].insert_rows(1, 3)
                book[name]["A1"] = f"{name} | 人工关账基准"
            book["基准总账"].append(["cn_studio", "2026-02", "1002", "银行存款", 100, 0, "人工总账", "独立复核", 0, 0])
            book.save(target)
            book.close()
            baseline = parse_shadow_close_workbook(target)
            self.assertEqual(baseline["entity_id"], "cn_studio")
            self.assertEqual(baseline["period"], "2026-02")
            self.assertEqual(baseline["row_count"], 1)
            self.assertEqual(baseline["rows"][0]["absolute_tolerance"], 0)
            self.assertEqual(baseline["rows"][0]["percent_tolerance"], 0)


if __name__ == "__main__":
    unittest.main()
