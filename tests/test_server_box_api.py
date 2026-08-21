import json
import io
import zipfile
import http.client
import hashlib
import hmac
import threading
import tempfile
import unittest
import os
from datetime import datetime, timedelta, timezone
from pathlib import Path
from http.server import ThreadingHTTPServer
from unittest.mock import patch
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from src.server import Handler, _validate_server_binding
from src.box_runtime import BoxRuntime
from src.pipeline_run_store import PipelineRunStore
from src.connector_sync import ConnectorSyncStore
from src.airwallex_webhooks import AirwallexWebhookStore
from src.pipeline_scheduler import pipeline_request_fingerprint, schedule_job_approval_fingerprint
from src.api_auth import hash_token
from src.ledger_store import LedgerStore
from src.finance_inbox import FinanceInboxStore
from src.workbook_templates import build_shadow_close_template
from src.tax_applicability_artifacts import (
    build_tax_applicability_workpaper,
    import_tax_applicability_review,
    review_tax_applicability_workpaper,
    write_tax_applicability_registry_receipt,
)
from src.pilot_readiness import (
    build_pilot_readiness_workpaper,
    review_pilot_readiness_workpaper,
)
from src.pilot_data_handoff import (
    build_pilot_data_handoff_workpaper,
    review_pilot_data_handoff_workpaper,
)
from src.trial_workspace import (
    TRIAL_WORKSPACE_ROOT_ENV,
    initialize_trial_workspace,
)
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]


def workbook_bytes(headers, row, title="明细"):
    book = Workbook()
    sheet = book.active
    sheet.title = title
    sheet.append(headers)
    sheet.append(row)
    body = io.BytesIO()
    book.save(body)
    return body.getvalue()


def multipart_bytes(fields, files, boundary="----FinanceImportBoundary"):
    chunks = []
    for name, value in fields.items():
        chunks.append(
            f'--{boundary}\r\nContent-Disposition: form-data; name="{name}"\r\n\r\n{value}\r\n'.encode("utf-8")
        )
    for name, filename, content in files:
        chunks.append(
            f'--{boundary}\r\nContent-Disposition: form-data; name="{name}"; filename="{filename}"\r\n'
            'Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n'.encode("utf-8")
        )
        chunks.append(content)
        chunks.append(b"\r\n")
    chunks.append(f"--{boundary}--\r\n".encode("utf-8"))
    return b"".join(chunks), boundary


def build_tax_review_registry(runtime, root):
    root = Path(root)
    review_dir = root / "reviews"
    review_dir.mkdir()
    answers = {
        "legal_form_and_pack_scope": "confirmed_in_scope",
        "tax_residency_and_permanent_establishment": "confirmed_in_scope",
        "direct_and_indirect_tax_registrations": "confirmed_complete",
        "fiscal_year_and_return_periods": "confirmed",
        "special_cross_border_and_group_regimes": "reviewed_no_additional_scope",
    }
    for entity_id in ("cn_studio", "sg_publisher"):
        workpaper = build_tax_applicability_workpaper(
            runtime, entity_id, prepared_by=f"{entity_id}-preparer",
            facts_as_of="2026-08-14",
        )
        for question in workpaper["entity"]["questions"]:
            question["answer"] = answers[question["question_id"]]
            question["evidence_references"] = [
                f"evidence://server-registry/{entity_id}/{question['question_id']}"
            ]
        workpaper_path = root / f"{entity_id}-workpaper.json"
        source_review = root / f"{entity_id}-review.json"
        workpaper_path.write_text(json.dumps(workpaper), encoding="utf-8")
        workpaper_path.chmod(0o600)
        review_tax_applicability_workpaper(
            runtime, workpaper_path, source_review,
            decision="approved-in-scope", actor=f"{entity_id}-reviewer",
            rationale="PRIVATE-SERVER-REGISTRY-RATIONALE",
            evidence_references=[f"advisor://server-registry/{entity_id}"],
        )
        import_tax_applicability_review(
            runtime, source_review, review_dir, as_of="2026-08-14",
        )
    receipt = root / "registry-receipt.json"
    write_tax_applicability_registry_receipt(
        runtime, review_dir, receipt,
        actor="registry-controller", as_of="2026-08-14",
    )
    return review_dir, receipt


def build_pilot_review(runtime, root):
    root = Path(root)
    workpaper = build_pilot_readiness_workpaper(
        runtime, period="2026-07", prepared_by="pilot-preparer",
    )
    workpaper["operator_principal"] = "pilot-operator"
    for entity in workpaper["entities"]:
        for domain in entity["data_domains"]:
            domain.update({
                "status": "ready", "acquisition_mode": "file_export",
                "period_coverage": ["2026-07"], "read_only_confirmed": True,
                "mapping_approved_by": "mapping-reviewer",
                "evidence_references": [
                    f"evidence://server-pilot/{entity['entity_id']}/{domain['domain']}"
                ],
            })
    workpaper["shadow_close_plan"].update({
        "planned": True, "baseline_owner": "baseline-owner",
        "evidence_references": ["workpaper://server-pilot/shadow-plan"],
    })
    source = root / "pilot-workpaper.json"
    reviewed = root / "pilot-reviewed.json"
    source.write_text(json.dumps(workpaper), encoding="utf-8")
    source.chmod(0o600)
    # Keep the lifecycle fixture bound to the API's explicit 2026-08-14
    # as_of date instead of letting the wall clock make it expire overnight.
    with patch("src.pilot_readiness.datetime", wraps=datetime) as clock:
        clock.now.return_value = datetime(2026, 8, 14, 12, tzinfo=timezone.utc)
        review_pilot_readiness_workpaper(
            runtime, source, reviewed, actor="pilot-reviewer",
            rationale="PRIVATE-SERVER-PILOT-RATIONALE confirms bounded controls.",
            evidence_references=["advisor://server-pilot/review"],
        )
    return reviewed


def build_pilot_handoff_review(runtime, root, pilot_review):
    root = Path(root)
    workpaper = build_pilot_data_handoff_workpaper(
        runtime, pilot_review,
        prepared_by="handoff-preparer",
        custodian_principal="handoff-custodian",
        as_of="2026-08-14",
    )
    for entity in workpaper["entities"]:
        for domain in entity["data_domains"]:
            domain.update({
                "status": "delivered", "transfer_mode": "local_only",
                "source_file_count": 1, "source_manifest_sha256": "c" * 64,
                "period_coverage": ["2026-07"],
                "contains_personal_data": "no", "privacy_control": "not_required",
                "source_owner": "handoff-source-owner",
                "access_approved_by": "handoff-access-approver",
                "evidence_references": [
                    f"evidence://server-handoff/{entity['entity_id']}/{domain['domain']}"
                ],
            })
    source = root / "handoff-workpaper.json"
    reviewed = root / "handoff-reviewed.json"
    source.write_text(json.dumps(workpaper), encoding="utf-8")
    source.chmod(0o600)
    with patch("src.pilot_data_handoff.datetime", wraps=datetime) as clock:
        clock.now.return_value = datetime(2026, 8, 14, 13, tzinfo=timezone.utc)
        review_pilot_data_handoff_workpaper(
            runtime, source, pilot_review, reviewed,
            actor="handoff-independent-reviewer",
            rationale="PRIVATE-SERVER-HANDOFF-RATIONALE confirms controlled intake.",
            evidence_references=["advisor://server-handoff/review"],
            as_of="2026-08-14",
        )
    return reviewed


class ServerBoxApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.auth_environment = patch.dict(os.environ, {
            "OPC_FINANCE_API_TOKEN": "", "OPC_FINANCE_API_AUTH_FILE": "",
        })
        cls.auth_environment.start()
        cls.server = ThreadingHTTPServer(("127.0.0.1", 0), Handler)
        cls.thread = threading.Thread(target=cls.server.serve_forever, daemon=True)
        cls.thread.start()
        cls.base_url = f"http://127.0.0.1:{cls.server.server_port}"

    @classmethod
    def tearDownClass(cls):
        cls.server.shutdown()
        cls.server.server_close()
        cls.thread.join(timeout=3)
        cls.auth_environment.stop()

    def _json(self, path, *, payload=None, token=None):
        data = None if payload is None else json.dumps(payload).encode("utf-8")
        headers = {"Content-Type": "application/json"} if data is not None else {}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        request = Request(
            self.base_url + path,
            data=data,
            headers=headers,
        )
        with urlopen(request, timeout=5) as response:
            return response.status, json.loads(response.read().decode("utf-8"))

    def test_statutory_bootstrap_is_entity_scoped(self):
        status, payload = self._json("/api/box?scope=statutory&entity_id=cn_studio")
        self.assertEqual(status, 200)
        self.assertEqual(payload["context"]["scope"]["entity"]["entity_id"], "cn_studio")
        self.assertTrue(payload["context"]["scope"]["books_must_remain_separate"])
        self.assertIn("agent.create_goal_draft", {
            item["service_id"] for item in payload["services"]
        })

    def test_connector_sync_api_is_read_only_secret_free_and_empty_by_default(self):
        with tempfile.TemporaryDirectory() as temp_dir, patch(
            "src.server.CONNECTOR_SYNCS", ConnectorSyncStore(Path(temp_dir) / "connector-sync"),
        ):
            status, payload = self._json("/api/box/connector-sync?limit=25")
        self.assertEqual(status, 200)
        self.assertEqual(payload["list_limit"], 25)
        self.assertEqual(payload["counts"], {
            "attempts": 0, "checkpoints": 0,
            "checkpoint_candidates": 0, "quarantine": 0,
        })
        self.assertFalse(payload["raw_requests_included"])
        self.assertFalse(payload["raw_responses_included"])
        self.assertFalse(payload["secret_values_included"])
        self.assertFalse(payload["external_actions_performed"])
        for query in ("?limit=0", "?limit=501", "?limit=bad", "?limit=1&limit=2"):
            request = Request(self.base_url + "/api/box/connector-sync" + query)
            with self.assertRaises(HTTPError) as raised:
                urlopen(request, timeout=5)
            self.assertEqual(raised.exception.code, 400)
            error = json.loads(raised.exception.read().decode("utf-8"))
            self.assertEqual(error["type"], "invalid_connector_sync_query")

    def test_airwallex_webhook_uses_hmac_not_bearer_and_is_durably_idempotent(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            policy_path = root / "auth-policy.json"
            policy_path.write_text(json.dumps({
                "schema_version": 1,
                "principals": [{
                    "principal_id": "admin", "token_sha256": hash_token("a" * 40),
                    "roles": ["admin"],
                }],
            }), encoding="utf-8")
            os.chmod(policy_path, 0o600)
            secret = "airwallex-webhook-secret"
            timestamp = str(int(datetime.now(timezone.utc).timestamp() * 1000))
            body = json.dumps({
                "id": "evt_server_001",
                "name": "spend.expense.updated",
                "account_id": "acct_cn_studio",
                "data": {
                    "id": "exp_server_001",
                    "legal_entity_id": "le_cn_studio",
                    "account_id": "acct_cn_studio",
                    "status": "APPROVED",
                },
                "created_at": datetime.now(timezone.utc).isoformat(),
                "version": "2026-07-17",
            }, separators=(",", ":")).encode()
            signature = hmac.new(
                secret.encode(), timestamp.encode() + body, hashlib.sha256,
            ).hexdigest()
            environment = {
                "OPC_FINANCE_API_TOKEN": "",
                "OPC_FINANCE_API_AUTH_FILE": str(policy_path),
                "OPC_AIRWALLEX_WEBHOOK_SECRET": secret,
                "OPC_AIRWALLEX_ENTITY_BINDINGS_JSON": json.dumps({
                    "cn_studio": {
                        "legal_entity_id": "le_cn_studio",
                        "account_id": "acct_cn_studio",
                        "environment": "sandbox",
                    },
                }),
            }
            store = AirwallexWebhookStore(root / "webhooks")
            airwallex_runtime = BoxRuntime(
                ROOT / "examples" / "boxes" / "sg_dtc_shopify_stripe_wise_airwallex_store.json",
                ROOT / "packs",
            )
            environment["OPC_AIRWALLEX_ENTITY_BINDINGS_JSON"] = json.dumps({
                "sg_store": {
                    "legal_entity_id": "le_cn_studio",
                    "account_id": "acct_cn_studio",
                    "environment": "sandbox",
                },
            })
            with patch.dict(os.environ, environment), patch(
                "src.server.AIRWALLEX_WEBHOOKS", store,
            ), patch("src.server.BOX_RUNTIME", airwallex_runtime):
                request = Request(
                    self.base_url + "/api/webhooks/airwallex/spend",
                    data=body,
                    headers={"x-timestamp": timestamp, "x-signature": signature},
                )
                with urlopen(request, timeout=5) as response:
                    first = json.loads(response.read().decode())
                self.assertEqual(response.status, 200)
                self.assertFalse(first["duplicate"])
                self.assertFalse(first["raw_expense_id_included"])
                with urlopen(request, timeout=5) as response:
                    duplicate = json.loads(response.read().decode())
                self.assertEqual(response.status, 200)
                self.assertTrue(duplicate["duplicate"])
                self.assertEqual(store.verify()["event_count"], 1)

                invalid = Request(
                    self.base_url + "/api/webhooks/airwallex/spend",
                    data=b"not-json",
                    headers={"x-timestamp": timestamp, "x-signature": "0" * 64},
                )
                with self.assertRaises(HTTPError) as raised:
                    urlopen(invalid, timeout=5)
                self.assertEqual(raised.exception.code, 401)
                error = json.loads(raised.exception.read().decode())
                self.assertEqual(error["type"], "invalid_webhook_signature")

    def test_planning_api_requires_entity_and_does_not_mix_legal_entities_or_currencies(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("plan_lines", [
                {"id": "P1", "entity_id": "cn_studio", "period": "2026-03", "scenario": "基准", "direction": "收入", "amount": 9999, "currency": "CNY", "anomalies": []},
                {"id": "P1", "entity_id": "sg_publisher", "period": "2026-03", "scenario": "基准", "direction": "收入", "amount": 500, "currency": "USD", "anomalies": []},
            ])
            ledger.save_dataset("bank_transactions", [
                {"id": "B1", "entity_id": "cn_studio", "transaction_date": "2026-02-28", "currency": "CNY", "balance": 99999},
                {"id": "B1", "entity_id": "sg_publisher", "transaction_date": "2026-02-28", "currency": "USD", "balance": 1000},
            ])
            with patch("src.server.LEDGER", ledger):
                status, result = self._json("/api/planning?entity_id=sg_publisher&period=2026-02")
                self.assertEqual(status, 200)
                self.assertEqual(result["entity_id"], "sg_publisher")
                self.assertEqual(result["functional_currency"], "USD")
                self.assertEqual(result["opening_cash"], 1000)
                self.assertEqual(result["forecast"][0]["inflows"], 500)
                with self.assertRaises(HTTPError) as raised:
                    urlopen(self.base_url + "/api/planning?period=2026-02", timeout=5)
                self.assertEqual(raised.exception.code, 400)

    def test_formal_planning_import_requires_and_persists_legal_entity(self):
        workbook = workbook_bytes(
            ["月份", "项目", "类别", "收支方向", "金额", "币种", "情景"],
            ["2026-03", "Global Game", "收入", "收入", 500, "USD", "基准"],
            title="预算",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            body, boundary = multipart_bytes({}, [("files", "plan.xlsx", workbook)])
            request = Request(
                self.base_url + "/api/planning-import", data=body,
                headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
            )
            with patch("src.server.LEDGER", ledger):
                with self.assertRaises(HTTPError) as raised:
                    urlopen(request, timeout=5)
                self.assertEqual(raised.exception.code, 400)
                self.assertIn("法律主体", raised.exception.read().decode("utf-8"))

                body, boundary = multipart_bytes(
                    {"entity_id": "sg_publisher"}, [("files", "plan.xlsx", workbook)],
                )
                request = Request(
                    self.base_url + "/api/planning-import", data=body,
                    headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
                )
                with urlopen(request, timeout=5) as response:
                    payload = json.loads(response.read().decode("utf-8"))
                self.assertEqual(payload["records"][0]["entity_id"], "sg_publisher")
                self.assertEqual(ledger.load_dataset("plan_lines")[0]["entity_id"], "sg_publisher")

    def test_invoice_import_rolls_verified_amount_to_scoped_purchase(self):
        workbook = workbook_bytes(
            ["发票号码", "开票日期", "销售方名称", "项目名称", "金额", "税额", "价税合计", "PO编号", "查验状态"],
            ["INV-001", "2026-03-20", "素材供应商", "视频制作", 60, 0, 60, "PO-CN-1", "已查验"],
            title="发票台账",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("purchases", [{
                "id": "P-CN-1", "entity_id": "cn_studio", "procurement_request_id": "REQ-CN-1",
                "po_number": "PO-CN-1", "vendor": "素材供应商", "item": "视频制作",
                "ordered_amount": 100, "accepted_amount": 100, "invoice_amount": 0,
                "paid_amount": 0, "currency": "CNY", "milestones": [{"id": "M1", "amount": 100}],
                "acceptance_history": [{"delivery_id": "DEL-CN-1"}], "anomalies": [],
            }, {
                "id": "P-SG-1", "entity_id": "sg_publisher", "po_number": "PO-SG-1",
                "vendor": "Overseas Vendor", "item": "Localization", "ordered_amount": 80,
                "accepted_amount": 80, "invoice_amount": 0, "paid_amount": 0,
                "currency": "USD", "anomalies": [],
            }])
            body, boundary = multipart_bytes(
                {"entity_id": "cn_studio"}, [("files", "invoice.xlsx", workbook)],
            )
            request = Request(
                self.base_url + "/api/invoice-import", data=body,
                headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
            )
            with patch("src.server.LEDGER", ledger), urlopen(request, timeout=5) as response:
                payload = json.loads(response.read().decode("utf-8"))
            self.assertEqual(payload["summary"]["payment_eligible_count"], 1)
            self.assertEqual(payload["records"][0]["purchase_match"]["accepted_delivery_ids"], ["DEL-CN-1"])
            purchases = {row["id"]: row for row in ledger.load_dataset("purchases")}
            self.assertEqual(purchases["P-CN-1"]["invoice_amount"], 60)
            self.assertEqual(purchases["P-CN-1"]["payment_eligible_amount"], 60)
            self.assertEqual(purchases["P-SG-1"]["invoice_amount"], 0)

    def test_settlement_import_stays_candidate_until_entity_scoped_review_releases_receivable(self):
        workbook = workbook_bytes(
            ["账期月份", "游戏名称", "平台", "渠道", "总流水", "分成基数", "分成比例", "结算金额", "甲方实收金额（结算币种）", "结算币种"],
            ["2026-07", "星海远征", "iOS", "App Store 中国区", 1000, 1000, 0.7, 700, 700, "CNY"],
            title="对外账单-国服",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("master_records", [
                {"id": "G1", "entity_id": "cn_studio", "record_type": "game", "code": "G001", "name": "星海远征", "active": True},
                {"id": "CH1", "entity_id": "cn_studio", "record_type": "channel", "code": "APP-CN", "name": "App Store 中国区", "project_code": "G001", "platform": "iOS", "currency": "CNY", "share_rate": 0.7, "settlement_formula": "share_base_x_rate", "contract_reference": "evidence://app-cn", "payment_days": 45, "active": True},
                {"id": "SG-CH", "entity_id": "sg_publisher", "record_type": "channel", "code": "APP-SG", "name": "App Store", "currency": "USD", "share_rate": 0.7, "contract_reference": "evidence://app-sg", "active": True},
            ])
            body, boundary = multipart_bytes(
                {"entity_id": "cn_studio"}, [("files", "settlement.xlsx", workbook)],
            )
            request = Request(
                self.base_url + "/api/import", data=body,
                headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
            )
            with patch("src.server.LEDGER", ledger), urlopen(request, timeout=5) as response:
                imported = json.loads(response.read().decode("utf-8"))
                self.assertEqual(imported["records"][0]["release_status"], "ready_for_review")
                self.assertEqual(ledger.load_dataset("settlements"), [])
                candidate_id = imported["records"][0]["id"]
                status, closed = self._json("/api/revenue-close?entity_id=cn_studio")
                self.assertEqual(status, 200)
                self.assertEqual(closed["summary"]["ready_for_review"], 1)
                status, reviewed = self._json("/api/revenue-close-review", payload={
                    "entity_id": "cn_studio", "candidate_ids": [candidate_id], "decision": "批准",
                    "actor": "业务负责人", "rationale": "已核对渠道后台与协议口径",
                })
                self.assertEqual(status, 200)
                self.assertEqual(reviewed["summary"]["released"], 1)
                self.assertEqual(ledger.load_dataset("settlements")[0]["entity_id"], "cn_studio")
                self.assertEqual(
                    {row["entity_id"] for row in ledger.load_dataset("master_records")},
                    {"cn_studio", "sg_publisher"},
                )

    def test_payment_authorization_chain_is_entity_target_currency_and_amount_scoped(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("purchases", [{
                "id": "PO-1", "entity_id": "cn_studio", "accepted_amount": 100,
                "invoice_amount": 100, "paid_amount": 0, "currency": "CNY", "anomalies": [],
            }])
            ledger.save_dataset("invoices", [{
                "id": "INV-1", "entity_id": "cn_studio", "total_amount": 100,
                "currency": "CNY", "verification_status": "已查验", "anomalies": [],
                "purchase_match": {"purchase_id": "PO-1"},
            }])
            ledger.save_dataset("bank_transactions", [{
                "id": "BANK-1", "entity_id": "cn_studio", "transaction_date": "2026-03-05",
                "transaction_id": "TX-1", "direction": "支出", "currency": "CNY", "amount": 100,
            }])
            with patch("src.server.LEDGER", ledger):
                status, created = self._json("/api/payment-request", payload={
                    "entity_id": "cn_studio", "target_type": "payable", "target_id": "PO-1",
                    "amount": 100, "period": "2026-03", "purpose": "支付已验收外包",
                    "evidence": ["验收单", "已查验发票"], "actor": "经办人",
                })
                self.assertEqual(status, 201)
                request_id = created["request"]["id"]
                status, approved = self._json("/api/payment-decision", payload={
                    "entity_id": "cn_studio", "request_id": request_id, "decision": "批准",
                    "actor": "负责人", "rationale": "验收发票账户均已复核",
                })
                self.assertEqual(status, 200)
                self.assertEqual(approved["request"]["status"], "已批准")

                status, allocated = self._json("/api/cash-allocation", payload={
                    "entity_id": "cn_studio", "transaction_id": "BANK-1", "target_type": "payable",
                    "target_id": "PO-1", "amount": 100, "authorization_reference": request_id,
                    "actor": "出纳", "note": "按已批准申请核销",
                })
                self.assertEqual(status, 201)
                self.assertFalse(allocated["allocation"]["authorization_gap"])

                ledger.save_dataset("bank_transactions", [{
                    "id": "BANK-2", "entity_id": "cn_studio", "transaction_date": "2026-03-06",
                    "transaction_id": "TX-2", "direction": "支出", "currency": "CNY", "amount": 1,
                }])
                with self.assertRaises(HTTPError) as raised:
                    self._json("/api/cash-allocation", payload={
                        "entity_id": "cn_studio", "transaction_id": "BANK-2", "target_type": "payable",
                        "target_id": "PO-1", "amount": 1, "authorization_reference": request_id,
                        "actor": "出纳", "note": "尝试重复使用授权",
                    })
                self.assertEqual(raised.exception.code, 400)
                self.assertIn("超过", raised.exception.read().decode("utf-8"))

    def test_collection_action_is_entity_scoped_idempotent_and_updates_planning_overlay(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("settlements", [{
                "id": "SET-1", "entity_id": "cn_studio", "period": "2026-01",
                "game": "国服游戏", "channel": "Apple", "currency": "CNY", "net_receivable": 500,
            }, {
                "id": "SET-1", "entity_id": "sg_publisher", "period": "2026-01",
                "game": "Global Game", "channel": "Apple", "currency": "USD", "net_receivable": 800,
            }])
            with patch("src.server.LEDGER", ledger):
                payload = {
                    "idempotency_key": "collection-cn-001", "entity_id": "cn_studio",
                    "settlement_id": "SET-1", "action_type": "回款承诺", "owner": "渠道运营",
                    "action_date": "2026-01-20", "promised_date": "2026-02-15",
                    "promised_amount": 300, "note": "渠道邮件确认下月付款", "actor": "财务",
                }
                status, created = self._json("/api/collection-action", payload=payload)
                self.assertEqual(status, 201)
                self.assertEqual(created["action"]["entity_id"], "cn_studio")
                status, replay = self._json("/api/collection-action", payload=payload)
                self.assertEqual(status, 200)
                self.assertTrue(replay["idempotent_replay"])
                status, planning = self._json("/api/planning?entity_id=cn_studio&period=2026-01")
                self.assertEqual(status, 200)
                self.assertEqual(planning["collection_commitment_total"], 300)
                with self.assertRaises(HTTPError) as raised:
                    self._json("/api/collection-action", payload={**payload, "entity_id": "sg_publisher", "idempotency_key": "collection-sg-001", "promised_amount": 900})
                self.assertEqual(raised.exception.code, 400)

    def test_procurement_request_uses_entity_budget_and_separate_approval(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("plan_lines", [{
                "id": "B1", "entity_id": "cn_studio", "project": "游戏A", "category": "素材制作",
                "period": "2026-03", "currency": "CNY", "direction": "支出", "amount": 100000,
                "scenario": "基准", "anomalies": [],
            }])
            payload = {
                "idempotency_key": "procurement-cn-001", "entity_id": "cn_studio",
                "project": "游戏A", "category": "素材制作", "description": "版本PV素材制作",
                "amount": 60000, "currency": "CNY", "period": "2026-03", "needed_by": "2026-03-20",
                "sourcing_method": "竞争比价", "selected_vendor": "供应商B",
                "quotes": [
                    {"vendor": "供应商A", "amount": 50000, "currency": "CNY"},
                    {"vendor": "供应商B", "amount": 60000, "currency": "CNY"},
                    {"vendor": "供应商C", "amount": 65000, "currency": "CNY"},
                ],
                "selection_rationale": "供应商B历史交付稳定且本次档期可以满足",
                "evidence": ["需求说明", "报价记录"], "actor": "制作人",
            }
            with patch("src.server.LEDGER", ledger):
                status, created = self._json("/api/procurement-request", payload=payload)
                self.assertEqual(status, 201)
                self.assertEqual(created["request"]["budget_snapshot"]["available_amount"], 100000)
                self.assertEqual(created["request"]["status"], "待批准")
                request_id = created["request"]["id"]
                status, approved = self._json("/api/procurement-request-decision", payload={
                    "entity_id": "cn_studio", "request_id": request_id, "decision": "批准",
                    "actor": "财务负责人", "rationale": "预算充足且供应商选择理由可以接受",
                })
                self.assertEqual(status, 200)
                self.assertEqual(approved["request"]["status"], "已批准")

    def test_procurement_request_to_order_delivery_and_acceptance_chain(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("plan_lines", [{
                "id": "B-CHAIN", "entity_id": "cn_studio", "project": "游戏A", "category": "素材制作",
                "period": "2026-03", "currency": "CNY", "direction": "支出", "amount": 100000,
                "scenario": "基准", "anomalies": [],
            }])
            with patch("src.server.LEDGER", ledger):
                _, created = self._json("/api/procurement-request", payload={
                    "idempotency_key": "pr-chain-001", "entity_id": "cn_studio",
                    "project": "游戏A", "category": "素材制作", "description": "版本美术外包",
                    "amount": 100000, "currency": "CNY", "period": "2026-03", "needed_by": "2026-03-20",
                    "sourcing_method": "竞争比价", "selected_vendor": "供应商A",
                    "quotes": [
                        {"vendor": "供应商A", "amount": 100000, "currency": "CNY"},
                        {"vendor": "供应商B", "amount": 108000, "currency": "CNY"},
                        {"vendor": "供应商C", "amount": 112000, "currency": "CNY"},
                    ], "evidence": ["需求说明", "三方报价"], "actor": "制作人",
                })
                request_id = created["request"]["id"]
                self._json("/api/procurement-request-decision", payload={
                    "entity_id": "cn_studio", "request_id": request_id, "decision": "批准",
                    "actor": "财务负责人", "rationale": "预算与三方报价均已核对无误",
                })
                status, ordered = self._json("/api/purchase-order", payload={
                    "idempotency_key": "po-chain-001", "entity_id": "cn_studio", "request_id": request_id,
                    "po_number": "PO-CHAIN-001", "order_date": "2026-03-01", "item": "版本美术外包",
                    "evidence": ["双方订单确认"], "actor": "采购经办",
                    "milestones": [{
                        "title": "首批素材", "amount": 100000, "due_date": "2026-03-20",
                        "acceptance_criteria": "源文件、预览图和清单均齐全", "owner": "美术负责人",
                    }],
                })
                self.assertEqual(status, 201)
                order = ordered["order"]
                self.assertEqual(order["procurement_request_id"], request_id)
                status, delivered = self._json("/api/purchase-delivery", payload={
                    "idempotency_key": "delivery-chain-001", "entity_id": "cn_studio",
                    "purchase_id": order["id"], "milestone_id": order["milestones"][0]["id"],
                    "delivered_amount": 100000, "delivery_date": "2026-03-18",
                    "evidence": ["交付清单", "文件哈希"], "actor": "供应商交付人",
                })
                self.assertEqual(status, 201)
                delivery = delivered["delivery"]
                status, accepted = self._json("/api/purchase-acceptance", payload={
                    "entity_id": "cn_studio", "purchase_id": order["id"], "delivery_id": delivery["id"],
                    "period": "2026-03", "decision": "全部验收", "actor": "美术负责人",
                    "evidence": ["验收截图"], "note": "已按订单标准核对全部素材",
                })
                self.assertEqual(status, 200)
                self.assertEqual(accepted["delivery"]["status"], "已验收")
                self.assertEqual(accepted["purchase"]["accepted_amount"], 100000)
                status, workflow = self._json("/api/procurement-workflow?entity_id=cn_studio")
                self.assertEqual(status, 200)
                self.assertEqual(workflow["summary"]["linked_orders"], 1)
                self.assertEqual(workflow["deliveries"][0]["purchase_id"], order["id"])

    def test_milestone_order_cannot_be_accepted_without_delivery_event(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("purchases", [{
                "id": "PO-NO-DELIVERY", "entity_id": "cn_studio", "po_number": "PO-NO-DELIVERY",
                "order_date": "2026-03-01", "ordered_amount": 100, "accepted_amount": 0,
                "invoice_amount": 0, "paid_amount": 0, "currency": "CNY", "milestones": [{
                    "id": "MS-1", "title": "交付", "amount": 100, "acceptance_criteria": "完整交付", "owner": "制作人",
                }], "anomalies": [],
            }])
            with patch("src.server.LEDGER", ledger):
                with self.assertRaises(HTTPError) as raised:
                    self._json("/api/purchase-acceptance", payload={
                        "entity_id": "cn_studio", "purchase_id": "PO-NO-DELIVERY", "period": "2026-03",
                        "decision": "全部验收", "actor": "制作人", "evidence": ["验收单"],
                    })
                self.assertEqual(raised.exception.code, 400)

    def test_vendor_bank_change_is_hashed_reviewed_and_bound_to_payment(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("purchases", [{
                "id": "PO-ACCOUNT", "entity_id": "cn_studio", "vendor": "供应商A",
                "accepted_amount": 100, "invoice_amount": 100, "paid_amount": 0,
                "currency": "CNY", "anomalies": [],
            }])
            ledger.save_dataset("invoices", [{
                "id": "INV-ACCOUNT", "entity_id": "cn_studio", "total_amount": 100,
                "currency": "CNY", "verification_status": "已查验", "anomalies": [],
                "purchase_match": {"purchase_id": "PO-ACCOUNT"},
            }])
            with patch("src.server.LEDGER", ledger):
                status, created = self._json("/api/vendor-bank-change", payload={
                    "idempotency_key": "vendor-bank-001", "entity_id": "cn_studio",
                    "vendor": "供应商A", "beneficiary_name": "供应商A有限公司",
                    "bank_name": "测试银行", "bank_country": "CN", "currency": "CNY",
                    "account_number": "6222 1234 5678 9012", "evidence": ["盖章账户函", "主数据联系人"],
                    "actor": "采购经办",
                })
                self.assertEqual(status, 201)
                change_id = created["change"]["id"]
                stored = ledger.load_dataset("vendor_bank_changes")[0]
                self.assertNotIn("account_number", stored)
                self.assertNotIn("6222123456789012", json.dumps(stored, ensure_ascii=False))
                status, reviewed = self._json("/api/vendor-bank-change-decision", payload={
                    "entity_id": "cn_studio", "change_id": change_id, "decision": "批准",
                    "actor": "资金复核", "rationale": "已回拨主数据联系人并核对银行证明",
                    "verification_method": "回拨主数据联系人", "verification_reference": "2026-08-14 回拨记录 REF-01",
                })
                self.assertEqual(status, 200)
                self.assertEqual(reviewed["change"]["status"], "已批准")
                status, payment = self._json("/api/payment-request", payload={
                    "entity_id": "cn_studio", "target_type": "payable", "target_id": "PO-ACCOUNT",
                    "amount": 100, "period": "2026-08", "purpose": "支付已验收外包",
                    "evidence": ["验收单", "已查验发票"], "bank_account_id": change_id,
                    "actor": "付款经办",
                })
                self.assertEqual(status, 201)
                self.assertEqual(payment["request"]["status"], "待批准")
                self.assertEqual(payment["request"]["vendor_bank_binding"]["account_masked"], "•••• 9012")

    def test_idempotency_key_does_not_replay_across_legal_entities(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("expense_claims", [{
                "id": "EXP-CN", "entity_id": "cn_studio", "idempotency_key": "same-client-key",
                "claimant": "员工A", "claim_date": "2026-03-01", "amount": 10, "currency": "CNY",
                "project": "CN Game", "category": "差旅", "purpose": "国内业务拜访",
                "evidence": ["发票"], "submitted_by": "员工A", "status": "待审批", "blockers": [],
            }])
            with patch("src.server.LEDGER", ledger):
                status, created = self._json("/api/expense-claim", payload={
                    "idempotency_key": "same-client-key", "entity_id": "sg_publisher",
                    "claimant": "Employee B", "claim_date": "2026-03-02", "amount": 20,
                    "currency": "USD", "project": "Global Game", "category": "Travel",
                    "purpose": "Overseas publisher meeting", "evidence": ["Receipt"], "actor": "Employee B",
                })
                self.assertEqual(status, 201)
                self.assertFalse(created["idempotent_replay"])
                self.assertEqual(created["claim"]["entity_id"], "sg_publisher")
                self.assertEqual(len(ledger.load_dataset("expense_claims")), 2)

    def test_non_game_box_rejects_game_demo_and_reference_agent_routes(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "cn_dtc_api_store.json", ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime):
            for path, expected_status, expected_type in (
                ("/api/sample?scenario=group", 404, "box_demo_unavailable"),
                ("/api/demo-workbook?scenario=group", 404, "box_demo_unavailable"),
                ("/api/agent-workspace?scenario=group", 409, "workbench_profile_mismatch"),
            ):
                with self.assertRaises(HTTPError) as raised:
                    urlopen(self.base_url + path, timeout=5)
                self.assertEqual(raised.exception.code, expected_status)
                body = json.loads(raised.exception.read().decode("utf-8"))
                self.assertEqual(body["type"], expected_type)

            request = Request(
                self.base_url + "/api/agent-run",
                data=json.dumps({"demo_scenario": "group"}).encode("utf-8"),
                headers={"Content-Type": "application/json"},
            )
            with self.assertRaises(HTTPError) as raised:
                urlopen(request, timeout=5)
            self.assertEqual(raised.exception.code, 409)
            body = json.loads(raised.exception.read().decode("utf-8"))
            self.assertEqual(body["type"], "workbench_profile_mismatch")

    def test_pipeline_catalog_and_preflight_api_are_box_scoped_and_read_only(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "cn_dtc_api_store.json", ROOT / "packs",
        )
        request_payload = json.loads(
            (ROOT / "examples" / "pipelines" / "commerce_channel_close_fixture.json").read_text(
                encoding="utf-8"
            )
        )
        with patch("src.server.BOX_RUNTIME", runtime):
            status, catalog = self._json("/api/box/pipelines")
            self.assertEqual(status, 200)
            self.assertEqual(
                {item["pipeline_id"] for item in catalog["pipelines"]},
                {
                    "finance.bank_statement_close",
                    "finance.trial_balance_review",
                    "finance.accounting_close_review",
                    "finance.first_close_discovery",
                    "finance.month_close_control",
                    "commerce.import_analyze", "commerce.channel_close",
                },
            )
            status, preflight = self._json(
                "/api/box/pipelines/preflight", payload=request_payload,
            )
            self.assertEqual(status, 200)
            self.assertTrue(preflight["ready_to_dispatch"])
            self.assertFalse(preflight["dispatch_performed"])
            self.assertFalse(preflight["source_access_performed"])

    def test_connector_readiness_is_pipeline_scoped_and_secret_free(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json",
            ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime), patch.dict(os.environ, {
            "OPC_SHOPIFY_ADMIN_TOKEN": "browser-must-never-see-this",
            "OPC_STRIPE_RESTRICTED_KEY": "browser-must-never-see-this-either",
        }):
            status, readiness = self._json("/api/box/connectors/readiness")
            self.assertEqual(status, 200)
            self.assertEqual(readiness["summary"]["network_connector_count"], 4)
            self.assertTrue({
                "shopify.orders", "shopify.monthly_order_evidence",
            } <= {
                item["connector_id"] for item in readiness["pipeline_connectors"]
            })
            self.assertTrue(all(
                item["configured"] for item in readiness["summary"]["required_env"]
            ))
            self.assertEqual(readiness["summary"]["network_provider_group_count"], 2)
            self.assertEqual(readiness["summary"]["blocked_provider_group_count"], 0)
            self.assertEqual(
                {item["pack_id"] for item in readiness["provider_groups"] if item["network_access"]},
                {"connector.shopify", "connector.stripe"},
            )
            self.assertTrue(all(
                item["diagnostic_status"]
                == "ready_to_initialize_private_access_probe_request"
                for item in readiness["provider_groups"] if item["network_access"]
            ))
            serialized = json.dumps(readiness)
            self.assertNotIn("browser-must-never-see-this", serialized)
            self.assertFalse(readiness["control_boundary"]["network_access_performed"])
            self.assertFalse(readiness["control_boundary"]["provider_access_probe_performed"])

    def test_connector_shadow_registry_api_is_server_mounted_and_safe(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "sg_dtc_shopify_stripe_wise_store.json",
            ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime), patch.dict(
            os.environ, {"OPC_CONNECTOR_SHADOW_REVIEW_DIR": ""}, clear=False,
        ):
            status, result = self._json("/api/box/connector-shadow?as_of=2026-08-14")
        self.assertEqual(status, 200)
        self.assertEqual(result["summary"]["activation_status"], "missing")
        self.assertEqual(result["summary"]["required_network_pack_count"], 3)
        self.assertFalse(result["summary"]["ready_for_connector_shadow_evidence"])
        self.assertFalse(result["control_boundary"]["paths_returned"])
        self.assertFalse(result["control_boundary"]["actors_returned"])
        self.assertFalse(result["control_boundary"]["external_actions_performed"])

    def test_production_readiness_api_aggregates_existing_gates_and_is_secret_free(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio_xero.json",
            ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime), patch.dict(os.environ, {
            "OPC_XERO_ACCESS_TOKEN": "api-must-never-return-this",
            "OPC_XERO_ENTITY_BINDINGS_JSON": "api-must-never-return-this-either",
            "OPC_PILOT_READINESS_REVIEW": "",
            "OPC_PILOT_DATA_HANDOFF_REVIEW": "",
            "OPC_PILOT_SHADOW_RUN_REGISTRATION": "",
            "OPC_PILOT_SHADOW_OBSERVATION_REVIEW": "",
            "OPC_PILOT_SHADOW_SERIES_REVIEW": "",
            "OPC_STABLE_PROMOTION_ROOT": "",
        }, clear=False):
            status, result = self._json(
                "/api/box/production-readiness?as_of=2026-08-14"
            )
        self.assertEqual(status, 200)
        self.assertEqual(result["summary"]["stage_count"], 11)
        connector = next(
            item for item in result["stages"]
            if item["stage_id"] == "connector_configuration"
        )
        self.assertTrue(connector["gate_passed"])
        self.assertFalse(connector["evidence_complete"])
        self.assertFalse(result["summary"]["ready_for_bounded_shadow"])
        self.assertFalse(result["summary"]["ready_for_external_filing"])
        serialized = json.dumps(result)
        self.assertNotIn("api-must-never-return-this", serialized)
        self.assertFalse(result["control_boundary"]["paths_returned"])
        self.assertFalse(result["control_boundary"]["external_actions_performed"])

    def test_activation_api_returns_only_dependency_aware_command_templates(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio_xero.json",
            ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime), patch.dict(os.environ, {
            "OPC_XERO_ACCESS_TOKEN": "activation-api-secret",
            "OPC_XERO_ENTITY_BINDINGS_JSON": "activation-api-secret-binding",
            "OPC_TAX_APPLICABILITY_REVIEW_DIR": "",
            "OPC_TAX_APPLICABILITY_REGISTRY_RECEIPT": "",
            "OPC_STABLE_PROMOTION_ROOT": "",
        }, clear=False):
            status, result = self._json("/api/box/activation?as_of=2026-08-14")
        self.assertEqual(status, 200)
        self.assertEqual(
            result["summary"]["current_wave_stage_ids"],
            ["tax_applicability", "connector_shadow_evidence"],
        )
        self.assertTrue(result["control_boundary"]["commands_are_templates_only"])
        self.assertFalse(result["control_boundary"]["commands_executed"])
        serialized = json.dumps(result)
        self.assertNotIn("activation-api-secret", serialized)
        self.assertNotIn("activation-api-secret-binding", serialized)

    def test_activation_api_projects_only_the_server_mounted_private_workspace(self):
        projected = {
            "schema_version": 1,
            "artifact_type": "initialized_first_customer_activation_status",
            "workspace": {"valid": True, "paths_returned": False},
            "activation": {
                "schema_version": 1,
                "artifact_type": "first_customer_activation_workspace",
                "summary": {"current_wave_stage_count": 0},
                "current_wave": [],
                "control_boundary": {
                    "private_paths_returned": False,
                    "external_actions_performed": False,
                },
            },
            "connector_access": {
                "summary": {"ready_for_bounded_shadow_dispatch": True},
            },
            "connector_access_alerts": {
                "alert_count": 1,
                "critical_count": 0,
                "warning_count": 1,
                "notifications_sent": False,
            },
            "control_boundary": {"private_paths_returned": False},
        }
        private_root = "/private/server-mounted-activation"
        with patch(
            "src.server.build_initialized_activation_status",
            return_value=projected,
        ) as builder, patch.dict(os.environ, {
            "OPC_ACTIVATION_WORKSPACE_ROOT": private_root,
        }, clear=False):
            status, result = self._json(
                "/api/box/activation?as_of=2026-08-16&activation_root=/attacker/path"
            )
        self.assertEqual(status, 200)
        self.assertEqual(result["artifact_type"], "first_customer_activation_workspace")
        self.assertEqual(result["initialized_workspace"], projected["workspace"])
        self.assertEqual(result["connector_access"], projected["connector_access"])
        self.assertEqual(
            result["connector_access_alerts"],
            projected["connector_access_alerts"],
        )
        self.assertTrue(
            result["control_boundary"]["server_mounted_activation_workspace_used"]
        )
        self.assertFalse(
            result["control_boundary"]["activation_root_accepted_from_request"]
        )
        self.assertEqual(builder.call_args.args[2], Path(private_root))
        self.assertEqual(builder.call_args.kwargs["as_of"], "2026-08-16")

    def test_trial_onboarding_api_is_available_only_for_bound_verified_trial(self):
        with patch.dict(os.environ, {TRIAL_WORKSPACE_ROOT_ENV: ""}, clear=False):
            status, unavailable = self._json("/api/box/trial-onboarding")
        self.assertEqual(status, 200)
        self.assertFalse(unavailable["available"])
        self.assertEqual(unavailable["reason"], "trial_workspace_not_configured")
        self.assertFalse(
            unavailable["control_boundary"]["trial_workspace_path_returned"]
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            trial = Path(temp_dir).resolve() / "trial"
            initialize_trial_workspace(
                profile="dtc",
                country="AE",
                packs_root=ROOT / "packs",
                destination_root=trial,
                actor="server-trial-founder",
                integrations=["shopify_stripe"],
            )
            runtime = BoxRuntime(trial / "box" / "box.json", ROOT / "packs")
            with patch("src.server.BOX_RUNTIME", runtime), patch.dict(
                os.environ, {TRIAL_WORKSPACE_ROOT_ENV: str(trial)}, clear=False,
            ):
                status, onboarding = self._json("/api/box/trial-onboarding")
            self.assertEqual(status, 200)
            self.assertTrue(onboarding["available"])
            self.assertEqual(onboarding["starter"]["starter_id"], "dtc.ae")
            self.assertEqual(onboarding["summary"]["journey_stage_count"], 5)
            self.assertTrue(onboarding["summary"]["demo_ready"])
            self.assertFalse(onboarding["summary"]["production_ready"])
            self.assertFalse(onboarding["control_boundary"]["paths_returned"])
            serialized = json.dumps(onboarding)
            self.assertNotIn(str(trial), serialized)
            self.assertNotIn("server-trial-founder", serialized)

    def test_pilot_readiness_workspace_uses_only_server_mounted_review_and_is_safe(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio.json", ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime), patch.dict(
            os.environ, {"OPC_PILOT_READINESS_REVIEW": ""}, clear=False,
        ):
            status, missing = self._json("/api/box/pilot-readiness")
        self.assertEqual(status, 200)
        self.assertEqual(missing["summary"]["activation_status"], "missing")
        self.assertFalse(missing["summary"]["ready_for_bounded_shadow"])
        self.assertEqual(missing["summary"]["entity_count"], 2)

        with tempfile.TemporaryDirectory() as temp_dir:
            reviewed = build_pilot_review(runtime, temp_dir)
            with patch("src.server.BOX_RUNTIME", runtime), patch.dict(os.environ, {
                "OPC_PILOT_READINESS_REVIEW": str(reviewed),
            }):
                status, workspace = self._json(
                    "/api/box/pilot-readiness?as_of=2026-08-14"
                )
        self.assertEqual(status, 200)
        self.assertEqual(workspace["summary"]["activation_status"], "current")
        self.assertTrue(workspace["summary"]["ready_for_bounded_shadow"])
        self.assertEqual(workspace["summary"]["total_data_domain_count"], 18)
        self.assertFalse(workspace["control_boundary"]["ready_for_statutory_release"])
        serialized = json.dumps(workspace)
        self.assertNotIn("PRIVATE-SERVER-PILOT", serialized)
        self.assertNotIn("pilot-reviewer", serialized)
        self.assertNotIn(str(temp_dir), serialized)

    def test_pilot_data_handoff_workspace_requires_paired_server_mounts_and_is_safe(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio.json", ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime), patch.dict(os.environ, {
            "OPC_PILOT_READINESS_REVIEW": "",
            "OPC_PILOT_DATA_HANDOFF_REVIEW": "",
        }, clear=False):
            status, missing = self._json("/api/box/pilot-data-handoff")
        self.assertEqual(status, 200)
        self.assertEqual(missing["summary"]["activation_status"], "missing")
        self.assertFalse(missing["summary"]["ready_for_controlled_data_intake"])

        with tempfile.TemporaryDirectory() as temp_dir:
            pilot_review = build_pilot_review(runtime, temp_dir)
            handoff_review = build_pilot_handoff_review(
                runtime, temp_dir, pilot_review,
            )
            with patch("src.server.BOX_RUNTIME", runtime), patch.dict(os.environ, {
                "OPC_PILOT_READINESS_REVIEW": str(pilot_review),
                "OPC_PILOT_DATA_HANDOFF_REVIEW": str(handoff_review),
            }):
                status, workspace = self._json(
                    "/api/box/pilot-data-handoff?as_of=2026-08-14"
                )
        self.assertEqual(status, 200)
        self.assertEqual(workspace["summary"]["activation_status"], "current")
        self.assertTrue(workspace["summary"]["ready_for_controlled_data_intake"])
        self.assertEqual(workspace["summary"]["total_data_domain_count"], 18)
        serialized = json.dumps(workspace)
        self.assertNotIn("PRIVATE-SERVER-HANDOFF", serialized)
        self.assertNotIn("handoff-independent-reviewer", serialized)
        self.assertNotIn("c" * 64, serialized)
        self.assertNotIn(str(temp_dir), serialized)

    def test_pilot_shadow_run_workspace_is_safe_and_missing_by_default(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio.json", ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime), patch.dict(os.environ, {
            "OPC_PILOT_READINESS_REVIEW": "",
            "OPC_PILOT_DATA_HANDOFF_REVIEW": "",
            "OPC_PILOT_SHADOW_RUN_REGISTRATION": "",
        }, clear=False):
            status, missing = self._json("/api/box/pilot-shadow-run")
        self.assertEqual(status, 200)
        self.assertEqual(missing["summary"]["activation_status"], "missing")
        self.assertEqual(missing["summary"]["entity_count"], 2)
        self.assertEqual(missing["summary"]["registered_entity_count"], 0)
        self.assertFalse(
            missing["summary"]["ready_for_first_shadow_observation"]
        )
        serialized = json.dumps(missing)
        self.assertNotIn("entity_runs", serialized)
        self.assertNotIn("registration_id", serialized)

    def test_pilot_shadow_observation_workspace_is_safe_and_missing_by_default(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio.json", ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime), patch.dict(os.environ, {
            "OPC_PILOT_READINESS_REVIEW": "",
            "OPC_PILOT_DATA_HANDOFF_REVIEW": "",
            "OPC_PILOT_SHADOW_RUN_REGISTRATION": "",
            "OPC_PILOT_SHADOW_OBSERVATION_REVIEW": "",
            "OPC_PILOT_SHADOW_ENTITY_REPORT_DIR": "",
            "OPC_PILOT_SHADOW_PORTFOLIO_REVIEW": "",
        }, clear=False):
            status, missing = self._json("/api/box/pilot-shadow-observation")
        self.assertEqual(status, 200)
        self.assertEqual(missing["summary"]["activation_status"], "missing")
        self.assertEqual(missing["summary"]["entity_count"], 2)
        self.assertEqual(missing["summary"]["reviewed_entity_count"], 0)
        self.assertFalse(missing["summary"]["ready_for_next_shadow_period"])
        self.assertTrue(missing["summary"]["portfolio_review_required"])
        serialized = json.dumps(missing)
        for forbidden in (
            '"entity_observations"', '"registration_actor"',
            '"source_attempt_id":', '"report_content_sha256"',
            '"manifest_content_sha256"',
        ):
            self.assertNotIn(forbidden, serialized)

    def test_pilot_shadow_series_workspace_is_safe_and_missing_by_default(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio.json", ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime), patch.dict(os.environ, {
            "OPC_PILOT_SHADOW_SERIES_REVIEW": "",
            "OPC_PILOT_SHADOW_SERIES_EVIDENCE_ROOT": "",
        }, clear=False):
            status, missing = self._json("/api/box/pilot-shadow-series")
        self.assertEqual(status, 200)
        self.assertEqual(missing["summary"]["activation_status"], "missing")
        self.assertEqual(missing["summary"]["entity_count"], 2)
        self.assertEqual(missing["summary"]["period_count"], 0)
        self.assertFalse(
            missing["summary"][
                "eligible_to_prepare_stable_promotion_evidence"
            ]
        )
        self.assertFalse(missing["summary"]["ready_for_stable_promotion"])
        serialized = json.dumps(missing)
        for forbidden in (
            "period_observations", "source_bundle_fingerprint",
            "separation_principals", "receipt_fingerprint",
        ):
            self.assertNotIn(forbidden, serialized)

    def test_pilot_shadow_period_index_uses_only_server_mounted_root(self):
        projected = {
            "schema_version": 1,
            "artifact_type": "pilot_shadow_period_workspace_index",
            "summary": {"activation_status": "current", "period_count": 1},
            "periods": [{"period": "2026-09", "event_count": 0}],
            "control_boundary": {"private_paths_returned": False},
        }
        private_root = "/private/server-mounted-activation"
        with patch(
            "src.server.build_pilot_shadow_period_workspace_index",
            return_value=projected,
        ) as builder, patch.dict(os.environ, {
            "OPC_ACTIVATION_WORKSPACE_ROOT": private_root,
        }, clear=False):
            status, result = self._json("/api/box/pilot-shadow-periods")
        self.assertEqual(status, 200)
        self.assertEqual(result, projected)
        self.assertEqual(builder.call_args.args[1], Path(private_root))

        with patch(
            "src.server.build_pilot_shadow_period_workspace_index",
            return_value={**projected, "summary": {
                "activation_status": "missing", "period_count": 0,
            }},
        ) as builder, patch.dict(os.environ, {
            "OPC_ACTIVATION_WORKSPACE_ROOT": "",
        }, clear=False):
            status, result = self._json(
                "/api/box/pilot-shadow-periods?activation_root=/attacker/path"
            )
        self.assertEqual(status, 200)
        self.assertEqual(result["summary"]["activation_status"], "missing")
        self.assertIsNone(builder.call_args.args[1])

    def test_tax_workspace_is_multi_entity_source_backed_and_read_only(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio.json", ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime):
            status, workspace = self._json(
                "/api/box/tax/workspace?period_year=2026&as_of=2026-08-13"
            )
            self.assertEqual(status, 200)
            self.assertEqual(workspace["summary"]["entity_count"], 2)
            self.assertEqual(workspace["summary"]["calendar_task_count"], 5)
            self.assertEqual(
                {item["entity"]["jurisdiction"] for item in workspace["entities"]},
                {"CN", "SG"},
            )
            self.assertTrue(all(item["official_sources"] for item in workspace["entities"]))
            self.assertFalse(workspace["control_boundary"]["tax_calculation_performed"])
            self.assertFalse(workspace["control_boundary"]["filing_performed"])
            self.assertFalse(workspace["control_boundary"]["external_submission_enabled"])
            self.assertFalse(
                workspace["control_boundary"][
                    "applicability_review_path_accepted_from_request"
                ]
            )

    def test_tax_workspace_uses_only_server_configured_review_directory(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio.json", ROOT / "packs",
        )
        with tempfile.TemporaryDirectory() as temp_dir, patch(
            "src.server.BOX_RUNTIME", runtime,
        ), patch.dict(os.environ, {
            "OPC_TAX_APPLICABILITY_REVIEW_DIR": temp_dir,
        }):
            status, workspace = self._json(
                "/api/box/tax/workspace?period_year=2026&as_of=2026-08-14"
            )
        self.assertEqual(status, 200)
        self.assertTrue(all(
            item["applicability_review_requirement"]["review_directory_configured"]
            and item["applicability_review_requirement"]["review"]["status"]
            == "not_attached"
            for item in workspace["entities"]
        ))
        self.assertFalse(
            workspace["applicability_review_registry"][
                "activation_receipt_configured"
            ]
        )
        self.assertFalse(
            workspace["applicability_review_registry"][
                "ready_for_calendar_release"
            ]
        )

    def test_tax_workspace_requires_server_configured_matching_registry_receipt(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio.json", ROOT / "packs",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            review_dir, receipt = build_tax_review_registry(runtime, temp_dir)
            with patch("src.server.BOX_RUNTIME", runtime), patch.dict(os.environ, {
                "OPC_TAX_APPLICABILITY_REVIEW_DIR": str(review_dir),
                "OPC_TAX_APPLICABILITY_REGISTRY_RECEIPT": str(receipt),
            }):
                status, workspace = self._json(
                    "/api/box/tax/workspace?period_year=2026&as_of=2026-08-14"
                )
        self.assertEqual(status, 200)
        registry = workspace["applicability_review_registry"]
        self.assertTrue(registry["activation_receipt_configured"])
        self.assertTrue(registry["activation_receipt_valid"])
        self.assertTrue(registry["ready_for_calendar_release"])
        self.assertTrue(registry["activation"]["registry_unchanged"])
        self.assertFalse(registry["activation"]["digital_signature_verified"])
        self.assertFalse(registry["activation"]["filing_authorization_granted"])
        self.assertNotIn("PRIVATE-SERVER", json.dumps(workspace))
        self.assertNotIn("registry-controller", json.dumps(workspace))

    def test_tax_workspace_accepts_bounded_date_only_anchor_preview(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "ie_dtc_shopify_stripe_ltd.json",
            ROOT / "packs",
        )
        query = urlencode({
            "period_year": "2026",
            "as_of": "2026-08-13",
            "anchors": json.dumps({
                "ie_store": {"cro_annual_return_date": "2026-09-30"},
            }),
        })
        with patch("src.server.BOX_RUNTIME", runtime):
            status, workspace = self._json(f"/api/box/tax/workspace?{query}")
        self.assertEqual(status, 200)
        task = next(
            task for task in workspace["entities"][0]["calendar"]["tasks"]
            if task["rule_id"] == "ie.cro.annual_return.calendar"
        )
        self.assertEqual(task["candidate_due_date"], "2026-11-25")
        self.assertFalse(workspace["anchor_preview"]["persistent_write_performed"])

    def test_tax_workspace_rejects_unknown_anchor_query(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "ie_dtc_shopify_stripe_ltd.json",
            ROOT / "packs",
        )
        query = urlencode({"anchors": json.dumps({
            "ie_store": {"company_number": "sensitive-value"},
        })})
        with patch("src.server.BOX_RUNTIME", runtime):
            with self.assertRaises(HTTPError) as raised:
                urlopen(self.base_url + f"/api/box/tax/workspace?{query}", timeout=5)
        self.assertEqual(raised.exception.code, 400)
        body = json.loads(raised.exception.read().decode("utf-8"))
        self.assertEqual(body["type"], "tax_workspace_error")

    def test_inbox_evidence_link_updates_purchase_without_accepting_it(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            inbox = FinanceInboxStore(Path(temp_dir) / "inbox")
            purchase = {
                "id": "PO-EVIDENCE-1", "po_number": "PO-EVIDENCE-1",
                "entity_id": "cn_studio", "vendor": "美术供应商", "item": "角色立绘",
                "ordered_amount": 100000, "accepted_amount": 0,
                "acceptance_status": "已交付待验收", "acceptance_evidence": [],
            }
            ledger.upsert_dataset("purchases", [purchase], "测试人", "测试")
            document = inbox.ingest(
                "角色立绘验收证明.pdf", b"%PDF-1.4 test evidence", "上传人", "cn_studio",
                [{"id": "cn_studio", "name": "星火游戏（上海）有限公司"}],
            )
            document["classification"]["document_type"] = "acceptance_evidence"
            document["recognition"] = {"evidence_only": True, "record_count": 0}
            document["status"] = "已提取待归档"
            inbox.save(document)
            with patch("src.server.LEDGER", ledger), patch("src.server.FINANCE_INBOX", inbox):
                status, payload = self._json("/api/inbox-link", payload={
                    "document_id": document["id"], "target_type": "purchase",
                    "target_id": purchase["id"], "entity_id": "cn_studio",
                    "actor": "项目负责人", "note": "证明文件已交付，仍由业务负责人决定验收",
                })
            self.assertEqual(status, 200)
            self.assertEqual(payload["purchase"]["accepted_amount"], 0)
            self.assertEqual(payload["purchase"]["acceptance_status"], "已交付待验收")
            self.assertIn(f"document:{document['id']}", payload["purchase"]["acceptance_evidence"])
            self.assertEqual(payload["document"]["status"], "已归档并关联")

    def test_first_close_readiness_is_entity_period_scoped_and_declaration_is_reviewed(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            inbox = FinanceInboxStore(Path(temp_dir) / "inbox")
            ledger.upsert_dataset("settlements", [{
                "id": "SET-1", "entity_id": "cn_studio", "period": "2026-01",
                "game": "ROR", "channel": "App Store", "currency": "CNY",
            }], "测试", "测试")
            ledger.upsert_dataset("bank_transactions", [{
                "id": "BANK-1", "entity_id": "cn_studio", "transaction_date": "2026-01-31",
            }], "测试", "测试")
            ledger.upsert_dataset("opening_balances", [{
                "id": "OPEN-1", "entity_id": "cn_studio", "period": "2026-01",
            }], "测试", "测试")
            ledger.upsert_dataset("master_records", [
                {"id": "GAME-1", "entity_id": "cn_studio", "record_type": "game", "status": "可用"},
                {"id": "CHANNEL-1", "entity_id": "cn_studio", "record_type": "channel", "status": "可用"},
            ], "测试", "测试")
            with patch("src.server.LEDGER", ledger), patch("src.server.FINANCE_INBOX", inbox), patch(
                "src.server.profile_gaps", return_value=[],
            ):
                status, initial = self._json("/api/onboarding?entity_id=cn_studio&period=2026-01")
                self.assertEqual(status, 200)
                self.assertIn("采购与验收", initial["first_close"]["blockers"])
                status, declared = self._json("/api/onboarding-declaration", payload={
                    "entity_id": "cn_studio", "period": "2026-01", "domain": "purchases",
                    "decision": "本期不适用", "actor": "复核人",
                    "rationale": "经业务负责人确认本月没有发生任何采购业务",
                    "evidence": ["本期采购清单"],
                })
                self.assertEqual(status, 200)
                item = next(row for row in declared["first_close"]["items"] if row["domain"] == "purchases")
                self.assertEqual(item["status"], "不适用")
                self.assertEqual(declared["declaration"]["entity_id"], "cn_studio")

    def test_live_agent_first_close_gate_aggregates_both_legal_entities(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            inbox = FinanceInboxStore(Path(temp_dir) / "inbox")
            with patch("src.server.LEDGER", ledger), patch("src.server.FINANCE_INBOX", inbox), patch(
                "src.server.profile_gaps", return_value=[],
            ):
                from src.server import _combined_first_close_readiness
                result = _combined_first_close_readiness("2026-01")
            self.assertEqual({row["entity_id"] for row in result["entities"]}, {"cn_studio", "sg_publisher"})
            self.assertTrue(any(item.startswith("cn_studio ·") for item in result["blockers"]))
            self.assertTrue(any(item.startswith("sg_publisher ·") for item in result["blockers"]))

    def test_reconciliation_evidence_links_to_settlement_without_changing_amount(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            inbox = FinanceInboxStore(Path(temp_dir) / "inbox")
            settlement = {
                "id": "SET-IOS-1", "entity_id": "cn_studio", "period": "2026-01",
                "game": "ROR", "channel": "App Store", "settlement_amount": 998.0,
            }
            ledger.upsert_dataset("settlements", [settlement], "测试", "测试")
            document = inbox.ingest(
                "IOS账单&经分数据核对.xlsx", workbook_bytes(
                    ["年月", "游戏", "日期", "平台", "渠道", "付费金额", "匹配-苹果账单", "差异"],
                    ["2026-01", "ROR", "2026-01-01", "ios", "国内", 1000, 998, -2],
                ), "上传人", "cn_studio",
                [{"id": "cn_studio", "name": "星火游戏（上海）有限公司"}],
            )
            document["classification"]["document_type"] = "settlement_reconciliation_evidence"
            document["recognition"] = {"evidence_only": True, "record_count": 1}
            document["status"] = "已提取待归档"
            inbox.save(document)
            with patch("src.server.LEDGER", ledger), patch("src.server.FINANCE_INBOX", inbox):
                status, payload = self._json("/api/inbox-link", payload={
                    "document_id": document["id"], "target_type": "settlement",
                    "target_id": settlement["id"], "entity_id": "cn_studio",
                    "actor": "财务负责人", "note": "核对iOS平台账单和经分日流水",
                })
            self.assertEqual(status, 200)
            self.assertEqual(payload["settlement"]["settlement_amount"], 998.0)
            self.assertIn(f"document:{document['id']}", payload["settlement"]["reconciliation_evidence"])

    def test_tax_workspace_rejects_invalid_period_without_dispatching_actions(self):
        with self.assertRaises(HTTPError) as raised:
            urlopen(
                self.base_url + "/api/box/tax/workspace?period_year=99",
                timeout=5,
            )
        self.assertEqual(raised.exception.code, 400)
        body = json.loads(raised.exception.read().decode("utf-8"))
        self.assertEqual(body["type"], "tax_workspace_error")

    def test_box_builder_options_and_preview_do_not_replace_active_runtime(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio.json", ROOT / "packs",
        )
        active_fingerprint = runtime.snapshot()["fingerprint"]
        with patch("src.server.BOX_RUNTIME", runtime):
            status, options = self._json("/api/box-builder/options")
            self.assertEqual(status, 200)
            self.assertEqual(
                {item["id"] for item in options["profiles"]},
                {"game", "dtc", "marketplace"},
            )
            self.assertEqual(options["schema_version"], 3)
            self.assertEqual(
                options["connector_binding_policy"]
                ["single_credential_connector_packs"],
                ["connector.shopify", "connector.stripe"],
            )
            self.assertTrue(options["starter_catalog"]["complete"])
            self.assertEqual(options["starter_catalog"]["ready_combination_count"], 45)
            self.assertEqual(options["starter_catalog"]["unavailable_combinations"], [])
            status, preview = self._json("/api/box-builder/preview", payload={
                "name": "新加坡平台电商候选",
                "business_type": "commerce",
                "channels": ["marketplace"],
                "data_mode": "demo",
                "entities": [{
                    "id": "sg_marketplace",
                    "name": "新加坡平台经营主体候选",
                    "tax_country": "SG",
                    "tax_pack": "jurisdiction.sg",
                    "functional_currency": "SGD",
                    "accounting_basis": "SFRS",
                    "fiscal_year_end": "12-31",
                    "tax_registrations": [],
                }],
            })
            self.assertEqual(status, 200)
            self.assertEqual(
                preview["candidate"]["product"]["workbench"]["profile"],
                "commerce_marketplace",
            )
            self.assertFalse(preview["control_boundary"]["active_runtime_changed"])
            self.assertEqual(runtime.snapshot()["fingerprint"], active_fingerprint)

            entity_bound_spec = json.loads(
                (
                    ROOT / "examples" / "box_specs"
                    / "dtc_cn_nl_us_entity_connectors.json"
                ).read_text(encoding="utf-8")
            )
            status, entity_bound_preview = self._json(
                "/api/box-builder/preview", payload=entity_bound_spec,
            )
            self.assertEqual(status, 200)
            self.assertEqual(
                entity_bound_preview["config"]["connector_bindings"],
                entity_bound_spec["connector_bindings"],
            )
            shopify_close = next(
                item for item in entity_bound_preview["candidate"]["pipelines"]
                if item["pipeline_id"] == "dtc.shopify_stripe_daily_close"
            )
            self.assertEqual(shopify_close["eligible_entity_ids"], ["nl_sales"])
            self.assertFalse(
                entity_bound_preview["control_boundary"]["connector_dispatch_performed"],
            )
            self.assertEqual(runtime.snapshot()["fingerprint"], active_fingerprint)

            request = Request(
                self.base_url + "/api/box-builder/bundle",
                data=json.dumps(preview["spec"]).encode("utf-8"),
                headers={"Content-Type": "application/json"},
            )
            with urlopen(request, timeout=10) as response:
                self.assertEqual(response.status, 200)
                self.assertEqual(response.headers.get_content_type(), "application/zip")
                self.assertRegex(
                    response.headers["Content-Disposition"],
                    r'attachment; filename="opc-finance-box-[0-9a-f]{12}\.zip"',
                )
                declared_sha256 = response.headers["X-OPC-Handoff-SHA256"]
                declared_runtime = response.headers["X-OPC-Runtime-Fingerprint"]
                declared_manifest_schema = response.headers["X-OPC-Manifest-Schema"]
                declared_manifest_file_count = response.headers[
                    "X-OPC-Manifest-File-Count"
                ]
                bundle = response.read()
            self.assertEqual(declared_sha256, hashlib.sha256(bundle).hexdigest())
            self.assertRegex(declared_runtime, r"^[0-9a-f]{64}$")
            with zipfile.ZipFile(io.BytesIO(bundle)) as archive:
                self.assertIn("bundle-manifest.json", archive.namelist())
                self.assertIn("compiled/box.lock.json", archive.namelist())
                self.assertIn("ACTIVATION.md", archive.namelist())
                manifest = json.loads(archive.read("bundle-manifest.json"))
                self.assertTrue(manifest["activation_guide_included"])
                self.assertEqual(declared_runtime, manifest["runtime_fingerprint"])
                self.assertEqual(int(declared_manifest_schema), manifest["schema_version"])
                self.assertEqual(
                    int(declared_manifest_file_count), manifest["file_count"],
                )
            self.assertEqual(runtime.snapshot()["fingerprint"], active_fingerprint)

    def test_optional_bearer_auth_protects_finance_api_but_not_health(self):
        token = "test-token-with-at-least-thirty-two-characters"
        with patch.dict(os.environ, {"OPC_FINANCE_API_TOKEN": token}):
            request = Request(self.base_url + "/api/box")
            with self.assertRaises(HTTPError) as raised:
                urlopen(request, timeout=5)
            self.assertEqual(raised.exception.code, 401)
            self.assertEqual(raised.exception.headers["WWW-Authenticate"], "Bearer")
            self.assertEqual(raised.exception.headers["Cache-Control"], "no-store, max-age=0")
            body = json.loads(raised.exception.read().decode("utf-8"))
            self.assertEqual(body["type"], "authentication_required")

            request = Request(
                self.base_url + "/api/box",
                headers={"Authorization": f"Bearer {token}"},
            )
            with urlopen(request, timeout=5) as response:
                self.assertEqual(response.status, 200)
                self.assertIn("context", json.loads(response.read().decode("utf-8")))

            with urlopen(self.base_url + "/api/health", timeout=5) as response:
                self.assertEqual(response.status, 200)

    def test_server_binding_rejects_unsafe_exposure_or_short_token(self):
        _validate_server_binding("127.0.0.1", None)
        _validate_server_binding("::1", None)
        with self.assertRaisesRegex(RuntimeError, "non-loopback"):
            _validate_server_binding("0.0.0.0", None)
        with self.assertRaisesRegex(RuntimeError, "at least 32"):
            _validate_server_binding("127.0.0.1", "short")
        _validate_server_binding("0.0.0.0", "x" * 32)

    def test_role_policy_enforces_operator_reviewer_separation_and_binds_actor(self):
        reader_token = "reader-token-abcdefghijklmnopqrstuvwxyz-123456"
        operator_token = "operator-token-abcdefghijklmnopqrstuvwxyz-123456"
        reviewer_token = "reviewer-token-abcdefghijklmnopqrstuvwxyz-123456"
        pipeline_request = json.loads(
            (ROOT / "examples" / "pipelines" / "shopify_stripe_daily_close_fixture.json").read_text(
                encoding="utf-8"
            )
        )
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json", ROOT / "packs",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            policy_path = root / "api-auth.json"
            policy_path.write_text(json.dumps({
                "schema_version": 1,
                "principals": [
                    {"principal_id": "finance_reader", "token_sha256": hash_token(reader_token), "roles": ["reader"]},
                    {"principal_id": "pipeline_bot", "token_sha256": hash_token(operator_token), "roles": ["operator"]},
                    {"principal_id": "independent_reviewer", "token_sha256": hash_token(reviewer_token), "roles": ["reviewer"]},
                ],
            }), encoding="utf-8")
            policy_path.chmod(0o600)
            store = PipelineRunStore(root / "runs")
            environment = {
                "OPC_FINANCE_API_TOKEN": "",
                "OPC_FINANCE_API_AUTH_FILE": str(policy_path),
            }
            with patch.dict(os.environ, environment), patch("src.server.BOX_RUNTIME", runtime), patch(
                "src.server.PIPELINE_RUNS", store,
            ):
                status, whoami = self._json("/api/auth/whoami", token=reader_token)
                self.assertEqual(status, 200)
                self.assertEqual(whoami["principal"]["principal_id"], "finance_reader")
                status, preflight = self._json(
                    "/api/box/pipelines/preflight", payload=pipeline_request, token=reader_token,
                )
                self.assertEqual(status, 200)
                self.assertTrue(preflight["ready_to_dispatch"], preflight)
                self.assertFalse(preflight["dispatch_performed"])
                status, builder_options = self._json(
                    "/api/box-builder/options", token=reader_token,
                )
                self.assertEqual(status, 200)
                us_tax_pack = next(
                    item for item in builder_options["jurisdictions"]
                    if item["country_code"] == "US"
                )
                status, candidate = self._json(
                    "/api/box-builder/preview",
                    payload={
                        "name": "reader-created-candidate",
                        "business_type": "commerce",
                        "channels": ["dtc"],
                        "entities": [{
                            "id": "us_store",
                            "name": "US candidate",
                            "tax_country": "US",
                            "tax_pack": us_tax_pack["id"],
                            "functional_currency": "USD",
                            "accounting_basis": "US_GAAP",
                            "tax_registrations": [],
                        }],
                    },
                    token=reader_token,
                )
                self.assertEqual(status, 200)
                self.assertFalse(candidate["control_boundary"]["active_runtime_changed"])
                status, monthly_periods = self._json(
                    "/api/box/pilot-shadow-periods", token=reader_token,
                )
                self.assertEqual(status, 200)
                self.assertEqual(
                    monthly_periods["summary"]["activation_status"], "missing",
                )
                request = Request(
                    self.base_url + "/api/box/pipelines/dispatch",
                    data=json.dumps(pipeline_request).encode("utf-8"),
                    headers={
                        "Content-Type": "application/json",
                        "Authorization": f"Bearer {reader_token}",
                    },
                )
                with self.assertRaises(HTTPError) as raised:
                    urlopen(request, timeout=5)
                self.assertEqual(raised.exception.code, 403)

                status, recorded = self._json(
                    "/api/box/pipeline-runs",
                    payload={"actor": "spoofed-actor", "request": pipeline_request},
                    token=operator_token,
                )
                self.assertEqual(status, 201)
                attempt_id = recorded["run_record"]["attempt_id"]
                self.assertEqual(recorded["run_record"]["actor"], "pipeline_bot")

                request = Request(
                    self.base_url + "/api/box/pipeline-run-reviews",
                    data=json.dumps({
                        "attempt_id": attempt_id,
                        "gate": "shopify_mapping_approval",
                        "decision": "approved",
                        "actor": "spoofed-reviewer",
                        "rationale": "证据已复核",
                    }).encode("utf-8"),
                    headers={
                        "Content-Type": "application/json",
                        "Authorization": f"Bearer {operator_token}",
                    },
                )
                with self.assertRaises(HTTPError) as raised:
                    urlopen(request, timeout=5)
                self.assertEqual(raised.exception.code, 403)

                status, reviewed = self._json(
                    "/api/box/pipeline-run-reviews",
                    payload={
                        "attempt_id": attempt_id,
                        "gate": "shopify_mapping_approval",
                        "decision": "approved",
                        "actor": "spoofed-reviewer",
                        "rationale": "证据已复核",
                    },
                    token=reviewer_token,
                )
                self.assertEqual(status, 201)
                review = reviewed["run"]["current_reviews"]["shopify_mapping_approval"]
                self.assertEqual(review["actor"], "independent_reviewer")

    def test_money_workflow_roles_and_idempotent_expense_submission(self):
        operator_token = "flow-operator-token-abcdefghijklmnopqrstuvwxyz-123456"
        reviewer_token = "flow-reviewer-token-abcdefghijklmnopqrstuvwxyz-123456"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            policy_path = root / "api-auth.json"
            policy_path.write_text(json.dumps({
                "schema_version": 1,
                "principals": [
                    {"principal_id": "finance_operator", "token_sha256": hash_token(operator_token), "roles": ["operator"]},
                    {"principal_id": "finance_reviewer", "token_sha256": hash_token(reviewer_token), "roles": ["reviewer"]},
                ],
            }), encoding="utf-8")
            policy_path.chmod(0o600)
            ledger = LedgerStore(root / "ledger")
            environment = {"OPC_FINANCE_API_TOKEN": "", "OPC_FINANCE_API_AUTH_FILE": str(policy_path)}
            payload = {
                "idempotency_key": "expense-cn-20260813-0001",
                "entity_id": "cn_studio", "claimant": "员工A", "claim_date": "2026-08-13",
                "amount": 800, "currency": "CNY", "project": "星火计划", "category": "差旅",
                "purpose": "参加渠道业务会议", "evidence": ["电子发票", "会议纪要"],
                "actor": "spoofed-actor",
            }
            with patch.dict(os.environ, environment), patch("src.server.LEDGER", ledger):
                status, first = self._json("/api/expense-claim", payload=payload, token=operator_token)
                self.assertEqual(status, 201)
                self.assertFalse(first["idempotent_replay"])
                self.assertEqual(first["claim"]["submitted_by"], "finance_operator")
                status, replay = self._json("/api/expense-claim", payload=payload, token=operator_token)
                self.assertEqual(status, 200)
                self.assertTrue(replay["idempotent_replay"])
                self.assertEqual(replay["claim"]["id"], first["claim"]["id"])
                self.assertEqual(len(ledger.load_dataset("expense_claims")), 1)

                request = Request(
                    self.base_url + "/api/expense-decision",
                    data=json.dumps({"claim_id": first["claim"]["id"], "decision": "批准", "rationale": "业务证据完整", "approved_amount": 800}).encode("utf-8"),
                    headers={"Content-Type": "application/json", "Authorization": f"Bearer {operator_token}"},
                )
                with self.assertRaises(HTTPError) as raised:
                    urlopen(request, timeout=5)
                self.assertEqual(raised.exception.code, 403)

                status, reviewed = self._json(
                    "/api/expense-decision",
                    payload={"claim_id": first["claim"]["id"], "decision": "批准", "rationale": "业务证据完整", "approved_amount": 800},
                    token=reviewer_token,
                )
                self.assertEqual(status, 200)
                self.assertEqual(reviewed["claim"]["status"], "已批准待付款")
                self.assertEqual(reviewed["claim"]["approval_history"][-1]["actor"], "finance_reviewer")

    def test_authenticated_money_action_requires_idempotency_key(self):
        operator_token = "required-key-operator-token-abcdefghijklmnopqrstuvwxyz-123456"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            policy_path = root / "api-auth.json"
            policy_path.write_text(json.dumps({
                "schema_version": 1,
                "principals": [
                    {"principal_id": "finance_operator", "token_sha256": hash_token(operator_token), "roles": ["operator"]},
                ],
            }), encoding="utf-8")
            policy_path.chmod(0o600)
            environment = {"OPC_FINANCE_API_TOKEN": "", "OPC_FINANCE_API_AUTH_FILE": str(policy_path)}
            with patch.dict(os.environ, environment), patch("src.server.LEDGER", LedgerStore(root / "ledger")):
                request = Request(
                    self.base_url + "/api/expense-claim",
                    data=json.dumps({"entity_id": "cn_studio"}).encode("utf-8"),
                    headers={"Content-Type": "application/json", "Authorization": f"Bearer {operator_token}"},
                )
                with self.assertRaises(HTTPError) as raised:
                    urlopen(request, timeout=5)
                self.assertEqual(raised.exception.code, 400)
                self.assertIn("idempotency_key", raised.exception.read().decode("utf-8"))

    def test_service_dispatch_preserves_draft_state_boundary(self):
        status, payload = self._json("/api/box/services/dispatch", payload={
            "service_id": "agent.create_goal_draft",
            "entity_ids": ["cn_studio"],
            "payload": {"objective": "完成七月月结", "period": "2026-07", "actor": "负责人"},
        })
        self.assertEqual(status, 200)
        self.assertEqual(payload["service"]["entity_ids"], ["cn_studio"])
        self.assertEqual(payload["output"]["output_status"], "draft_not_persisted")
        self.assertFalse(payload["output"]["state_changed"])

    def test_service_dispatch_evaluates_cfo_metric_and_types_stale_scope_error(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "global_game_studio.json",
            ROOT / "packs",
        )
        request = {
            "service_id": "core.evaluate_cfo_metrics",
            "entity_id": "cn_studio",
            "payload": {
                "runtime_fingerprint": runtime.snapshot()["fingerprint"],
                "period": "2026-07",
                "currency": "CNY",
                "metric_type_ids": ["game_platform_net_revenue"],
                "operand_values": {
                    "platform_gross_settlement": "1000",
                    "refunds_chargebacks_and_platform_fees": "175.5",
                },
                "confirmed_control_type_ids": [
                    "platform_statement_scope_confirmed",
                    "settlement_currency_confirmed",
                ],
            },
        }
        status, payload = self._json("/api/box/services/dispatch", payload=request)
        self.assertEqual(status, 200)
        self.assertEqual(payload["output"]["metric_results"][0]["value"], "824.5")
        self.assertFalse(payload["output"]["external_actions_performed"])

        request["payload"]["runtime_fingerprint"] = "a" * 64
        invalid = Request(
            self.base_url + "/api/box/services/dispatch",
            data=json.dumps(request).encode("utf-8"),
            headers={"Content-Type": "application/json"},
        )
        with self.assertRaises(HTTPError) as raised:
            urlopen(invalid, timeout=5)
        self.assertEqual(raised.exception.code, 400)
        body = json.loads(raised.exception.read().decode("utf-8"))
        self.assertEqual(body["type"], "invalid_service_request")

    def test_unselected_pack_service_is_forbidden(self):
        request = Request(
            self.base_url + "/api/box/services/dispatch",
            data=json.dumps({"service_id": "commerce.analyze", "payload": {}}).encode("utf-8"),
            headers={"Content-Type": "application/json"},
        )
        with self.assertRaises(HTTPError) as raised:
            urlopen(request, timeout=5)
        self.assertEqual(raised.exception.code, 403)
        body = json.loads(raised.exception.read().decode("utf-8"))
        self.assertEqual(body["type"], "service_forbidden")

    def test_json_body_limit_is_enforced_before_reading_payload(self):
        connection = http.client.HTTPConnection(
            "127.0.0.1", self.server.server_port, timeout=5
        )
        connection.putrequest("POST", "/api/box/services/dispatch")
        connection.putheader("Content-Type", "application/json")
        connection.putheader("Content-Length", str(29 * 1024 * 1024))
        connection.endheaders()
        response = connection.getresponse()
        body = json.loads(response.read().decode("utf-8"))
        connection.close()
        self.assertEqual(response.status, 413)
        self.assertEqual(body["type"], "payload_too_large")

    def test_json_body_limit_also_covers_legacy_finance_endpoint(self):
        connection = http.client.HTTPConnection("127.0.0.1", self.server.server_port, timeout=5)
        connection.putrequest("POST", "/api/finance-ops")
        connection.putheader("Content-Type", "application/json")
        connection.putheader("Content-Length", str(29 * 1024 * 1024))
        connection.endheaders()
        response = connection.getresponse()
        body = json.loads(response.read().decode("utf-8"))
        connection.close()
        self.assertEqual(response.status, 413)
        self.assertEqual(body["type"], "payload_too_large")

    def test_box_builder_has_a_tighter_limit_before_reading_payload(self):
        connection = http.client.HTTPConnection("127.0.0.1", self.server.server_port, timeout=5)
        connection.putrequest("POST", "/api/box-builder/preview")
        connection.putheader("Content-Type", "application/json")
        connection.putheader("Content-Length", str(257 * 1024))
        connection.endheaders()
        response = connection.getresponse()
        body = json.loads(response.read().decode("utf-8"))
        connection.close()
        self.assertEqual(response.status, 413)
        self.assertEqual(body["type"], "payload_too_large")

    def test_invalid_content_length_is_rejected_on_legacy_finance_endpoint(self):
        connection = http.client.HTTPConnection("127.0.0.1", self.server.server_port, timeout=5)
        connection.putrequest("POST", "/api/finance-ops")
        connection.putheader("Content-Type", "application/json")
        connection.putheader("Content-Length", "not-a-number")
        connection.endheaders()
        response = connection.getresponse()
        body = json.loads(response.read().decode("utf-8"))
        connection.close()
        self.assertEqual(response.status, 400)
        self.assertEqual(body["type"], "invalid_content_length")

    def test_cn_tax_filing_assist_rejects_overseas_entity(self):
        request = Request(self.base_url + "/api/tax-filing-assist?period=2026-02&entity_id=sg_publisher")
        with self.assertRaises(HTTPError) as raised:
            urlopen(request, timeout=5)
        self.assertEqual(raised.exception.code, 400)
        body = json.loads(raised.exception.read().decode("utf-8"))
        self.assertIn("新加坡主体税务工作区", body["error"])

    def test_finance_ops_post_filters_records_and_tax_pack_by_entity(self):
        status, payload = self._json("/api/finance-ops", payload={
            "entity_id": "sg_publisher", "period": "2026-02",
            "records": [
                {"entity_id": "cn_studio", "period": "2026-02", "scope": "国内", "game": "CN",
                 "channel": "App Store 中国区", "currency": "CNY", "settlement_amount": 999},
                {"entity_id": "sg_publisher", "period": "2026-02", "scope": "海外", "game": "Global",
                 "channel": "App Store", "currency": "USD", "settlement_amount": 100},
            ],
        })
        self.assertEqual(status, 200)
        self.assertEqual(payload["entity_id"], "sg_publisher")
        self.assertEqual(payload["tax_pack"]["jurisdiction"], "SG")
        self.assertEqual(payload["tax_pack"]["returns_workspace"]["returns"], [])
        self.assertEqual({item["original_currency"] for item in payload["vouchers"]}, {"USD"})

    def test_accountant_pack_requires_and_preserves_single_legal_entity(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("settlements", [
                {"id": "CN-1", "entity_id": "cn_studio", "period": "2026-02", "scope": "国内",
                 "game": "CN", "channel": "iOS", "currency": "CNY", "settlement_amount": 100},
                {"id": "SG-1", "entity_id": "sg_publisher", "period": "2026-02", "scope": "海外",
                 "game": "Global", "channel": "iOS", "currency": "USD", "settlement_amount": 20},
            ])
            with patch("src.server.LEDGER", ledger):
                with self.assertRaises(HTTPError) as raised:
                    urlopen(self.base_url + "/api/accountant-pack?period=2026-02", timeout=5)
                self.assertEqual(raised.exception.code, 400)
                with urlopen(
                    self.base_url + "/api/accountant-pack?entity_id=sg_publisher&period=2026-02",
                    timeout=5,
                ) as response:
                    body = response.read()
            with zipfile.ZipFile(io.BytesIO(body)) as archive:
                manifest = json.loads(archive.read("00_交付清单.json"))
                listing = archive.read("08_数据集清单.csv").decode("utf-8-sig")
                self.assertEqual(manifest["entity_id"], "sg_publisher")
                self.assertEqual(manifest["jurisdiction"], "SG")
                self.assertTrue(manifest["books_must_remain_separate"])
                self.assertIn("settlements,1", listing)

    def test_game_revenue_policy_requires_legal_entity(self):
        payload = {
            "game": "Global", "channel": "App Store", "revenue_stream": "IAP",
            "presentation": "净额法", "recognition_method": "即时确认", "effective_from": "2026-02",
            "actor": "财务用户", "evidence": ["渠道协议"],
            "role_facts": {"controls_pricing": False, "responsible_for_fulfillment": False,
                           "bears_refund_risk": False, "controls_virtual_goods": False},
        }
        request = Request(
            self.base_url + "/api/game-revenue-policy", data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"},
        )
        with self.assertRaises(HTTPError) as raised:
            urlopen(request, timeout=5)
        self.assertEqual(raised.exception.code, 400)

    def test_overseas_entity_cannot_post_before_local_ledger_mapping_review(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            with patch("src.server.LEDGER", ledger):
                request = Request(
                    self.base_url + "/api/post-vouchers",
                    data=json.dumps({
                        "entity_id": "sg_publisher", "period": "2026-02", "actor": "复核人",
                    }).encode("utf-8"),
                    headers={"Content-Type": "application/json"},
                )
                with self.assertRaises(HTTPError) as raised:
                    urlopen(request, timeout=5)
        self.assertEqual(raised.exception.code, 400)
        body = json.loads(raised.exception.read().decode("utf-8"))
        self.assertIn("当地会计批准", body["error"])

    def test_local_accountant_can_approve_versioned_overseas_ledger_mapping(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            with patch("src.server.LEDGER", ledger):
                status, payload = self._json("/api/ledger-adapter-review", payload={
                    "entity_id": "sg_publisher", "decision": "批准",
                    "actor": "SG accounting firm",
                    "rationale": "已核对适用准则、本位币和全部科目角色映射",
                    "evidence": ["signed chart mapping"],
                })
                self.assertEqual(status, 201)
                self.assertTrue(payload["ledger_adapter"]["posting_ready"])
                status, finance = self._json("/api/finance-ops", payload={
                    "entity_id": "sg_publisher", "period": "2026-02", "records": [{
                        "entity_id": "sg_publisher", "period": "2026-02", "game": "Global",
                        "channel": "App Store", "currency": "USD", "settlement_amount": 100,
                    }],
                })
                self.assertEqual(status, 200)
                self.assertTrue(finance["ledger_adapter"]["posting_ready"])
                self.assertEqual(finance["vouchers"][0]["debit"][0]["account_code"], "1200")

    def test_bank_reconciliation_review_is_entity_scoped_persisted_and_returned_by_finance_api(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("bank_transactions", [{
                "id": "CN-BANK-1", "entity_id": "cn_studio", "transaction_date": "2026-02-28",
                "account_masked": "6222****0001", "currency": "CNY", "amount": 100,
                "direction": "收入", "status": "高置信匹配", "balance": 1100,
            }, {
                "id": "SG-BANK-1", "entity_id": "sg_publisher", "transaction_date": "2026-02-28",
                "account_masked": "DBS****0001", "currency": "USD", "amount": 20,
                "direction": "收入", "status": "高置信匹配", "balance": 220,
            }])
            with patch("src.server.LEDGER", ledger):
                status, response = self._json("/api/bank-reconciliation-review", payload={
                    "entity_id": "cn_studio", "period": "2026-02",
                    "account_masked": "6222****0001", "currency": "CNY",
                    "ledger_ending_balance": 1100, "decision": "确认",
                    "actor": "财务负责人", "rationale": "已核对银行对账单和总账余额",
                    "evidence": ["银行对账单", "总账银行明细"],
                })
                self.assertEqual(status, 201)
                self.assertEqual(response["review"]["entity_id"], "cn_studio")
                self.assertEqual(len(ledger.load_dataset("bank_reconciliation_reviews")), 1)
                status, finance = self._json("/api/finance-ops", payload={
                    "entity_id": "cn_studio", "period": "2026-02",
                    "bank_transactions": [ledger.load_dataset("bank_transactions")[0]],
                })
                self.assertEqual(status, 200)
                self.assertTrue(finance["bank_reconciliation"]["complete"])
                self.assertEqual(finance["bank_reconciliation"]["accounts"][0]["account_masked"], "6222****0001")

                with self.assertRaises(HTTPError) as raised:
                    self._json("/api/bank-reconciliation-review", payload={
                        "entity_id": "cn_studio", "period": "2026-02",
                        "account_masked": "DBS****0001", "currency": "USD",
                        "ledger_ending_balance": 220, "decision": "确认",
                        "actor": "财务负责人", "rationale": "尝试跨主体选择海外银行账户",
                        "evidence": ["银行对账单"],
                    })
                self.assertEqual(raised.exception.code, 400)

    def test_sg_asset_and_accrual_endpoints_store_functional_currency_and_entity(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            with patch("src.server.LEDGER", ledger):
                status, asset = self._json("/api/asset-card", payload={
                    "entity_id": "sg_publisher", "name": "Build server", "asset_type": "固定资产",
                    "acquisition_date": "2026-02-10", "original_cost": 7200, "currency": "CNY",
                    "fx_rate": 1 / 7.2, "useful_months": 36, "residual_value": 0,
                    "project": "Global Game", "vendor": "Vendor", "evidence": ["invoice", "acceptance"],
                    "actor": "finance",
                })
                self.assertEqual(status, 201)
                self.assertEqual(asset["asset_card"]["entity_id"], "sg_publisher")
                self.assertEqual(asset["asset_card"]["functional_currency"], "USD")
                self.assertEqual(asset["asset_card"]["functional_cost"], 1000)
                status, accrual = self._json("/api/accrual", payload={
                    "entity_id": "sg_publisher", "period": "2026-02", "description": "Art outsourcing",
                    "amount": 500, "currency": "USD", "expense_role": "cost_of_sales",
                    "counterparty": "Vendor", "project": "Global Game", "evidence": ["acceptance"],
                    "actor": "finance",
                })
                self.assertEqual(status, 201)
                self.assertEqual(accrual["accrual"]["entity_id"], "sg_publisher")
                self.assertEqual(accrual["accrual"]["functional_currency"], "USD")
                self.assertEqual(accrual["accrual"]["functional_amount"], 500)

    def test_accounting_review_rejects_cross_entity_item_selection(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = LedgerStore(Path(temp_dir) / "ledger")
            ledger.save_dataset("accruals", [{
                "id": "A1", "entity_id": "sg_publisher", "period": "2026-02",
                "status": "待会计复核", "blockers": [], "amount": 100,
            }])
            with patch("src.server.LEDGER", ledger):
                request = Request(
                    self.base_url + "/api/accrual-review",
                    data=json.dumps({
                        "entity_id": "cn_studio", "item_id": "A1", "decision": "批准",
                        "actor": "reviewer", "rationale": "证据和主体均已核对",
                    }).encode("utf-8"), headers={"Content-Type": "application/json"},
                )
                with self.assertRaises(HTTPError) as raised:
                    urlopen(request, timeout=5)
                self.assertEqual(raised.exception.code, 400)
                self.assertIn("不属于所选法律主体", raised.exception.read().decode("utf-8"))

    def test_shadow_close_import_and_read_are_entity_scoped_and_read_only(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = build_shadow_close_template(root / "shadow-close.xlsx")
            workbook = load_workbook(workbook_path)
            workbook["基准总账"].append([
                "cn_studio", "2026-02", "1002", "银行存款", 100, 0,
                "人工总账", "独立复核", 0, 0,
            ])
            workbook.save(workbook_path)
            workbook.close()
            boundary = "----ShadowCloseBoundary"
            body = (
                f"--{boundary}\r\nContent-Disposition: form-data; name=\"actor\"\r\n\r\n伪造用户\r\n"
                f"--{boundary}\r\nContent-Disposition: form-data; name=\"files\"; filename=\"shadow-close.xlsx\"\r\n"
                "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
            ).encode("utf-8") + workbook_path.read_bytes() + f"\r\n--{boundary}--\r\n".encode("utf-8")
            ledger = LedgerStore(root / "ledger")
            request = Request(
                self.base_url + "/api/shadow-close-import", data=body,
                headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
            )
            with patch("src.server.LEDGER", ledger):
                with urlopen(request, timeout=5) as response:
                    payload = json.loads(response.read().decode("utf-8"))
                self.assertEqual(payload["baseline"]["entity_id"], "cn_studio")
                self.assertEqual(payload["baseline"]["row_count"], 1)
                status, readback = self._json("/api/shadow-close?entity_id=cn_studio&period=2026-02")
                self.assertEqual(status, 200)
                self.assertEqual(readback["baseline"]["entity_id"], "cn_studio")
                self.assertIn("Shadow close 是只读验证", readback["report"]["guardrail"])
                self.assertRegex(
                    readback["report"]["runtime_fingerprint"], r"^[0-9a-f]{64}$",
                )
                self.assertEqual(ledger.load_dataset("opening_balances"), [])

    def test_stripe_pipeline_api_runs_with_explicit_box_runtime(self):
        payload = json.loads(
            (ROOT / "examples" / "pipelines" / "stripe_daily_close_fixture.json").read_text(
                encoding="utf-8"
            )
        )
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "cn_dtc_stripe_store.json", ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime):
            status, result = self._json("/api/box/pipelines/dispatch", payload=payload)
        self.assertEqual(status, 200)
        self.assertTrue(result["ready"])
        self.assertEqual(result["pipeline"]["pipeline_id"], "stripe.daily_close")
        self.assertFalse(result["external_actions_performed"])

    def test_shopify_stripe_pipeline_api_preserves_candidate_boundary(self):
        payload = json.loads(
            (ROOT / "examples" / "pipelines" / "shopify_stripe_daily_close_fixture.json").read_text(
                encoding="utf-8"
            )
        )
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json", ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime):
            status, result = self._json("/api/box/pipelines/dispatch", payload=payload)
        self.assertEqual(status, 200)
        self.assertTrue(result["ready"])
        self.assertEqual(result["pipeline"]["pipeline_id"], "dtc.shopify_stripe_daily_close")
        self.assertTrue(result["founder_briefing"]["candidate_only"])
        self.assertFalse(result["external_actions_performed"])
        self.assertFalse(result["network_access_performed"])

    def test_marketplace_pipeline_api_preserves_contract_and_inventory_gates(self):
        payload = json.loads(
            (ROOT / "examples" / "pipelines" / "marketplace_channel_close_fixture.json").read_text(
                encoding="utf-8"
            )
        )
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "cn_marketplace_store.json", ROOT / "packs",
        )
        with patch("src.server.BOX_RUNTIME", runtime):
            status, result = self._json("/api/box/pipelines/dispatch", payload=payload)
        self.assertEqual(status, 200)
        self.assertTrue(result["ready"])
        self.assertEqual(result["pipeline"]["pipeline_id"], "marketplace.channel_close")
        self.assertIn("marketplace_contract_mapping", result["pipeline"]["required_review_gates"])
        self.assertIn("marketplace_inventory_mapping", result["pipeline"]["required_review_gates"])
        self.assertTrue(result["founder_briefing"]["inventory_adjustment_prohibited"])
        self.assertFalse(result["external_actions_performed"])

    def test_pipeline_run_api_records_lists_and_reads_scoped_attempt(self):
        payload = json.loads(
            (ROOT / "examples" / "pipelines" / "shopify_stripe_daily_close_fixture.json").read_text(
                encoding="utf-8"
            )
        )
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json", ROOT / "packs",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            store = PipelineRunStore(Path(temp_dir) / "runs")
            with patch("src.server.BOX_RUNTIME", runtime), patch("src.server.PIPELINE_RUNS", store):
                status, response = self._json("/api/box/pipeline-runs", payload={
                    "actor": "HTTP 测试负责人", "request": payload,
                })
                self.assertEqual(status, 201)
                self.assertTrue(response["pipeline_result"]["ready"])
                attempt_id = response["run_record"]["attempt_id"]
                status, listed = self._json(
                    "/api/box/pipeline-runs?pipeline_id=dtc.shopify_stripe_daily_close&entity_id=cn_dtc_company"
                )
                self.assertEqual(status, 200)
                self.assertEqual([item["attempt_id"] for item in listed["runs"]], [attempt_id])
                status, detail = self._json(f"/api/box/pipeline-runs/{attempt_id}")
                self.assertEqual(status, 200)
                self.assertEqual(detail["run"]["attempt_id"], attempt_id)
                self.assertFalse(detail["run"]["secret_values_persisted"])
                self.assertEqual(detail["run"]["review_status"], "pending_review")
                status, queue = self._json("/api/box/pipeline-review-queue")
                self.assertEqual(status, 200)
                self.assertEqual(len(queue["review_tasks"]), 3)
                for gate in detail["run"]["required_review_gates"]:
                    status, reviewed = self._json("/api/box/pipeline-run-reviews", payload={
                        "attempt_id": attempt_id,
                        "gate": gate,
                        "decision": "approved",
                        "actor": "HTTP 复核人",
                        "rationale": "证据已完成复核",
                        "evidence_references": [f"evidence://{gate}"],
                    })
                    self.assertEqual(status, 201)
                self.assertTrue(reviewed["run"]["review_complete"])
                self.assertTrue(reviewed["run"]["release_candidate"])
                status, detail = self._json(f"/api/box/pipeline-runs/{attempt_id}")
                self.assertEqual(status, 200)
                self.assertEqual(len(detail["run"]["review_history"]), 3)
                status, queue = self._json("/api/box/pipeline-review-queue")
                self.assertEqual(status, 200)
                self.assertEqual(queue["review_tasks"], [])
                status, integrity = self._json("/api/box/pipeline-run-integrity")
                self.assertEqual(status, 200)
                self.assertTrue(integrity["integrity"]["valid"])
                self.assertEqual(integrity["integrity"]["event_count"], 4)
                self.assertEqual(integrity["integrity"]["attempt_count_for_box"], 1)
                self.assertEqual(integrity["integrity"]["review_event_count_for_box"], 3)

                request = Request(
                    self.base_url + "/api/box/pipeline-run-reviews",
                    data=json.dumps({
                        "attempt_id": attempt_id, "gate": "not_a_gate",
                        "decision": "approved", "actor": "HTTP 复核人", "rationale": "无",
                    }).encode("utf-8"),
                    headers={"Content-Type": "application/json"},
                )
                with self.assertRaises(HTTPError) as raised:
                    urlopen(request, timeout=5)
                self.assertEqual(raised.exception.code, 400)

    def test_pipeline_run_api_rejects_invalid_wrapper_without_writing(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json", ROOT / "packs",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            store = PipelineRunStore(Path(temp_dir) / "runs")
            with patch("src.server.BOX_RUNTIME", runtime), patch("src.server.PIPELINE_RUNS", store):
                request = Request(
                    self.base_url + "/api/box/pipeline-runs",
                    data=json.dumps({"actor": "test"}).encode("utf-8"),
                    headers={"Content-Type": "application/json"},
                )
                with self.assertRaises(HTTPError) as raised:
                    urlopen(request, timeout=5)
                self.assertEqual(raised.exception.code, 400)
                self.assertFalse(store.events_file.exists())

    def test_pipeline_schedule_api_is_read_only_until_operator_runs_due_job(self):
        runtime = BoxRuntime(
            ROOT / "examples" / "boxes" / "cn_dtc_shopify_stripe_store.json", ROOT / "packs",
        )
        with patch("src.server.PIPELINE_SCHEDULE_FILE", None):
            status, response = self._json("/api/box/pipeline-schedule")
            self.assertEqual(status, 200)
            self.assertFalse(response["configured"])
            self.assertFalse(response["dispatch_performed"])
            status, observability = self._json("/api/box/pipeline-observability")
            self.assertEqual(status, 200)
            self.assertFalse(observability["schedule_configured"])
            self.assertFalse(observability["external_actions_performed"])
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            scheduled = (datetime.now(timezone.utc) - timedelta(minutes=1)).replace(second=0, microsecond=0)
            request_path = root / "request.json"
            request_path.write_bytes(
                (ROOT / "examples" / "pipelines" / "shopify_stripe_daily_close_fixture.json").read_bytes()
            )
            schedule_path = root / "schedule.json"
            schedule_job = {
                    "job_id": "http-dtc-close", "enabled": True,
                    "pipeline_id": "dtc.shopify_stripe_daily_close",
                    "entity_id": "cn_dtc_company", "request_file": "request.json",
                    "request_fingerprint": pipeline_request_fingerprint(
                        json.loads(request_path.read_text(encoding="utf-8"))
                    ),
                    "cadence": {"kind": "daily", "local_time": scheduled.strftime("%H:%M")},
                    "execution_window_minutes": 60, "max_attempts": 2,
                    "retry_delay_minutes": 15, "lease_seconds": 900,
                    "operator": "http_scheduler", "alert_owner": "finance_owner",
                    "approved_by": "schedule_reviewer", "approved_at": scheduled.isoformat(),
                    "approval_fingerprint": None,
            }
            schedule_job["approval_fingerprint"] = schedule_job_approval_fingerprint(schedule_job)
            schedule_path.write_text(json.dumps({
                "schema_version": 2, "timezone": "UTC", "jobs": [schedule_job],
            }), encoding="utf-8")
            store = PipelineRunStore(root / "runs")
            with (
                patch("src.server.BOX_RUNTIME", runtime),
                patch("src.server.PIPELINE_RUNS", store),
                patch("src.server.PIPELINE_SCHEDULE_FILE", str(schedule_path)),
            ):
                status, response = self._json("/api/box/pipeline-schedule")
                self.assertEqual(status, 200)
                self.assertTrue(response["configured"])
                self.assertFalse(response["schedule_path_returned"])
                self.assertEqual(response["schedule"]["jobs"][0]["status"], "due")
                self.assertFalse(store.events_file.exists())
                status, response = self._json(
                    "/api/box/pipeline-schedule/run",
                    payload={"actor": "http_scheduler", "job_id": "http-dtc-close"},
                )
                self.assertEqual(status, 201)
                self.assertEqual(response["counts"]["dispatched"], 1)
                status, observability = self._json("/api/box/pipeline-observability")
                self.assertEqual(status, 200)
                self.assertTrue(observability["schedule_configured"])
                self.assertTrue(observability["ledger"]["integrity_valid"])
                with urlopen(
                    self.base_url + "/api/box/pipeline-observability?format=prometheus",
                    timeout=5,
                ) as metrics_response:
                    metrics = metrics_response.read().decode("utf-8")
                    self.assertIn("text/plain", metrics_response.headers["Content-Type"])
                self.assertIn("opc_finance_pipeline_ledger_integrity 1", metrics)
                self.assertNotIn(str(root), metrics)
                self.assertFalse(response["external_actions_performed"])
                status, response = self._json("/api/box/pipeline-schedule")
                self.assertEqual(status, 200)
                self.assertEqual(response["schedule"]["jobs"][0]["status"], "completed")

    def test_pipeline_api_rejects_unknown_contract(self):
        request = Request(
            self.base_url + "/api/box/pipelines/dispatch",
            data=json.dumps({"pipeline_id": "unknown", "payload": {}}).encode("utf-8"),
            headers={"Content-Type": "application/json"},
        )
        with self.assertRaises(HTTPError) as raised:
            urlopen(request, timeout=5)
        self.assertEqual(raised.exception.code, 400)
        body = json.loads(raised.exception.read().decode("utf-8"))
        self.assertEqual(body["type"], "invalid_pipeline_request")

    def test_inbox_api_returns_configured_legal_entities(self):
        status, payload = self._json("/api/inbox?limit=1")
        self.assertEqual(status, 200)
        self.assertEqual(
            {item["id"] for item in payload["entities"]},
            {"cn_studio", "sg_publisher"},
        )


if __name__ == "__main__":
    unittest.main()
